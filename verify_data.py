#!/usr/bin/env python3
"""
데이터 검증 스크립트
2025년 데이터의 공급가액 기준 총 매출과 검사목적별 매출을 계산합니다.
실서버에서 실행하여 UI에 표시된 값과 비교하세요.
"""

from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict

DATA_DIR = Path("data")

def verify_2025_data():
    """2025년 데이터 검증"""
    data_path = DATA_DIR / "2025"

    if not data_path.exists():
        print(f"ERROR: {data_path} 폴더가 없습니다.")
        return

    files = sorted(data_path.glob("*.xlsx"))
    if not files:
        print(f"ERROR: {data_path}에 xlsx 파일이 없습니다.")
        return

    print("=" * 60)
    print("2025년 데이터 검증 (공급가액 기준)")
    print("=" * 60)

    total_sales = 0
    total_count = 0
    by_purpose = defaultdict(lambda: {'sales': 0, 'count': 0})

    for f in files:
        print(f"\n파일 처리 중: {f.name}")
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # 컬럼 인덱스 찾기
            sales_idx = None
            purpose_idx = None
            for i, h in enumerate(headers):
                if h == '공급가액':
                    sales_idx = i
                if h == '검사목적':
                    purpose_idx = i

            if sales_idx is None:
                print(f"  WARNING: '공급가액' 컬럼을 찾을 수 없습니다.")
                print(f"  사용 가능한 컬럼: {headers}")
                wb.close()
                continue

            row_count = 0
            file_sales = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                sales = row[sales_idx] if sales_idx < len(row) else 0
                purpose = row[purpose_idx] if purpose_idx and purpose_idx < len(row) else ''

                # 공급가액 처리
                if sales is None:
                    sales = 0
                elif isinstance(sales, str):
                    sales = float(sales.replace(',', '').replace('원', '')) if sales.strip() else 0
                else:
                    sales = float(sales)

                # 검사목적 처리
                purpose = str(purpose or '').strip()
                if not purpose:
                    purpose = '미분류'

                total_sales += sales
                total_count += 1
                file_sales += sales
                row_count += 1

                by_purpose[purpose]['sales'] += sales
                by_purpose[purpose]['count'] += 1

            wb.close()
            print(f"  처리 건수: {row_count:,}건, 공급가액 합계: {file_sales/100000000:.2f}억")

        except Exception as e:
            print(f"  ERROR: {e}")

    # 결과 출력
    print("\n" + "=" * 60)
    print("검증 결과 요약")
    print("=" * 60)
    print(f"\n총 건수: {total_count:,}건")
    print(f"총 매출 (공급가액): {total_sales:,.0f}원")
    print(f"총 매출 (억): {total_sales/100000000:.2f}억")

    if total_count > 0:
        avg_price = total_sales / total_count
        print(f"평균 단가: {avg_price:,.0f}원 ({avg_price/10000:.1f}만)")

    print("\n" + "-" * 60)
    print("검사목적별 매출 (공급가액 기준)")
    print("-" * 60)

    # 매출 기준 정렬
    sorted_purposes = sorted(by_purpose.items(), key=lambda x: x[1]['sales'], reverse=True)

    for purpose, data in sorted_purposes:
        sales = data['sales']
        count = data['count']
        avg = sales / count if count > 0 else 0
        print(f"  {purpose:15s}: {count:>7,}건, {sales/100000000:>8.2f}억, 평균 {avg/10000:>6.1f}만")

    print("\n" + "-" * 60)
    print("검사목적별 합계 확인")
    print("-" * 60)
    purpose_total_sales = sum(d['sales'] for d in by_purpose.values())
    purpose_total_count = sum(d['count'] for d in by_purpose.values())
    print(f"목적별 합계: {purpose_total_count:,}건, {purpose_total_sales/100000000:.2f}억")
    print(f"전체 합계:   {total_count:,}건, {total_sales/100000000:.2f}억")

    if abs(purpose_total_sales - total_sales) < 1:
        print("✓ 검증 성공: 합계 일치")
    else:
        print("✗ 검증 실패: 합계 불일치")

if __name__ == "__main__":
    verify_2025_data()
