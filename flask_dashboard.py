"""
경영지표 대시보드 (Flask 버전)
- 오래된 CPU에서도 작동
- Chart.js 사용
- 연도 비교, 검사목적 필터, 업체별 분석, 부적합항목 분석
- AI 분석 (Google Gemini API)
"""
from flask import Flask, render_template_string, jsonify, request
import os
from pathlib import Path
from datetime import datetime
import json
import subprocess
import secrets
import hashlib

app = Flask(__name__)

# 터미널 인증 설정
TERMINAL_PASSWORD = "biofl2024"  # 터미널 접속 비밀번호
terminal_sessions = {}  # 세션 토큰 저장

# Gemini API 설정 (여러 키로 429 에러 대응)
GEMINI_API_KEYS = [
    os.environ.get('GEMINI_API_KEY', ''),
    os.environ.get('GEMINI_API_KEY_2', 'AIzaSyA7saUcePkpMh3olwkKKG7z-u1XXcDc7u4'),  # 경영지표1
    os.environ.get('GEMINI_API_KEY_3', 'AIzaSyCo8k3H7Pi128OuBgcupa7jlcm-hH1q68g'),  # 경영지표2
]
GEMINI_API_KEYS = [k for k in GEMINI_API_KEYS if k]  # 빈 키 제거
current_api_key_index = 0  # 현재 사용 중인 키 인덱스

# Claude API 설정
CLAUDE_API_KEY = os.environ.get('CLAUDE_API_KEY', '')  # 환경변수에서 로드
CLAUDE_MODEL = "claude-opus-4-20250514"  # Opus 4 - 최고 성능 모델
USE_CLAUDE = bool(CLAUDE_API_KEY)  # API 키가 있으면 Claude 사용

# 경로 설정 - 절대 경로 사용
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"  # 상대 경로로 변경
CACHE_FILE = BASE_DIR / "data_cache.pkl"  # 파일 캐시 경로
SQLITE_DB = DATA_DIR / "business_data.db"  # SQLite 데이터베이스 경로

# 데이터 캐시 (메모리에 저장)
DATA_CACHE = {}
CACHE_TIME = {}
FILE_MTIME = {}  # 파일 수정 시간 추적
AI_SUMMARY_CACHE = {}  # AI용 데이터 요약 캐시
USE_SQLITE = True  # SQLite 사용 여부


def init_sqlite_db():
    """SQLite 데이터베이스 초기화"""
    import sqlite3

    conn = sqlite3.connect(str(SQLITE_DB))
    cursor = conn.cursor()

    # 기본 데이터 테이블 (연도별 Excel 데이터)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS excel_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT,
            접수번호 TEXT,
            접수일자 TEXT,
            발행일 TEXT,
            검체유형 TEXT,
            업체명 TEXT,
            의뢰인명 TEXT,
            업체주소 TEXT,
            영업담당 TEXT,
            검사목적 TEXT,
            총금액 REAL,
            raw_data TEXT
        )
    ''')

    # food_item 데이터 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS food_item_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year TEXT,
            접수일자 TEXT,
            발행일 TEXT,
            검체유형 TEXT,
            업체명 TEXT,
            의뢰인명 TEXT,
            업체주소 TEXT,
            항목명 TEXT,
            규격 TEXT,
            항목담당 TEXT,
            결과입력자 TEXT,
            입력일 TEXT,
            분석일 TEXT,
            항목단위 TEXT,
            시험결과 TEXT,
            시험치 TEXT,
            성적서결과 TEXT,
            판정 TEXT,
            검사목적 TEXT,
            긴급여부 TEXT,
            항목수수료 REAL,
            영업담당 TEXT
        )
    ''')

    # 메타데이터 테이블 (파일 수정 시간 추적)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS file_metadata (
            file_path TEXT PRIMARY KEY,
            mtime REAL,
            row_count INTEGER
        )
    ''')

    # 토큰 사용량 추적 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS token_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            year_month TEXT,
            model TEXT,
            input_tokens INTEGER,
            output_tokens INTEGER,
            total_tokens INTEGER,
            cost_usd REAL,
            cost_krw REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 인덱스 생성 (빠른 검색용)
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_excel_year ON excel_data(year)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_excel_manager ON excel_data(영업담당)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_excel_purpose ON excel_data(검사목적)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_food_year ON food_item_data(year)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_food_manager ON food_item_data(영업담당)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_food_purpose ON food_item_data(검사목적)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_food_item ON food_item_data(항목명)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_token_yearmonth ON token_usage(year_month)')

    conn.commit()
    conn.close()
    print("[SQLITE] 데이터베이스 초기화 완료")


# 토큰 비용 설정 (USD per 1M tokens)
TOKEN_COSTS = {
    'gemini-2.0-flash': {'input': 0.075, 'output': 0.30},  # Gemini 2.0 Flash
    'claude-3-haiku': {'input': 0.80, 'output': 4.00},
    'claude-3-sonnet': {'input': 3.00, 'output': 15.00},
    'claude-3-opus': {'input': 15.00, 'output': 75.00},
    'claude-sonnet-4-20250514': {'input': 3.00, 'output': 15.00},  # Claude Sonnet 4
    'claude-opus-4-20250514': {'input': 15.00, 'output': 75.00},  # Claude Opus 4
    'claude-3-5-haiku-20241022': {'input': 0.80, 'output': 4.00},  # Claude 3.5 Haiku
}
USD_TO_KRW = 1450  # 환율


def call_claude_api(prompt, system_prompt=None, max_tokens=1024):
    """Claude API 호출 함수"""
    import urllib.request
    import json

    url = "https://api.anthropic.com/v1/messages"

    messages = [{"role": "user", "content": prompt}]

    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": max_tokens,
        "messages": messages
    }

    if system_prompt:
        payload["system"] = system_prompt

    headers = {
        "x-api-key": CLAUDE_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }

    try:
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode('utf-8'),
            headers=headers,
            method='POST'
        )

        with urllib.request.urlopen(req, timeout=60) as response:
            result = json.loads(response.read().decode('utf-8'))

        # 토큰 사용량 기록
        usage = result.get('usage', {})
        input_tokens = usage.get('input_tokens', 0)
        output_tokens = usage.get('output_tokens', 0)
        record_token_usage(CLAUDE_MODEL, input_tokens, output_tokens)
        print(f"[Claude] 토큰 사용: 입력={input_tokens}, 출력={output_tokens}")

        # 응답 텍스트 추출
        content = result.get('content', [])
        if content and len(content) > 0:
            return {'success': True, 'text': content[0].get('text', ''), 'usage': usage}
        else:
            return {'success': False, 'error': '응답 없음'}

    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8') if e.fp else str(e)
        print(f"[Claude] HTTP 오류: {e.code} - {error_body}")
        return {'success': False, 'error': f'API 오류 {e.code}: {error_body}'}
    except Exception as e:
        print(f"[Claude] 오류: {e}")
        return {'success': False, 'error': str(e)}


def record_token_usage(model, input_tokens, output_tokens):
    """토큰 사용량 기록"""
    import sqlite3
    from datetime import datetime

    total_tokens = input_tokens + output_tokens

    # 비용 계산
    cost_info = TOKEN_COSTS.get(model, {'input': 0.075, 'output': 0.30})
    cost_usd = (input_tokens * cost_info['input'] / 1_000_000) + (output_tokens * cost_info['output'] / 1_000_000)
    cost_krw = cost_usd * USD_TO_KRW

    today = datetime.now().strftime('%Y-%m-%d')
    year_month = datetime.now().strftime('%Y-%m')

    try:
        conn = sqlite3.connect(str(SQLITE_DB))
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO token_usage (date, year_month, model, input_tokens, output_tokens, total_tokens, cost_usd, cost_krw)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (today, year_month, model, input_tokens, output_tokens, total_tokens, cost_usd, cost_krw))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[TOKEN] 사용량 기록 오류: {e}")


def get_token_usage_stats():
    """토큰 사용량 통계 조회"""
    import sqlite3
    from datetime import datetime

    current_month = datetime.now().strftime('%Y-%m')
    # 저번달 계산
    now = datetime.now()
    if now.month == 1:
        last_month = f"{now.year - 1}-12"
    else:
        last_month = f"{now.year}-{now.month - 1:02d}"

    try:
        conn = sqlite3.connect(str(SQLITE_DB))
        cursor = conn.cursor()

        # 이번달 통계
        cursor.execute('''
            SELECT COALESCE(SUM(total_tokens), 0), COALESCE(SUM(cost_usd), 0), COALESCE(SUM(cost_krw), 0)
            FROM token_usage WHERE year_month = ?
        ''', (current_month,))
        this_month = cursor.fetchone()

        # 저번달 통계
        cursor.execute('''
            SELECT COALESCE(SUM(total_tokens), 0), COALESCE(SUM(cost_usd), 0), COALESCE(SUM(cost_krw), 0)
            FROM token_usage WHERE year_month = ?
        ''', (last_month,))
        prev_month = cursor.fetchone()

        conn.close()

        return {
            'this_month': {
                'tokens': int(this_month[0]),
                'cost_usd': round(this_month[1], 4),
                'cost_krw': round(this_month[2], 0)
            },
            'last_month': {
                'tokens': int(prev_month[0]),
                'cost_usd': round(prev_month[1], 4),
                'cost_krw': round(prev_month[2], 0)
            }
        }
    except Exception as e:
        print(f"[TOKEN] 통계 조회 오류: {e}")
        return {
            'this_month': {'tokens': 0, 'cost_usd': 0, 'cost_krw': 0},
            'last_month': {'tokens': 0, 'cost_usd': 0, 'cost_krw': 0}
        }


def check_sqlite_needs_update():
    """SQLite DB 업데이트 필요 여부 확인"""
    import sqlite3

    if not SQLITE_DB.exists():
        return True

    try:
        conn = sqlite3.connect(str(SQLITE_DB))
        cursor = conn.cursor()

        # 모든 Excel 파일의 현재 mtime 확인
        for year in ['2024', '2025']:
            # 기본 데이터
            data_path = DATA_DIR / str(year)
            if data_path.exists():
                for f in sorted(data_path.glob("*.xlsx")):
                    file_path = str(f)
                    current_mtime = f.stat().st_mtime

                    cursor.execute('SELECT mtime FROM file_metadata WHERE file_path = ?', (file_path,))
                    row = cursor.fetchone()

                    if not row or row[0] < current_mtime:
                        conn.close()
                        print(f"[SQLITE] 업데이트 필요: {f.name}")
                        return True

            # food_item 데이터
            food_path = DATA_DIR / "food_item" / str(year)
            if food_path.exists():
                for f in sorted(food_path.glob("*.xlsx")):
                    file_path = str(f)
                    current_mtime = f.stat().st_mtime

                    cursor.execute('SELECT mtime FROM file_metadata WHERE file_path = ?', (file_path,))
                    row = cursor.fetchone()

                    if not row or row[0] < current_mtime:
                        conn.close()
                        print(f"[SQLITE] 업데이트 필요: {f.name}")
                        return True

        conn.close()
        return False

    except Exception as e:
        print(f"[SQLITE] 체크 오류: {e}")
        return True


def convert_excel_to_sqlite():
    """Excel 파일을 SQLite로 변환"""
    import sqlite3
    import time
    from openpyxl import load_workbook

    print("[SQLITE] Excel → SQLite 변환 시작...")
    start_time = time.time()

    init_sqlite_db()
    conn = sqlite3.connect(str(SQLITE_DB))
    cursor = conn.cursor()

    total_records = 0

    for year in ['2024', '2025']:
        # 기본 데이터 변환
        data_path = DATA_DIR / str(year)
        if data_path.exists():
            for f in sorted(data_path.glob("*.xlsx")):
                file_path = str(f)
                current_mtime = f.stat().st_mtime

                # 이미 변환된 파일인지 확인
                cursor.execute('SELECT mtime FROM file_metadata WHERE file_path = ?', (file_path,))
                row = cursor.fetchone()
                if row and row[0] >= current_mtime:
                    print(f"[SQLITE] {f.name} 스킵 (이미 최신)")
                    continue

                # 기존 데이터 삭제 후 재삽입
                cursor.execute('DELETE FROM excel_data WHERE year = ? AND raw_data LIKE ?',
                              (year, f'%{f.name}%'))

                try:
                    wb = load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]

                    batch = []
                    for row_data in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row_data))
                        batch.append((
                            year,
                            str(row_dict.get('접수번호', '')),
                            str(row_dict.get('접수일자', '')),
                            str(row_dict.get('발행일', '')),
                            str(row_dict.get('검체유형', '')),
                            str(row_dict.get('업체명', '')),
                            str(row_dict.get('의뢰인명', '')),
                            str(row_dict.get('업체주소', '')),
                            str(row_dict.get('영업담당', '')),
                            str(row_dict.get('검사목적', '')),
                            float(row_dict.get('총금액', 0) or 0),
                            json.dumps(row_dict, ensure_ascii=False, default=str)
                        ))

                    cursor.executemany('''
                        INSERT INTO excel_data
                        (year, 접수번호, 접수일자, 발행일, 검체유형, 업체명, 의뢰인명, 업체주소, 영업담당, 검사목적, 총금액, raw_data)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', batch)

                    # 메타데이터 업데이트
                    cursor.execute('''
                        INSERT OR REPLACE INTO file_metadata (file_path, mtime, row_count)
                        VALUES (?, ?, ?)
                    ''', (file_path, current_mtime, len(batch)))

                    wb.close()
                    total_records += len(batch)
                    print(f"[SQLITE] {f.name}: {len(batch)}건 변환")

                except Exception as e:
                    print(f"[SQLITE ERROR] {f.name}: {e}")

        # food_item 데이터 변환
        food_path = DATA_DIR / "food_item" / str(year)
        if food_path.exists():
            for f in sorted(food_path.glob("*.xlsx")):
                file_path = str(f)
                current_mtime = f.stat().st_mtime

                cursor.execute('SELECT mtime FROM file_metadata WHERE file_path = ?', (file_path,))
                row = cursor.fetchone()
                if row and row[0] >= current_mtime:
                    print(f"[SQLITE] food_item {f.name} 스킵 (이미 최신)")
                    continue

                # 파일명 기반으로 삭제 (월별 데이터)
                month = f.stem.split('_')[-1] if '_' in f.stem else f.stem
                cursor.execute('DELETE FROM food_item_data WHERE year = ?', (year,))

                try:
                    wb = load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]

                    required_columns = ['접수일자', '발행일', '검체유형', '업체명', '의뢰인명', '업체주소',
                                       '항목명', '규격', '항목담당', '결과입력자', '입력일', '분석일',
                                       '항목단위', '시험결과', '시험치', '성적서결과', '판정', '검사목적',
                                       '긴급여부', '항목수수료', '영업담당']

                    col_indices = {}
                    for i, h in enumerate(headers):
                        if h in required_columns:
                            col_indices[h] = i

                    batch = []
                    for row_data in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for col_name, idx in col_indices.items():
                            row_dict[col_name] = row_data[idx] if idx < len(row_data) else None

                        batch.append((
                            year,
                            str(row_dict.get('접수일자', '') or ''),
                            str(row_dict.get('발행일', '') or ''),
                            str(row_dict.get('검체유형', '') or ''),
                            str(row_dict.get('업체명', '') or ''),
                            str(row_dict.get('의뢰인명', '') or ''),
                            str(row_dict.get('업체주소', '') or ''),
                            str(row_dict.get('항목명', '') or ''),
                            str(row_dict.get('규격', '') or ''),
                            str(row_dict.get('항목담당', '') or ''),
                            str(row_dict.get('결과입력자', '') or ''),
                            str(row_dict.get('입력일', '') or ''),
                            str(row_dict.get('분석일', '') or ''),
                            str(row_dict.get('항목단위', '') or ''),
                            str(row_dict.get('시험결과', '') or ''),
                            str(row_dict.get('시험치', '') or ''),
                            str(row_dict.get('성적서결과', '') or ''),
                            str(row_dict.get('판정', '') or ''),
                            str(row_dict.get('검사목적', '') or ''),
                            str(row_dict.get('긴급여부', '') or ''),
                            float(row_dict.get('항목수수료', 0) or 0),
                            str(row_dict.get('영업담당', '') or '')
                        ))

                    cursor.executemany('''
                        INSERT INTO food_item_data
                        (year, 접수일자, 발행일, 검체유형, 업체명, 의뢰인명, 업체주소, 항목명, 규격,
                         항목담당, 결과입력자, 입력일, 분석일, 항목단위, 시험결과, 시험치, 성적서결과,
                         판정, 검사목적, 긴급여부, 항목수수료, 영업담당)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', batch)

                    cursor.execute('''
                        INSERT OR REPLACE INTO file_metadata (file_path, mtime, row_count)
                        VALUES (?, ?, ?)
                    ''', (file_path, current_mtime, len(batch)))

                    wb.close()
                    total_records += len(batch)
                    print(f"[SQLITE] food_item {f.name}: {len(batch)}건 변환")

                except Exception as e:
                    print(f"[SQLITE ERROR] food_item {f.name}: {e}")

    conn.commit()
    conn.close()

    elapsed = time.time() - start_time
    print(f"[SQLITE] 변환 완료! 총 {total_records:,}건, {elapsed:.1f}초 소요")


def load_excel_data_sqlite(year):
    """SQLite에서 데이터 로드 (빠름)"""
    import sqlite3
    import time

    start_time = time.time()

    conn = sqlite3.connect(str(SQLITE_DB))
    cursor = conn.cursor()

    cursor.execute('SELECT raw_data FROM excel_data WHERE year = ?', (str(year),))
    rows = cursor.fetchall()

    data = []
    for row in rows:
        try:
            data.append(json.loads(row[0]))
        except:
            pass

    conn.close()

    elapsed = time.time() - start_time
    print(f"[SQLITE] {year}년 데이터 로드: {len(data):,}건, {elapsed:.2f}초")

    return data


def load_food_item_data_sqlite(year):
    """SQLite에서 food_item 데이터 로드 (빠름)"""
    import sqlite3
    import time

    start_time = time.time()

    conn = sqlite3.connect(str(SQLITE_DB))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('''
        SELECT 접수일자, 발행일, 검체유형, 업체명, 의뢰인명, 업체주소, 항목명, 규격,
               항목담당, 결과입력자, 입력일, 분석일, 항목단위, 시험결과, 시험치,
               성적서결과, 판정, 검사목적, 긴급여부, 항목수수료, 영업담당
        FROM food_item_data WHERE year = ?
    ''', (str(year),))

    rows = cursor.fetchall()
    data = [dict(row) for row in rows]

    conn.close()

    elapsed = time.time() - start_time
    print(f"[SQLITE] food_item {year}년 데이터 로드: {len(data):,}건, {elapsed:.2f}초")

    return data


def get_data_files_mtime():
    """모든 데이터 파일의 최신 수정 시간 반환"""
    latest_mtime = 0
    for year in ['2024', '2025']:
        data_path = DATA_DIR / str(year)
        if data_path.exists():
            for f in data_path.glob("*.xlsx"):
                mtime = f.stat().st_mtime
                if mtime > latest_mtime:
                    latest_mtime = mtime
        food_path = DATA_DIR / "food_item" / str(year)
        if food_path.exists():
            for f in food_path.glob("*.xlsx"):
                mtime = f.stat().st_mtime
                if mtime > latest_mtime:
                    latest_mtime = mtime
    return latest_mtime


def load_cache_from_file():
    """파일에서 캐시 로드 (서버 시작 시)"""
    global DATA_CACHE, CACHE_TIME, FILE_MTIME, AI_SUMMARY_CACHE
    import pickle

    if not CACHE_FILE.exists():
        print("[CACHE] 캐시 파일 없음 - 새로 생성 필요")
        return False

    try:
        # 데이터 파일 수정 시간 확인
        current_mtime = get_data_files_mtime()
        cache_mtime = CACHE_FILE.stat().st_mtime

        # 캐시가 데이터보다 오래된 경우 무효화
        if current_mtime > cache_mtime:
            print(f"[CACHE] 데이터 파일이 캐시보다 최신 - 다시 로드 필요")
            return False

        with open(CACHE_FILE, 'rb') as f:
            cached = pickle.load(f)

        DATA_CACHE = cached.get('DATA_CACHE', {})
        CACHE_TIME = cached.get('CACHE_TIME', {})
        FILE_MTIME = cached.get('FILE_MTIME', {})
        AI_SUMMARY_CACHE = cached.get('AI_SUMMARY_CACHE', {})

        # 캐시 시간 업데이트 (현재 시간 기준으로)
        import time
        current_time = time.time()
        for key in CACHE_TIME:
            CACHE_TIME[key] = current_time

        total_records = sum(len(v) for v in DATA_CACHE.values() if isinstance(v, list))
        print(f"[CACHE] 파일에서 캐시 로드 완료 ({total_records:,}건)")
        return True

    except Exception as e:
        print(f"[CACHE] 파일 캐시 로드 실패: {e}")
        return False


def save_cache_to_file():
    """캐시를 파일로 저장"""
    import pickle

    try:
        cached = {
            'DATA_CACHE': DATA_CACHE,
            'CACHE_TIME': CACHE_TIME,
            'FILE_MTIME': FILE_MTIME,
            'AI_SUMMARY_CACHE': AI_SUMMARY_CACHE
        }
        with open(CACHE_FILE, 'wb') as f:
            pickle.dump(cached, f)
        print(f"[CACHE] 파일로 캐시 저장 완료")
    except Exception as e:
        print(f"[CACHE] 파일 캐시 저장 실패: {e}")

# 설정
MANAGER_TO_BRANCH = {
    # 본사/마케팅
    "본사접수": "본사",
    "마케팅": "마케팅",
    # 서울센터
    "조봉현": "서울센터", "오석현": "서울센터", "오세중": "서울센터", "장동주": "서울센터",
    # 경북센터
    "엄상흠": "경북센터",
    # 충청지사
    "장동욱": "충청지사", "박은태": "충청지사", "지병훈": "충청지사",
    # 전라지사
    "이강현": "전라지사",
    # 기타지사
    "엄은정": "기타지사", "정유경": "기타지사", "심태보": "기타지사", "이성복": "기타지사", "도준구": "기타지사", "ISA": "기타지사",
}

# 부서별 매핑 (메인 대시보드 부서별 카드용)
MANAGER_TO_DEPARTMENT = {
    "본사접수": "본사",
    "마케팅": "마케팅",
    # 직영(영업부)
    "오세중": "영업부", "장동주": "영업부", "조봉현": "영업부", "오석현": "영업부", "엄상흠": "영업부",
    # 지사
    "장동욱": "지사", "박은태": "지사", "지병훈": "지사",
    "엄은정": "지사", "정유경": "지사",
    "이강현": "지사", "도준구": "지사", "이성복": "지사",
    "ISA": "지사",
}

# 지사에 포함될 담당자 목록 (지사 카드용)
BRANCH_MEMBERS = {"장동욱", "박은태", "지병훈", "엄은정", "정유경", "이강현", "도준구", "이성복", "ISA"}

# 개인별 분석에서 제외할 영업담당 (외부 기관 등)
EXCLUDED_MANAGERS = {"IBK", "미지정"}

def load_excel_data(year, use_cache=True):
    """데이터 로드 (SQLite 우선, 없으면 Excel)"""
    import time

    # 캐시 확인 (1시간 유효)
    cache_key = str(year)
    if use_cache and cache_key in DATA_CACHE:
        cache_age = time.time() - CACHE_TIME.get(cache_key, 0)
        if cache_age < 3600:  # 1시간
            print(f"[CACHE] {year}년 데이터 캐시 사용 ({len(DATA_CACHE[cache_key])}건)")
            return DATA_CACHE[cache_key]

    # SQLite 사용 (DB가 존재하면)
    if USE_SQLITE and SQLITE_DB.exists():
        all_data = load_excel_data_sqlite(year)
        DATA_CACHE[cache_key] = all_data
        CACHE_TIME[cache_key] = time.time()
        return all_data

    # 기존 Excel 로드 방식 (폴백)
    from openpyxl import load_workbook

    data_path = DATA_DIR / str(year)
    if not data_path.exists():
        return []

    print(f"[LOAD] {year}년 데이터 로딩 시작 (Excel)...")
    start_time = time.time()

    all_data = []
    files = sorted(data_path.glob("*.xlsx"))

    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                all_data.append(row_dict)
            wb.close()
            print(f"[LOAD] {f.name} 완료")
        except Exception as e:
            print(f"[ERROR] Loading {f}: {e}")

    elapsed = time.time() - start_time
    print(f"[LOAD] {year}년 완료: {len(all_data)}건, {elapsed:.1f}초 소요")

    # 캐시 저장
    DATA_CACHE[cache_key] = all_data
    CACHE_TIME[cache_key] = time.time()

    return all_data

def load_food_item_data(year, use_cache=True):
    """food_item 데이터 로드 (SQLite 우선, 없으면 Excel)"""
    import time

    cache_key = f"food_item_{year}"
    if use_cache and cache_key in DATA_CACHE:
        cache_age = time.time() - CACHE_TIME.get(cache_key, 0)
        if cache_age < 3600:
            print(f"[CACHE] food_item {year}년 데이터 캐시 사용 ({len(DATA_CACHE[cache_key])}건)")
            return DATA_CACHE[cache_key]

    # SQLite 사용 (DB가 존재하면)
    if USE_SQLITE and SQLITE_DB.exists():
        all_data = load_food_item_data_sqlite(year)
        DATA_CACHE[cache_key] = all_data
        CACHE_TIME[cache_key] = time.time()
        return all_data

    # 기존 Excel 로드 방식 (폴백)
    from openpyxl import load_workbook

    data_path = DATA_DIR / "food_item" / str(year)
    if not data_path.exists():
        print(f"[WARN] food_item {year}년 폴더 없음: {data_path}")
        return []

    print(f"[LOAD] food_item {year}년 데이터 로딩 시작 (Excel)...")
    start_time = time.time()

    # 필요한 컬럼만 로드
    required_columns = ['접수일자', '발행일', '검체유형', '업체명', '의뢰인명', '업체주소',
                       '항목명', '규격', '항목담당', '결과입력자', '입력일', '분석일',
                       '항목단위', '시험결과', '시험치', '성적서결과', '판정', '검사목적',
                       '긴급여부', '항목수수료', '영업담당']

    all_data = []
    files = sorted(data_path.glob("*.xlsx"))

    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # 컬럼 인덱스 매핑
            col_indices = {}
            for i, h in enumerate(headers):
                if h in required_columns:
                    col_indices[h] = i

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_name, idx in col_indices.items():
                    row_dict[col_name] = row[idx] if idx < len(row) else None
                all_data.append(row_dict)
            wb.close()
            print(f"[LOAD] food_item {f.name} 완료")
        except Exception as e:
            print(f"[ERROR] Loading food_item {f}: {e}")

    elapsed = time.time() - start_time
    print(f"[LOAD] food_item {year}년 완료: {len(all_data)}건, {elapsed:.1f}초 소요")

    DATA_CACHE[cache_key] = all_data
    CACHE_TIME[cache_key] = time.time()

    return all_data


def check_data_changed(year):
    """데이터 파일 변경 감지"""
    data_path = DATA_DIR / str(year)
    if not data_path.exists():
        return False

    files = sorted(data_path.glob("*.xlsx"))
    current_mtimes = {}

    for f in files:
        current_mtimes[str(f)] = f.stat().st_mtime

    cache_key = f"mtime_{year}"
    old_mtimes = FILE_MTIME.get(cache_key, {})

    if current_mtimes != old_mtimes:
        FILE_MTIME[cache_key] = current_mtimes
        return True

    return False


def get_ai_data_summary(force_refresh=False):
    """AI 분석용 데이터 요약 생성 (캐시됨)"""
    import time

    cache_key = 'ai_summary'

    # 데이터 변경 확인
    data_changed = check_data_changed('2024') or check_data_changed('2025')

    # 캐시 유효성 확인 (1시간 또는 데이터 변경 시)
    if not force_refresh and cache_key in AI_SUMMARY_CACHE:
        cache_age = time.time() - AI_SUMMARY_CACHE.get('_time', 0)
        if cache_age < 3600 and not data_changed:
            print(f"[AI-CACHE] 요약 캐시 사용 (나이: {cache_age:.0f}초)")
            return AI_SUMMARY_CACHE[cache_key]

    print(f"[AI-CACHE] 데이터 요약 생성 중...")
    start_time = time.time()

    # 데이터 로드
    food_2024 = load_food_item_data('2024')
    food_2025 = load_food_item_data('2025')

    # 요약 통계 계산
    summary = {
        '2024': {'total_count': 0, 'total_fee': 0, 'by_purpose': {}, 'by_sample_type': {},
                 'by_manager': {}, 'by_item': {}, 'monthly': {}},
        '2025': {'total_count': 0, 'total_fee': 0, 'by_purpose': {}, 'by_sample_type': {},
                 'by_manager': {}, 'by_item': {}, 'monthly': {}},
        'filter_values': {'purposes': set(), 'sample_types': set(), 'items': set(), 'managers': set()}
    }

    for year, data in [('2024', food_2024), ('2025', food_2025)]:
        for row in data:
            purpose = str(row.get('검사목적', '') or '').strip()
            sample_type = str(row.get('검체유형', '') or '').strip()
            item_name = str(row.get('항목명', '') or '').strip()
            manager = str(row.get('영업담당', '') or '').strip() or '미지정'
            fee = row.get('항목수수료', 0) or 0
            date = row.get('접수일자')

            if isinstance(fee, str):
                fee = float(fee.replace(',', '').replace('원', '')) if fee else 0

            summary[year]['total_count'] += 1
            summary[year]['total_fee'] += fee

            # 목적별
            if purpose:
                if purpose not in summary[year]['by_purpose']:
                    summary[year]['by_purpose'][purpose] = {'count': 0, 'fee': 0}
                summary[year]['by_purpose'][purpose]['count'] += 1
                summary[year]['by_purpose'][purpose]['fee'] += fee
                summary['filter_values']['purposes'].add(purpose)

            # 검체유형별
            if sample_type:
                if sample_type not in summary[year]['by_sample_type']:
                    summary[year]['by_sample_type'][sample_type] = {'count': 0, 'fee': 0}
                summary[year]['by_sample_type'][sample_type]['count'] += 1
                summary[year]['by_sample_type'][sample_type]['fee'] += fee
                summary['filter_values']['sample_types'].add(sample_type)

            # 영업담당별
            if manager not in summary[year]['by_manager']:
                summary[year]['by_manager'][manager] = {'count': 0, 'fee': 0}
            summary[year]['by_manager'][manager]['count'] += 1
            summary[year]['by_manager'][manager]['fee'] += fee
            summary['filter_values']['managers'].add(manager)

            # 항목별 (TOP 50만)
            if item_name:
                if item_name not in summary[year]['by_item']:
                    summary[year]['by_item'][item_name] = {'count': 0, 'fee': 0}
                summary[year]['by_item'][item_name]['count'] += 1
                summary[year]['by_item'][item_name]['fee'] += fee
                summary['filter_values']['items'].add(item_name)

            # 월별
            if date and hasattr(date, 'month'):
                m = date.month
                if m not in summary[year]['monthly']:
                    summary[year]['monthly'][m] = {'count': 0, 'fee': 0}
                summary[year]['monthly'][m]['count'] += 1
                summary[year]['monthly'][m]['fee'] += fee

    # set을 sorted list로 변환
    summary['filter_values']['purposes'] = sorted(summary['filter_values']['purposes'])
    summary['filter_values']['sample_types'] = sorted(summary['filter_values']['sample_types'])
    summary['filter_values']['items'] = sorted(summary['filter_values']['items'])[:100]  # 상위 100개만
    # ISA, IBK 등 제외 대상은 필터 목록에서 제외
    summary['filter_values']['managers'] = sorted([m for m in summary['filter_values']['managers'] if m not in EXCLUDED_MANAGERS])

    # 항목별 데이터 정렬 (상위 50개만 유지)
    for year in ['2024', '2025']:
        sorted_items = sorted(summary[year]['by_item'].items(),
                             key=lambda x: x[1]['fee'], reverse=True)[:50]
        summary[year]['by_item'] = dict(sorted_items)

    elapsed = time.time() - start_time
    print(f"[AI-CACHE] 요약 생성 완료: {elapsed:.1f}초 소요")

    AI_SUMMARY_CACHE[cache_key] = summary
    AI_SUMMARY_CACHE['_time'] = time.time()

    return summary


def process_food_item_data(data, purpose_filter=None, sample_type_filter=None,
                           item_filter=None, manager_filter=None):
    """검사항목 데이터 처리"""
    by_item = {}  # 항목별 데이터
    by_item_month = {}  # 항목별-월별 데이터
    by_item_analyzer = {}  # 항목별-분석자 데이터
    by_sample_type_item = {}  # 검체유형별-항목 데이터
    by_manager_item = {}  # 영업담당별-항목 데이터
    by_manager_fee = {}  # 영업담당별-수수료 데이터
    by_month_fee = {}  # 월별-수수료 데이터
    by_purpose_sample_type = {}  # 검사목적별-검체유형 매핑
    by_purpose_sample_type_item = {}  # 검사목적+검체유형별-항목 매핑

    purposes = set()
    sample_types = set()
    items = set()
    managers = set()
    analyzers = set()

    total_fee = 0
    total_count = 0

    for row in data:
        purpose = str(row.get('검사목적', '') or '').strip()
        sample_type = str(row.get('검체유형', '') or '').strip()
        item_name = str(row.get('항목명', '') or '').strip()
        manager = str(row.get('영업담당', '') or '').strip() or '미지정'
        analyzer = str(row.get('결과입력자', '') or '').strip() or '미지정'
        fee = row.get('항목수수료', 0) or 0
        date = row.get('접수일자')

        if isinstance(fee, str):
            fee = float(fee.replace(',', '').replace('원', '')) if fee else 0

        # 목록 수집
        if purpose: purposes.add(purpose)
        if sample_type: sample_types.add(sample_type)
        if item_name: items.add(item_name)
        if manager and manager != '미지정': managers.add(manager)
        if analyzer and analyzer != '미지정': analyzers.add(analyzer)

        # 검사목적별-검체유형 매핑 수집
        if purpose and sample_type:
            if purpose not in by_purpose_sample_type:
                by_purpose_sample_type[purpose] = set()
            by_purpose_sample_type[purpose].add(sample_type)

        # 검사목적+검체유형별-항목 매핑 수집 (잔류농약, 항생물질 제외)
        if purpose and sample_type and item_name:
            if not (sample_type.startswith('잔류농약') or sample_type.startswith('항생물질')):
                key = f"{purpose}|{sample_type}"
                if key not in by_purpose_sample_type_item:
                    by_purpose_sample_type_item[key] = set()
                by_purpose_sample_type_item[key].add(item_name)

        # 필터 적용
        if purpose_filter and purpose_filter != '전체' and purpose != purpose_filter:
            continue
        # 검체유형 필터 (와일드카드 지원)
        if sample_type_filter and sample_type_filter != '전체':
            if sample_type_filter.endswith('*'):
                # 와일드카드 패턴: "잔류농약*" -> 잔류농약으로 시작하는 모든 유형 매칭
                prefix = sample_type_filter[:-1]  # '*' 제거
                if not sample_type.startswith(prefix):
                    continue
            elif sample_type != sample_type_filter:
                continue
        if item_filter and item_filter != '전체' and item_name != item_filter:
            continue
        if manager_filter and manager_filter != '전체' and manager != manager_filter:
            continue

        # 월 추출
        month = 0
        if date:
            if hasattr(date, 'month'):
                month = date.month
            else:
                try:
                    month = int(str(date).split('-')[1])
                except:
                    month = 0

        total_fee += fee
        total_count += 1

        # 항목별 집계
        if item_name:
            if item_name not in by_item:
                by_item[item_name] = {'count': 0, 'fee': 0}
            by_item[item_name]['count'] += 1
            by_item[item_name]['fee'] += fee

            # 항목별-월별
            if month > 0:
                if item_name not in by_item_month:
                    by_item_month[item_name] = {}
                if month not in by_item_month[item_name]:
                    by_item_month[item_name][month] = 0
                by_item_month[item_name][month] += 1

            # 항목별-분석자
            if item_name not in by_item_analyzer:
                by_item_analyzer[item_name] = {}
            if analyzer not in by_item_analyzer[item_name]:
                by_item_analyzer[item_name][analyzer] = {'count': 0, 'fee': 0}
            by_item_analyzer[item_name][analyzer]['count'] += 1
            by_item_analyzer[item_name][analyzer]['fee'] += fee

        # 검체유형별-항목
        if sample_type:
            if sample_type not in by_sample_type_item:
                by_sample_type_item[sample_type] = {}
            if item_name:
                if item_name not in by_sample_type_item[sample_type]:
                    by_sample_type_item[sample_type][item_name] = {'count': 0, 'fee': 0}
                by_sample_type_item[sample_type][item_name]['count'] += 1
                by_sample_type_item[sample_type][item_name]['fee'] += fee

        # 영업담당별 집계
        if manager not in by_manager_item:
            by_manager_item[manager] = {'count': 0, 'fee': 0, 'items': {}}
        by_manager_item[manager]['count'] += 1
        by_manager_item[manager]['fee'] += fee
        if item_name:
            if item_name not in by_manager_item[manager]['items']:
                by_manager_item[manager]['items'][item_name] = {'count': 0, 'fee': 0}
            by_manager_item[manager]['items'][item_name]['count'] += 1
            by_manager_item[manager]['items'][item_name]['fee'] += fee

        # 월별 수수료
        if month > 0:
            if month not in by_month_fee:
                by_month_fee[month] = {'count': 0, 'fee': 0}
            by_month_fee[month]['count'] += 1
            by_month_fee[month]['fee'] += fee

    # 결과 정리
    by_item_sorted = sorted(by_item.items(), key=lambda x: x[1]['count'], reverse=True)
    by_manager_sorted = sorted(by_manager_item.items(), key=lambda x: x[1]['fee'], reverse=True)

    return {
        'by_item': by_item_sorted,
        'by_item_month': {k: list(v.items()) for k, v in by_item_month.items()},
        'by_item_analyzer': {k: sorted(v.items(), key=lambda x: x[1]['count'], reverse=True)
                            for k, v in by_item_analyzer.items()},
        'by_sample_type_item': {k: sorted(v.items(), key=lambda x: x[1]['count'], reverse=True)
                               for k, v in by_sample_type_item.items()},
        'by_manager_item': by_manager_sorted,
        'by_month_fee': list(by_month_fee.items()),
        'purposes': sorted(purposes),
        'sample_types': sorted(sample_types),
        'items': sorted(items),
        'managers': sorted(managers),
        'analyzers': sorted(analyzers),
        'total_fee': total_fee,
        'total_count': total_count,
        'by_purpose_sample_type': {k: sorted(v) for k, v in by_purpose_sample_type.items()},
        'by_purpose_sample_type_item': {k: sorted(v) for k, v in by_purpose_sample_type_item.items()}
    }

def extract_region(address):
    """주소에서 시/도, 시/군/구 추출"""
    if not address:
        return None, None

    addr = str(address).strip()
    if not addr:
        return None, None

    # 시/도 추출
    sido = None
    sigungu = None

    # 광역시/특별시/도 패턴
    sido_patterns = [
        '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
        '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
    ]

    for pattern in sido_patterns:
        if pattern in addr:
            sido = pattern
            break

    # 시/군/구 추출 (첫 번째 시/군/구 단위)
    import re
    # 시, 군, 구 패턴 매칭
    match = re.search(r'([가-힣]+(?:시|군|구))', addr)
    if match:
        sigungu = match.group(1)
        # 시도명이 시군구에 포함되어 있으면 다음 매칭 찾기
        if sido and (sigungu == sido + '시' or sigungu == sido + '도'):
            matches = re.findall(r'([가-힣]+(?:시|군|구))', addr)
            if len(matches) > 1:
                sigungu = matches[1]

    return sido, sigungu

def process_data(data, purpose_filter=None):
    """데이터 처리"""
    by_manager = {}
    by_branch = {}
    by_month = {}
    by_client = {}
    by_purpose = {}
    by_defect = {}
    by_defect_month = {}
    by_defect_purpose = {}  # 부적합-검사목적별 데이터
    by_defect_purpose_month = {}  # 부적합-검사목적별-월별 데이터
    by_purpose_month = {}  # 목적별-월별 데이터
    by_region = {}  # 지역별 데이터
    by_region_manager = {}  # 지역-담당자별 데이터
    by_purpose_manager = {}  # 목적별-담당자 데이터
    by_purpose_region = {}  # 목적별-지역 데이터
    by_sample_type = {}  # 검체유형별 데이터
    by_sample_type_month = {}  # 검체유형별-월별 데이터
    by_sample_type_manager = {}  # 검체유형별-담당자 데이터
    by_sample_type_purpose = {}  # 검체유형별-목적 데이터
    by_urgent_month = {}  # 월별 긴급 데이터
    by_branch_month_clients = {}  # 지사별 월별 거래처 (중복 분석용)
    by_department = {}  # 부서별 데이터 (본사, 마케팅, 영업부, 지사)
    purposes = set()
    sample_types = set()  # 검체유형 목록
    total_sales = 0
    total_count = 0

    # 주소 컬럼 자동 감지
    address_columns = ['거래처 주소', '채품지주소', '채품장소', '주소', '시료주소', '업체주소', '거래처주소', '검체주소', '시료채취장소']

    for row in data:
        purpose = str(row.get('검사목적', '') or '').strip()
        purposes.add(purpose) if purpose else None

        # 검사목적 필터 적용
        if purpose_filter and purpose_filter != '전체' and purpose != purpose_filter:
            continue

        manager = row.get('영업담당', '미지정')
        sales = row.get('공급가액', 0) or 0
        date = row.get('접수일자')
        client = str(row.get('거래처', '') or '').strip() or '미지정'
        defect = str(row.get('부적합항목', '') or '').strip()
        sample_type = str(row.get('검체유형', '') or '').strip()
        urgent_raw = str(row.get('긴급여부', '') or '').strip()
        # '일반'이 아니고 값이 있으면 모두 긴급으로 처리
        is_urgent = urgent_raw and urgent_raw != '일반'
        if sample_type:
            sample_types.add(sample_type)

        if isinstance(sales, str):
            sales = float(sales.replace(',', '').replace('원', '')) if sales else 0

        # 매니저별
        if manager not in by_manager:
            by_manager[manager] = {'sales': 0, 'count': 0, 'clients': {}, 'urgent': 0, 'urgent_by_purpose': {}, 'by_purpose': {}}
        by_manager[manager]['sales'] += sales
        by_manager[manager]['count'] += 1
        # 검사목적별 매출/건수 추가
        if purpose:
            if purpose not in by_manager[manager]['by_purpose']:
                by_manager[manager]['by_purpose'][purpose] = {'sales': 0, 'count': 0}
            by_manager[manager]['by_purpose'][purpose]['sales'] += sales
            by_manager[manager]['by_purpose'][purpose]['count'] += 1
        if is_urgent:
            by_manager[manager]['urgent'] += 1
            # 검사목적별 긴급 건수 추가
            if purpose:
                if purpose not in by_manager[manager]['urgent_by_purpose']:
                    by_manager[manager]['urgent_by_purpose'][purpose] = 0
                by_manager[manager]['urgent_by_purpose'][purpose] += 1
        if client not in by_manager[manager]['clients']:
            by_manager[manager]['clients'][client] = {'sales': 0, 'count': 0}
        by_manager[manager]['clients'][client]['sales'] += sales
        by_manager[manager]['clients'][client]['count'] += 1

        # 지사별
        branch = MANAGER_TO_BRANCH.get(manager, '기타')
        if branch not in by_branch:
            by_branch[branch] = {'sales': 0, 'count': 0, 'managers': set(), 'by_purpose': {}}
        by_branch[branch]['sales'] += sales
        by_branch[branch]['count'] += 1
        by_branch[branch]['managers'].add(manager)

        # 팀별 검사목적별 데이터
        if purpose:
            if purpose not in by_branch[branch]['by_purpose']:
                by_branch[branch]['by_purpose'][purpose] = {'sales': 0, 'count': 0}
            by_branch[branch]['by_purpose'][purpose]['sales'] += sales
            by_branch[branch]['by_purpose'][purpose]['count'] += 1

        # 부서별 (본사, 마케팅, 영업부, 지사)
        department = MANAGER_TO_DEPARTMENT.get(manager, '기타')
        if department not in by_department:
            by_department[department] = {'sales': 0, 'count': 0}
        by_department[department]['sales'] += sales
        by_department[department]['count'] += 1

        # 월별
        month = 0
        if date:
            if hasattr(date, 'month'):
                month = date.month
            else:
                try:
                    month = int(str(date).split('-')[1])
                except:
                    month = 0

        if month > 0:
            if month not in by_month:
                by_month[month] = {'sales': 0, 'count': 0, 'byPurpose': {}, 'byManager': {}, 'byBranch': {}}
            by_month[month]['sales'] += sales
            by_month[month]['count'] += 1

            # 월별 검사목적별 데이터
            if purpose:
                if purpose not in by_month[month]['byPurpose']:
                    by_month[month]['byPurpose'][purpose] = {'sales': 0, 'count': 0}
                by_month[month]['byPurpose'][purpose]['sales'] += sales
                by_month[month]['byPurpose'][purpose]['count'] += 1

            # 월별 담당자별 데이터
            if manager not in by_month[month]['byManager']:
                by_month[month]['byManager'][manager] = {'sales': 0, 'count': 0, 'byPurpose': {}}
            by_month[month]['byManager'][manager]['sales'] += sales
            by_month[month]['byManager'][manager]['count'] += 1
            # 월별 담당자별 검사목적 데이터
            if purpose:
                if purpose not in by_month[month]['byManager'][manager]['byPurpose']:
                    by_month[month]['byManager'][manager]['byPurpose'][purpose] = {'sales': 0, 'count': 0}
                by_month[month]['byManager'][manager]['byPurpose'][purpose]['sales'] += sales
                by_month[month]['byManager'][manager]['byPurpose'][purpose]['count'] += 1

            # 월별 팀별 데이터
            if branch not in by_month[month]['byBranch']:
                by_month[month]['byBranch'][branch] = {'sales': 0, 'count': 0, 'byPurpose': {}}
            by_month[month]['byBranch'][branch]['sales'] += sales
            by_month[month]['byBranch'][branch]['count'] += 1
            # 월별 팀별 검사목적 데이터
            if purpose:
                if purpose not in by_month[month]['byBranch'][branch]['byPurpose']:
                    by_month[month]['byBranch'][branch]['byPurpose'][purpose] = {'sales': 0, 'count': 0}
                by_month[month]['byBranch'][branch]['byPurpose'][purpose]['sales'] += sales
                by_month[month]['byBranch'][branch]['byPurpose'][purpose]['count'] += 1

            # 월별 긴급 데이터
            if month not in by_urgent_month:
                by_urgent_month[month] = {'sales': 0, 'count': 0}
            if is_urgent:
                by_urgent_month[month]['sales'] += sales
                by_urgent_month[month]['count'] += 1

            # 지사별 월별 거래처 (중복 분석용)
            if branch not in by_branch_month_clients:
                by_branch_month_clients[branch] = {}
            if month not in by_branch_month_clients[branch]:
                by_branch_month_clients[branch][month] = set()
            if client and client != '미지정':
                by_branch_month_clients[branch][month].add(client)

        # 거래처별
        if client not in by_client:
            by_client[client] = {'sales': 0, 'count': 0, 'purposes': {}, 'managers': {}, 'months': set()}
        by_client[client]['sales'] += sales
        by_client[client]['count'] += 1
        # 거래처별 담당자 집계
        if manager and manager != '미지정':
            if manager not in by_client[client]['managers']:
                by_client[client]['managers'][manager] = {'sales': 0, 'count': 0}
            by_client[client]['managers'][manager]['sales'] += sales
            by_client[client]['managers'][manager]['count'] += 1
        # 거래처별 거래 월 추적
        if month > 0:
            by_client[client]['months'].add(month)
        if purpose:
            if purpose not in by_client[client]['purposes']:
                by_client[client]['purposes'][purpose] = {'sales': 0, 'count': 0}
            by_client[client]['purposes'][purpose]['sales'] += sales
            by_client[client]['purposes'][purpose]['count'] += 1

        # 검사목적별
        if purpose:
            if purpose not in by_purpose:
                by_purpose[purpose] = {'sales': 0, 'count': 0}
            by_purpose[purpose]['sales'] += sales
            by_purpose[purpose]['count'] += 1

            # 목적별-담당자 데이터
            if purpose not in by_purpose_manager:
                by_purpose_manager[purpose] = {}
            if manager not in by_purpose_manager[purpose]:
                by_purpose_manager[purpose][manager] = {'sales': 0, 'count': 0}
            by_purpose_manager[purpose][manager]['sales'] += sales
            by_purpose_manager[purpose][manager]['count'] += 1

            # 목적별-월별 데이터
            if month > 0:
                if purpose not in by_purpose_month:
                    by_purpose_month[purpose] = {}
                if month not in by_purpose_month[purpose]:
                    by_purpose_month[purpose][month] = {'sales': 0, 'count': 0, 'by_manager': {}}
                by_purpose_month[purpose][month]['sales'] += sales
                by_purpose_month[purpose][month]['count'] += 1
                # 담당자별 월별 목적 데이터
                if manager not in by_purpose_month[purpose][month]['by_manager']:
                    by_purpose_month[purpose][month]['by_manager'][manager] = {'sales': 0, 'count': 0}
                by_purpose_month[purpose][month]['by_manager'][manager]['sales'] += sales
                by_purpose_month[purpose][month]['by_manager'][manager]['count'] += 1

        # 부적합항목별
        if defect:
            if defect not in by_defect:
                by_defect[defect] = {'count': 0}
            by_defect[defect]['count'] += 1

            # 부적합항목 월별
            if month > 0:
                if defect not in by_defect_month:
                    by_defect_month[defect] = {}
                if month not in by_defect_month[defect]:
                    by_defect_month[defect][month] = 0
                by_defect_month[defect][month] += 1

            # 부적합항목-검사목적별
            if purpose:
                if purpose not in by_defect_purpose:
                    by_defect_purpose[purpose] = {}
                if defect not in by_defect_purpose[purpose]:
                    by_defect_purpose[purpose][defect] = {'count': 0}
                by_defect_purpose[purpose][defect]['count'] += 1

                # 부적합항목-검사목적별-월별
                if month > 0:
                    if purpose not in by_defect_purpose_month:
                        by_defect_purpose_month[purpose] = {}
                    if defect not in by_defect_purpose_month[purpose]:
                        by_defect_purpose_month[purpose][defect] = {}
                    if month not in by_defect_purpose_month[purpose][defect]:
                        by_defect_purpose_month[purpose][defect][month] = 0
                    by_defect_purpose_month[purpose][defect][month] += 1

        # 검체유형별
        if sample_type:
            if sample_type not in by_sample_type:
                by_sample_type[sample_type] = {'sales': 0, 'count': 0}
            by_sample_type[sample_type]['sales'] += sales
            by_sample_type[sample_type]['count'] += 1

            # 검체유형별-담당자 데이터
            if sample_type not in by_sample_type_manager:
                by_sample_type_manager[sample_type] = {}
            if manager not in by_sample_type_manager[sample_type]:
                by_sample_type_manager[sample_type][manager] = {'sales': 0, 'count': 0, 'by_purpose': {}}
            by_sample_type_manager[sample_type][manager]['sales'] += sales
            by_sample_type_manager[sample_type][manager]['count'] += 1
            # 담당자별 목적 데이터 추가
            if purpose:
                if purpose not in by_sample_type_manager[sample_type][manager]['by_purpose']:
                    by_sample_type_manager[sample_type][manager]['by_purpose'][purpose] = {'sales': 0, 'count': 0}
                by_sample_type_manager[sample_type][manager]['by_purpose'][purpose]['sales'] += sales
                by_sample_type_manager[sample_type][manager]['by_purpose'][purpose]['count'] += 1

            # 검체유형별-목적 데이터
            if purpose:
                if sample_type not in by_sample_type_purpose:
                    by_sample_type_purpose[sample_type] = {}
                if purpose not in by_sample_type_purpose[sample_type]:
                    by_sample_type_purpose[sample_type][purpose] = {'sales': 0, 'count': 0}
                by_sample_type_purpose[sample_type][purpose]['sales'] += sales
                by_sample_type_purpose[sample_type][purpose]['count'] += 1

            # 검체유형별-월별 데이터
            if month > 0:
                if sample_type not in by_sample_type_month:
                    by_sample_type_month[sample_type] = {}
                if month not in by_sample_type_month[sample_type]:
                    by_sample_type_month[sample_type][month] = {'sales': 0, 'count': 0, 'by_manager': {}, 'by_purpose': {}}
                by_sample_type_month[sample_type][month]['sales'] += sales
                by_sample_type_month[sample_type][month]['count'] += 1
                # 담당자별 월별 검체유형 데이터
                if manager not in by_sample_type_month[sample_type][month]['by_manager']:
                    by_sample_type_month[sample_type][month]['by_manager'][manager] = {'sales': 0, 'count': 0}
                by_sample_type_month[sample_type][month]['by_manager'][manager]['sales'] += sales
                by_sample_type_month[sample_type][month]['by_manager'][manager]['count'] += 1
                # 목적별 월별 검체유형 데이터
                if purpose:
                    if purpose not in by_sample_type_month[sample_type][month]['by_purpose']:
                        by_sample_type_month[sample_type][month]['by_purpose'][purpose] = {'sales': 0, 'count': 0}
                    by_sample_type_month[sample_type][month]['by_purpose'][purpose]['sales'] += sales
                    by_sample_type_month[sample_type][month]['by_purpose'][purpose]['count'] += 1

        # 지역별 분석
        address = None
        for col in address_columns:
            if row.get(col):
                address = row.get(col)
                break

        sido, sigungu = extract_region(address)

        if sido:
            region_key = sido
            if sigungu:
                region_key = f"{sido} {sigungu}"

            # 지역별 통계
            if region_key not in by_region:
                by_region[region_key] = {'sales': 0, 'count': 0, 'sido': sido, 'sigungu': sigungu or '', 'managers': {}}
            by_region[region_key]['sales'] += sales
            by_region[region_key]['count'] += 1

            # 지역-담당자별 통계
            if manager not in by_region[region_key]['managers']:
                by_region[region_key]['managers'][manager] = {'sales': 0, 'count': 0}
            by_region[region_key]['managers'][manager]['sales'] += sales
            by_region[region_key]['managers'][manager]['count'] += 1

            # 담당자-지역별 통계
            if manager not in by_region_manager:
                by_region_manager[manager] = {}
            if region_key not in by_region_manager[manager]:
                by_region_manager[manager][region_key] = {'sales': 0, 'count': 0, 'sido': sido, 'sigungu': sigungu or ''}
            by_region_manager[manager][region_key]['sales'] += sales
            by_region_manager[manager][region_key]['count'] += 1

            # 목적별-지역 데이터
            if purpose:
                if purpose not in by_purpose_region:
                    by_purpose_region[purpose] = {}
                if region_key not in by_purpose_region[purpose]:
                    by_purpose_region[purpose][region_key] = {'sales': 0, 'count': 0}
                by_purpose_region[purpose][region_key]['sales'] += sales
                by_purpose_region[purpose][region_key]['count'] += 1

        total_sales += sales
        total_count += 1

    # 정렬 (EXCLUDED_MANAGERS 제외)
    sorted_managers = sorted(
        [(m, d) for m, d in by_manager.items() if m not in EXCLUDED_MANAGERS],
        key=lambda x: x[1]['sales'], reverse=True
    )
    sorted_branches = sorted(by_branch.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_clients = sorted(by_client.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_purposes = sorted(by_purpose.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_defects = sorted(by_defect.items(), key=lambda x: x[1]['count'], reverse=True)

    # 매니저별 TOP 10 거래처
    manager_top_clients = {}
    for mgr, data in sorted_managers:
        clients = sorted(data['clients'].items(), key=lambda x: x[1]['sales'], reverse=True)[:10]
        manager_top_clients[mgr] = clients

    # 고효율 업체 (높은 단가)
    high_efficiency = [(c, d) for c, d in sorted_clients if d['count'] > 0]
    high_efficiency = sorted(high_efficiency, key=lambda x: x[1]['sales'] / x[1]['count'] if x[1]['count'] > 0 else 0, reverse=True)[:20]

    # 대량 업체 (많은 건수)
    high_volume = sorted(by_client.items(), key=lambda x: x[1]['count'], reverse=True)[:20]

    # 지역별 정렬 (매출 기준)
    sorted_regions = sorted(by_region.items(), key=lambda x: x[1]['sales'], reverse=True)

    # 지역별 TOP 담당자
    region_top_managers = {}
    for region, data in sorted_regions:
        managers = sorted(data['managers'].items(), key=lambda x: x[1]['sales'], reverse=True)
        region_top_managers[region] = [
            {'name': m, 'sales': d['sales'], 'count': d['count']}
            for m, d in managers[:5]
        ]

    # 담당자별 지역 분포
    manager_regions = {}
    for mgr, regions in by_region_manager.items():
        sorted_mgr_regions = sorted(regions.items(), key=lambda x: x[1]['sales'], reverse=True)
        manager_regions[mgr] = [
            {'region': r, 'sales': d['sales'], 'count': d['count'], 'sido': d['sido'], 'sigungu': d['sigungu']}
            for r, d in sorted_mgr_regions[:10]
        ]

    # 목적별 담당자 데이터 정리
    purpose_managers = {}
    for purpose, managers in by_purpose_manager.items():
        sorted_pm = sorted(managers.items(), key=lambda x: x[1]['sales'], reverse=True)
        purpose_managers[purpose] = [
            {'name': m, 'sales': d['sales'], 'count': d['count']}
            for m, d in sorted_pm[:20]
        ]

    # 목적별 지역 데이터 정리
    purpose_regions = {}
    for purpose, regions in by_purpose_region.items():
        sorted_pr = sorted(regions.items(), key=lambda x: x[1]['sales'], reverse=True)
        purpose_regions[purpose] = [
            {'region': r, 'sales': d['sales'], 'count': d['count']}
            for r, d in sorted_pr[:20]
        ]

    # 검체유형별 정렬
    sorted_sample_types = sorted(by_sample_type.items(), key=lambda x: x[1]['sales'], reverse=True)

    # 검체유형별 담당자 데이터 정리
    sample_type_managers = {}
    for st, managers in by_sample_type_manager.items():
        sorted_stm = sorted(managers.items(), key=lambda x: x[1]['sales'], reverse=True)
        sample_type_managers[st] = [
            {'name': m, 'sales': d['sales'], 'count': d['count'], 'by_purpose': d.get('by_purpose', {})}
            for m, d in sorted_stm[:20]
        ]

    # 검체유형별 목적 데이터 정리
    sample_type_purposes = {}
    for st, purposes_data in by_sample_type_purpose.items():
        sorted_stp = sorted(purposes_data.items(), key=lambda x: x[1]['sales'], reverse=True)
        sample_type_purposes[st] = [
            {'name': p, 'sales': d['sales'], 'count': d['count']}
            for p, d in sorted_stp[:20]
        ]

    # 지사별 월별 거래처 중복률 계산
    branch_client_retention = {}
    for branch, month_clients in by_branch_month_clients.items():
        months = sorted(month_clients.keys())
        retention_data = []
        all_clients = set()
        for month in months:
            clients = month_clients[month]
            # 이전 달과의 중복률 계산
            if all_clients:
                overlap = len(clients & all_clients)
                retention_rate = (overlap / len(all_clients) * 100) if all_clients else 0
            else:
                overlap = 0
                retention_rate = 0
            retention_data.append({
                'month': month,
                'total': len(clients),
                'overlap': overlap,
                'retention': round(retention_rate, 1),
                'new': len(clients - all_clients) if all_clients else len(clients)
            })
            all_clients.update(clients)
        branch_client_retention[branch] = retention_data

    # 전체 월별 거래처 중복률 (모든 지사 합산)
    all_month_clients = {}
    for branch, month_clients in by_branch_month_clients.items():
        for month, clients in month_clients.items():
            if month not in all_month_clients:
                all_month_clients[month] = set()
            all_month_clients[month].update(clients)

    total_retention = []
    cumulative_clients = set()
    for month in sorted(all_month_clients.keys()):
        clients = all_month_clients[month]
        if cumulative_clients:
            overlap = len(clients & cumulative_clients)
            retention_rate = (overlap / len(cumulative_clients) * 100) if cumulative_clients else 0
        else:
            overlap = 0
            retention_rate = 0
        total_retention.append({
            'month': month,
            'total': len(clients),
            'overlap': overlap,
            'retention': round(retention_rate, 1),
            'new': len(clients - cumulative_clients) if cumulative_clients else len(clients),
            'cumulative': len(cumulative_clients | clients)
        })
        cumulative_clients.update(clients)

    return {
        'by_manager': [(m, {'sales': d['sales'], 'count': d['count'], 'urgent': d.get('urgent', 0), 'urgent_by_purpose': d.get('urgent_by_purpose', {}), 'by_purpose': d.get('by_purpose', {})}) for m, d in sorted_managers],
        'by_branch': [(k, {'sales': v['sales'], 'count': v['count'], 'managers': len(v['managers']), 'by_purpose': v.get('by_purpose', {})})
                      for k, v in sorted_branches],
        'by_month': sorted(by_month.items()),
        'by_urgent_month': sorted(by_urgent_month.items()),
        'by_client': [(c, {
            'sales': d['sales'],
            'count': d['count'],
            'avg': d['sales']/d['count'] if d['count'] > 0 else 0,
            'manager': max(d.get('managers', {}).items(), key=lambda x: x[1]['sales'])[0] if d.get('managers') else '미지정',
            'purpose': max(d.get('purposes', {}).items(), key=lambda x: x[1]['sales'])[0] if d.get('purposes') else '',
            'tradeMonths': len(d.get('months', set())),
            'purposes': d.get('purposes', {})
        }) for c, d in sorted_clients[:100]],
        'by_purpose': sorted_purposes,
        'by_defect': sorted_defects[:30],
        'by_defect_month': {d: sorted(months.items()) for d, months in by_defect_month.items()},
        'by_defect_purpose': {p: sorted(defects.items(), key=lambda x: x[1]['count'], reverse=True)[:30] for p, defects in by_defect_purpose.items()},
        'by_defect_purpose_month': {p: {d: sorted(months.items()) for d, months in defects.items()} for p, defects in by_defect_purpose_month.items()},
        'by_purpose_month': {p: {m: {'sales': d['sales'], 'count': d['count'], 'by_manager': d.get('by_manager', {})} for m, d in months.items()} for p, months in by_purpose_month.items()},
        'manager_top_clients': manager_top_clients,
        'high_efficiency': [(c, {'sales': d['sales'], 'count': d['count'], 'avg': d['sales']/d['count'] if d['count'] > 0 else 0})
                           for c, d in high_efficiency],
        'high_volume': [(c, {'sales': d['sales'], 'count': d['count'], 'avg': d['sales']/d['count'] if d['count'] > 0 else 0})
                       for c, d in high_volume],
        'by_region': [(r, {'sales': d['sales'], 'count': d['count'], 'sido': d['sido'], 'sigungu': d['sigungu']})
                      for r, d in sorted_regions[:50]],
        'region_top_managers': region_top_managers,
        'manager_regions': manager_regions,
        'purpose_managers': purpose_managers,
        'purpose_regions': purpose_regions,
        'purposes': sorted(list(purposes)),
        'by_sample_type': sorted_sample_types,
        'by_sample_type_month': {st: {m: {'sales': d['sales'], 'count': d['count'], 'by_manager': d.get('by_manager', {}), 'by_purpose': d.get('by_purpose', {})} for m, d in months.items()} for st, months in by_sample_type_month.items()},
        'sample_type_managers': sample_type_managers,
        'sample_type_purposes': sample_type_purposes,
        'sample_types': sorted(list(sample_types)),
        'branch_client_retention': branch_client_retention,
        'total_client_retention': total_retention,
        'by_department': by_department,
        'total_sales': total_sales,
        'total_count': total_count
    }

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>경영지표 대시보드</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }

        :root {
            --primary: #6366f1;
            --primary-dark: #4f46e5;
            --primary-light: #e0e7ff;
            --success: #10b981;
            --success-light: #d1fae5;
            --danger: #ef4444;
            --danger-light: #fee2e2;
            --warning: #f59e0b;
            --warning-light: #fef3c7;
            --info: #06b6d4;
            --info-light: #cffafe;
            --purple: #8b5cf6;
            --purple-light: #ede9fe;
            --pink: #ec4899;
            --pink-light: #fce7f3;
            --orange: #f97316;
            --orange-light: #ffedd5;
            --teal: #14b8a6;
            --teal-light: #ccfbf1;
            --rose: #f43f5e;
            --rose-light: #ffe4e6;
            --sky: #0ea5e9;
            --sky-light: #e0f2fe;
            --lime: #84cc16;
            --lime-light: #ecfccb;
            --amber: #f59e0b;
            --amber-light: #fef3c7;
            --cyan: #06b6d4;
            --cyan-light: #cffafe;
            --gray-50: #f8fafc;
            --gray-100: #f1f5f9;
            --gray-200: #e2e8f0;
            --gray-300: #cbd5e1;
            --gray-400: #94a3b8;
            --gray-500: #64748b;
            --gray-600: #475569;
            --gray-700: #334155;
            --gray-800: #1e293b;
            --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
            --shadow: 0 1px 3px 0 rgb(0 0 0 / 0.1), 0 1px 2px -1px rgb(0 0 0 / 0.1);
            --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
        }

        body {
            font-family: 'Pretendard', -apple-system, BlinkMacSystemFont, 'Malgun Gothic', sans-serif;
            background: var(--gray-100);
            color: var(--gray-800);
            min-height: 100vh;
        }

        /* 헤더 */
        .header {
            background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
            padding: 16px 24px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            position: sticky;
            top: 0;
            z-index: 100;
            box-shadow: var(--shadow-md);
        }

        .header-left {
            display: flex;
            align-items: center;
            gap: 12px;
        }

        .logo {
            width: 40px;
            height: 40px;
            background: rgba(255,255,255,0.2);
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 20px;
            color: white;
        }

        .header-title {
            color: white;
            font-size: 22px;
            font-weight: 700;
        }

        .header-right {
            display: flex;
            align-items: center;
            gap: 12px;
        }

        .token-badge {
            background: rgba(255,255,255,0.15);
            padding: 8px 14px;
            border-radius: 10px;
            color: white;
            font-size: 12px;
            line-height: 1.4;
        }

        .token-badge .current { font-weight: 600; }
        .token-badge .prev { opacity: 0.7; font-size: 11px; }

        /* 메인 컨테이너 */
        .main-container {
            max-width: 1800px;
            margin: 0 auto;
            padding: 24px;
        }

        /* 필터 섹션 */
        .filter-section {
            background: white;
            border-radius: 16px;
            padding: 20px 24px;
            margin-bottom: 20px;
            box-shadow: var(--shadow);
        }

        .filter-row {
            display: flex;
            align-items: center;
            gap: 16px;
            flex-wrap: wrap;
        }

        .filter-group {
            display: flex;
            align-items: center;
            gap: 8px;
            background: var(--gray-50);
            padding: 8px 14px;
            border-radius: 10px;
            border: 1px solid var(--gray-200);
        }

        .filter-label {
            font-size: 13px;
            color: var(--gray-500);
            font-weight: 500;
        }

        .filter-select {
            padding: 8px 12px;
            border: 1px solid var(--gray-200);
            border-radius: 8px;
            font-size: 14px;
            color: var(--gray-700);
            background: white;
            cursor: pointer;
            min-width: 90px;
        }

        .filter-select:focus {
            outline: none;
            border-color: var(--primary);
            box-shadow: 0 0 0 3px var(--primary-light);
        }

        .filter-btn {
            padding: 6px 14px;
            border: 1px solid var(--gray-200);
            border-radius: 6px;
            background: white;
            font-size: 13px;
            color: var(--gray-600);
            cursor: pointer;
            transition: all 0.2s;
        }
        .filter-btn:hover {
            background: var(--gray-50);
            border-color: var(--gray-300);
        }
        .filter-btn.active {
            background: var(--primary);
            border-color: var(--primary);
            color: white;
        }

        .filter-divider {
            width: 1px;
            height: 32px;
            background: var(--gray-200);
        }

        .filter-checkbox {
            display: flex;
            align-items: center;
            gap: 6px;
            cursor: pointer;
            padding: 8px 12px;
            background: var(--gray-50);
            border-radius: 8px;
            border: 1px solid var(--gray-200);
        }

        .filter-checkbox input {
            width: 16px;
            height: 16px;
            accent-color: var(--primary);
        }

        .filter-checkbox span {
            font-size: 13px;
            color: var(--gray-600);
        }

        .btn-search {
            background: var(--primary);
            color: white;
            border: none;
            padding: 10px 24px;
            border-radius: 10px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 8px;
            transition: all 0.2s;
            margin-left: auto;
        }

        .btn-search:hover {
            background: var(--primary-dark);
            transform: translateY(-1px);
            box-shadow: var(--shadow-md);
        }

        .btn-search:disabled {
            opacity: 0.6;
            cursor: not-allowed;
            transform: none;
        }

        /* 탭 카드 그리드 */
        .tab-cards {
            display: flex;
            gap: 10px;
            margin-bottom: 24px;
            flex-wrap: wrap;
        }

        .tab-card {
            background: white;
            border-radius: 12px;
            padding: 14px 16px;
            text-align: center;
            cursor: pointer;
            transition: all 0.2s;
            border: 2px solid transparent;
            box-shadow: var(--shadow-sm);
            min-width: 90px;
            flex: 1;
            max-width: 120px;
        }

        .tab-card:hover {
            transform: translateY(-3px);
            box-shadow: var(--shadow-md);
            border-color: var(--gray-200);
        }

        .tab-card.active {
            border-color: var(--primary);
            background: var(--primary-light);
        }

        .tab-card.active .tab-icon {
            background: var(--primary);
            color: white;
        }

        .tab-card.active .tab-label {
            color: var(--primary-dark);
            font-weight: 600;
        }

        .tab-icon {
            width: 40px;
            height: 40px;
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
            margin: 0 auto 8px;
            background: var(--gray-100);
            color: var(--gray-500);
            transition: all 0.2s;
        }

        .tab-label {
            font-size: 12px;
            color: var(--gray-600);
            font-weight: 500;
            white-space: nowrap;
        }

        /* 특수 탭 스타일 */
        .tab-card.main-tab {
            background: linear-gradient(135deg, #0ea5e9, #06b6d4);
        }
        .tab-card.main-tab .tab-icon {
            background: rgba(255,255,255,0.2);
            color: white;
        }
        .tab-card.main-tab .tab-label {
            color: white;
        }
        .tab-card.main-tab.active {
            border-color: rgba(255,255,255,0.5);
        }

        .tab-card.ai-tab {
            background: linear-gradient(135deg, #6366f1, #8b5cf6);
        }
        .tab-card.ai-tab .tab-icon {
            background: rgba(255,255,255,0.2);
            color: white;
        }
        .tab-card.ai-tab .tab-label {
            color: white;
        }

        .tab-card.info-tab {
            background: linear-gradient(135deg, #10b981, #34d399);
        }
        .tab-card.info-tab .tab-icon {
            background: rgba(255,255,255,0.2);
            color: white;
        }
        .tab-card.info-tab .tab-label {
            color: white;
        }

        .tab-card.terminal-tab {
            background: linear-gradient(135deg, #1e293b, #334155);
        }
        .tab-card.terminal-tab .tab-icon {
            background: rgba(16, 185, 129, 0.2);
            color: #10b981;
        }
        .tab-card.terminal-tab .tab-label {
            color: #10b981;
        }

        /* KPI 카드 섹션 */
        .kpi-section {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 20px;
            margin-bottom: 24px;
        }

        .kpi-card {
            background: white;
            border-radius: 16px;
            padding: 24px;
            box-shadow: var(--shadow);
            position: relative;
            overflow: hidden;
        }

        .kpi-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
        }

        .kpi-card.sales::before { background: linear-gradient(90deg, #6366f1, #8b5cf6); }
        .kpi-card.count::before { background: linear-gradient(90deg, #10b981, #34d399); }
        .kpi-card.price::before { background: linear-gradient(90deg, #f59e0b, #fbbf24); }
        .kpi-card.goal::before { background: linear-gradient(90deg, #ec4899, #f472b6); }

        .kpi-header {
            display: flex;
            align-items: flex-start;
            justify-content: space-between;
            margin-bottom: 12px;
        }

        .kpi-icon {
            width: 48px;
            height: 48px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 20px;
        }

        .kpi-card.sales .kpi-icon { background: var(--primary-light); color: var(--primary); }
        .kpi-card.count .kpi-icon { background: var(--success-light); color: var(--success); }
        .kpi-card.price .kpi-icon { background: var(--warning-light); color: var(--warning); }
        .kpi-card.goal .kpi-icon { background: var(--pink-light); color: var(--pink); }

        .kpi-trend {
            display: flex;
            align-items: center;
            gap: 4px;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }

        .kpi-trend.up { background: var(--success-light); color: var(--success); }
        .kpi-trend.down { background: var(--danger-light); color: var(--danger); }

        .kpi-label {
            font-size: 14px;
            color: var(--gray-500);
            margin-bottom: 6px;
        }

        .kpi-value {
            font-size: 28px;
            font-weight: 700;
            color: var(--gray-800);
        }

        .kpi-compare {
            font-size: 13px;
            color: var(--gray-400);
            margin-top: 8px;
            padding-top: 8px;
            border-top: 1px dashed var(--gray-200);
        }

        .kpi-compare span {
            color: var(--gray-600);
            font-weight: 500;
        }

        /* KPI 카드 호버 오버레이 */
        .kpi-overlay {
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(2px);
            border-radius: 16px;
            padding: 24px;
            opacity: 0;
            visibility: hidden;
            transition: all 0.25s ease;
            display: flex;
            flex-direction: column;
            justify-content: center;
            z-index: 10;
            border: 2px solid var(--gray-200);
        }

        .kpi-card:hover .kpi-overlay:not(:empty) {
            opacity: 1;
            visibility: visible;
        }

        .kpi-overlay .overlay-year-badge {
            position: absolute;
            top: 12px;
            right: 12px;
            background: var(--gray-600);
            color: white;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }

        .kpi-overlay .overlay-label {
            font-size: 12px;
            color: var(--gray-500);
            margin-bottom: 6px;
        }

        .kpi-overlay .overlay-value {
            font-size: 26px;
            font-weight: 700;
            color: var(--gray-700);
            margin-bottom: 8px;
        }

        .kpi-overlay .overlay-change {
            font-size: 13px;
            color: var(--gray-500);
        }

        .kpi-overlay .overlay-change .up { color: var(--success); font-weight: 600; }
        .kpi-overlay .overlay-change .down { color: var(--danger); font-weight: 600; }

        /* 검사 목적별 카드 섹션 */
        .purpose-kpi-section {
            margin-bottom: 24px;
        }

        .section-title-bar {
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 16px;
        }

        .section-title {
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 18px;
            font-weight: 600;
            color: var(--gray-800);
        }

        .section-badge {
            background: var(--primary-light);
            color: var(--primary);
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 600;
        }

        .purpose-kpi-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
            gap: 16px;
        }

        /* 부서별 카드 스타일 */
        .dept-card {
            background: white;
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            border: 1px solid var(--gray-100);
            position: relative;
            overflow: hidden;
            transition: all 0.2s ease;
        }

        .dept-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(0,0,0,0.1);
        }

        .dept-card-header {
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 16px;
        }

        .dept-icon {
            width: 44px;
            height: 44px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 20px;
        }

        .dept-name {
            font-size: 18px;
            font-weight: 700;
            color: var(--gray-800);
        }

        .dept-card-body {
            display: flex;
            flex-direction: column;
            gap: 10px;
        }

        .dept-stat {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .dept-label {
            font-size: 13px;
            color: var(--gray-500);
        }

        .dept-value {
            font-size: 15px;
            font-weight: 600;
            color: var(--gray-800);
        }

        .dept-card-compare {
            margin-top: 16px;
            padding-top: 12px;
            border-top: 1px dashed var(--gray-200);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .dept-card-compare .compare-label {
            font-size: 12px;
            color: var(--gray-400);
        }

        .dept-card-compare .compare-value {
            font-size: 14px;
            font-weight: 700;
        }

        .dept-ratio {
            font-size: 14px;
            font-weight: 700;
            color: var(--primary);
            background: var(--primary-light);
            padding: 2px 8px;
            border-radius: 10px;
            margin-left: 8px;
        }

        .dept-overlay {
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.85);
            border-radius: 16px;
            display: none;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            color: white;
            padding: 16px;
            opacity: 0;
            transition: opacity 0.3s ease;
        }

        .dept-card:hover .dept-overlay.active {
            display: flex;
            opacity: 1;
        }

        .dept-overlay .overlay-title {
            font-size: 13px;
            color: rgba(255,255,255,0.7);
            margin-bottom: 8px;
        }

        .dept-overlay .overlay-value {
            font-size: 28px;
            font-weight: 700;
        }

        .dept-overlay .overlay-value.positive { color: #4ade80; }
        .dept-overlay .overlay-value.negative { color: #f87171; }

        .dept-overlay .overlay-detail {
            font-size: 12px;
            color: rgba(255,255,255,0.6);
            margin-top: 8px;
        }

        @media (max-width: 1200px) {
            .department-cards {
                grid-template-columns: repeat(2, 1fr) !important;
            }
        }

        @media (max-width: 768px) {
            .department-cards {
                grid-template-columns: 1fr !important;
            }
        }

        .purpose-kpi-card {
            background: white;
            border-radius: 16px;
            padding: 20px;
            box-shadow: var(--shadow);
            position: relative;
            overflow: hidden;
            cursor: pointer;
            transition: all 0.2s;
        }

        .purpose-kpi-card:hover {
            transform: translateY(-3px);
            box-shadow: var(--shadow-md);
        }

        .purpose-kpi-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
        }

        .purpose-kpi-card[data-color="blue"]::before { background: var(--primary); }
        .purpose-kpi-card[data-color="green"]::before { background: var(--success); }
        .purpose-kpi-card[data-color="orange"]::before { background: var(--orange); }
        .purpose-kpi-card[data-color="purple"]::before { background: var(--purple); }
        .purpose-kpi-card[data-color="pink"]::before { background: var(--pink); }
        .purpose-kpi-card[data-color="info"]::before { background: var(--info); }
        .purpose-kpi-card[data-color="teal"]::before { background: var(--teal); }
        .purpose-kpi-card[data-color="amber"]::before { background: var(--amber); }
        .purpose-kpi-card[data-color="rose"]::before { background: var(--rose); }
        .purpose-kpi-card[data-color="sky"]::before { background: var(--sky); }
        .purpose-kpi-card[data-color="lime"]::before { background: var(--lime); }
        .purpose-kpi-card[data-color="cyan"]::before { background: var(--cyan); }
        .purpose-kpi-card[data-color="danger"]::before { background: var(--danger); }

        .purpose-kpi-card[data-color="blue"] .purpose-kpi-icon { background: var(--primary-light); color: var(--primary); }
        .purpose-kpi-card[data-color="green"] .purpose-kpi-icon { background: var(--success-light); color: var(--success); }
        .purpose-kpi-card[data-color="orange"] .purpose-kpi-icon { background: var(--orange-light); color: var(--orange); }
        .purpose-kpi-card[data-color="purple"] .purpose-kpi-icon { background: var(--purple-light); color: var(--purple); }
        .purpose-kpi-card[data-color="pink"] .purpose-kpi-icon { background: var(--pink-light); color: var(--pink); }
        .purpose-kpi-card[data-color="info"] .purpose-kpi-icon { background: var(--info-light); color: var(--info); }
        .purpose-kpi-card[data-color="teal"] .purpose-kpi-icon { background: var(--teal-light); color: var(--teal); }
        .purpose-kpi-card[data-color="amber"] .purpose-kpi-icon { background: var(--amber-light); color: var(--amber); }
        .purpose-kpi-card[data-color="rose"] .purpose-kpi-icon { background: var(--rose-light); color: var(--rose); }
        .purpose-kpi-card[data-color="sky"] .purpose-kpi-icon { background: var(--sky-light); color: var(--sky); }
        .purpose-kpi-card[data-color="lime"] .purpose-kpi-icon { background: var(--lime-light); color: var(--lime); }
        .purpose-kpi-card[data-color="cyan"] .purpose-kpi-icon { background: var(--cyan-light); color: var(--cyan); }
        .purpose-kpi-card[data-color="danger"] .purpose-kpi-icon { background: var(--danger-light); color: var(--danger); }

        .purpose-kpi-card[data-color="danger"] .purpose-kpi-value { color: var(--danger); }

        .purpose-kpi-header {
            display: flex;
            align-items: flex-start;
            justify-content: space-between;
            margin-bottom: 12px;
        }

        .purpose-kpi-icon {
            width: 40px;
            height: 40px;
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
        }

        .purpose-kpi-trend {
            display: flex;
            align-items: center;
            gap: 4px;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }

        .purpose-kpi-trend.up { background: var(--success-light); color: var(--success); }
        .purpose-kpi-trend.down { background: var(--danger-light); color: var(--danger); }

        .purpose-kpi-name {
            font-size: 13px;
            color: var(--gray-500);
            margin-bottom: 6px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .purpose-kpi-value {
            font-size: 22px;
            font-weight: 700;
            color: var(--gray-800);
            margin-bottom: 4px;
        }

        .purpose-kpi-sub {
            font-size: 12px;
            color: var(--gray-400);
        }

        .purpose-kpi-sub span {
            color: var(--gray-600);
            font-weight: 500;
        }

        /* 전년도 오버레이 스타일 */
        .purpose-kpi-overlay {
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(2px);
            border-radius: 16px;
            padding: 20px;
            opacity: 0;
            visibility: hidden;
            transition: all 0.25s ease;
            display: flex;
            flex-direction: column;
            justify-content: center;
            z-index: 10;
            border: 2px solid var(--gray-200);
        }

        .purpose-kpi-card:hover .purpose-kpi-overlay {
            opacity: 1;
            visibility: visible;
        }

        .overlay-year-badge {
            position: absolute;
            top: 12px;
            right: 12px;
            background: var(--gray-600);
            color: white;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }

        .overlay-label {
            font-size: 12px;
            color: var(--gray-500);
            margin-bottom: 4px;
        }

        .overlay-name {
            font-size: 13px;
            color: var(--gray-600);
            margin-bottom: 8px;
            font-weight: 500;
        }

        .overlay-value {
            font-size: 22px;
            font-weight: 700;
            color: var(--gray-700);
            margin-bottom: 4px;
        }

        .overlay-sub {
            font-size: 12px;
            color: var(--gray-400);
        }

        .overlay-change {
            margin-top: 10px;
            padding-top: 10px;
            border-top: 1px dashed var(--gray-200);
            font-size: 12px;
            color: var(--gray-500);
        }

        .overlay-change .up { color: var(--success); font-weight: 600; }
        .overlay-change .down { color: var(--danger); font-weight: 600; }

        /* ====== 개인별 탭 전용 스타일 ====== */
        .personal-kpi-section {
            margin-bottom: 24px;
        }

        /* 효율성 분석 4분면 범례 */
        .quadrant-legend {
            display: flex;
            gap: 16px;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }
        .q-item {
            font-size: 12px;
            padding: 4px 10px;
            border-radius: 6px;
            background: var(--gray-100);
            color: var(--gray-600);
        }
        .q-item.q1 { background: rgba(37, 99, 235, 0.15); color: #1d4ed8; }   /* 고건수·고매출: 파란색 */
        .q-item.q2 { background: rgba(6, 182, 212, 0.15); color: #0891b2; }   /* 저건수·고매출: 청록색 */
        .q-item.q3 { background: rgba(249, 115, 22, 0.15); color: #ea580c; }  /* 고건수·저매출: 주황색 */
        .q-item.q4 { background: rgba(220, 38, 38, 0.15); color: #dc2626; }   /* 저건수·저매출: 빨간색 */

        /* 다중 선택 드롭다운 */
        .multi-select-container {
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .preset-btns {
            display: flex;
            gap: 4px;
        }
        .preset-btn {
            padding: 4px 10px;
            border: 1px solid var(--gray-200);
            background: white;
            border-radius: 6px;
            font-size: 11px;
            cursor: pointer;
            transition: all 0.2s;
        }
        .preset-btn:hover { background: var(--gray-50); }
        .preset-btn.active {
            background: var(--primary);
            color: white;
            border-color: var(--primary);
        }
        .multi-select-dropdown { position: relative; }
        .multi-select-btn {
            padding: 6px 12px;
            border: 1px solid var(--gray-200);
            background: white;
            border-radius: 8px;
            font-size: 12px;
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 4px;
        }
        .selected-count {
            color: var(--primary);
            font-weight: 600;
        }
        .multi-select-list {
            position: absolute;
            top: 100%;
            right: 0;
            margin-top: 4px;
            background: white;
            border: 1px solid var(--gray-200);
            border-radius: 8px;
            box-shadow: var(--shadow-md);
            z-index: 100;
            max-height: 250px;
            overflow-y: auto;
            min-width: 180px;
        }
        .multi-select-item {
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 8px 12px;
            cursor: pointer;
            transition: background 0.15s;
            font-size: 13px;
        }
        .multi-select-item:hover { background: var(--gray-50); }
        .multi-select-item input { accent-color: var(--primary); }
        .selected-tags {
            display: flex;
            flex-wrap: wrap;
            gap: 6px;
            padding: 8px 16px 0;
            min-height: 28px;
        }
        .selected-tag {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 2px 8px;
            background: var(--primary-light);
            color: var(--primary);
            border-radius: 12px;
            font-size: 11px;
            font-weight: 500;
        }
        .selected-tag .remove {
            cursor: pointer;
            font-size: 14px;
            line-height: 1;
        }
        .selected-tag .remove:hover { color: var(--danger); }

        /* 차트 컨트롤 */
        .chart-controls {
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .sort-toggle-btn {
            padding: 6px 12px;
            border: 1px solid var(--gray-200);
            background: white;
            border-radius: 6px;
            font-size: 11px;
            cursor: pointer;
            transition: all 0.2s;
        }
        .sort-toggle-btn:hover { background: var(--gray-50); }

        /* 긴급 배지 */
        .urgent-badge {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 3px 8px;
            background: var(--danger-light);
            color: var(--danger);
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }

        /* 정렬 가능 테이블 */
        .sortable-table th.sortable {
            cursor: pointer;
            user-select: none;
            transition: background 0.2s;
        }
        .sortable-table th.sortable:hover { background: var(--gray-100); }

        /* 모달 스타일 */
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.5);
            backdrop-filter: blur(4px);
            z-index: 1000;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .modal-content {
            background: white;
            border-radius: 20px;
            width: 90%;
            max-width: 800px;
            max-height: 85vh;
            overflow-y: auto;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        }
        .modal-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 20px 24px;
            border-bottom: 1px solid var(--gray-200);
            position: sticky;
            top: 0;
            background: white;
            z-index: 10;
        }
        .modal-header h3 {
            font-size: 18px;
            font-weight: 600;
            color: var(--gray-800);
        }
        .modal-close {
            width: 36px;
            height: 36px;
            border-radius: 50%;
            border: none;
            background: var(--gray-100);
            color: var(--gray-600);
            font-size: 18px;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s;
        }
        .modal-close:hover {
            background: var(--danger-light);
            color: var(--danger);
        }
        .modal-body { padding: 24px; }
        .modal-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 24px;
        }
        .modal-section h4 {
            font-size: 14px;
            font-weight: 600;
            color: var(--gray-700);
            margin-bottom: 12px;
        }
        .modal-client-list {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        .modal-client-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px 12px;
            background: var(--gray-50);
            border-radius: 8px;
        }
        .modal-client-name {
            font-size: 13px;
            color: var(--gray-700);
            font-weight: 500;
        }
        .modal-client-value {
            font-size: 13px;
            color: var(--primary);
            font-weight: 600;
        }
        .modal-region-tags {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
        }
        .region-tag {
            padding: 6px 12px;
            background: var(--primary-light);
            color: var(--primary);
            border-radius: 20px;
            font-size: 12px;
            font-weight: 500;
        }

        /* SVG Korea Map Styles */
        .region-path {
            fill: #dbeafe;
            stroke: #fff;
            stroke-width: 2;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .region-path:hover {
            fill: #93c5fd;
            transform: scale(1.02);
            filter: drop-shadow(0 4px 8px rgba(0,0,0,0.15));
        }
        .region-path.selected {
            fill: #3b82f6;
            stroke: #1e40af;
            stroke-width: 3;
        }
        .region-path.level-1 { fill: #dbeafe; }
        .region-path.level-2 { fill: #93c5fd; }
        .region-path.level-3 { fill: #3b82f6; }
        .region-path.level-4 { fill: #1e3a8a; }
        .map-label {
            font-size: 11px;
            font-weight: 600;
            fill: #374151;
            pointer-events: none;
            text-anchor: middle;
        }
        .map-label.small {
            font-size: 9px;
        }

        /* Region KPI Overlay */
        .region-kpi-overlay {
            position: absolute;
            top: 100%;
            left: 0;
            width: 280px;
            background: white;
            border: 2px solid var(--primary);
            border-radius: 12px;
            padding: 16px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            z-index: 1000;
            font-size: 12px;
        }

        /* Region Detail Panel */
        .region-detail-section {
            margin-bottom: 20px;
        }
        .region-detail-section:last-child {
            margin-bottom: 0;
        }
        .region-detail-title {
            font-size: 13px;
            font-weight: 600;
            color: var(--gray-700);
            margin-bottom: 10px;
            padding-bottom: 6px;
            border-bottom: 1px solid var(--gray-100);
        }
        .region-stat-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 12px;
            margin-bottom: 16px;
        }
        .region-stat-item {
            background: var(--gray-50);
            padding: 12px;
            border-radius: 8px;
            text-align: center;
        }
        .region-stat-value {
            font-size: 18px;
            font-weight: 700;
            color: var(--gray-800);
        }
        .region-stat-label {
            font-size: 11px;
            color: var(--gray-500);
            margin-top: 4px;
        }
        .region-manager-list {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        .region-manager-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 8px 12px;
            background: var(--gray-50);
            border-radius: 6px;
        }
        .region-ai-opinion {
            background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
            border: 1px solid #7dd3fc;
            border-radius: 10px;
            padding: 14px;
            font-size: 13px;
            line-height: 1.6;
            color: var(--gray-700);
        }
        .region-top-clients {
            display: flex;
            flex-direction: column;
            gap: 6px;
        }
        .region-client-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 6px 0;
            border-bottom: 1px dashed var(--gray-100);
        }
        .region-client-item:last-child {
            border-bottom: none;
        }
        .heatmap-cell {
            padding: 4px 8px;
            border-radius: 4px;
            text-align: center;
            font-weight: 600;
        }
        .heatmap-high { background: #dcfce7; color: #166534; }
        .heatmap-medium { background: #fef9c3; color: #854d0e; }
        .heatmap-low { background: #fee2e2; color: #991b1b; }
        .region-distribution {
            display: flex;
            gap: 4px;
            flex-wrap: wrap;
        }
        .region-chip {
            padding: 2px 8px;
            background: var(--primary-light);
            color: var(--primary);
            border-radius: 12px;
            font-size: 11px;
        }

        /* 상세 버튼 */
        .btn-detail {
            padding: 6px 12px;
            background: var(--gray-100);
            border: none;
            border-radius: 6px;
            font-size: 12px;
            color: var(--gray-600);
            cursor: pointer;
            transition: all 0.2s;
        }
        .btn-detail:hover {
            background: var(--primary-light);
            color: var(--primary);
        }
        .text-center { text-align: center; }

        /* 콘텐츠 영역 */
        .content-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 24px;
        }

        .card {
            background: white;
            border-radius: 16px;
            box-shadow: var(--shadow);
            overflow: hidden;
        }

        .card-header {
            padding: 20px 24px;
            border-bottom: 1px solid var(--gray-100);
            display: flex;
            align-items: center;
            justify-content: space-between;
        }

        .card-title {
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 16px;
            font-weight: 600;
            color: var(--gray-800);
        }

        .card-badge {
            background: var(--primary-light);
            color: var(--primary);
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }

        .card-body {
            padding: 24px;
        }

        /* 차트 */
        .chart-container {
            position: relative;
            height: 320px;
        }

        .chart-legend {
            display: flex;
            gap: 20px;
            margin-bottom: 16px;
        }

        .legend-item {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 13px;
            color: var(--gray-600);
        }

        .legend-color {
            width: 12px;
            height: 12px;
            border-radius: 3px;
        }

        /* 테이블 */
        .data-table {
            width: 100%;
            border-collapse: collapse;
        }

        .data-table th {
            padding: 12px 16px;
            text-align: left;
            font-size: 12px;
            font-weight: 600;
            color: var(--gray-500);
            background: var(--gray-50);
            border-bottom: 1px solid var(--gray-200);
        }

        .data-table th.sortable {
            cursor: pointer;
            user-select: none;
            position: relative;
            padding-right: 24px;
        }

        .data-table th.sortable:hover {
            background: var(--gray-100);
        }

        .data-table th.sortable::after {
            content: '⇅';
            position: absolute;
            right: 8px;
            color: var(--gray-300);
            font-size: 12px;
        }

        .data-table th.sortable.asc::after {
            content: '↑';
            color: var(--primary);
        }

        .data-table th.sortable.desc::after {
            content: '↓';
            color: var(--primary);
        }

        .data-table td {
            padding: 14px 16px;
            font-size: 14px;
            border-bottom: 1px solid var(--gray-100);
        }

        .data-table tbody tr:hover {
            background: var(--gray-50);
        }

        .data-table .text-right {
            text-align: right;
        }

        .progress-cell {
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .progress-bar {
            flex: 1;
            height: 8px;
            background: var(--gray-200);
            border-radius: 4px;
            overflow: hidden;
        }

        .progress-fill {
            height: 100%;
            background: var(--primary);
            border-radius: 4px;
        }

        .progress-value {
            font-size: 12px;
            color: var(--gray-500);
            min-width: 40px;
            text-align: right;
        }

        .change-badge {
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }

        .change-badge.positive {
            background: var(--success-light);
            color: var(--success);
        }

        .change-badge.negative {
            background: var(--danger-light);
            color: var(--danger);
        }

        /* 탭 콘텐츠 */
        .tab-content {
            display: none;
        }

        .tab-content.active {
            display: block;
        }

        /* KPI 섹션 표시 제어 */
        .kpi-section.hidden {
            display: none;
        }

        /* 토스트 */
        .toast {
            position: fixed;
            top: 20px;
            right: 20px;
            padding: 14px 24px;
            background: var(--success);
            color: white;
            border-radius: 12px;
            box-shadow: var(--shadow-md);
            z-index: 1000;
            display: none;
            font-size: 14px;
        }

        .toast.error { background: var(--danger); }
        .toast.loading { background: var(--primary); }

        /* 반응형 */
        @media (max-width: 1400px) {
            .kpi-section { grid-template-columns: repeat(2, 1fr); }
            .content-grid { grid-template-columns: 1fr; }
            .purpose-kpi-grid { grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); }
        }

        @media (max-width: 768px) {
            .tab-cards { overflow-x: auto; flex-wrap: nowrap; padding-bottom: 10px; }
            .tab-card { flex: 0 0 auto; }
            .filter-row { flex-direction: column; align-items: stretch; }
            .btn-search { margin-left: 0; justify-content: center; }
            .kpi-section { grid-template-columns: 1fr; }
        }

        /* 스크롤바 */
        ::-webkit-scrollbar { width: 6px; height: 6px; }
        ::-webkit-scrollbar-track { background: var(--gray-100); }
        ::-webkit-scrollbar-thumb { background: var(--gray-300); border-radius: 3px; }
        ::-webkit-scrollbar-thumb:hover { background: var(--gray-400); }

        .scroll-table { max-height: 400px; overflow-y: auto; }

        /* AI 분석 섹션 */
        .ai-section { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; padding: 30px; margin-bottom: 24px; }
        .ai-header { color: white; margin-bottom: 20px; }
        .ai-header h2 { font-size: 24px; margin-bottom: 8px; }
        .ai-header p { opacity: 0.9; font-size: 14px; }
        .ai-input-container { background: white; border-radius: 16px; padding: 20px; }
        .ai-input-wrapper { display: flex; gap: 12px; margin-bottom: 16px; }
        .ai-input { flex: 1; padding: 14px 18px; border: 2px solid var(--gray-200); border-radius: 12px; font-size: 15px; outline: none; }
        .ai-input:focus { border-color: var(--primary); }
        .ai-btn { padding: 14px 28px; background: var(--primary); color: white; border: none; border-radius: 12px; font-weight: 600; cursor: pointer; transition: all 0.2s; }
        .ai-btn:hover { background: var(--primary-dark); }
        .ai-btn:disabled { opacity: 0.6; cursor: not-allowed; }
        .ai-examples { display: flex; gap: 8px; flex-wrap: wrap; }
        .ai-example { background: var(--gray-100); padding: 6px 12px; border-radius: 20px; font-size: 12px; color: var(--gray-600); cursor: pointer; transition: all 0.2s; }
        .ai-example:hover { background: var(--primary-light); color: var(--primary); }
        .ai-result { margin-top: 20px; padding: 20px; background: var(--gray-50); border-radius: 12px; display: none; }
        .ai-result.show { display: block; }
        .ai-result-table { width: 100%; }
        .ai-result-table th, .ai-result-table td { padding: 10px; text-align: left; border-bottom: 1px solid var(--gray-200); }
        .ai-insight { margin-top: 16px; padding: 12px 16px; background: var(--warning-light); border-radius: 8px; font-size: 14px; }
    </style>
</head>
<body>
    <div id="toast" class="toast"></div>

    <!-- 헤더 -->
    <header class="header">
        <div class="header-left">
            <div class="logo">📊</div>
            <h1 class="header-title">경영지표 대시보드</h1>
        </div>
        <div class="header-right">
            <div class="token-badge">
                <div class="current">이번달: <span id="thisMonthTokens">0</span> 토큰 | ₩<span id="thisMonthKRW">0</span></div>
                <div class="prev">저번달: <span id="lastMonthTokens">0</span> 토큰 | ₩<span id="lastMonthKRW">0</span></div>
            </div>
        </div>
    </header>

    <main class="main-container">
        <!-- 필터 섹션 -->
        <section class="filter-section">
            <div class="filter-row">
                <div class="filter-group">
                    📅
                    <span class="filter-label">조회기간</span>
                    <select id="yearSelect" class="filter-select">
                        <option value="2025">2025년</option>
                        <option value="2024">2024년</option>
                    </select>
                    <select id="monthSelect" class="filter-select">
                        <option value="">전체</option>
                        <option value="1">1월</option>
                        <option value="2">2월</option>
                        <option value="3">3월</option>
                        <option value="4">4월</option>
                        <option value="5">5월</option>
                        <option value="6">6월</option>
                        <option value="7">7월</option>
                        <option value="8">8월</option>
                        <option value="9">9월</option>
                        <option value="10">10월</option>
                        <option value="11">11월</option>
                        <option value="12">12월</option>
                    </select>
                </div>

                <div class="filter-divider"></div>

                <label class="filter-checkbox">
                    <input type="checkbox" id="compareCheck">
                    <span>전년비교</span>
                </label>

                <div class="filter-group" id="compareYearGroup" style="display: none;">
                    <select id="compareYearSelect" class="filter-select">
                        <option value="2024">2024년</option>
                        <option value="2023">2023년</option>
                    </select>
                </div>

                <div class="filter-divider"></div>

                <div class="filter-group">
                    🎯
                    <select id="purposeSelect" class="filter-select" style="min-width: 180px;">
                        <option value="전체">검사목적: 전체</option>
                    </select>
                </div>

                <button id="btnSearch" class="btn-search" onclick="loadData()">
                    🔍 조회하기
                </button>
            </div>
        </section>

        <!-- 탭 카드 -->
        <section class="tab-cards">
            <div class="tab-card main-tab active" onclick="showTab('main')">
                <div class="tab-icon">🏠</div>
                <div class="tab-label">메인</div>
            </div>
            <div class="tab-card" onclick="showTab('personal')">
                <div class="tab-icon">👤</div>
                <div class="tab-label">개인별</div>
            </div>
            <div class="tab-card" onclick="showTab('team')">
                <div class="tab-icon">👥</div>
                <div class="tab-label">팀별</div>
            </div>
            <div class="tab-card" onclick="showTab('monthly')">
                <div class="tab-icon">📆</div>
                <div class="tab-label">월별</div>
            </div>
            <div class="tab-card" onclick="showTab('client')">
                <div class="tab-icon">🏢</div>
                <div class="tab-label">업체별</div>
            </div>
            <div class="tab-card" onclick="showTab('region')">
                <div class="tab-icon">📍</div>
                <div class="tab-label">지역별</div>
            </div>
            <div class="tab-card" onclick="showTab('purpose')">
                <div class="tab-icon">🎯</div>
                <div class="tab-label">목적별</div>
            </div>
            <div class="tab-card" onclick="showTab('sampleType')">
                <div class="tab-icon">🧪</div>
                <div class="tab-label">유형</div>
            </div>
            <div class="tab-card" onclick="showTab('defect')">
                <div class="tab-icon">⚠️</div>
                <div class="tab-label">부적합</div>
            </div>
            <div class="tab-card" onclick="showTab('foodItem')">
                <div class="tab-icon">🔬</div>
                <div class="tab-label">검사항목</div>
            </div>
            <div class="tab-card ai-tab" onclick="showTab('aiAnalysis')">
                <div class="tab-icon">🤖</div>
                <div class="tab-label">AI 분석</div>
            </div>
            <div class="tab-card info-tab" onclick="showTab('companyInfo')">
                <div class="tab-icon">🏛️</div>
                <div class="tab-label">기업 정보</div>
            </div>
            <div class="tab-card terminal-tab" onclick="showTab('webTerminal')">
                <div class="tab-icon">💻</div>
                <div class="tab-label">터미널</div>
            </div>
        </section>

        <!-- KPI 카드 -->
        <section class="kpi-section" id="kpiSection">
            <div class="kpi-card sales">
                <div class="kpi-header">
                    <div class="kpi-icon">💰</div>
                    <div class="kpi-trend up" id="salesTrend" style="visibility: hidden;"><span>0%</span></div>
                </div>
                <div class="kpi-label">총 매출</div>
                <div class="kpi-value" id="totalSales">-</div>
                <div class="kpi-compare" id="compareTotalSales" style="display: none;"></div>
                <div class="kpi-overlay" id="salesOverlay"></div>
            </div>

            <div class="kpi-card count">
                <div class="kpi-header">
                    <div class="kpi-icon">📋</div>
                    <div class="kpi-trend up" id="countTrend" style="visibility: hidden;"><span>0%</span></div>
                </div>
                <div class="kpi-label">총 건수</div>
                <div class="kpi-value" id="totalCount">-</div>
                <div class="kpi-compare" id="compareTotalCount" style="display: none;"></div>
                <div class="kpi-overlay" id="countOverlay"></div>
            </div>

            <div class="kpi-card price">
                <div class="kpi-header">
                    <div class="kpi-icon">🏷️</div>
                    <div class="kpi-trend up" id="priceTrend" style="visibility: hidden;"><span>0%</span></div>
                </div>
                <div class="kpi-label">평균 단가</div>
                <div class="kpi-value" id="avgPrice">-</div>
                <div class="kpi-compare" id="compareAvgPrice" style="display: none;"></div>
                <div class="kpi-overlay" id="priceOverlay"></div>
            </div>

            <div class="kpi-card goal">
                <div class="kpi-header">
                    <div class="kpi-icon">🏆</div>
                </div>
                <div class="kpi-label">목표 달성률</div>
                <div class="kpi-value" id="goalRate">-</div>
                <div class="kpi-compare">목표: <span id="goalTarget">70억</span></div>
                <div class="kpi-compare" id="goalCompare" style="display: none;"></div>
                <div class="kpi-overlay" id="goalOverlay"></div>
            </div>
        </section>

        <!-- 메인 탭 콘텐츠 -->
        <div id="main" class="tab-content active">
            <!-- 부서별 현황 카드 -->
            <section class="department-section" style="margin-bottom: 24px;">
                <div class="section-title-bar">
                    <div class="section-title">🏢 부서별 현황</div>
                </div>
                <div class="department-cards" style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-top: 16px;">
                    <!-- 본사 카드 -->
                    <div class="dept-card" id="deptCardBonsa">
                        <div class="dept-card-header">
                            <div class="dept-icon" style="background: linear-gradient(135deg, #6366f1, #8b5cf6);">🏛️</div>
                            <div class="dept-name">본사 <span class="dept-ratio" id="deptBonsaRatio">-</span></div>
                        </div>
                        <div class="dept-card-body">
                            <div class="dept-stat">
                                <span class="dept-label">총 매출</span>
                                <span class="dept-value" id="deptBonsaSales">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">건수</span>
                                <span class="dept-value" id="deptBonsaCount">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">평균단가</span>
                                <span class="dept-value" id="deptBonsaAvg">-</span>
                            </div>
                        </div>
                        <div class="dept-overlay" id="deptBonsaOverlay"></div>
                    </div>

                    <!-- 마케팅 카드 -->
                    <div class="dept-card" id="deptCardMarketing">
                        <div class="dept-card-header">
                            <div class="dept-icon" style="background: linear-gradient(135deg, #10b981, #059669);">📢</div>
                            <div class="dept-name">마케팅 <span class="dept-ratio" id="deptMarketingRatio">-</span></div>
                        </div>
                        <div class="dept-card-body">
                            <div class="dept-stat">
                                <span class="dept-label">총 매출</span>
                                <span class="dept-value" id="deptMarketingSales">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">건수</span>
                                <span class="dept-value" id="deptMarketingCount">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">평균단가</span>
                                <span class="dept-value" id="deptMarketingAvg">-</span>
                            </div>
                        </div>
                        <div class="dept-overlay" id="deptMarketingOverlay"></div>
                    </div>

                    <!-- 영업부 카드 -->
                    <div class="dept-card" id="deptCardSales">
                        <div class="dept-card-header">
                            <div class="dept-icon" style="background: linear-gradient(135deg, #f59e0b, #d97706);">💼</div>
                            <div class="dept-name">영업부 <span class="dept-ratio" id="deptSalesRatio">-</span></div>
                        </div>
                        <div class="dept-card-body">
                            <div class="dept-stat">
                                <span class="dept-label">총 매출</span>
                                <span class="dept-value" id="deptSalesSales">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">건수</span>
                                <span class="dept-value" id="deptSalesCount">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">평균단가</span>
                                <span class="dept-value" id="deptSalesAvg">-</span>
                            </div>
                        </div>
                        <div class="dept-overlay" id="deptSalesOverlay"></div>
                    </div>

                    <!-- 지사 카드 -->
                    <div class="dept-card" id="deptCardBranch">
                        <div class="dept-card-header">
                            <div class="dept-icon" style="background: linear-gradient(135deg, #ec4899, #db2777);">🏬</div>
                            <div class="dept-name">지사 <span class="dept-ratio" id="deptBranchRatio">-</span></div>
                        </div>
                        <div class="dept-card-body">
                            <div class="dept-stat">
                                <span class="dept-label">총 매출</span>
                                <span class="dept-value" id="deptBranchSales">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">건수</span>
                                <span class="dept-value" id="deptBranchCount">-</span>
                            </div>
                            <div class="dept-stat">
                                <span class="dept-label">평균단가</span>
                                <span class="dept-value" id="deptBranchAvg">-</span>
                            </div>
                        </div>
                        <div class="dept-overlay" id="deptBranchOverlay"></div>
                    </div>
                </div>
            </section>

            <section class="purpose-kpi-section" id="purposeKpiSection">
                <div class="section-title-bar">
                    <div class="section-title">🎯 검사 목적별 현황</div>
                    <div class="section-badge" id="purposeCount">0개 목적</div>
                </div>
                <div class="purpose-kpi-grid" id="purposeGrid"></div>
            </section>
        </div>

        <!-- 개인별 탭 -->
        <div id="personal" class="tab-content">
            <!-- 개인별 전용 KPI 카드 -->
            <div class="kpi-section personal-kpi-section">
                <div class="kpi-card sales">
                    <div class="kpi-header">
                        <div class="kpi-icon">👥</div>
                    </div>
                    <div class="kpi-label">총 영업담당자</div>
                    <div class="kpi-value" id="personalTotalManagers">-</div>
                    <div class="kpi-compare">활동 중인 담당자</div>
                </div>
                <div class="kpi-card count">
                    <div class="kpi-header">
                        <div class="kpi-icon">💵</div>
                    </div>
                    <div class="kpi-label">평균 매출</div>
                    <div class="kpi-value" id="personalAvgSales">-</div>
                    <div class="kpi-compare">담당자당 평균</div>
                </div>
                <div class="kpi-card price">
                    <div class="kpi-header">
                        <div class="kpi-icon">🚀</div>
                        <div class="kpi-trend up" id="topGrowthTrend" style="visibility: hidden;">↑ 0%</div>
                    </div>
                    <div class="kpi-label">최고 성장자</div>
                    <div class="kpi-value" id="personalTopGrowth" style="font-size: 20px;">-</div>
                    <div class="kpi-compare" id="personalTopGrowthRate">전년 대비</div>
                </div>
                <div class="kpi-card goal">
                    <div class="kpi-header">
                        <div class="kpi-icon">🚨</div>
                    </div>
                    <div class="kpi-label">긴급 최고 요청자 TOP 5</div>
                    <div class="kpi-value" id="personalUrgentTop" style="font-size: 13px; line-height: 1.5;">-</div>
                </div>
            </div>

            <!-- 효율성 분석 + 월별 추이 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 효율성 분석 (건수 vs 매출)</div>
                        <div class="card-badge">4분면 분석</div>
                    </div>
                    <div class="card-body">
                        <div class="quadrant-legend">
                            <span class="q-item q1">🌟 고건수·고매출</span>
                            <span class="q-item q2">💎 저건수·고매출</span>
                            <span class="q-item q3">📈 고건수·저매출</span>
                            <span class="q-item q4">⚠️ 저건수·저매출</span>
                        </div>
                        <div class="chart-container" style="height: 300px;"><canvas id="efficiencyChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📆 월별 매출 추이</div>
                        <div class="multi-select-container">
                            <div class="preset-btns">
                                <button class="preset-btn active" onclick="setMonthlyPreset('all')">전체</button>
                                <button class="preset-btn" onclick="setMonthlyPreset('top3')">TOP 3</button>
                            </div>
                            <div class="multi-select-dropdown">
                                <button class="multi-select-btn" onclick="toggleMultiSelect()">
                                    담당자 선택 <span class="selected-count" id="selectedCount">(0)</span> ▼
                                </button>
                                <div class="multi-select-list" id="managerSelectList" style="display: none;"></div>
                            </div>
                        </div>
                    </div>
                    <div class="selected-tags" id="selectedManagerTags"></div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 280px;"><canvas id="managerMonthlyChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 매출 TOP 15 + 건당 매출 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 영업담당별 매출 TOP 15</div>
                        <div style="display: flex; align-items: center; gap: 10px;">
                            <select id="managerChartPurposeFilter" class="filter-select" style="padding: 6px 12px; border-radius: 6px; border: 1px solid #e2e8f0;" onchange="updateManagerChart()">
                                <option value="전체">전체 검사목적</option>
                            </select>
                            <div class="card-badge" id="managerChartBadge">2025년</div>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="managerLegend" style="display: none;"></div>
                        <div class="chart-container"><canvas id="managerChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">💰 건당 매출 (평균단가)</div>
                        <div class="chart-controls">
                            <select id="perCasePurposeSelect" class="filter-select" style="min-width: 140px;" onchange="updatePerCaseChart()">
                                <option value="전체">전체 목적</option>
                            </select>
                            <button class="sort-toggle-btn" id="perCaseSortBtn" onclick="togglePerCaseSort()">내림차순 ▼</button>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container"><canvas id="perCaseChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 긴급 접수 + 일 방문 거래처 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🚨 긴급 접수 건수</div>
                        <select id="urgentPurposeSelect" class="filter-select" style="min-width: 180px;" onchange="updateUrgentChart()">
                            <option value="전체">검사목적: 전체</option>
                        </select>
                    </div>
                    <div class="card-body">
                        <div style="display: flex; gap: 12px; margin-bottom: 8px; font-size: 11px;">
                            <span style="color: #ef4444;">● 상위 (80%↑)</span>
                            <span style="color: #f59e0b;">● 중위 (50%↑)</span>
                            <span style="color: #6366f1;">● 하위</span>
                        </div>
                        <div class="chart-container" style="height: 260px;"><canvas id="urgentChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🏢 일 방문 거래처 수</div>
                        <div class="card-badge">접수일 기준</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="dailyClientLegend" style="display: none;"></div>
                        <div class="chart-container" style="height: 280px;"><canvas id="dailyClientChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 긴급 월별 추이 + 긴급 건당 단가 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 긴급 월별 추이</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 280px;"><canvas id="urgentMonthlyChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">💰 긴급 건당 단가</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 280px;"><canvas id="urgentUnitPriceChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 영업담당별 상세 테이블 -->
            <div class="card">
                <div class="card-header">
                    <div class="card-title">📋 영업담당별 상세</div>
                    <div style="display: flex; align-items: center; gap: 10px;">
                        <select id="managerPurposeFilter" class="filter-select" style="padding: 6px 12px; border-radius: 6px; border: 1px solid #e2e8f0;" onchange="updateManagerTable()">
                            <option value="전체">전체 검사목적</option>
                        </select>
                        <div class="card-badge" id="managerTableBadge">0명</div>
                    </div>
                </div>
                <div class="card-body">
                    <div class="scroll-table" style="max-height: 500px;">
                        <table class="data-table sortable-table" id="managerTable">
                            <thead id="managerTableHead"></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- 담당자 상세 모달 -->
            <div id="managerModal" class="modal-overlay" style="display: none;">
                <div class="modal-content">
                    <div class="modal-header">
                        <h3 id="modalManagerName">담당자 상세</h3>
                        <button class="modal-close" onclick="closeManagerModal()">✕</button>
                    </div>
                    <div class="modal-body">
                        <div class="modal-grid">
                            <div class="modal-section">
                                <h4>📊 주요 거래 업체 TOP 5</h4>
                                <div id="modalTopClients" class="modal-client-list"></div>
                            </div>
                            <div class="modal-section">
                                <h4>🎯 검사 목적별 비중</h4>
                                <div style="height: 200px;"><canvas id="modalPurposeCanvas"></canvas></div>
                            </div>
                        </div>
                        <div class="modal-section" style="margin-top: 20px;">
                            <h4>📍 담당 지역 분포</h4>
                            <div id="modalRegions" class="modal-region-tags"></div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 팀별 탭 -->
        <div id="team" class="tab-content">
            <!-- 팀별 전용 KPI 카드 -->
            <div class="kpi-section team-kpi-section">
                <div class="kpi-card sales">
                    <div class="kpi-header">
                        <div class="kpi-icon">🏢</div>
                    </div>
                    <div class="kpi-label">총 팀 수</div>
                    <div class="kpi-value" id="teamTotalBranches">-</div>
                    <div class="kpi-compare">활동 중인 팀</div>
                </div>
                <div class="kpi-card count">
                    <div class="kpi-header">
                        <div class="kpi-icon">💵</div>
                    </div>
                    <div class="kpi-label">팀 평균 매출</div>
                    <div class="kpi-value" id="teamAvgSales">-</div>
                    <div class="kpi-compare">팀당 평균</div>
                </div>
                <div class="kpi-card price">
                    <div class="kpi-header">
                        <div class="kpi-icon">🏆</div>
                    </div>
                    <div class="kpi-label">최고 성과 팀</div>
                    <div class="kpi-value" id="teamTopBranch" style="font-size: 20px;">-</div>
                    <div class="kpi-compare" id="teamTopBranchSales">-</div>
                </div>
                <div class="kpi-card goal">
                    <div class="kpi-header">
                        <div class="kpi-icon">🚀</div>
                        <div class="kpi-trend up" id="teamTopGrowthTrend" style="visibility: hidden;">↑ 0%</div>
                    </div>
                    <div class="kpi-label">최고 성장 팀</div>
                    <div class="kpi-value" id="teamTopGrowth" style="font-size: 20px;">-</div>
                    <div class="kpi-compare" id="teamTopGrowthRate">전년 대비</div>
                </div>
            </div>

            <!-- 팀별 매출 TOP + 건당 매출 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 팀별 매출 현황</div>
                        <div style="display: flex; align-items: center; gap: 10px;">
                            <select id="branchChartPurposeFilter" class="filter-select" style="padding: 6px 12px; border-radius: 6px; border: 1px solid #e2e8f0;" onchange="updateBranchChart()">
                                <option value="전체">전체 검사목적</option>
                            </select>
                            <div class="card-badge" id="branchChartBadge">2025년</div>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="branchLegend" style="display: none;"></div>
                        <div class="chart-container"><canvas id="branchChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">💰 팀별 건당 매출 (평균단가)</div>
                        <div class="chart-controls">
                            <select id="branchPerCasePurposeSelect" class="filter-select" style="min-width: 140px;" onchange="updateBranchPerCaseChart()">
                                <option value="전체">전체 목적</option>
                            </select>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="branchPerCaseLegend" style="display: none;"></div>
                        <div class="chart-container"><canvas id="branchPerCaseChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 효율성 분석 + 월별 추이 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 팀별 효율성 분석</div>
                        <div class="card-badge">건수 vs 매출</div>
                    </div>
                    <div class="card-body">
                        <div class="quadrant-legend">
                            <span class="q-item q1">🌟 고건수·고매출</span>
                            <span class="q-item q2">💎 저건수·고매출</span>
                            <span class="q-item q3">📈 고건수·저매출</span>
                            <span class="q-item q4">⚠️ 저건수·저매출</span>
                        </div>
                        <div class="chart-container" style="height: 300px;"><canvas id="branchEfficiencyChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📆 팀별 월별 추이</div>
                        <div class="chart-controls" style="display: flex; gap: 8px; align-items: center;">
                            <button class="filter-btn active" onclick="setBranchMonthlyFilter('all')" id="branchMonthlyAll">전체</button>
                            <button class="filter-btn" onclick="setBranchMonthlyFilter('top3')" id="branchMonthlyTop3">TOP 3</button>
                            <select id="branchMonthlySelect" class="filter-select" style="min-width: 120px;" onchange="setBranchMonthlyFilter('select')">
                                <option value="">팀 선택</option>
                            </select>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="branchMonthlyLegend" style="display: none;"></div>
                        <div class="chart-container" style="height: 280px;"><canvas id="branchMonthlyChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 팀별 상세 테이블 -->
            <div class="card" style="margin-bottom: 24px;">
                <div class="card-header">
                    <div class="card-title">📋 팀별 상세</div>
                    <div style="display: flex; align-items: center; gap: 10px;">
                        <select id="branchTablePurposeFilter" class="filter-select" style="padding: 6px 12px; border-radius: 6px; border: 1px solid #e2e8f0;" onchange="updateBranchTable()">
                            <option value="전체">전체 검사목적</option>
                        </select>
                        <div class="card-badge" id="branchTableBadge">0개 팀</div>
                    </div>
                </div>
                <div class="card-body">
                    <div class="scroll-table">
                        <table class="data-table" id="branchTable">
                            <thead id="branchTableHead"><tr><th>팀명</th><th class="text-right">매출액</th><th class="text-right">건수</th><th class="text-right">평균단가</th><th class="text-right">담당자수</th><th>비중</th></tr></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- 거래처 중복 분석 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🔄 월별 거래처 중복 현황</div>
                        <div class="card-badge">기존 거래처 vs 신규</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 300px;"><canvas id="clientRetentionChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 거래처 리텐션율 추이</div>
                        <div class="card-badge">이전달 대비 유지율</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 300px;"><canvas id="retentionRateChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 팀별 거래처 리텐션 테이블 -->
            <div class="card" style="margin-bottom: 24px;">
                <div class="card-header">
                    <div class="card-title">📋 팀별 거래처 현황</div>
                    <div class="card-badge" id="branchRetentionBadge">0개</div>
                </div>
                <div class="card-body">
                    <div class="scroll-table">
                        <table class="data-table" id="branchRetentionTable">
                            <thead>
                                <tr>
                                    <th>팀명</th>
                                    <th class="text-right">누적 거래처</th>
                                    <th class="text-right">1월</th>
                                    <th class="text-right">2월</th>
                                    <th class="text-right">3월</th>
                                    <th class="text-right">4월</th>
                                    <th class="text-right">5월</th>
                                    <th class="text-right">6월</th>
                                    <th class="text-right">7월</th>
                                    <th class="text-right">8월</th>
                                    <th class="text-right">9월</th>
                                    <th class="text-right">10월</th>
                                    <th class="text-right">11월</th>
                                    <th class="text-right">12월</th>
                                </tr>
                            </thead>
                            <tbody id="branchRetentionBody"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- 월별 탭 -->
        <div id="monthly" class="tab-content">
            <!-- 월별 KPI 카드 -->
            <section class="kpi-section monthly-kpi-section">
                <div class="kpi-card sales">
                    <div class="kpi-header"><div class="kpi-icon">🏆</div></div>
                    <div class="kpi-label">최고 매출월</div>
                    <div class="kpi-value" id="monthlyMaxMonth">-</div>
                    <div class="kpi-compare" id="monthlyMaxValue">-</div>
                </div>
                <div class="kpi-card count">
                    <div class="kpi-header"><div class="kpi-icon">📉</div></div>
                    <div class="kpi-label">최저 매출월</div>
                    <div class="kpi-value" id="monthlyMinMonth">-</div>
                    <div class="kpi-compare" id="monthlyMinValue">-</div>
                </div>
                <div class="kpi-card price">
                    <div class="kpi-header"><div class="kpi-icon">📊</div></div>
                    <div class="kpi-label">월평균 매출</div>
                    <div class="kpi-value" id="monthlyAvgSales">-</div>
                    <div class="kpi-compare" id="monthlyAvgCount">-</div>
                </div>
                <div class="kpi-card goal">
                    <div class="kpi-header"><div class="kpi-icon">📅</div></div>
                    <div class="kpi-label">YTD 누적</div>
                    <div class="kpi-value" id="monthlyYtdSales">-</div>
                    <div class="kpi-compare" id="monthlyYtdCount">-</div>
                </div>
            </section>

            <!-- 매출/건수 추이 차트 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 월별 매출 추이</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-legend" id="monthlyLegend" style="display: none;"></div>
                        <div class="chart-container" style="height: 350px;"><canvas id="monthlyChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 월별 건수 추이</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 350px;"><canvas id="monthlyCountChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 분기별/평균단가/전년비교 차트 -->
            <div class="content-grid" style="grid-template-columns: repeat(3, 1fr); margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 분기별 매출</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 300px;"><canvas id="quarterlyChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">💰 월별 평균단가</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 300px;"><canvas id="monthlyAvgPriceChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 전년 대비</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 300px;"><canvas id="yoyChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 목적별 월별 히트맵 -->
            <div class="card" style="margin-bottom: 24px;">
                <div class="card-header">
                    <div class="card-title">🔥 검사목적별 월별 히트맵</div>
                </div>
                <div class="card-body">
                    <div class="scroll-table">
                        <table class="data-table heatmap-table" id="purposeHeatmapTable">
                            <thead><tr id="heatmapHeader"><th>검사목적</th></tr></thead>
                            <tbody id="heatmapBody"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <!-- 월별 상세 테이블 -->
            <div class="card">
                <div class="card-header">
                    <div class="card-title">📋 월별 상세</div>
                    <div class="card-badge" id="monthlyTableBadge">12개월</div>
                </div>
                <div class="card-body">
                    <div class="scroll-table">
                        <table class="data-table" id="monthlyDetailTable">
                            <thead><tr><th>월</th><th class="text-right">매출액</th><th class="text-right">건수</th><th class="text-right">평균단가</th><th class="text-right">비중</th><th>상세</th></tr></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- 월 상세 모달 -->
        <div id="monthModal" class="modal" style="display: none;">
            <div class="modal-content" style="max-width: 900px;">
                <div class="modal-header">
                    <span class="modal-title" id="monthModalTitle">월 상세</span>
                    <span class="modal-close" onclick="closeMonthModal()">&times;</span>
                </div>
                <div class="modal-body">
                    <div class="modal-grid" style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                        <div>
                            <h4 style="margin-bottom: 10px;">검사목적별 구성</h4>
                            <div style="height: 250px;"><canvas id="monthPurposeChart"></canvas></div>
                        </div>
                        <div>
                            <h4 style="margin-bottom: 10px;">담당자별 구성</h4>
                            <div style="height: 250px;"><canvas id="monthManagerChart"></canvas></div>
                        </div>
                    </div>
                    <div style="margin-top: 20px;">
                        <h4 style="margin-bottom: 10px;">주요 지표</h4>
                        <div id="monthDetailStats" style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px;"></div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 업체별 탭 -->
        <div id="client" class="tab-content">
            <!-- 업체 현황 KPI 카드 -->
            <section class="kpi-section client-kpi-section" style="grid-template-columns: repeat(5, 1fr);">
                <div class="kpi-card sales" style="border-top: 4px solid var(--primary);">
                    <div class="kpi-header"><div class="kpi-icon">🏢</div></div>
                    <div class="kpi-label">총 거래업체</div>
                    <div class="kpi-value" id="clientTotalCount">-</div>
                    <div class="kpi-compare" id="clientTotalCompare">전년: -</div>
                </div>
                <div class="kpi-card count" style="border-top: 4px solid var(--success);">
                    <div class="kpi-header"><div class="kpi-icon">🆕</div></div>
                    <div class="kpi-label">신규 업체</div>
                    <div class="kpi-value" id="clientNewCount" style="color: var(--success);">-</div>
                    <div class="kpi-compare">올해 첫 거래</div>
                </div>
                <div class="kpi-card price" style="border-top: 4px solid var(--primary);">
                    <div class="kpi-header"><div class="kpi-icon">🔄</div></div>
                    <div class="kpi-label">유지 업체</div>
                    <div class="kpi-value" id="clientRetainedCount">-</div>
                    <div class="kpi-compare">전년부터 지속</div>
                </div>
                <div class="kpi-card goal" style="border-top: 4px solid var(--danger);">
                    <div class="kpi-header"><div class="kpi-icon">📤</div></div>
                    <div class="kpi-label">이탈 업체</div>
                    <div class="kpi-value" id="clientChurnedCount" style="color: var(--danger);">-</div>
                    <div class="kpi-compare">올해 거래 없음</div>
                </div>
                <div class="kpi-card" style="border-top: 4px solid var(--warning);">
                    <div class="kpi-header"><div class="kpi-icon">⭐</div></div>
                    <div class="kpi-label">VIP 업체</div>
                    <div class="kpi-value" id="clientVipCount" style="color: var(--warning);">-</div>
                    <div class="kpi-compare">매출 1억 이상</div>
                </div>
            </section>

            <!-- 담당자 영업력 KPI -->
            <section class="manager-kpi-section" style="margin-bottom: 24px;">
                <div class="section-title-bar" style="margin-bottom: 12px;">
                    <div class="section-title">🎯 담당자 영업력 (업체 확보/성장)</div>
                </div>
                <div class="manager-kpi-grid" style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px;">
                    <div class="manager-kpi-card" id="kpiClientKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #6366f1; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">👑</div>
                        <div style="font-size: 13px; color: #64748b;">업체 보유왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiClientKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiClientKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiClientKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiNewKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #10b981; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">🌱</div>
                        <div style="font-size: 13px; color: #64748b;">신규 확보왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiNewKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiNewKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiNewKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiGrowthKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #8b5cf6; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">💰</div>
                        <div style="font-size: 13px; color: #64748b;">성장 기여왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiGrowthKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiGrowthKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiGrowthKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiVipKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #f59e0b; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">💎</div>
                        <div style="font-size: 13px; color: #64748b;">VIP 확보왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiVipKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiVipKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiVipKingOverlay" style="display: none;"></div>
                    </div>
                </div>
            </section>

            <!-- 담당자 관리력 KPI -->
            <section class="manager-kpi-section" style="margin-bottom: 24px;">
                <div class="section-title-bar" style="margin-bottom: 12px;">
                    <div class="section-title">🤝 담당자 관리력 (유지/활성화)</div>
                </div>
                <div class="manager-kpi-grid" style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px;">
                    <div class="manager-kpi-card" id="kpiRetentionKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #06b6d4; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">🔄</div>
                        <div style="font-size: 13px; color: #64748b;">유지율 TOP</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiRetentionKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiRetentionKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiRetentionKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiSteadyKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #14b8a6; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">📅</div>
                        <div style="font-size: 13px; color: #64748b;">꾸준 거래왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiSteadyKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiSteadyKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiSteadyKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiActiveKing" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #ec4899; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">🔥</div>
                        <div style="font-size: 13px; color: #64748b;">활성 관리왕</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiActiveKingName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiActiveKingValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiActiveKingOverlay" style="display: none;"></div>
                    </div>
                    <div class="manager-kpi-card" id="kpiChurnWarning" style="background: white; border-radius: 12px; padding: 16px; border-left: 4px solid #ef4444; cursor: pointer; position: relative;">
                        <div style="font-size: 24px; margin-bottom: 8px;">⚠️</div>
                        <div style="font-size: 13px; color: #64748b;">이탈 주의</div>
                        <div style="font-size: 18px; font-weight: 700; color: #1e293b;" id="kpiChurnWarningName">-</div>
                        <div style="font-size: 12px; color: #94a3b8;" id="kpiChurnWarningValue">-</div>
                        <div class="manager-kpi-overlay" id="kpiChurnWarningOverlay" style="display: none;"></div>
                    </div>
                </div>
            </section>

            <!-- 매출/건수 TOP 10 차트 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🏆 매출 TOP 10</div>
                        <div class="card-badge" id="clientSalesChartBadge">-</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 350px;"><canvas id="clientSalesChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 건수 TOP 10</div>
                        <div class="card-badge" id="clientCountChartBadge">-</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 350px;"><canvas id="clientCountChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 유지 거래처 / 신규&이탈 테이블 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🔄 유지 거래처 (전년 대비 성장)</div>
                        <div class="card-badge" id="retainedTableBadge">0개</div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 300px;">
                            <table class="data-table" id="retainedClientTable">
                                <thead><tr><th>업체명</th><th>담당자</th><th class="text-right">올해</th><th class="text-right">전년</th><th class="text-right">증감</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title" id="newChurnTableTitle">🆕 신규 업체</div>
                        <div style="display: flex; gap: 8px;">
                            <button class="filter-btn active" id="btnNewClients" onclick="setClientTableMode('new')">신규 <span id="newClientsBtnCount">0</span>개</button>
                            <button class="filter-btn" id="btnChurnedClients" onclick="setClientTableMode('churned')">이탈 <span id="churnedClientsBtnCount">0</span>개</button>
                        </div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 300px;">
                            <table class="data-table" id="newChurnClientTable">
                                <thead id="newChurnTableHead"><tr><th>업체명</th><th>담당자</th><th class="text-right">매출액</th><th class="text-right">건수</th><th>주요 검사</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- 검사목적별 / 담당자별 거래처 현황 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📋 검사목적별 거래처 현황</div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 300px;">
                            <table class="data-table" id="clientByPurposeTable">
                                <thead><tr><th>검사목적</th><th class="text-right">업체수</th><th class="text-right">총매출</th><th class="text-right">평균매출</th><th>주요업체</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">👤 담당자별 거래처 현황</div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 300px;">
                            <table class="data-table" id="clientByManagerTable">
                                <thead><tr><th>담당자</th><th class="text-right">업체수</th><th class="text-right">신규</th><th class="text-right">유지</th><th class="text-right">이탈</th><th class="text-right">총매출</th><th class="text-right">성장률</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 지역별 탭 -->
        <div id="region" class="tab-content">
            <!-- 지역 KPI 카드 -->
            <section class="kpi-section region-kpi-section" style="grid-template-columns: repeat(4, 1fr); margin-bottom: 24px;">
                <div class="kpi-card region-kpi" id="kpiMainRegion" style="border-top: 4px solid var(--primary); position: relative; cursor: pointer;">
                    <div class="kpi-header"><div class="kpi-icon">🏆</div></div>
                    <div class="kpi-label">주력 지역</div>
                    <div class="kpi-value" id="mainRegionName" style="font-size: 20px;">-</div>
                    <div class="kpi-compare" id="mainRegionValue">-</div>
                    <div class="region-kpi-overlay" id="mainRegionOverlay" style="display: none;"></div>
                </div>
                <div class="kpi-card region-kpi" id="kpiGrowthRegion" style="border-top: 4px solid var(--success); position: relative; cursor: pointer;">
                    <div class="kpi-header"><div class="kpi-icon">📈</div></div>
                    <div class="kpi-label">성장 지역</div>
                    <div class="kpi-value" id="growthRegionName" style="font-size: 20px; color: var(--success);">-</div>
                    <div class="kpi-compare" id="growthRegionValue">-</div>
                    <div class="region-kpi-overlay" id="growthRegionOverlay" style="display: none;"></div>
                </div>
                <div class="kpi-card region-kpi" id="kpiNewRegion" style="border-top: 4px solid var(--info); position: relative; cursor: pointer;">
                    <div class="kpi-header"><div class="kpi-icon">🆕</div></div>
                    <div class="kpi-label">신규 진출</div>
                    <div class="kpi-value" id="newRegionName" style="font-size: 20px; color: var(--info);">-</div>
                    <div class="kpi-compare" id="newRegionValue">-</div>
                    <div class="region-kpi-overlay" id="newRegionOverlay" style="display: none;"></div>
                </div>
                <div class="kpi-card region-kpi" id="kpiWeakRegion" style="border-top: 4px solid var(--warning); position: relative; cursor: pointer;">
                    <div class="kpi-header"><div class="kpi-icon">🎯</div></div>
                    <div class="kpi-label">공략 필요</div>
                    <div class="kpi-value" id="weakRegionName" style="font-size: 20px; color: var(--warning);">-</div>
                    <div class="kpi-compare" id="weakRegionValue">-</div>
                    <div class="region-kpi-overlay" id="weakRegionOverlay" style="display: none;"></div>
                </div>
            </section>

            <!-- 지도 + 상세 패널 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card" style="min-height: 500px;">
                    <div class="card-header">
                        <div class="card-title">🗺️ 전국 지역별 매출 현황</div>
                        <div style="display: flex; gap: 8px;">
                            <span style="font-size: 11px; color: #94a3b8;">클릭하면 상세 정보</span>
                        </div>
                    </div>
                    <div class="card-body" style="display: flex; justify-content: center; align-items: center;">
                        <div id="koreaMapContainer" style="width: 100%; max-width: 450px; position: relative;">
                            <!-- SVG Korea Map -->
                            <svg id="koreaMap" viewBox="0 0 400 550" style="width: 100%; height: auto;">
                                <!-- 강원 -->
                                <path id="map-강원" d="M250,80 L320,60 L350,100 L340,160 L290,180 L240,160 L230,120 Z"
                                    class="region-path" data-region="강원"/>
                                <!-- 경기 -->
                                <path id="map-경기" d="M160,100 L230,120 L240,160 L220,200 L170,210 L130,180 L140,130 Z"
                                    class="region-path" data-region="경기"/>
                                <!-- 서울 -->
                                <path id="map-서울" d="M170,140 L200,135 L205,165 L175,170 Z"
                                    class="region-path" data-region="서울"/>
                                <!-- 인천 -->
                                <path id="map-인천" d="M120,140 L145,135 L150,170 L125,175 Z"
                                    class="region-path" data-region="인천"/>
                                <!-- 충북 -->
                                <path id="map-충북" d="M220,200 L290,180 L300,230 L260,270 L200,260 L190,220 Z"
                                    class="region-path" data-region="충북"/>
                                <!-- 세종 -->
                                <path id="map-세종" d="M165,235 L185,230 L190,255 L170,260 Z"
                                    class="region-path" data-region="세종"/>
                                <!-- 대전 -->
                                <path id="map-대전" d="M185,265 L210,260 L215,290 L190,295 Z"
                                    class="region-path" data-region="대전"/>
                                <!-- 충남 -->
                                <path id="map-충남" d="M100,200 L170,210 L190,220 L200,260 L170,290 L100,280 L80,240 Z"
                                    class="region-path" data-region="충남"/>
                                <!-- 전북 -->
                                <path id="map-전북" d="M100,290 L180,295 L200,340 L160,380 L90,360 L70,320 Z"
                                    class="region-path" data-region="전북"/>
                                <!-- 경북 -->
                                <path id="map-경북" d="M260,270 L300,230 L360,250 L370,330 L310,370 L250,350 L240,300 Z"
                                    class="region-path" data-region="경북"/>
                                <!-- 대구 -->
                                <path id="map-대구" d="M275,340 L305,335 L310,365 L280,370 Z"
                                    class="region-path" data-region="대구"/>
                                <!-- 울산 -->
                                <path id="map-울산" d="M340,380 L370,375 L375,410 L345,415 Z"
                                    class="region-path" data-region="울산"/>
                                <!-- 경남 -->
                                <path id="map-경남" d="M200,380 L250,350 L310,370 L340,420 L280,460 L200,440 L180,400 Z"
                                    class="region-path" data-region="경남"/>
                                <!-- 부산 -->
                                <path id="map-부산" d="M300,450 L340,440 L355,480 L310,490 Z"
                                    class="region-path" data-region="부산"/>
                                <!-- 광주 -->
                                <path id="map-광주" d="M105,385 L135,380 L140,410 L110,415 Z"
                                    class="region-path" data-region="광주"/>
                                <!-- 전남 -->
                                <path id="map-전남" d="M60,360 L160,380 L180,430 L150,480 L60,470 L40,420 Z"
                                    class="region-path" data-region="전남"/>
                                <!-- 제주 -->
                                <path id="map-제주" d="M60,520 L150,515 L155,545 L55,550 Z"
                                    class="region-path" data-region="제주"/>

                                <!-- Region Labels -->
                                <text x="290" y="120" class="map-label" data-region="강원">강원</text>
                                <text x="185" y="165" class="map-label" data-region="경기">경기</text>
                                <text x="183" y="155" class="map-label small" data-region="서울">서울</text>
                                <text x="130" y="160" class="map-label small" data-region="인천">인천</text>
                                <text x="245" y="230" class="map-label" data-region="충북">충북</text>
                                <text x="173" y="250" class="map-label small" data-region="세종">세종</text>
                                <text x="195" y="282" class="map-label small" data-region="대전">대전</text>
                                <text x="130" y="250" class="map-label" data-region="충남">충남</text>
                                <text x="130" y="340" class="map-label" data-region="전북">전북</text>
                                <text x="305" y="300" class="map-label" data-region="경북">경북</text>
                                <text x="288" y="357" class="map-label small" data-region="대구">대구</text>
                                <text x="352" y="400" class="map-label small" data-region="울산">울산</text>
                                <text x="255" y="410" class="map-label" data-region="경남">경남</text>
                                <text x="318" y="472" class="map-label small" data-region="부산">부산</text>
                                <text x="117" y="402" class="map-label small" data-region="광주">광주</text>
                                <text x="100" y="430" class="map-label" data-region="전남">전남</text>
                                <text x="100" y="535" class="map-label" data-region="제주">제주</text>
                            </svg>
                            <!-- 범례 -->
                            <div id="mapLegend" style="position: absolute; bottom: 10px; right: 10px; background: white; padding: 10px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); font-size: 11px;">
                                <div style="font-weight: 600; margin-bottom: 6px;">매출 규모</div>
                                <div style="display: flex; align-items: center; gap: 4px; margin-bottom: 4px;">
                                    <div style="width: 16px; height: 16px; background: #1e3a8a; border-radius: 3px;"></div>
                                    <span>10억 이상</span>
                                </div>
                                <div style="display: flex; align-items: center; gap: 4px; margin-bottom: 4px;">
                                    <div style="width: 16px; height: 16px; background: #3b82f6; border-radius: 3px;"></div>
                                    <span>5억 ~ 10억</span>
                                </div>
                                <div style="display: flex; align-items: center; gap: 4px; margin-bottom: 4px;">
                                    <div style="width: 16px; height: 16px; background: #93c5fd; border-radius: 3px;"></div>
                                    <span>1억 ~ 5억</span>
                                </div>
                                <div style="display: flex; align-items: center; gap: 4px;">
                                    <div style="width: 16px; height: 16px; background: #dbeafe; border-radius: 3px;"></div>
                                    <span>1억 미만</span>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                <div class="card" id="regionDetailCard" style="min-height: 500px;">
                    <div class="card-header">
                        <div class="card-title" id="regionDetailTitle">📍 지역 상세 정보</div>
                        <div class="card-badge" id="regionDetailBadge">지역 선택</div>
                    </div>
                    <div class="card-body" id="regionDetailBody" style="padding: 16px;">
                        <div style="text-align: center; color: #94a3b8; padding: 60px 20px;">
                            <div style="font-size: 48px; margin-bottom: 16px;">🗺️</div>
                            <div style="font-size: 14px;">좌측 지도에서 지역을 클릭하면<br>상세 정보가 표시됩니다.</div>
                        </div>
                    </div>
                </div>
            </div>

            <!-- 매출 차트 + 성장률 차트 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📊 지역별 매출 순위</div>
                        <div class="card-badge" id="regionSalesChartBadge">-</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 350px;"><canvas id="regionSalesChart"></canvas></div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">📈 지역별 성장률</div>
                        <div class="card-badge" id="regionGrowthChartBadge">전년 대비</div>
                    </div>
                    <div class="card-body">
                        <div class="chart-container" style="height: 350px;"><canvas id="regionGrowthChart"></canvas></div>
                    </div>
                </div>
            </div>

            <!-- 히트맵 테이블 + 지역 TOP 업체 -->
            <div class="content-grid" style="margin-bottom: 24px;">
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🌡️ 지역별 현황 히트맵</div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 350px;">
                            <table class="data-table" id="regionHeatmapTable">
                                <thead><tr><th>지역</th><th class="text-right">매출액</th><th class="text-right">건수</th><th class="text-right">성장률</th><th>비중</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <div class="card-header">
                        <div class="card-title">🏢 지역별 TOP 업체</div>
                    </div>
                    <div class="card-body">
                        <div class="scroll-table" style="max-height: 350px;">
                            <table class="data-table" id="regionTopClientTable">
                                <thead><tr><th>지역</th><th>업체명</th><th class="text-right">매출액</th><th>담당자</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- 담당자별 지역 분포 -->
            <div class="card" style="margin-bottom: 24px;">
                <div class="card-header">
                    <div class="card-title">👤 담당자별 지역 분포</div>
                    <div class="card-badge" id="managerRegionBadge">-</div>
                </div>
                <div class="card-body">
                    <div class="scroll-table" style="max-height: 350px;">
                        <table class="data-table" id="managerRegionTable">
                            <thead><tr><th>담당자</th><th>주력 지역</th><th class="text-right">지역수</th><th class="text-right">총매출</th><th>지역 분포</th></tr></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- 목적별 탭 -->
        <div id="purpose" class="tab-content">
            <div class="content-grid">
                <div class="card">
                    <div class="card-header"><div class="card-title">📊 목적별 월별 추이</div></div>
                    <div class="card-body"><div class="chart-container"><canvas id="purposeMonthlyChart"></canvas></div></div>
                </div>
                <div class="card">
                    <div class="card-header"><div class="card-title">📋 목적별 상세</div></div>
                    <div class="card-body">
                        <div class="scroll-table">
                            <table class="data-table" id="purposeTable">
                                <thead><tr><th>검사목적</th><th class="text-right">매출액</th><th class="text-right">건수</th><th>비중</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 유형별 탭 -->
        <div id="sampleType" class="tab-content">
            <section class="purpose-kpi-section">
                <div class="section-title-bar">
                    <div class="section-title">🧪 검체 유형별 현황</div>
                    <div class="section-badge" id="sampleTypeCount">0개 유형</div>
                </div>
                <div class="purpose-kpi-grid" id="sampleTypeGrid"></div>
            </section>
            <div class="content-grid">
                <div class="card">
                    <div class="card-header"><div class="card-title">🥧 유형별 매출 비중</div></div>
                    <div class="card-body"><div class="chart-container"><canvas id="sampleTypeChart"></canvas></div></div>
                </div>
                <div class="card">
                    <div class="card-header"><div class="card-title">📋 유형별 상세</div></div>
                    <div class="card-body">
                        <div class="scroll-table">
                            <table class="data-table" id="sampleTypeTable">
                                <thead><tr><th>검체유형</th><th class="text-right">매출액</th><th class="text-right">건수</th><th>비중</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 부적합 탭 -->
        <div id="defect" class="tab-content">
            <div class="content-grid">
                <div class="card">
                    <div class="card-header"><div class="card-title">📊 부적합 항목별 현황</div></div>
                    <div class="card-body"><div class="chart-container"><canvas id="defectChart"></canvas></div></div>
                </div>
                <div class="card">
                    <div class="card-header"><div class="card-title">📋 부적합 상세</div></div>
                    <div class="card-body">
                        <div class="scroll-table">
                            <table class="data-table" id="defectTable">
                                <thead><tr><th>부적합항목</th><th class="text-right">건수</th><th>비중</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 검사항목 탭 -->
        <div id="foodItem" class="tab-content">
            <div class="content-grid">
                <div class="card">
                    <div class="card-header"><div class="card-title">📊 검사항목 TOP 15</div></div>
                    <div class="card-body"><div class="chart-container"><canvas id="foodItemChart"></canvas></div></div>
                </div>
                <div class="card">
                    <div class="card-header"><div class="card-title">📋 검사항목별 상세</div><div class="card-badge" id="foodItemTableBadge">0개</div></div>
                    <div class="card-body">
                        <div class="scroll-table">
                            <table class="data-table" id="foodItemTable">
                                <thead><tr><th>항목명</th><th class="text-right">매출액</th><th class="text-right">건수</th><th>비중</th></tr></thead>
                                <tbody></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- AI 분석 탭 -->
        <div id="aiAnalysis" class="tab-content">
            <section class="ai-section">
                <div class="ai-header">
                    <h2>🤖 AI 데이터 분석</h2>
                    <p>자연어로 질문하면 데이터를 분석해드립니다.</p>
                </div>
                <div class="ai-input-container">
                    <div class="ai-input-wrapper">
                        <input type="text" id="aiQueryInput" class="ai-input" placeholder="예: 2025년 1월과 2024년 1월 매출 비교해줘">
                        <button onclick="runAiAnalysis()" class="ai-btn" id="aiBtn">분석하기</button>
                    </div>
                    <div class="ai-examples">
                        <span class="ai-example" onclick="setAiQuery('월별 매출')">월별 매출</span>
                        <span class="ai-example" onclick="setAiQuery('2025년 vs 2024년 비교')">연도 비교</span>
                        <span class="ai-example" onclick="setAiQuery('영업담당 TOP 10')">TOP 담당자</span>
                    </div>
                    <div class="ai-result" id="aiResult">
                        <div id="aiLoading" style="text-align: center; display: none;">⏳ AI가 분석 중입니다...</div>
                        <div id="aiError" style="color: var(--danger); display: none;"></div>
                        <div id="aiContent"></div>
                    </div>
                </div>
            </section>
        </div>

        <!-- 기업 정보 탭 -->
        <div id="companyInfo" class="tab-content">
            <div class="card">
                <div class="card-header"><div class="card-title">🏛️ 기업 정보 관리</div></div>
                <div class="card-body">
                    <p style="color: var(--gray-500); text-align: center; padding: 40px;">기업 정보 관리 기능</p>
                </div>
            </div>
        </div>

        <!-- 터미널 탭 -->
        <div id="webTerminal" class="tab-content">
            <div class="card" style="background: #1e293b;">
                <div class="card-header" style="border-color: #334155;"><div class="card-title" style="color: #10b981;">💻 웹 터미널</div></div>
                <div class="card-body">
                    <div style="background: #0f172a; padding: 20px; border-radius: 12px; font-family: monospace; color: #10b981; min-height: 300px;">
                        <p>$ 터미널 기능</p>
                    </div>
                </div>
            </div>
        </div>
    </main>

    <script>
        // 전역 변수
        let charts = {};
        let currentData = null;
        let compareData = null;
        let currentTab = 'main';
        let managerTableSort = { column: null, direction: 'desc' };

        // 유틸리티 함수
        function formatCurrency(value) {
            const sign = value < 0 ? '-' : '';
            const absValue = Math.abs(value);
            if (absValue >= 100000000) return sign + (absValue/100000000).toFixed(1) + '억';
            if (absValue >= 10000) return sign + (absValue/10000).toFixed(0) + '만';
            return Math.round(value).toLocaleString();
        }

        function showToast(message, type = 'success', duration = 3000) {
            const toast = document.getElementById('toast');
            toast.textContent = message;
            toast.className = 'toast ' + type;
            toast.style.display = 'block';
            if (type !== 'loading') setTimeout(() => toast.style.display = 'none', duration);
        }

        function hideToast() { document.getElementById('toast').style.display = 'none'; }

        // 탭 전환
        function showTab(tabId) {
            currentTab = tabId;
            document.querySelectorAll('.tab-card').forEach(c => c.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            const card = document.querySelector(`.tab-card[onclick="showTab('${tabId}')"]`);
            if (card) card.classList.add('active');
            const content = document.getElementById(tabId);
            if (content) content.classList.add('active');
            document.getElementById('kpiSection').classList.toggle('hidden', tabId !== 'main');
        }

        // 비교 체크박스
        document.getElementById('compareCheck').addEventListener('change', function() {
            document.getElementById('compareYearGroup').style.display = this.checked ? 'flex' : 'none';
        });

        // 토큰 사용량 로드
        async function loadTokenUsage() {
            try {
                const res = await fetch('/api/token-usage');
                const data = await res.json();
                if (data.this_month) {
                    document.getElementById('thisMonthTokens').textContent = data.this_month.total_tokens.toLocaleString();
                    document.getElementById('thisMonthKRW').textContent = Math.round(data.this_month.total_cost_krw).toLocaleString();
                }
                if (data.last_month) {
                    document.getElementById('lastMonthTokens').textContent = data.last_month.total_tokens.toLocaleString();
                    document.getElementById('lastMonthKRW').textContent = Math.round(data.last_month.total_cost_krw).toLocaleString();
                }
            } catch (e) { console.log('토큰 로드 실패'); }
        }

        // 데이터 로드 (실제 API 호출)
        async function loadData() {
            const btn = document.getElementById('btnSearch');
            btn.disabled = true;
            btn.innerHTML = '⏳ 로딩중...';
            showToast('데이터를 불러오는 중...', 'loading');

            try {
                const year = document.getElementById('yearSelect').value;
                const month = document.getElementById('monthSelect').value;
                const purpose = document.getElementById('purposeSelect').value;
                const compareCheck = document.getElementById('compareCheck').checked;
                const compareYear = document.getElementById('compareYearSelect').value;

                let url = `/api/data?year=${year}`;
                if (month) url += `&month=${month}`;
                if (purpose !== '전체') url += `&purpose=${encodeURIComponent(purpose)}`;

                const res = await fetch(url);
                currentData = await res.json();
                currentData.year = year;

                // 비교 데이터 로드
                if (compareCheck) {
                    let compUrl = `/api/data?year=${compareYear}`;
                    if (month) compUrl += `&month=${month}`;
                    if (purpose !== '전체') compUrl += `&purpose=${encodeURIComponent(purpose)}`;
                    const compRes = await fetch(compUrl);
                    compareData = await compRes.json();
                    compareData.year = compareYear;
                } else {
                    compareData = null;
                }

                updateAll();
                hideToast();
                showToast(`${year}년 데이터 로드 완료`, 'success');
            } catch (e) {
                hideToast();
                showToast('데이터 로드 실패: ' + e.message, 'error');
            }

            btn.disabled = false;
            btn.innerHTML = '🔍 조회하기';
        }

        function updateAll() {
            updateSummary();
            updatePurposeGrid();
            updateDepartmentCards();  // 부서별 카드 업데이트
            updatePersonalTab();  // 개인별 탭 전체 업데이트
            updateTeamTab();      // 팀별 탭 전체 업데이트
            updateManagerChart();
            updateBranchChart();
            updateMonthlyTab();   // 월별 탭 전체 업데이트
            updateManagerTable();
            updateBranchTable();
            updateClientTab();
            updateRegionTab();
            updateSampleTypeTab();
            updateDefectTab();
            updatePurposeTab();
        }

        function updateSummary() {
            const totalSales = currentData.total_sales || 0;
            const totalCount = currentData.total_count || 0;
            const avgPrice = totalCount > 0 ? totalSales / totalCount : 0;
            const goalTarget = 7000000000;
            const goalRate = ((totalSales / goalTarget) * 100).toFixed(1);

            document.getElementById('totalSales').textContent = formatCurrency(totalSales);
            document.getElementById('totalCount').textContent = totalCount.toLocaleString() + '건';
            document.getElementById('avgPrice').textContent = formatCurrency(avgPrice);
            document.getElementById('goalRate').textContent = goalRate + '%';
            document.getElementById('managerChartBadge').textContent = currentData.year + '년';

            if (compareData) {
                const compSales = compareData.total_sales || 0;
                const compCount = compareData.total_count || 0;
                const compAvg = compCount > 0 ? compSales / compCount : 0;

                const salesDiff = compSales > 0 ? ((totalSales - compSales) / compSales * 100).toFixed(1) : 0;
                const countDiff = compCount > 0 ? ((totalCount - compCount) / compCount * 100).toFixed(1) : 0;
                const priceDiff = compAvg > 0 ? ((avgPrice - compAvg) / compAvg * 100).toFixed(1) : 0;

                updateTrendBadge('salesTrend', salesDiff);
                updateTrendBadge('countTrend', countDiff);
                updateTrendBadge('priceTrend', priceDiff);

                document.getElementById('compareTotalSales').innerHTML = `${compareData.year}년: <span>${formatCurrency(compSales)}</span>`;
                document.getElementById('compareTotalSales').style.display = 'block';
                document.getElementById('compareTotalCount').innerHTML = `${compareData.year}년: <span>${compCount.toLocaleString()}건</span>`;
                document.getElementById('compareTotalCount').style.display = 'block';
                document.getElementById('compareAvgPrice').innerHTML = `${compareData.year}년: <span>${formatCurrency(compAvg)}</span>`;
                document.getElementById('compareAvgPrice').style.display = 'block';

                // KPI 카드 호버 오버레이 업데이트
                const salesUp = parseFloat(salesDiff) >= 0;
                document.getElementById('salesOverlay').innerHTML = `
                    <div class="overlay-year-badge">${compareData.year}년</div>
                    <div class="overlay-label">전년도 총 매출</div>
                    <div class="overlay-value">${formatCurrency(compSales)}</div>
                    <div class="overlay-change">증감: <span class="${salesUp ? 'up' : 'down'}">${salesUp ? '+' : ''}${salesDiff}%</span></div>
                `;

                const countUp = parseFloat(countDiff) >= 0;
                document.getElementById('countOverlay').innerHTML = `
                    <div class="overlay-year-badge">${compareData.year}년</div>
                    <div class="overlay-label">전년도 총 건수</div>
                    <div class="overlay-value">${compCount.toLocaleString()}건</div>
                    <div class="overlay-change">증감: <span class="${countUp ? 'up' : 'down'}">${countUp ? '+' : ''}${countDiff}%</span></div>
                `;

                const priceUp = parseFloat(priceDiff) >= 0;
                document.getElementById('priceOverlay').innerHTML = `
                    <div class="overlay-year-badge">${compareData.year}년</div>
                    <div class="overlay-label">전년도 평균 단가</div>
                    <div class="overlay-value">${formatCurrency(compAvg)}</div>
                    <div class="overlay-change">증감: <span class="${priceUp ? 'up' : 'down'}">${priceUp ? '+' : ''}${priceDiff}%</span></div>
                `;

                // 목표달성률에 전년대비 성장률 표시
                document.getElementById('goalCompare').innerHTML = `성장률: <span class="${salesUp ? 'up' : 'down'}" style="color: var(--${salesUp ? 'success' : 'danger'}); font-weight: 600;">${salesUp ? '+' : ''}${salesDiff}%</span>`;
                document.getElementById('goalCompare').style.display = 'block';

                // 목표달성률 오버레이 (증감률 크게, 금액 차이 표시)
                const salesDiffAmount = totalSales - compSales;
                const diffAmountStr = (salesDiffAmount >= 0 ? '+' : '') + formatCurrency(salesDiffAmount);
                document.getElementById('goalOverlay').innerHTML = `
                    <div class="overlay-year-badge">${compareData.year}년 대비</div>
                    <div class="overlay-label">전년대비 성장률</div>
                    <div class="overlay-value" style="color: var(--${salesUp ? 'success' : 'danger'});">${salesUp ? '+' : ''}${salesDiff}%</div>
                    <div class="overlay-sub">전년 매출: ${formatCurrency(compSales)}</div>
                    <div class="overlay-change">차이: <span class="${salesUp ? 'up' : 'down'}">${diffAmountStr}</span></div>
                `;
            } else {
                ['compareTotalSales', 'compareTotalCount', 'compareAvgPrice', 'goalCompare'].forEach(id => {
                    const el = document.getElementById(id);
                    if (el) el.style.display = 'none';
                });
                ['salesTrend', 'countTrend', 'priceTrend'].forEach(id => document.getElementById(id).style.visibility = 'hidden');
                // 오버레이 비우기
                ['salesOverlay', 'countOverlay', 'priceOverlay', 'goalOverlay'].forEach(id => document.getElementById(id).innerHTML = '');
            }
        }

        function updateTrendBadge(id, diff) {
            const el = document.getElementById(id);
            el.style.visibility = 'visible';
            const isUp = parseFloat(diff) >= 0;
            el.className = 'kpi-trend ' + (isUp ? 'up' : 'down');
            el.innerHTML = `<span>${isUp ? '↑' : '↓'} ${isUp ? '+' : ''}${diff}%</span>`;
        }

        // 부서별 카드 업데이트
        function updateDepartmentCards() {
            const dept = currentData.by_department || {};
            const totalSales = currentData.total_sales || 1;
            const compareDept = compareData ? (compareData.by_department || {}) : {};

            // 부서별 데이터 매핑
            const deptMapping = [
                { key: '본사', prefix: 'Bonsa' },
                { key: '마케팅', prefix: 'Marketing' },
                { key: '영업부', prefix: 'Sales' },
                { key: '지사', prefix: 'Branch' }
            ];

            deptMapping.forEach(({ key, prefix }) => {
                const data = dept[key] || { sales: 0, count: 0 };
                const sales = data.sales || 0;
                const count = data.count || 0;
                const avg = count > 0 ? sales / count : 0;
                const ratio = totalSales > 0 ? (sales / totalSales * 100) : 0;

                // 값 업데이트
                const salesEl = document.getElementById(`dept${prefix}Sales`);
                const countEl = document.getElementById(`dept${prefix}Count`);
                const avgEl = document.getElementById(`dept${prefix}Avg`);
                const ratioEl = document.getElementById(`dept${prefix}Ratio`);

                if (salesEl) salesEl.textContent = formatCurrency(sales);
                if (countEl) countEl.textContent = count.toLocaleString() + '건';
                if (avgEl) avgEl.textContent = formatCurrency(avg);
                if (ratioEl) ratioEl.textContent = ratio.toFixed(1) + '%';

                // 오버레이 (전년 대비)
                const overlayEl = document.getElementById(`dept${prefix}Overlay`);

                if (compareData && compareDept[key]) {
                    const compSales = compareDept[key].sales || 0;
                    const compCount = compareDept[key].count || 0;
                    if (compSales > 0) {
                        const growth = ((sales - compSales) / compSales * 100).toFixed(1);
                        const isPositive = parseFloat(growth) >= 0;
                        const diff = sales - compSales;
                        if (overlayEl) {
                            overlayEl.className = 'dept-overlay active';
                            overlayEl.innerHTML = `
                                <div class="overlay-title">전년 대비</div>
                                <div class="overlay-value ${isPositive ? 'positive' : 'negative'}">${isPositive ? '+' : ''}${growth}%</div>
                                <div class="overlay-detail">${compareData.year}년: ${formatCurrency(compSales)}</div>
                                <div class="overlay-detail">차이: ${isPositive ? '+' : ''}${formatCurrency(diff)}</div>
                            `;
                        }
                    } else {
                        if (overlayEl) overlayEl.className = 'dept-overlay';
                    }
                } else {
                    if (overlayEl) overlayEl.className = 'dept-overlay';
                }
            });
        }

        function updatePurposeGrid() {
            const grid = document.getElementById('purposeGrid');
            const purposes = currentData.by_purpose || [];
            const colors = ['blue', 'green', 'orange', 'purple', 'pink', 'info', 'teal', 'amber', 'rose', 'sky', 'lime', 'cyan'];
            const icons = ['📋', '🥗', '⏰', '🥜', '🧬', '📄', '⚗️', '🏷️', '📤', '🌱', '☢️', '🔬', '💊', '🌙', '🥕', '❌'];

            document.getElementById('purposeCount').textContent = purposes.length + '개 목적';

            const compareMap = compareData ? Object.fromEntries(compareData.by_purpose || []) : {};

            grid.innerHTML = purposes.map((p, i) => {
                const name = p[0], sales = p[1].sales, count = p[1].count;
                const isCancel = name === '접수취소';
                const cardColor = isCancel ? 'danger' : colors[i % colors.length];
                const salesValue = isCancel ? '-' + formatCurrency(Math.abs(sales)) : formatCurrency(sales);

                let changeHtml = '';
                let overlayHtml = '';

                if (compareData && compareMap[name]) {
                    const compSales = compareMap[name].sales || 0;
                    const compCount = compareMap[name].count || 0;
                    const compSalesValue = isCancel ? '-' + formatCurrency(Math.abs(compSales)) : formatCurrency(compSales);

                    if (Math.abs(compSales) > 0) {
                        const diff = ((sales - compSales) / Math.abs(compSales) * 100).toFixed(1);
                        const isUp = parseFloat(diff) >= 0;
                        changeHtml = `<div class="purpose-kpi-trend ${isUp ? 'up' : 'down'}">${isUp ? '↑' : '↓'} ${isUp ? '+' : ''}${diff}%</div>`;

                        // 호버 오버레이 생성 (새 구조)
                        overlayHtml = `
                            <div class="purpose-kpi-overlay">
                                <div class="overlay-year-badge">${compareData.year}년</div>
                                <div class="overlay-label">전년도 실적</div>
                                <div class="overlay-name">${name}</div>
                                <div class="overlay-value">${compSalesValue}</div>
                                <div class="overlay-sub">건수: ${compCount.toLocaleString()}건</div>
                                <div class="overlay-change">증감: <span class="${isUp ? 'up' : 'down'}">${isUp ? '+' : ''}${diff}%</span></div>
                            </div>
                        `;
                    }
                }

                return `
                    <div class="purpose-kpi-card" data-color="${cardColor}" onclick="selectPurpose('${name}')">
                        <div class="purpose-kpi-header">
                            <div class="purpose-kpi-icon">${icons[i % icons.length]}</div>
                            ${changeHtml}
                        </div>
                        <div class="purpose-kpi-name">${name}</div>
                        <div class="purpose-kpi-value">${salesValue}</div>
                        <div class="purpose-kpi-sub">건수: <span>${count.toLocaleString()}건</span></div>
                        ${overlayHtml}
                    </div>
                `;
            }).join('');
        }

        function selectPurpose(name) {
            document.getElementById('purposeSelect').value = name;
            showToast(`"${name}" 선택됨`, 'success');
        }

        // ====== 개인별 탭 관련 변수 및 함수 ======
        let managerSortColumn = 'sales';
        let managerSortOrder = 'desc';
        let perCaseSortOrder = 'desc';
        let selectedManagers = [];
        let monthlyPreset = 'all';

        function updatePersonalTab() {
            const managers = currentData.by_manager || [];
            if (managers.length === 0) return;

            const totalManagers = managers.length;
            const totalSales = managers.reduce((sum, m) => sum + (m[1].sales || 0), 0);
            const avgSales = totalSales / totalManagers;

            // KPI 카드 업데이트
            document.getElementById('personalTotalManagers').textContent = totalManagers + '명';
            document.getElementById('personalAvgSales').textContent = formatCurrency(avgSales);

            // 긴급 최고 요청자 TOP 5
            const urgentTop5 = [...managers]
                .sort((a, b) => (b[1].urgent || 0) - (a[1].urgent || 0))
                .slice(0, 5);

            document.getElementById('personalUrgentTop').innerHTML = urgentTop5.map((m, i) =>
                `<div style="display: flex; gap: 8px; align-items: center; margin-bottom: 3px;">
                    <span style="min-width: 70px;">${i + 1}. ${m[0]}</span>
                    <span style="color: var(--danger); font-weight: 600;">${m[1].urgent || 0}건</span>
                </div>`
            ).join('');

            // 최고 성장자 (전년 비교 시)
            if (compareData && compareData.by_manager) {
                const compareMap = Object.fromEntries(compareData.by_manager);
                const withGrowth = managers.map(m => {
                    const compSales = compareMap[m[0]]?.sales || 0;
                    const growth = compSales > 0 ? ((m[1].sales - compSales) / compSales * 100) : 0;
                    return { name: m[0], growth };
                }).sort((a, b) => b.growth - a.growth);

                if (withGrowth.length > 0) {
                    document.getElementById('personalTopGrowth').textContent = withGrowth[0].name;
                    document.getElementById('personalTopGrowthRate').textContent = '전년 대비 +' + withGrowth[0].growth.toFixed(1) + '%';
                    document.getElementById('topGrowthTrend').style.visibility = 'visible';
                    document.getElementById('topGrowthTrend').innerHTML = '↑ +' + withGrowth[0].growth.toFixed(1) + '%';
                }
            } else {
                document.getElementById('personalTopGrowth').textContent = '-';
                document.getElementById('personalTopGrowthRate').textContent = '전년 비교 필요';
                document.getElementById('topGrowthTrend').style.visibility = 'hidden';
            }

            // 다중 선택 체크박스 목록 생성
            initManagerMultiSelect();

            // 건당 매출/긴급 목적 드롭다운 초기화
            initPerCasePurposeSelect();
            initUrgentPurposeSelect();
            initManagerPurposeFilter();
            initManagerChartPurposeFilter();

            // 차트들 업데이트
            updateEfficiencyChart();
            updateManagerMonthlyChart();
            updatePerCaseChart();
            updateUrgentChart();
            updateUrgentMonthlyChart();
            updateUrgentUnitPriceChart();
            updateDailyClientChart();
        }

        function initManagerMultiSelect() {
            const managers = currentData.by_manager || [];
            const listEl = document.getElementById('managerSelectList');
            if (!listEl) return;
            listEl.innerHTML = managers.map(m =>
                `<label class="multi-select-item">
                    <input type="checkbox" value="${m[0]}" onchange="onManagerCheckChange()">
                    <span>${m[0]}</span>
                </label>`
            ).join('');
            selectedManagers = [];
            updateSelectedCount();
            updateSelectedTags();
        }

        function onManagerCheckChange() {
            const checkboxes = document.querySelectorAll('#managerSelectList input:checked');
            selectedManagers = Array.from(checkboxes).map(cb => cb.value);
            document.querySelectorAll('.preset-btn').forEach(btn => btn.classList.remove('active'));
            monthlyPreset = 'custom';
            updateSelectedCount();
            updateSelectedTags();
            updateManagerMonthlyChart();
        }

        function updateSelectedCount() {
            const el = document.getElementById('selectedCount');
            if (el) el.textContent = `(${selectedManagers.length})`;
        }

        function updateSelectedTags() {
            const tagsEl = document.getElementById('selectedManagerTags');
            if (!tagsEl) return;
            if (monthlyPreset === 'all') {
                tagsEl.innerHTML = '<span class="selected-tag">📊 전체 합계</span>';
            } else if (monthlyPreset === 'top3') {
                const top3 = (currentData.by_manager || []).slice(0, 3).map(m => m[0]);
                tagsEl.innerHTML = top3.map(name => `<span class="selected-tag">${name}</span>`).join('');
            } else {
                tagsEl.innerHTML = selectedManagers.map(name =>
                    `<span class="selected-tag">${name} <span class="remove" onclick="removeManager('${name}')">×</span></span>`
                ).join('');
            }
        }

        function removeManager(name) {
            selectedManagers = selectedManagers.filter(n => n !== name);
            const cb = document.querySelector(`#managerSelectList input[value="${name}"]`);
            if (cb) cb.checked = false;
            updateSelectedCount();
            updateSelectedTags();
            updateManagerMonthlyChart();
        }

        function toggleMultiSelect() {
            const list = document.getElementById('managerSelectList');
            if (list) list.style.display = list.style.display === 'none' ? 'block' : 'none';
        }

        function setMonthlyPreset(preset) {
            monthlyPreset = preset;
            document.querySelectorAll('.preset-btn').forEach(btn => btn.classList.remove('active'));
            const activeBtn = document.querySelector(`.preset-btn[onclick="setMonthlyPreset('${preset}')"]`);
            if (activeBtn) activeBtn.classList.add('active');
            document.querySelectorAll('#managerSelectList input').forEach(cb => cb.checked = false);
            selectedManagers = [];
            updateSelectedCount();
            updateSelectedTags();
            updateManagerMonthlyChart();
        }

        function initPerCasePurposeSelect() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => purposes.add(p[0]));
            const select = document.getElementById('perCasePurposeSelect');
            if (select) {
                select.innerHTML = '<option value="전체">전체 목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체' && p !== '접수취소').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function initUrgentPurposeSelect() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('urgentPurposeSelect');
            if (select) {
                select.innerHTML = '<option value="전체">검사목적: 전체</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function initManagerPurposeFilter() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('managerPurposeFilter');
            if (select) {
                select.innerHTML = '<option value="전체">전체 검사목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function togglePerCaseSort() {
            perCaseSortOrder = perCaseSortOrder === 'desc' ? 'asc' : 'desc';
            const btn = document.getElementById('perCaseSortBtn');
            if (btn) btn.textContent = perCaseSortOrder === 'desc' ? '내림차순 ▼' : '오름차순 ▲';
            updatePerCaseChart();
        }

        // 효율성 분석 산점도
        function updateEfficiencyChart() {
            const ctx = document.getElementById('efficiencyChart');
            if (!ctx) return;
            if (charts.efficiency) charts.efficiency.destroy();

            const managers = currentData.by_manager || [];
            if (managers.length === 0) return;

            const avgCount = managers.reduce((sum, m) => sum + (m[1].count || 0), 0) / managers.length;
            const avgSales = managers.reduce((sum, m) => sum + (m[1].sales || 0), 0) / managers.length;

            // 4분면별 색상 (더 명확한 구분)
            const quadrantColors = {
                q1: 'rgba(37, 99, 235, 0.85)',   // 고건수·고매출: 진한 파란색
                q2: 'rgba(6, 182, 212, 0.85)',   // 저건수·고매출: 청록색(시안)
                q3: 'rgba(249, 115, 22, 0.85)',  // 고건수·저매출: 주황색
                q4: 'rgba(220, 38, 38, 0.85)',   // 저건수·저매출: 빨간색
            };

            const data = managers.map(m => {
                const isHighCount = (m[1].count || 0) >= avgCount;
                const isHighSales = (m[1].sales || 0) >= avgSales;
                let color;
                if (isHighCount && isHighSales) color = quadrantColors.q1;
                else if (!isHighCount && isHighSales) color = quadrantColors.q2;
                else if (isHighCount && !isHighSales) color = quadrantColors.q3;
                else color = quadrantColors.q4;
                return { x: m[1].count || 0, y: m[1].sales || 0, name: m[0], color };
            });

            // 데이터셋 구성 (현재 연도)
            const datasets = [{
                label: currentData.year + '년',
                data: data.map(d => ({ x: d.x, y: d.y })),
                backgroundColor: data.map(d => d.color),
                borderColor: data.map(d => d.color.replace('0.85', '1')),
                borderWidth: 2,
                pointRadius: 12,
                pointHoverRadius: 16,
            }];

            // 전년도 비교 데이터 추가
            if (compareData && compareData.by_manager) {
                const compManagers = compareData.by_manager || [];
                const compData = compManagers.map(m => ({
                    x: m[1].count || 0,
                    y: m[1].sales || 0,
                    name: m[0]
                }));
                datasets.push({
                    label: compareData.year + '년',
                    data: compData.map(d => ({ x: d.x, y: d.y })),
                    backgroundColor: 'rgba(168, 85, 247, 0.4)',  // 보라색 (전년도)
                    borderColor: 'rgba(168, 85, 247, 0.8)',
                    borderWidth: 2,
                    pointRadius: 9,
                    pointHoverRadius: 12,
                    pointStyle: 'triangle',
                });
            }

            charts.efficiency = new Chart(ctx.getContext('2d'), {
                type: 'scatter',
                data: { datasets },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: compareData ? true : false, position: 'top' },
                        tooltip: {
                            callbacks: {
                                label: (context) => {
                                    const idx = context.dataIndex;
                                    const dsIdx = context.datasetIndex;
                                    const mgrs = dsIdx === 0 ? managers : (compareData?.by_manager || []);
                                    const m = mgrs[idx];
                                    if (!m) return '';
                                    const year = dsIdx === 0 ? currentData.year : compareData?.year;
                                    return [m[0] + ' (' + year + '년)', '매출: ' + formatCurrency(m[1].sales || 0), '건수: ' + (m[1].count || 0).toLocaleString() + '건'];
                                }
                            }
                        }
                    },
                    scales: {
                        x: { title: { display: true, text: '건수' }, grid: { color: 'rgba(0,0,0,0.05)' } },
                        y: { title: { display: true, text: '매출 (공급가액)' }, ticks: { callback: v => formatCurrency(v) }, grid: { color: 'rgba(0,0,0,0.05)' } }
                    }
                }
            });
        }

        // 담당자별 월별 매출 추이
        function updateManagerMonthlyChart() {
            const ctx = document.getElementById('managerMonthlyChart');
            if (!ctx) return;
            if (charts.managerMonthly) charts.managerMonthly.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const colors = ['#6366f1', '#10b981', '#f59e0b', '#ec4899', '#06b6d4', '#8b5cf6', '#ef4444', '#14b8a6', '#f97316', '#84cc16'];
            const managers = currentData.by_manager || [];

            if (monthlyPreset === 'all') {
                // 전체 합계 - 월별 데이터 사용 (매출, 건수, 검사목적 포함)
                const monthMap = Object.fromEntries(currentData.by_month || []);
                const monthlyInfo = labels.map((_, i) => {
                    const m = monthMap[i+1] || {};
                    return {
                        sales: m.sales || 0,
                        count: m.count || 0,
                        perCase: (m.count > 0) ? (m.sales / m.count) : 0,
                        byPurpose: m.byPurpose || {}
                    };
                });
                const totalMonthly = monthlyInfo.map(m => m.sales);
                const nonZeroSales = totalMonthly.filter(v => v > 0);
                const ownAvg = nonZeroSales.length > 0 ? nonZeroSales.reduce((a,b) => a+b, 0) / nonZeroSales.length : 0;

                // 검사목적별 월평균 계산
                const purposeAvg = {};
                const allPurposes = new Set();
                monthlyInfo.forEach(m => Object.keys(m.byPurpose).forEach(p => allPurposes.add(p)));
                allPurposes.forEach(purpose => {
                    const values = monthlyInfo.map(m => m.byPurpose[purpose]?.sales || 0);
                    const nonZero = values.filter(v => v > 0);
                    purposeAvg[purpose] = nonZero.length > 0 ? nonZero.reduce((a,b) => a+b, 0) / nonZero.length : 0;
                });

                // 데이터셋 구성
                const datasets = [{
                    label: currentData.year + '년 전체',
                    data: totalMonthly,
                    monthlyInfo,
                    ownAvg,
                    borderColor: '#6366f1',
                    backgroundColor: 'rgba(99, 102, 241, 0.2)',
                    fill: true,
                    tension: 0.4,
                    pointRadius: 8,
                    pointHoverRadius: 12,
                    pointStyle: totalMonthly.map(v => v < ownAvg ? 'triangle' : 'circle'),
                    pointBackgroundColor: totalMonthly.map(v => v < ownAvg ? '#ef4444' : '#6366f1'),
                    isComparison: false,
                }];

                // 전년도 비교 데이터 추가
                if (compareData && compareData.by_month) {
                    const compMonthMap = Object.fromEntries(compareData.by_month || []);
                    const compMonthlyInfo = labels.map((_, i) => {
                        const m = compMonthMap[i+1] || {};
                        return {
                            sales: m.sales || 0,
                            count: m.count || 0,
                            perCase: (m.count > 0) ? (m.sales / m.count) : 0,
                            byPurpose: m.byPurpose || {}
                        };
                    });
                    datasets.push({
                        label: compareData.year + '년 전체',
                        data: compMonthlyInfo.map(m => m.sales),
                        monthlyInfo: compMonthlyInfo,
                        borderColor: 'rgba(156, 163, 175, 0.8)',
                        backgroundColor: 'rgba(156, 163, 175, 0.1)',
                        fill: false,
                        tension: 0.4,
                        pointRadius: 4,
                        pointBackgroundColor: 'rgba(156, 163, 175, 0.8)',
                        borderDash: [5, 5],
                        isComparison: true,
                    });
                }

                charts.managerMonthly = new Chart(ctx.getContext('2d'), {
                    type: 'line',
                    data: { labels, datasets },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {
                            legend: { display: compareData ? true : false, position: 'top' },
                            tooltip: {
                                callbacks: {
                                    label: function(context) {
                                        const ds = context.dataset;
                                        const monthIdx = context.dataIndex;
                                        const info = ds.monthlyInfo?.[monthIdx];
                                        if (!info) return ds.label + ': ' + formatCurrency(context.raw);

                                        let result = [
                                            ds.label + ': ' + formatCurrency(info.sales),
                                            '  건수: ' + info.count.toLocaleString() + '건',
                                            '  건당: ' + formatCurrency(info.perCase)
                                        ];

                                        // 자체 월평균 대비 및 검사목적별 증감 (현재 연도만)
                                        if (!ds.isComparison && ds.ownAvg) {
                                            const avg = ds.ownAvg;
                                            const diff = info.sales - avg;
                                            const diffPct = avg > 0 ? ((diff / avg) * 100).toFixed(1) : 0;

                                            result.push('─────────');
                                            result.push(`월평균: ${formatCurrency(avg)}`);
                                            if (diff >= 0) {
                                                result.push(`📈 월평균 대비 +${diffPct}%`);
                                                const increases = Object.entries(info.byPurpose || {})
                                                    .map(([p, d]) => ({ name: p, sales: d.sales, avg: purposeAvg[p] || 0, diff: d.sales - (purposeAvg[p] || 0) }))
                                                    .filter(d => d.diff > 0)
                                                    .sort((a, b) => b.diff - a.diff)
                                                    .slice(0, 3);
                                                if (increases.length > 0) {
                                                    result.push('▲ 증가 요인:');
                                                    increases.forEach(d => {
                                                        const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                        result.push(`  • ${d.name}: +${formatCurrency(d.diff)} (+${pct}%)`);
                                                    });
                                                }
                                            } else {
                                                result.push(`📉 월평균 대비 ${diffPct}%`);
                                                const decreases = Object.entries(purposeAvg)
                                                    .map(p => ({ name: p[0], avg: p[1], sales: info.byPurpose?.[p[0]]?.sales || 0 }))
                                                    .map(d => ({ ...d, diff: d.sales - d.avg }))
                                                    .filter(d => d.diff < 0)
                                                    .sort((a, b) => a.diff - b.diff)
                                                    .slice(0, 3);
                                                if (decreases.length > 0) {
                                                    result.push('▼ 감소 요인:');
                                                    decreases.forEach(d => {
                                                        const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                        result.push(`  • ${d.name}: ${formatCurrency(d.diff)} (${pct}%)`);
                                                    });
                                                }
                                            }
                                        }

                                        return result;
                                    }
                                }
                            }
                        },
                        scales: { y: { ticks: { callback: v => formatCurrency(v) } } }
                    }
                });
            } else if (monthlyPreset === 'top3') {
                // TOP 3 - by_month의 byManager 실제 데이터 사용
                const monthMap = Object.fromEntries(currentData.by_month || []);
                const top3Labels = managers.slice(0, 3).map(m => m[0]);

                // 현재 연도 데이터셋 (매출, 건수, 검사목적, 자체 월평균 포함)
                const datasets = top3Labels.map((name, i) => {
                    const monthlyInfo = labels.map((_, mi) => {
                        const monthData = monthMap[mi+1];
                        const mgrData = monthData?.byManager?.[name];
                        const sales = mgrData?.sales || 0;
                        const count = mgrData?.count || 0;
                        const byPurpose = mgrData?.byPurpose || {};
                        return { sales, count, perCase: count > 0 ? sales / count : 0, byPurpose };
                    });
                    const salesArr = monthlyInfo.map(d => d.sales);
                    const nonZeroSales = salesArr.filter(v => v > 0);
                    const ownAvg = nonZeroSales.length > 0 ? nonZeroSales.reduce((a,b) => a+b, 0) / nonZeroSales.length : 0;
                    return {
                        label: name,
                        data: salesArr,
                        monthlyInfo,
                        ownAvg,
                        borderColor: colors[i],
                        backgroundColor: colors[i] + '20',
                        fill: false,
                        tension: 0.4,
                        pointRadius: 5,
                        isComparison: false,
                    };
                });

                // 담당자별 검사목적별 월평균 계산
                const managerPurposeAvg = {};
                datasets.filter(ds => !ds.isComparison).forEach(ds => {
                    managerPurposeAvg[ds.label] = {};
                    const allPurposes = new Set();
                    ds.monthlyInfo.forEach(m => Object.keys(m.byPurpose).forEach(p => allPurposes.add(p)));
                    allPurposes.forEach(purpose => {
                        const values = ds.monthlyInfo.map(m => m.byPurpose[purpose]?.sales || 0);
                        const nonZero = values.filter(v => v > 0);
                        managerPurposeAvg[ds.label][purpose] = nonZero.length > 0 ? nonZero.reduce((a,b) => a+b, 0) / nonZero.length : 0;
                    });
                });

                // 전년도 비교 데이터 추가
                if (compareData && compareData.by_month) {
                    const compMonthMap = Object.fromEntries(compareData.by_month || []);
                    top3Labels.forEach((name, i) => {
                        const monthlyInfo = labels.map((_, mi) => {
                            const monthData = compMonthMap[mi+1];
                            const mgrData = monthData?.byManager?.[name];
                            const sales = mgrData?.sales || 0;
                            const count = mgrData?.count || 0;
                            const byPurpose = mgrData?.byPurpose || {};
                            return { sales, count, perCase: count > 0 ? sales / count : 0, byPurpose };
                        });
                        datasets.push({
                            label: name + ' (' + compareData.year + ')',
                            data: monthlyInfo.map(d => d.sales),
                            monthlyInfo,
                            borderColor: colors[i] + '60',
                            backgroundColor: 'transparent',
                            fill: false,
                            tension: 0.4,
                            pointRadius: 3,
                            borderDash: [5, 5],
                            isComparison: true,
                        });
                    });
                }

                charts.managerMonthly = new Chart(ctx.getContext('2d'), {
                    type: 'line',
                    data: { labels, datasets },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {
                            legend: { position: 'top' },
                            tooltip: {
                                callbacks: {
                                    label: function(context) {
                                        const ds = context.dataset;
                                        const monthIdx = context.dataIndex;
                                        const info = ds.monthlyInfo?.[monthIdx];
                                        if (!info) return ds.label + ': ' + formatCurrency(context.raw);

                                        let result = [
                                            ds.label + ': ' + formatCurrency(info.sales),
                                            '  건수: ' + info.count.toLocaleString() + '건',
                                            '  건당: ' + formatCurrency(info.perCase)
                                        ];

                                        // 자체 월평균 대비 및 검사목적별 증감 (현재 연도만)
                                        if (!ds.isComparison && ds.ownAvg) {
                                            const ownAvg = ds.ownAvg;
                                            const diff = info.sales - ownAvg;
                                            const diffPct = ownAvg > 0 ? ((diff / ownAvg) * 100).toFixed(1) : 0;

                                            result.push('─────────');
                                            result.push(`월평균: ${formatCurrency(ownAvg)}`);
                                            const purposeAvg = managerPurposeAvg[ds.label] || {};
                                            if (diff >= 0) {
                                                result.push(`📈 월평균 대비 +${diffPct}%`);
                                                const increases = Object.entries(info.byPurpose || {})
                                                    .map(([p, d]) => ({ name: p, sales: d.sales, avg: purposeAvg[p] || 0, diff: d.sales - (purposeAvg[p] || 0) }))
                                                    .filter(d => d.diff > 0)
                                                    .sort((a, b) => b.diff - a.diff)
                                                    .slice(0, 3);
                                                if (increases.length > 0) {
                                                    result.push('▲ 증가 요인:');
                                                    increases.forEach(d => {
                                                        const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                        result.push(`  • ${d.name}: +${formatCurrency(d.diff)} (+${pct}%)`);
                                                    });
                                                }
                                            } else {
                                                result.push(`📉 월평균 대비 ${diffPct}%`);
                                                const decreases = Object.entries(purposeAvg)
                                                    .map(p => ({ name: p[0], avg: p[1], sales: info.byPurpose?.[p[0]]?.sales || 0 }))
                                                    .map(d => ({ ...d, diff: d.sales - d.avg }))
                                                    .filter(d => d.diff < 0)
                                                    .sort((a, b) => a.diff - b.diff)
                                                    .slice(0, 3);
                                                if (decreases.length > 0) {
                                                    result.push('▼ 감소 요인:');
                                                    decreases.forEach(d => {
                                                        const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                        result.push(`  • ${d.name}: ${formatCurrency(d.diff)} (${pct}%)`);
                                                    });
                                                }
                                            }
                                        }

                                        return result;
                                    }
                                }
                            }
                        },
                        scales: { y: { ticks: { callback: v => formatCurrency(v) } } }
                    }
                });
            } else {
                if (selectedManagers.length === 0) {
                    charts.managerMonthly = new Chart(ctx.getContext('2d'), {
                        type: 'line',
                        data: { labels, datasets: [] },
                        options: {
                            responsive: true,
                            maintainAspectRatio: false,
                            plugins: { title: { display: true, text: '담당자를 선택해주세요', font: { size: 14 } } }
                        }
                    });
                } else {
                    const monthMap = Object.fromEntries(currentData.by_month || []);

                    // 현재 연도 데이터셋 (매출, 건수, 검사목적, 자체 월평균 포함)
                    const datasets = selectedManagers.map((name, i) => {
                        const monthlyInfo = labels.map((_, mi) => {
                            const monthData = monthMap[mi+1];
                            const mgrData = monthData?.byManager?.[name];
                            const sales = mgrData?.sales || 0;
                            const count = mgrData?.count || 0;
                            const byPurpose = mgrData?.byPurpose || {};
                            return { sales, count, perCase: count > 0 ? sales / count : 0, byPurpose };
                        });
                        const salesArr = monthlyInfo.map(d => d.sales);
                        const nonZeroSales = salesArr.filter(v => v > 0);
                        const ownAvg = nonZeroSales.length > 0 ? nonZeroSales.reduce((a,b) => a+b, 0) / nonZeroSales.length : 0;
                        return {
                            label: name,
                            data: salesArr,
                            monthlyInfo,
                            ownAvg,
                            borderColor: colors[i % colors.length],
                            backgroundColor: colors[i % colors.length] + '20',
                            fill: false,
                            tension: 0.4,
                            pointRadius: 5,
                            isComparison: false,
                        };
                    });

                    // 담당자별 검사목적별 월평균 계산
                    const managerPurposeAvg = {};
                    datasets.forEach(ds => {
                        managerPurposeAvg[ds.label] = {};
                        const allPurposes = new Set();
                        ds.monthlyInfo.forEach(m => Object.keys(m.byPurpose).forEach(p => allPurposes.add(p)));
                        allPurposes.forEach(purpose => {
                            const values = ds.monthlyInfo.map(m => m.byPurpose[purpose]?.sales || 0);
                            const nonZero = values.filter(v => v > 0);
                            managerPurposeAvg[ds.label][purpose] = nonZero.length > 0 ? nonZero.reduce((a,b) => a+b, 0) / nonZero.length : 0;
                        });
                    });

                    // 전년도 비교 데이터 추가
                    if (compareData && compareData.by_month) {
                        const compMonthMap = Object.fromEntries(compareData.by_month || []);
                        selectedManagers.forEach((name, i) => {
                            const monthlyInfo = labels.map((_, mi) => {
                                const monthData = compMonthMap[mi+1];
                                const mgrData = monthData?.byManager?.[name];
                                const sales = mgrData?.sales || 0;
                                const count = mgrData?.count || 0;
                                const byPurpose = mgrData?.byPurpose || {};
                                return { sales, count, perCase: count > 0 ? sales / count : 0, byPurpose };
                            });
                            datasets.push({
                                label: name + ' (' + compareData.year + ')',
                                data: monthlyInfo.map(d => d.sales),
                                monthlyInfo,
                                borderColor: colors[i % colors.length] + '60',
                                backgroundColor: 'transparent',
                                fill: false,
                                tension: 0.4,
                                pointRadius: 3,
                                borderDash: [5, 5],
                                isComparison: true,
                            });
                        });
                    }

                    charts.managerMonthly = new Chart(ctx.getContext('2d'), {
                        type: 'line',
                        data: { labels, datasets },
                        options: {
                            responsive: true,
                            maintainAspectRatio: false,
                            plugins: {
                                legend: { position: 'top' },
                                tooltip: {
                                    callbacks: {
                                        label: function(context) {
                                            const ds = context.dataset;
                                            const monthIdx = context.dataIndex;
                                            const info = ds.monthlyInfo?.[monthIdx];
                                            if (!info) return ds.label + ': ' + formatCurrency(context.raw);

                                            let result = [
                                                ds.label + ': ' + formatCurrency(info.sales),
                                                '  건수: ' + info.count.toLocaleString() + '건',
                                                '  건당: ' + formatCurrency(info.perCase)
                                            ];

                                            // 자체 월평균 대비 및 검사목적별 증감 (현재 연도만)
                                            if (!ds.isComparison && ds.ownAvg) {
                                                const ownAvg = ds.ownAvg;
                                                const diff = info.sales - ownAvg;
                                                const diffPct = ownAvg > 0 ? ((diff / ownAvg) * 100).toFixed(1) : 0;

                                                result.push('─────────');
                                                result.push(`월평균: ${formatCurrency(ownAvg)}`);
                                                const purposeAvg = managerPurposeAvg[ds.label] || {};
                                                if (diff >= 0) {
                                                    result.push(`📈 월평균 대비 +${diffPct}%`);
                                                    const increases = Object.entries(info.byPurpose || {})
                                                        .map(([p, d]) => ({ name: p, sales: d.sales, avg: purposeAvg[p] || 0, diff: d.sales - (purposeAvg[p] || 0) }))
                                                        .filter(d => d.diff > 0)
                                                        .sort((a, b) => b.diff - a.diff)
                                                        .slice(0, 3);
                                                    if (increases.length > 0) {
                                                        result.push('▲ 증가 요인:');
                                                        increases.forEach(d => {
                                                            const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                            result.push(`  • ${d.name}: +${formatCurrency(d.diff)} (+${pct}%)`);
                                                        });
                                                    }
                                                } else {
                                                    result.push(`📉 월평균 대비 ${diffPct}%`);
                                                    const decreases = Object.entries(purposeAvg)
                                                        .map(p => ({ name: p[0], avg: p[1], sales: info.byPurpose?.[p[0]]?.sales || 0 }))
                                                        .map(d => ({ ...d, diff: d.sales - d.avg }))
                                                        .filter(d => d.diff < 0)
                                                        .sort((a, b) => a.diff - b.diff)
                                                        .slice(0, 3);
                                                    if (decreases.length > 0) {
                                                        result.push('▼ 감소 요인:');
                                                        decreases.forEach(d => {
                                                            const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                            result.push(`  • ${d.name}: ${formatCurrency(d.diff)} (${pct}%)`);
                                                        });
                                                    }
                                                }
                                            }

                                            return result;
                                        }
                                    }
                                }
                            },
                            scales: { y: { ticks: { callback: v => formatCurrency(v) } } }
                        }
                    });
                }
            }
        }

        // 건당 매출 차트
        function updatePerCaseChart() {
            const ctx = document.getElementById('perCaseChart');
            if (!ctx) return;
            if (charts.perCase) charts.perCase.destroy();

            const selectedPurpose = document.getElementById('perCasePurposeSelect')?.value || '전체';
            const managers = currentData.by_manager || [];

            // 검사목적별 필터 적용 + 검사목적별 데이터 포함
            let chartData = managers.map(m => {
                let sales = 0, count = 0;
                const byPurpose = m[1].by_purpose || {};

                if (selectedPurpose === '전체') {
                    sales = m[1].sales || 0;
                    count = m[1].count || 0;
                } else {
                    const purposeData = byPurpose[selectedPurpose];
                    if (purposeData) {
                        sales = purposeData.sales || 0;
                        count = purposeData.count || 0;
                    }
                }
                const avgPrice = count > 0 ? sales / count : 0;

                // 검사목적별 건당 매출 계산
                const purposeAvgPrices = {};
                Object.entries(byPurpose).forEach(([purpose, data]) => {
                    if (data.count > 0) {
                        purposeAvgPrices[purpose] = {
                            avgPrice: data.sales / data.count,
                            sales: data.sales,
                            count: data.count
                        };
                    }
                });

                return {
                    name: m[0],
                    avgPrice,
                    sales,
                    count,
                    byPurpose,
                    purposeAvgPrices
                };
            }).filter(d => d.avgPrice > 0);

            chartData.sort((a, b) => perCaseSortOrder === 'desc' ? b.avgPrice - a.avgPrice : a.avgPrice - b.avgPrice);
            const avgAll = chartData.reduce((s, d) => s + d.avgPrice, 0) / (chartData.length || 1);
            const totalSalesAll = chartData.reduce((s, d) => s + d.sales, 0);
            const totalCountAll = chartData.reduce((s, d) => s + d.count, 0);

            // 전체 담당자 검사목적별 평균 건당 매출 계산
            const purposeGlobalAvg = {};
            const purposeTotals = {};
            chartData.forEach(d => {
                Object.entries(d.purposeAvgPrices).forEach(([purpose, data]) => {
                    if (!purposeTotals[purpose]) {
                        purposeTotals[purpose] = { totalSales: 0, totalCount: 0, managerCount: 0 };
                    }
                    purposeTotals[purpose].totalSales += data.sales;
                    purposeTotals[purpose].totalCount += data.count;
                    purposeTotals[purpose].managerCount++;
                });
            });
            Object.entries(purposeTotals).forEach(([purpose, data]) => {
                purposeGlobalAvg[purpose] = data.totalCount > 0 ? data.totalSales / data.totalCount : 0;
            });

            // 고단가/저단가 기준
            const HIGH_PRICE_THRESHOLD = 150000; // 15만원
            const LOW_PRICE_THRESHOLD = 50000;   // 5만원

            // 전체 평균 고단가/저단가 비중 계산
            let totalHighCount = 0, totalLowCount = 0, totalAllCount = 0;
            chartData.forEach(d => {
                Object.entries(d.purposeAvgPrices).forEach(([purpose, data]) => {
                    totalAllCount += data.count;
                    if (data.avgPrice >= HIGH_PRICE_THRESHOLD) {
                        totalHighCount += data.count;
                    }
                    if (data.avgPrice <= LOW_PRICE_THRESHOLD) {
                        totalLowCount += data.count;
                    }
                });
            });
            const avgHighRatio = totalAllCount > 0 ? (totalHighCount / totalAllCount * 100) : 0;
            const avgLowRatio = totalAllCount > 0 ? (totalLowCount / totalAllCount * 100) : 0;

            // 데이터셋 구성 (현재 연도)
            const datasets = [{
                label: currentData.year + '년 건당 매출',
                data: chartData.map(d => d.avgPrice),
                backgroundColor: chartData.map(d => d.avgPrice >= avgAll ? 'rgba(16, 185, 129, 0.7)' : 'rgba(245, 158, 11, 0.7)'),
                borderRadius: 6,
            }];

            // 전년도 데이터 맵
            const compManagerMap = compareData ? Object.fromEntries((compareData.by_manager || []).map(m => [m[0], m[1]])) : {};

            // 전년도 비교 데이터 추가 (검사목적 필터 적용)
            const compChartData = [];
            if (compareData && compareData.by_manager) {
                chartData.forEach(d => {
                    const comp = compManagerMap[d.name];
                    if (!comp) {
                        compChartData.push({ avgPrice: 0, count: 0 });
                        return;
                    }
                    let compSales = 0, compCount = 0;
                    if (selectedPurpose === '전체') {
                        compSales = comp.sales || 0;
                        compCount = comp.count || 0;
                    } else {
                        const purposeData = comp.by_purpose?.[selectedPurpose];
                        if (purposeData) {
                            compSales = purposeData.sales || 0;
                            compCount = purposeData.count || 0;
                        }
                    }
                    compChartData.push({ avgPrice: compCount > 0 ? compSales / compCount : 0, count: compCount });
                });

                datasets.push({
                    label: compareData.year + '년 건당 매출',
                    data: compChartData.map(d => d.avgPrice),
                    backgroundColor: 'rgba(156, 163, 175, 0.5)',
                    borderRadius: 6,
                });
            }

            // 외부 HTML 툴팁 생성 함수
            const getOrCreatePerCaseTooltip = (chart) => {
                let tooltipEl = document.getElementById('perCaseChartTooltip');
                if (!tooltipEl) {
                    tooltipEl = document.createElement('div');
                    tooltipEl.id = 'perCaseChartTooltip';
                    tooltipEl.style.cssText = `
                        position: fixed;
                        background: rgba(30, 41, 59, 0.95);
                        border: 1px solid rgba(99, 102, 241, 0.5);
                        border-radius: 12px;
                        padding: 16px;
                        pointer-events: none;
                        z-index: 99999;
                        font-size: 12px;
                        color: #e2e8f0;
                        box-shadow: 0 10px 40px rgba(0,0,0,0.3);
                        max-width: 320px;
                        transition: opacity 0.2s ease;
                    `;
                    document.body.appendChild(tooltipEl);
                }
                return tooltipEl;
            };

            // 외부 툴팁 핸들러
            const externalTooltipHandler = (context) => {
                const { chart, tooltip } = context;
                const tooltipEl = getOrCreatePerCaseTooltip(chart);

                if (tooltip.opacity === 0) {
                    tooltipEl.style.opacity = 0;
                    return;
                }

                if (tooltip.body) {
                    const dataIndex = tooltip.dataPoints[0].dataIndex;
                    const datasetIndex = tooltip.dataPoints[0].datasetIndex;
                    const d = chartData[dataIndex];

                    let html = '';

                    // 헤더
                    html += `<div style="font-size: 14px; font-weight: bold; color: #fff; margin-bottom: 12px; padding-bottom: 8px; border-bottom: 1px solid rgba(255,255,255,0.2);">👤 ${d.name}</div>`;

                    if (datasetIndex !== 0 && compChartData[dataIndex]) {
                        // 전년도 데이터
                        const compD = compChartData[dataIndex];
                        html += `<div>${compareData.year}년 건당: ${formatCurrency(Math.round(compD.avgPrice))}</div>`;
                    } else {
                        // 현재 연도 데이터 - 상세 오버레이

                        // 1. 기본 지표
                        html += `<div style="margin-bottom: 4px;">💰 건당 매출: <strong>${formatCurrency(Math.round(d.avgPrice))}</strong></div>`;
                        html += `<div style="margin-bottom: 4px;">📋 총 거래 건수: <strong>${d.count.toLocaleString()}건</strong></div>`;
                        html += `<div style="margin-bottom: 4px;">📊 총 매출액: <strong>${(d.sales / 100000000).toFixed(2)}억</strong></div>`;

                        // 전체 평균 대비
                        const diffFromAvg = ((d.avgPrice - avgAll) / avgAll * 100);
                        const diffIcon = diffFromAvg >= 0 ? '📈' : '📉';
                        const diffSign = diffFromAvg >= 0 ? '+' : '';
                        const diffColor = diffFromAvg >= 0 ? '#10b981' : '#ef4444';
                        html += `<div style="margin-bottom: 8px;">${diffIcon} 전체 평균(${formatCurrency(Math.round(avgAll))}) 대비: <span style="color: ${diffColor}; font-weight: bold;">${diffSign}${diffFromAvg.toFixed(1)}%</span></div>`;

                        // 2. 단가 구성 분석
                        html += `<div style="color: #94a3b8; margin: 12px 0 8px; padding-top: 8px; border-top: 1px dashed rgba(255,255,255,0.2);">── 단가 구성 분석 ──</div>`;

                        // 담당자별 고단가/저단가 비중 계산
                        let mgrHighCount = 0, mgrLowCount = 0, mgrTotalCount = 0;
                        Object.entries(d.purposeAvgPrices).forEach(([purpose, data]) => {
                            mgrTotalCount += data.count;
                            if (data.avgPrice >= HIGH_PRICE_THRESHOLD) mgrHighCount += data.count;
                            if (data.avgPrice <= LOW_PRICE_THRESHOLD) mgrLowCount += data.count;
                        });
                        const mgrHighRatio = mgrTotalCount > 0 ? (mgrHighCount / mgrTotalCount * 100) : 0;
                        const mgrLowRatio = mgrTotalCount > 0 ? (mgrLowCount / mgrTotalCount * 100) : 0;
                        const highDiff = mgrHighRatio - avgHighRatio;
                        const lowDiff = mgrLowRatio - avgLowRatio;

                        const highDiffColor = highDiff >= 0 ? '#10b981' : '#f59e0b';
                        const lowDiffColor = lowDiff <= 0 ? '#10b981' : '#f59e0b';
                        html += `<div style="margin-bottom: 4px;">🔺 고단가(15만↑) 비중: ${mgrHighRatio.toFixed(1)}% <span style="color: ${highDiffColor};">(평균 대비 ${highDiff >= 0 ? '+' : ''}${highDiff.toFixed(1)}%p)</span></div>`;
                        html += `<div style="margin-bottom: 8px;">🔻 저단가(5만↓) 비중: ${mgrLowRatio.toFixed(1)}% <span style="color: ${lowDiffColor};">(평균 대비 ${lowDiff >= 0 ? '+' : ''}${lowDiff.toFixed(1)}%p)</span></div>`;

                        // 3. 강점 검사목적
                        const strengths = Object.entries(d.purposeAvgPrices)
                            .map(([purpose, data]) => {
                                const globalAvg = purposeGlobalAvg[purpose] || 0;
                                const diff = globalAvg > 0 ? ((data.avgPrice - globalAvg) / globalAvg * 100) : 0;
                                return { purpose, avgPrice: data.avgPrice, globalAvg, diff, count: data.count };
                            })
                            .filter(item => item.diff > 0 && item.count >= 3)
                            .sort((a, b) => b.diff - a.diff)
                            .slice(0, 3);

                        if (strengths.length > 0) {
                            html += `<div style="color: #10b981; margin: 12px 0 6px; font-weight: 600;">▲ 강점 검사목적 (평균 대비 높음)</div>`;
                            strengths.forEach(s => {
                                html += `<div style="margin-left: 8px; margin-bottom: 2px;">• ${s.purpose}: ${formatCurrency(Math.round(s.avgPrice))} <span style="color: #10b981;">(+${s.diff.toFixed(0)}%)</span></div>`;
                            });
                        }

                        // 4. 개선 기회
                        const improvements = Object.entries(d.purposeAvgPrices)
                            .map(([purpose, data]) => {
                                const globalAvg = purposeGlobalAvg[purpose] || 0;
                                const diff = globalAvg > 0 ? ((data.avgPrice - globalAvg) / globalAvg * 100) : 0;
                                return { purpose, avgPrice: data.avgPrice, globalAvg, diff, count: data.count };
                            })
                            .filter(item => item.diff < -10 && item.count >= 3)
                            .sort((a, b) => a.diff - b.diff)
                            .slice(0, 3);

                        if (improvements.length > 0) {
                            html += `<div style="color: #f59e0b; margin: 12px 0 6px; font-weight: 600;">▼ 개선 기회</div>`;
                            improvements.forEach(s => {
                                html += `<div style="margin-left: 8px; margin-bottom: 2px;">• ${s.purpose}: ${formatCurrency(Math.round(s.avgPrice))} <span style="color: #f59e0b;">(${s.diff.toFixed(0)}%)</span></div>`;
                            });
                        }
                    }

                    tooltipEl.innerHTML = html;
                }

                // 위치 계산 (화면 밖으로 나가지 않도록)
                const { offsetLeft: positionX, offsetTop: positionY } = chart.canvas;
                const canvasRect = chart.canvas.getBoundingClientRect();

                let left = canvasRect.left + tooltip.caretX + 15;
                let top = canvasRect.top + tooltip.caretY - 10;

                // 우측 경계 체크
                const tooltipWidth = tooltipEl.offsetWidth || 320;
                if (left + tooltipWidth > window.innerWidth - 20) {
                    left = canvasRect.left + tooltip.caretX - tooltipWidth - 15;
                }

                // 하단 경계 체크
                const tooltipHeight = tooltipEl.offsetHeight || 300;
                if (top + tooltipHeight > window.innerHeight - 20) {
                    top = window.innerHeight - tooltipHeight - 20;
                }

                // 상단 경계 체크
                if (top < 10) top = 10;

                tooltipEl.style.opacity = 1;
                tooltipEl.style.left = left + 'px';
                tooltipEl.style.top = top + 'px';
            };

            charts.perCase = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: chartData.map(d => d.name),
                    datasets
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: compareData ? true : false, position: 'top' },
                        tooltip: {
                            enabled: false,
                            external: externalTooltipHandler
                        }
                    },
                    scales: { y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } } }
                }
            });
        }

        // 긴급 접수 건수 차트
        function updateUrgentChart() {
            const ctx = document.getElementById('urgentChart');
            if (!ctx) return;
            if (charts.urgent) charts.urgent.destroy();

            const selectedPurpose = document.getElementById('urgentPurposeSelect')?.value || '전체';
            const managers = currentData.by_manager || [];

            // 월 수 계산 (실제 데이터가 있는 월)
            const byMonth = currentData.by_month || [];
            const monthCount = byMonth.length || 12;

            // 전체 평균 계산 (목적별 긴급 건수)
            const purposeAvgMap = {};
            let totalUrgent = 0;
            let managerWithUrgent = 0;
            managers.forEach(m => {
                const urgentByPurpose = m[1].urgent_by_purpose || {};
                const urgent = m[1].urgent || 0;
                if (urgent > 0) {
                    managerWithUrgent++;
                    totalUrgent += urgent;
                }
                Object.entries(urgentByPurpose).forEach(([purpose, count]) => {
                    if (!purposeAvgMap[purpose]) purposeAvgMap[purpose] = { total: 0, count: 0 };
                    purposeAvgMap[purpose].total += count;
                    purposeAvgMap[purpose].count++;
                });
            });
            // 평균 계산
            Object.keys(purposeAvgMap).forEach(purpose => {
                purposeAvgMap[purpose].avg = purposeAvgMap[purpose].total / (managerWithUrgent || 1);
            });
            const overallAvg = totalUrgent / (managerWithUrgent || 1);

            // 검사목적별 필터 적용
            const urgentData = managers.map(m => {
                let urgentCount = 0;
                const urgentByPurpose = m[1].urgent_by_purpose || {};
                if (selectedPurpose === '전체') {
                    urgentCount = m[1].urgent || 0;
                } else {
                    urgentCount = urgentByPurpose[selectedPurpose] || 0;
                }
                return {
                    name: m[0],
                    urgent: urgentCount,
                    urgentByPurpose: urgentByPurpose,
                    totalUrgent: m[1].urgent || 0,
                    monthlyAvg: (m[1].urgent || 0) / monthCount
                };
            }).sort((a, b) => b.urgent - a.urgent);

            const maxUrgent = Math.max(...urgentData.map(d => d.urgent)) || 1;

            // 데이터셋 구성
            const datasets = [{
                label: currentData.year + '년 긴급',
                data: urgentData.map(d => d.urgent),
                backgroundColor: urgentData.map(d => {
                    const ratio = d.urgent / maxUrgent;
                    if (ratio >= 0.8) return 'rgba(239, 68, 68, 0.8)';
                    if (ratio >= 0.5) return 'rgba(245, 158, 11, 0.8)';
                    return 'rgba(99, 102, 241, 0.8)';
                }),
                borderRadius: 6,
            }];

            // 전년도 비교 데이터 추가
            if (compareData && compareData.by_manager) {
                const compManagerMap = Object.fromEntries((compareData.by_manager || []).map(m => [m[0], m[1]]));
                datasets.push({
                    label: compareData.year + '년 긴급',
                    data: urgentData.map(d => {
                        const comp = compManagerMap[d.name];
                        if (!comp) return 0;
                        if (selectedPurpose === '전체') {
                            return comp.urgent || 0;
                        } else {
                            const purposeUrgent = comp.urgent_by_purpose || {};
                            return purposeUrgent[selectedPurpose] || 0;
                        }
                    }),
                    backgroundColor: 'rgba(156, 163, 175, 0.5)',
                    borderRadius: 6,
                });
            }

            charts.urgent = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: urgentData.map(d => d.name),
                    datasets
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: compareData ? true : false, position: 'top' },
                        tooltip: {
                            callbacks: {
                                title: function(context) {
                                    return context[0].label;
                                },
                                label: function(context) {
                                    if (context.datasetIndex > 0) {
                                        return context.dataset.label + ': ' + context.raw + '건';
                                    }
                                    const idx = context.dataIndex;
                                    const d = urgentData[idx];
                                    const lines = [];
                                    lines.push('긴급 건수: ' + d.urgent + '건');
                                    lines.push('월평균: ' + d.monthlyAvg.toFixed(1) + '건/월');

                                    // 전체 평균 대비
                                    const diffFromAvg = d.totalUrgent - overallAvg;
                                    const diffPct = overallAvg > 0 ? ((d.totalUrgent - overallAvg) / overallAvg * 100).toFixed(0) : 0;
                                    lines.push('전체 평균(' + overallAvg.toFixed(0) + '건) 대비: ' + (diffFromAvg >= 0 ? '+' : '') + diffFromAvg.toFixed(0) + '건 (' + (diffFromAvg >= 0 ? '+' : '') + diffPct + '%)');

                                    return lines;
                                },
                                afterBody: function(context) {
                                    if (context[0].datasetIndex > 0) return [];
                                    const idx = context[0].dataIndex;
                                    const d = urgentData[idx];
                                    const lines = ['', '── 검사목적별 비교 ──'];

                                    // 평균 대비 높은 목적과 낮은 목적 분석
                                    const higherPurposes = [];
                                    const lowerPurposes = [];

                                    Object.entries(d.urgentByPurpose).forEach(([purpose, count]) => {
                                        const avg = purposeAvgMap[purpose]?.avg || 0;
                                        if (avg > 0) {
                                            const diff = count - avg;
                                            const pct = (diff / avg * 100).toFixed(0);
                                            if (diff > 0) {
                                                higherPurposes.push({ purpose, count, avg, diff, pct: '+' + pct });
                                            } else if (diff < 0) {
                                                lowerPurposes.push({ purpose, count, avg, diff, pct });
                                            }
                                        }
                                    });

                                    // 평균보다 긴급이 많은 목적들 (상위 3개)
                                    if (higherPurposes.length > 0) {
                                        higherPurposes.sort((a, b) => b.diff - a.diff);
                                        lines.push('▲ 평균 대비 높음:');
                                        higherPurposes.slice(0, 3).forEach(p => {
                                            lines.push('  ' + p.purpose + ': ' + p.count + '건 (평균 ' + p.avg.toFixed(0) + '건, ' + p.pct + '%)');
                                        });
                                    }

                                    // 평균보다 긴급이 적은 목적들 (상위 3개)
                                    if (lowerPurposes.length > 0) {
                                        lowerPurposes.sort((a, b) => a.diff - b.diff);
                                        lines.push('▼ 평균 대비 낮음:');
                                        lowerPurposes.slice(0, 3).forEach(p => {
                                            lines.push('  ' + p.purpose + ': ' + p.count + '건 (평균 ' + p.avg.toFixed(0) + '건, ' + p.pct + '%)');
                                        });
                                    }

                                    if (higherPurposes.length === 0 && lowerPurposes.length === 0) {
                                        lines.push('(목적별 데이터 없음)');
                                    }

                                    return lines;
                                }
                            }
                        }
                    },
                    scales: { y: { beginAtZero: true } }
                }
            });
        }

        // 긴급 월별 추이 차트
        function updateUrgentMonthlyChart() {
            const ctx = document.getElementById('urgentMonthlyChart');
            if (!ctx) return;
            if (charts.urgentMonthly) charts.urgentMonthly.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const urgentMonthMap = Object.fromEntries(currentData.by_urgent_month || []);
            const urgentMonthly = labels.map((_, i) => urgentMonthMap[i+1]?.count || 0);

            const datasets = [{
                label: currentData.year + '년 긴급',
                data: urgentMonthly,
                borderColor: '#ef4444',
                backgroundColor: 'rgba(239, 68, 68, 0.2)',
                fill: true,
                tension: 0.4,
                pointRadius: 5,
                pointBackgroundColor: '#ef4444',
            }];

            // 전년도 비교 데이터
            if (compareData && compareData.by_urgent_month) {
                const compMap = Object.fromEntries(compareData.by_urgent_month || []);
                datasets.push({
                    label: compareData.year + '년 긴급',
                    data: labels.map((_, i) => compMap[i+1]?.count || 0),
                    borderColor: 'rgba(156, 163, 175, 0.8)',
                    backgroundColor: 'rgba(156, 163, 175, 0.1)',
                    fill: false,
                    tension: 0.4,
                    pointRadius: 4,
                    borderDash: [5, 5],
                });
            }

            charts.urgentMonthly = new Chart(ctx.getContext('2d'), {
                type: 'line',
                data: { labels, datasets },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: compareData ? true : false, position: 'top' } },
                    scales: { y: { beginAtZero: true } }
                }
            });
        }

        // 긴급 건당 단가 차트
        function updateUrgentUnitPriceChart() {
            const ctx = document.getElementById('urgentUnitPriceChart');
            if (!ctx) return;
            if (charts.urgentUnitPrice) charts.urgentUnitPrice.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const urgentMonthMap = Object.fromEntries(currentData.by_urgent_month || []);
            const urgentUnitPrices = labels.map((_, i) => {
                const d = urgentMonthMap[i+1];
                return d && d.count > 0 ? d.sales / d.count : 0;
            });

            const datasets = [{
                label: currentData.year + '년 긴급 건당 단가',
                data: urgentUnitPrices,
                borderColor: '#f59e0b',
                backgroundColor: 'rgba(245, 158, 11, 0.2)',
                fill: true,
                tension: 0.4,
                pointRadius: 5,
                pointBackgroundColor: '#f59e0b',
            }];

            // 전년도 비교 데이터
            if (compareData && compareData.by_urgent_month) {
                const compMap = Object.fromEntries(compareData.by_urgent_month || []);
                datasets.push({
                    label: compareData.year + '년 긴급 건당 단가',
                    data: labels.map((_, i) => {
                        const d = compMap[i+1];
                        return d && d.count > 0 ? d.sales / d.count : 0;
                    }),
                    borderColor: 'rgba(156, 163, 175, 0.8)',
                    backgroundColor: 'rgba(156, 163, 175, 0.1)',
                    fill: false,
                    tension: 0.4,
                    pointRadius: 4,
                    borderDash: [5, 5],
                });
            }

            charts.urgentUnitPrice = new Chart(ctx.getContext('2d'), {
                type: 'line',
                data: { labels, datasets },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: compareData ? true : false, position: 'top' } },
                    scales: { y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } } }
                }
            });
        }

        // 일 방문 거래처 수 차트
        function updateDailyClientChart() {
            const ctx = document.getElementById('dailyClientChart');
            if (!ctx) return;
            if (charts.dailyClient) charts.dailyClient.destroy();

            const managers = currentData.by_manager || [];
            // 거래처 수는 클라이언트 데이터에서 추정 (데이터가 없으면 건수 기준으로 추정)
            const chartData = managers.map(m => ({
                name: m[0],
                avgDailyClients: Math.round((m[1].count || 0) / 250 * 10) / 10  // 연간 영업일 250일 기준
            })).sort((a, b) => b.avgDailyClients - a.avgDailyClients);

            const avgAll = chartData.reduce((s, d) => s + d.avgDailyClients, 0) / (chartData.length || 1);

            // 전년도 비교 데이터
            const datasets = [{
                label: currentData.year + '년',
                data: chartData.map(d => d.avgDailyClients),
                backgroundColor: 'rgba(99, 102, 241, 0.8)',
                borderRadius: 6,
            }];

            if (compareData && compareData.by_manager) {
                const compareMap = Object.fromEntries(compareData.by_manager || []);
                datasets.push({
                    label: compareData.year + '년',
                    data: chartData.map(d => {
                        const compData = compareMap[d.name];
                        return compData ? Math.round((compData.count || 0) / 250 * 10) / 10 : 0;
                    }),
                    backgroundColor: 'rgba(139, 92, 246, 0.5)',
                    borderRadius: 6,
                });
                document.getElementById('dailyClientLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background: rgba(99, 102, 241, 0.8);"></div><span>${currentData.year}년</span></div><div class="legend-item"><div class="legend-color" style="background: rgba(139, 92, 246, 0.5);"></div><span>${compareData.year}년</span></div>`;
                document.getElementById('dailyClientLegend').style.display = 'flex';
            } else {
                document.getElementById('dailyClientLegend').style.display = 'none';
            }

            charts.dailyClient = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: chartData.map(d => d.name),
                    datasets
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: false } },
                    scales: { y: { beginAtZero: true, title: { display: true, text: '일평균 건수' } } }
                }
            });
        }

        // 담당자 상세 모달
        function showManagerDetail(managerName) {
            const managers = currentData.by_manager || [];
            const manager = managers.find(m => m[0] === managerName);
            if (!manager) return;

            // 현재 선택된 검사목적 필터 확인
            const purposeFilter = document.getElementById('managerPurposeFilter')?.value || '전체';
            const isPurposeFiltered = purposeFilter !== '전체';

            // 모달 제목 (필터 적용 시 표시)
            const titleSuffix = isPurposeFiltered ? ` (${purposeFilter})` : '';
            document.getElementById('modalManagerName').textContent = managerName + ' 상세' + titleSuffix;

            // 담당자별 주요 거래 업체 (manager_top_clients 사용)
            // 필터 적용 시 해당 목적의 업체만 표시
            let managerClients = currentData.manager_top_clients?.[managerName] || [];
            if (isPurposeFiltered) {
                // by_purpose 데이터가 있는 업체 필터링
                const byPurpose = manager[1].by_purpose || {};
                const purposeData = byPurpose[purposeFilter];
                if (purposeData) {
                    // 해당 목적에서의 매출 기준으로 정렬된 업체 표시
                    // (현재 manager_top_clients에서 해당 목적 데이터 필터링)
                    managerClients = managerClients.filter(c => {
                        // 업체별 목적 데이터 확인 (by_client에서)
                        const clientData = (currentData.by_client || []).find(cl => cl[0] === c[0]);
                        return clientData && clientData[1].purposes && clientData[1].purposes[purposeFilter];
                    });
                }
            }
            if (managerClients.length > 0) {
                document.getElementById('modalTopClients').innerHTML = managerClients.slice(0, 5).map(c => `
                    <div class="modal-client-item">
                        <span class="modal-client-name">${c[0]}</span>
                        <span class="modal-client-value">${formatCurrency(c[1].sales)}</span>
                    </div>
                `).join('');
            } else {
                document.getElementById('modalTopClients').innerHTML = '<div style="color: var(--gray-400); font-size: 13px;">데이터 없음</div>';
            }

            // 담당자별 검사 목적별 비중 차트 (purpose_managers 사용)
            const modalCtx = document.getElementById('modalPurposeCanvas');
            if (charts.modalPurpose) charts.modalPurpose.destroy();

            // 목적별 담당자 데이터에서 해당 담당자 데이터 추출
            const managerPurposes = [];
            const purposeManagers = currentData.purpose_managers || {};

            if (isPurposeFiltered) {
                // 필터 적용 시: 해당 목적의 세부 정보만 표시 (by_purpose 사용)
                const byPurpose = manager[1].by_purpose || {};
                const purposeData = byPurpose[purposeFilter];
                if (purposeData) {
                    managerPurposes.push({
                        name: purposeFilter,
                        sales: purposeData.sales,
                        count: purposeData.count
                    });
                }
            } else {
                // 전체: 모든 목적별 데이터
                for (const [purpose, managers] of Object.entries(purposeManagers)) {
                    const mgrData = managers.find(m => m.name === managerName);
                    if (mgrData) {
                        managerPurposes.push({ name: purpose, sales: mgrData.sales });
                    }
                }
            }
            managerPurposes.sort((a, b) => b.sales - a.sales);
            const topPurposes = managerPurposes.slice(0, 5);

            if (topPurposes.length > 0) {
                charts.modalPurpose = new Chart(modalCtx.getContext('2d'), {
                    type: 'doughnut',
                    data: {
                        labels: topPurposes.map(p => p.name),
                        datasets: [{ data: topPurposes.map(p => p.sales), backgroundColor: ['#6366f1', '#10b981', '#f59e0b', '#ec4899', '#06b6d4'] }]
                    },
                    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { font: { size: 11 } } } } }
                });
            }

            // 담당자별 담당 지역 (manager_regions 사용)
            const managerRegions = currentData.manager_regions?.[managerName] || [];
            if (managerRegions.length > 0) {
                document.getElementById('modalRegions').innerHTML = managerRegions.slice(0, 5).map(r =>
                    `<span class="region-tag">${r.region} (${formatCurrency(r.sales)})</span>`
                ).join('');
            } else {
                document.getElementById('modalRegions').innerHTML = '<span style="color: var(--gray-400); font-size: 13px;">데이터 없음</span>';
            }

            document.getElementById('managerModal').style.display = 'flex';
        }

        function closeManagerModal() {
            document.getElementById('managerModal').style.display = 'none';
        }

        // ====== 팀별 탭 관련 함수 ======
        function updateTeamTab() {
            const branches = currentData.by_branch || [];
            if (branches.length === 0) return;

            const totalBranches = branches.length;
            const totalSales = branches.reduce((sum, b) => sum + (b[1].sales || 0), 0);
            const avgSales = totalSales / totalBranches;

            // KPI 카드 업데이트
            document.getElementById('teamTotalBranches').textContent = totalBranches + '개';
            document.getElementById('teamAvgSales').textContent = formatCurrency(avgSales);

            // 최고 성과 팀
            const topBranch = branches.reduce((max, b) => (b[1].sales > (max[1]?.sales || 0)) ? b : max, branches[0]);
            document.getElementById('teamTopBranch').textContent = topBranch[0];
            document.getElementById('teamTopBranchSales').textContent = '매출: ' + formatCurrency(topBranch[1].sales);

            // 최고 성장 팀 (전년 비교 시)
            if (compareData && compareData.by_branch) {
                const compareMap = Object.fromEntries(compareData.by_branch);
                const withGrowth = branches.map(b => {
                    const compSales = compareMap[b[0]]?.sales || 0;
                    const growth = compSales > 0 ? ((b[1].sales - compSales) / compSales * 100) : 0;
                    return { name: b[0], growth };
                }).sort((a, b) => b.growth - a.growth);

                if (withGrowth.length > 0) {
                    document.getElementById('teamTopGrowth').textContent = withGrowth[0].name;
                    document.getElementById('teamTopGrowthRate').textContent = '전년 대비 +' + withGrowth[0].growth.toFixed(1) + '%';
                    document.getElementById('teamTopGrowthTrend').style.visibility = 'visible';
                    document.getElementById('teamTopGrowthTrend').innerHTML = '↑ +' + withGrowth[0].growth.toFixed(1) + '%';
                }
            } else {
                document.getElementById('teamTopGrowth').textContent = '-';
                document.getElementById('teamTopGrowthRate').textContent = '전년 비교 필요';
                document.getElementById('teamTopGrowthTrend').style.visibility = 'hidden';
            }

            // 드롭다운 초기화
            initBranchChartPurposeFilter();
            initBranchPerCasePurposeSelect();
            initBranchTablePurposeFilter();

            // 차트들 업데이트
            updateBranchChart();
            updateBranchPerCaseChart();
            updateBranchEfficiencyChart();
            updateBranchMonthlyChart();
            updateBranchTable();
            updateClientRetentionChart();
            updateRetentionRateChart();
            updateBranchRetentionTable();
        }

        // 팀별 목적 필터 초기화
        function initBranchChartPurposeFilter() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('branchChartPurposeFilter');
            if (select) {
                select.innerHTML = '<option value="전체">전체 검사목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function initBranchPerCasePurposeSelect() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('branchPerCasePurposeSelect');
            if (select) {
                select.innerHTML = '<option value="전체">전체 목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function initBranchTablePurposeFilter() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('branchTablePurposeFilter');
            if (select) {
                select.innerHTML = '<option value="전체">전체 검사목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        // 팀별 건당 매출 차트
        function updateBranchPerCaseChart() {
            const ctx = document.getElementById('branchPerCaseChart');
            if (!ctx) return;
            if (charts.branchPerCase) charts.branchPerCase.destroy();

            const selectedPurpose = document.getElementById('branchPerCasePurposeSelect')?.value || '전체';
            const branches = currentData.by_branch || [];
            if (branches.length === 0) return;

            // 검사목적별 필터 적용
            const branchData = branches.map(b => {
                let sales = 0, count = 0;
                if (selectedPurpose === '전체') {
                    sales = b[1].sales || 0;
                    count = b[1].count || 0;
                } else {
                    const purposeData = b[1].by_purpose?.[selectedPurpose];
                    if (purposeData) {
                        sales = purposeData.sales || 0;
                        count = purposeData.count || 0;
                    }
                }
                const avgPrice = count > 0 ? sales / count : 0;
                return { name: b[0], avgPrice, sales, count };
            }).filter(d => d.avgPrice > 0).sort((a, b) => b.avgPrice - a.avgPrice);

            const avgAll = branchData.reduce((s, d) => s + d.avgPrice, 0) / (branchData.length || 1);
            const totalSales = branchData.reduce((s, d) => s + d.sales, 0);
            const totalCount = branchData.reduce((s, d) => s + d.count, 0);
            const avgAvgPrice = branchData.length > 0 ? branchData.reduce((s, d) => s + d.avgPrice, 0) / branchData.length : 0;

            const legendEl = document.getElementById('branchPerCaseLegend');

            if (compareData) {
                // 비교 데이터 처리 (검사목적 필터 적용)
                const compareBranches = compareData.by_branch || [];
                const compareMap = {};
                let compTotalSales = 0, compTotalCount = 0;
                compareBranches.forEach(b => {
                    let sales = 0, count = 0;
                    if (selectedPurpose === '전체') {
                        sales = b[1].sales || 0;
                        count = b[1].count || 0;
                    } else {
                        const purposeData = b[1].by_purpose?.[selectedPurpose];
                        if (purposeData) {
                            sales = purposeData.sales || 0;
                            count = purposeData.count || 0;
                        }
                    }
                    const avgPrice = count > 0 ? sales / count : 0;
                    compareMap[b[0]] = { avgPrice, sales, count };
                    compTotalSales += sales;
                    compTotalCount += count;
                });

                legendEl.innerHTML = `
                    <div class="legend-item"><div class="legend-color" style="background: rgba(16, 185, 129, 0.8);"></div><span>${currentData.year}년</span></div>
                    <div class="legend-item"><div class="legend-color" style="background: rgba(245, 158, 11, 0.6);"></div><span>${compareData.year}년</span></div>
                    <div style="margin-left: auto; display: flex; gap: 20px; font-size: 12px; color: #666;">
                        <span>총매출: <strong>${formatCurrency(totalSales)}</strong></span>
                        <span>총건수: <strong>${totalCount.toLocaleString()}건</strong></span>
                        <span>평균단가: <strong>${formatCurrency(avgAvgPrice)}</strong></span>
                    </div>`;
                legendEl.style.display = 'flex';

                charts.branchPerCase = new Chart(ctx.getContext('2d'), {
                    type: 'bar',
                    data: {
                        labels: branchData.map(b => b.name),
                        datasets: [
                            { label: currentData.year + '년', data: branchData.map(b => b.avgPrice), backgroundColor: 'rgba(16, 185, 129, 0.8)', borderRadius: 6 },
                            { label: compareData.year + '년', data: branchData.map(b => compareMap[b.name]?.avgPrice || 0), backgroundColor: 'rgba(245, 158, 11, 0.6)', borderRadius: 6 }
                        ]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: false } },
                        scales: {
                            y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } },
                            x: { grid: { display: false } }
                        }
                    }
                });
            } else {
                legendEl.innerHTML = `
                    <div style="display: flex; gap: 20px; font-size: 12px; color: #666;">
                        <span>총매출: <strong>${formatCurrency(totalSales)}</strong></span>
                        <span>총건수: <strong>${totalCount.toLocaleString()}건</strong></span>
                        <span>평균단가: <strong>${formatCurrency(avgAvgPrice)}</strong></span>
                    </div>`;
                legendEl.style.display = 'flex';

                charts.branchPerCase = new Chart(ctx.getContext('2d'), {
                    type: 'bar',
                    data: {
                        labels: branchData.map(b => b.name),
                        datasets: [{
                            label: '건당 매출',
                            data: branchData.map(b => b.avgPrice),
                            backgroundColor: branchData.map(d => d.avgPrice >= avgAll ? 'rgba(16, 185, 129, 0.7)' : 'rgba(245, 158, 11, 0.7)'),
                            borderRadius: 6
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: false } },
                        scales: {
                            y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } },
                            x: { grid: { display: false } }
                        }
                    }
                });
            }
        }

        // 지사별 효율성 분석 산점도
        function updateBranchEfficiencyChart() {
            const ctx = document.getElementById('branchEfficiencyChart');
            if (!ctx) return;
            if (charts.branchEfficiency) charts.branchEfficiency.destroy();

            const branches = currentData.by_branch || [];
            if (branches.length === 0) return;

            const avgCount = branches.reduce((sum, b) => sum + (b[1].count || 0), 0) / branches.length;
            const avgSales = branches.reduce((sum, b) => sum + (b[1].sales || 0), 0) / branches.length;

            const data = branches.map(b => {
                const isHighCount = (b[1].count || 0) >= avgCount;
                const isHighSales = (b[1].sales || 0) >= avgSales;
                let color;
                if (isHighCount && isHighSales) color = 'rgba(16, 185, 129, 0.8)';
                else if (!isHighCount && isHighSales) color = 'rgba(99, 102, 241, 0.8)';
                else if (isHighCount && !isHighSales) color = 'rgba(245, 158, 11, 0.8)';
                else color = 'rgba(239, 68, 68, 0.8)';
                return { x: b[1].count || 0, y: b[1].sales || 0, name: b[0], color };
            });

            charts.branchEfficiency = new Chart(ctx.getContext('2d'), {
                type: 'scatter',
                data: {
                    datasets: [{
                        data: data.map(d => ({ x: d.x, y: d.y })),
                        backgroundColor: data.map(d => d.color),
                        pointRadius: 15,
                        pointHoverRadius: 20,
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                label: (context) => {
                                    const idx = context.dataIndex;
                                    const b = branches[idx];
                                    return [b[0], '매출: ' + formatCurrency(b[1].sales || 0), '건수: ' + (b[1].count || 0).toLocaleString() + '건'];
                                }
                            }
                        }
                    },
                    scales: {
                        x: { title: { display: true, text: '건수' }, grid: { color: 'rgba(0,0,0,0.05)' } },
                        y: { title: { display: true, text: '매출 (공급가액)' }, ticks: { callback: v => formatCurrency(v) }, grid: { color: 'rgba(0,0,0,0.05)' } }
                    }
                }
            });
        }

        // 지사별 월별 추이
        let branchMonthlyFilter = 'all';
        let branchMonthlySelected = '';

        function initBranchMonthlySelect() {
            const branches = currentData.by_branch || [];
            const select = document.getElementById('branchMonthlySelect');
            if (select) {
                select.innerHTML = '<option value="">팀 선택</option>' +
                    branches.map(b => `<option value="${b[0]}">${b[0]}</option>`).join('');
            }
        }

        function setBranchMonthlyFilter(type) {
            branchMonthlyFilter = type;
            document.getElementById('branchMonthlyAll').classList.toggle('active', type === 'all');
            document.getElementById('branchMonthlyTop3').classList.toggle('active', type === 'top3');
            if (type === 'select') {
                branchMonthlySelected = document.getElementById('branchMonthlySelect').value;
            } else {
                document.getElementById('branchMonthlySelect').value = '';
                branchMonthlySelected = '';
            }
            updateBranchMonthlyChart();
        }

        function updateBranchMonthlyChart() {
            const ctx = document.getElementById('branchMonthlyChart');
            if (!ctx) return;
            if (charts.branchMonthly) charts.branchMonthly.destroy();

            // 드롭다운 초기화
            initBranchMonthlySelect();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const colors = ['#6366f1', '#10b981', '#f59e0b', '#ec4899', '#06b6d4', '#8b5cf6', '#ef4444', '#14b8a6'];
            let branches = [...(currentData.by_branch || [])];
            const monthMap = Object.fromEntries(currentData.by_month || []);

            // 필터 적용
            if (branchMonthlyFilter === 'top3') {
                branches = branches.slice(0, 3);
            } else if (branchMonthlyFilter === 'select' && branchMonthlySelected) {
                branches = branches.filter(b => b[0] === branchMonthlySelected);
            }

            // 팀별 월별 데이터 - 실제 byBranch 데이터 사용 (매출, 건수, 검사목적 포함)
            const branchMonthlyData = branches.map(b => {
                const branchName = b[0];
                const monthlyInfo = labels.map((_, mi) => {
                    const monthData = monthMap[mi+1];
                    const branchData = monthData?.byBranch?.[branchName];
                    const sales = branchData?.sales || 0;
                    const count = branchData?.count || 0;
                    const byPurpose = branchData?.byPurpose || {};
                    return { sales, count, perCase: count > 0 ? sales / count : 0, byPurpose };
                });
                const salesArr = monthlyInfo.map(d => d.sales);
                const nonZeroSales = salesArr.filter(v => v > 0);
                const ownAvg = nonZeroSales.length > 0 ? nonZeroSales.reduce((a,b) => a+b, 0) / nonZeroSales.length : 0;
                return { name: branchName, data: salesArr, monthlyInfo, ownAvg };
            });

            // 팀별 검사목적별 월평균 계산 (증감 비교용)
            const branchPurposeAvg = {};
            branchMonthlyData.forEach(b => {
                branchPurposeAvg[b.name] = {};
                const allPurposes = new Set();
                b.monthlyInfo.forEach(m => Object.keys(m.byPurpose).forEach(p => allPurposes.add(p)));
                allPurposes.forEach(purpose => {
                    const values = b.monthlyInfo.map(m => m.byPurpose[purpose]?.sales || 0);
                    const nonZero = values.filter(v => v > 0);
                    branchPurposeAvg[b.name][purpose] = nonZero.length > 0 ? nonZero.reduce((a,b) => a+b, 0) / nonZero.length : 0;
                });
            });

            // 전체 월별 평균 계산
            const monthlyAvg = labels.map((_, mi) => {
                const monthData = monthMap[mi+1];
                if (!monthData || !monthData.byBranch) return 0;
                const branchSales = Object.values(monthData.byBranch).map(b => b.sales || 0);
                return branchSales.length > 0 ? branchSales.reduce((a,b) => a+b, 0) / branchSales.length : 0;
            });

            // 데이터셋 생성 (자체 월평균 포함)
            const datasets = branchMonthlyData.map((b, i) => ({
                label: b.name,
                data: b.data,
                monthlyInfo: b.monthlyInfo,
                ownAvg: b.ownAvg,
                borderColor: colors[i % colors.length],
                backgroundColor: colors[i % colors.length],
                fill: false,
                tension: 0.4,
                pointRadius: 8,
                pointHoverRadius: 12,
                pointStyle: b.data.map(v => v < b.ownAvg ? 'triangle' : 'circle'),
                pointBackgroundColor: b.data.map(v => v < b.ownAvg ? '#ef4444' : colors[i % colors.length]),
                pointBorderColor: b.data.map(v => v < b.ownAvg ? '#ef4444' : colors[i % colors.length]),
                borderWidth: 2,
            }));

            // 평균선 추가
            datasets.push({
                label: '평균',
                data: monthlyAvg,
                borderColor: '#94a3b8',
                borderDash: [5, 5],
                borderWidth: 2,
                pointRadius: 0,
                fill: false,
            });

            // 전년도 비교 데이터 추가
            if (compareData && compareData.by_month) {
                const compMonthMap = Object.fromEntries(compareData.by_month || []);
                branchMonthlyData.forEach((b, i) => {
                    const monthlyInfo = labels.map((_, mi) => {
                        const monthData = compMonthMap[mi+1];
                        const sales = monthData?.byBranch?.[b.name]?.sales || 0;
                        const count = monthData?.byBranch?.[b.name]?.count || 0;
                        return { sales, count, perCase: count > 0 ? sales / count : 0 };
                    });
                    datasets.push({
                        label: b.name + ' (' + compareData.year + ')',
                        data: monthlyInfo.map(d => d.sales),
                        monthlyInfo,
                        borderColor: colors[i % colors.length] + '50',
                        backgroundColor: 'transparent',
                        fill: false,
                        tension: 0.4,
                        pointRadius: 3,
                        borderDash: [3, 3],
                        borderWidth: 1.5,
                        isComparison: true,
                    });
                });
            }

            charts.branchMonthly = new Chart(ctx.getContext('2d'), {
                type: 'line',
                data: { labels, datasets },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { position: 'top', labels: { usePointStyle: true } },
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    const ds = context.dataset;
                                    const label = ds.label || '';
                                    const value = context.raw || 0;
                                    const monthIdx = context.dataIndex;
                                    const info = ds.monthlyInfo?.[monthIdx];

                                    if (label === '평균') return `${label}: ${formatCurrency(value)}`;

                                    // 매출, 건수, 건당 단가 표시
                                    let result = [`${label}: ${formatCurrency(value)}`];
                                    if (info) {
                                        result.push(`  건수: ${info.count.toLocaleString()}건`);
                                        result.push(`  건당: ${formatCurrency(info.perCase)}`);
                                    }

                                    // 자체 월평균 대비 및 검사목적별 증감 (현재 연도만)
                                    if (!ds.isComparison && info && ds.ownAvg) {
                                        const ownAvg = ds.ownAvg;
                                        const diff = value - ownAvg;
                                        const diffPct = ownAvg > 0 ? ((diff / ownAvg) * 100).toFixed(1) : 0;

                                        result.push('─────────');
                                        result.push(`월평균: ${formatCurrency(ownAvg)}`);
                                        if (diff >= 0) {
                                            result.push(`📈 월평균 대비 +${diffPct}%`);
                                            // 평균보다 높은 검사목적 (증가 요인)
                                            const purposeAvg = branchPurposeAvg[label] || {};
                                            const increases = Object.entries(info.byPurpose || {})
                                                .map(([p, d]) => ({ name: p, sales: d.sales, avg: purposeAvg[p] || 0, diff: d.sales - (purposeAvg[p] || 0) }))
                                                .filter(d => d.diff > 0)
                                                .sort((a, b) => b.diff - a.diff)
                                                .slice(0, 3);
                                            if (increases.length > 0) {
                                                result.push('▲ 증가 요인:');
                                                increases.forEach(d => {
                                                    const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                    result.push(`  • ${d.name}: +${formatCurrency(d.diff)} (+${pct}%)`);
                                                });
                                            }
                                        } else {
                                            result.push(`📉 월평균 대비 ${diffPct}%`);
                                            // 평균보다 낮은 검사목적 (감소 요인)
                                            const purposeAvg = branchPurposeAvg[label] || {};
                                            const decreases = Object.entries(purposeAvg)
                                                .map(p => ({ name: p[0], avg: p[1], sales: info.byPurpose?.[p[0]]?.sales || 0 }))
                                                .map(d => ({ ...d, diff: d.sales - d.avg }))
                                                .filter(d => d.diff < 0)
                                                .sort((a, b) => a.diff - b.diff)
                                                .slice(0, 3);
                                            if (decreases.length > 0) {
                                                result.push('▼ 감소 요인:');
                                                decreases.forEach(d => {
                                                    const pct = d.avg > 0 ? ((d.diff / d.avg) * 100).toFixed(0) : 0;
                                                    result.push(`  • ${d.name}: ${formatCurrency(d.diff)} (${pct}%)`);
                                                });
                                            }
                                        }
                                    }

                                    return result;
                                }
                            }
                        }
                    },
                    scales: {
                        y: {
                            ticks: { callback: v => formatCurrency(v) },
                            grid: { color: 'rgba(0,0,0,0.05)' }
                        },
                        x: { grid: { display: false } }
                    }
                }
            });
        }

        // 월별 거래처 중복 현황 (Stacked Bar: 기존 vs 신규)
        function updateClientRetentionChart() {
            const ctx = document.getElementById('clientRetentionChart');
            if (!ctx) return;
            if (charts.clientRetention) charts.clientRetention.destroy();

            const retention = currentData.total_client_retention || [];
            const labels = retention.map(d => d.month + '월');
            const overlap = retention.map(d => d.overlap);
            const newClients = retention.map(d => d.new);

            charts.clientRetention = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels,
                    datasets: [
                        { label: '기존 거래처', data: overlap, backgroundColor: 'rgba(99, 102, 241, 0.8)', borderRadius: 4 },
                        { label: '신규 거래처', data: newClients, backgroundColor: 'rgba(16, 185, 129, 0.8)', borderRadius: 4 }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { position: 'top' } },
                    scales: {
                        x: { stacked: true },
                        y: { stacked: true, title: { display: true, text: '거래처 수' } }
                    }
                }
            });
        }

        // 거래처 리텐션율 추이 (Line Chart)
        function updateRetentionRateChart() {
            const ctx = document.getElementById('retentionRateChart');
            if (!ctx) return;
            if (charts.retentionRate) charts.retentionRate.destroy();

            const retention = currentData.total_client_retention || [];
            const labels = retention.map(d => d.month + '월');
            const rates = retention.map(d => d.retention);
            const totals = retention.map(d => d.total);

            charts.retentionRate = new Chart(ctx.getContext('2d'), {
                type: 'line',
                data: {
                    labels,
                    datasets: [
                        {
                            label: '리텐션율 (%)',
                            data: rates,
                            borderColor: '#6366f1',
                            backgroundColor: 'rgba(99, 102, 241, 0.1)',
                            fill: true,
                            tension: 0.4,
                            yAxisID: 'y'
                        },
                        {
                            label: '월별 거래처 수',
                            data: totals,
                            borderColor: '#10b981',
                            backgroundColor: 'rgba(16, 185, 129, 0.1)',
                            fill: false,
                            tension: 0.4,
                            yAxisID: 'y1'
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { position: 'top' } },
                    scales: {
                        y: {
                            type: 'linear',
                            position: 'left',
                            title: { display: true, text: '리텐션율 (%)' },
                            min: 0,
                            max: 100
                        },
                        y1: {
                            type: 'linear',
                            position: 'right',
                            title: { display: true, text: '거래처 수' },
                            grid: { drawOnChartArea: false }
                        }
                    }
                }
            });
        }

        // 지사별 거래처 리텐션 테이블
        function updateBranchRetentionTable() {
            const tbody = document.getElementById('branchRetentionBody');
            if (!tbody) return;

            const branchRetention = currentData.branch_client_retention || {};
            const branches = Object.keys(branchRetention).sort();

            document.getElementById('branchRetentionBadge').textContent = branches.length + '개';

            let html = '';
            for (const branch of branches) {
                const data = branchRetention[branch] || [];
                const monthMap = Object.fromEntries(data.map(d => [d.month, d]));

                // 누적 거래처 수 계산
                let cumulative = 0;
                for (const d of data) {
                    cumulative += d.new;
                }

                html += `<tr><td><strong>${branch}</strong></td><td class="text-right">${cumulative}</td>`;
                for (let m = 1; m <= 12; m++) {
                    const d = monthMap[m];
                    if (d) {
                        const color = d.retention > 50 ? '#10b981' : d.retention > 30 ? '#f59e0b' : '#ef4444';
                        html += `<td class="text-right"><span style="color:${color}">${d.total}</span><br><small style="color:#888">(+${d.new})</small></td>`;
                    } else {
                        html += '<td class="text-right">-</td>';
                    }
                }
                html += '</tr>';
            }
            tbody.innerHTML = html;
        }

        function updateManagerChart() {
            const purposeFilter = document.getElementById('managerChartPurposeFilter')?.value || '전체';
            let managers = [];

            // 검사목적 필터 적용
            if (purposeFilter === '전체') {
                managers = currentData.by_manager || [];
            } else {
                const purposeManagerData = currentData.purpose_managers?.[purposeFilter] || [];
                managers = purposeManagerData.map(m => [m.name, { sales: m.sales, count: m.count }]);
            }

            // 전체 담당자의 검사목적별 평균 계산
            const allManagers = currentData.by_manager || [];
            const purposeAvgMap = {};
            let totalSales = 0;
            let managerCount = 0;
            allManagers.forEach(m => {
                const byPurpose = m[1].by_purpose || {};
                const sales = m[1].sales || 0;
                if (sales > 0) {
                    managerCount++;
                    totalSales += sales;
                }
                Object.entries(byPurpose).forEach(([purpose, data]) => {
                    if (!purposeAvgMap[purpose]) purposeAvgMap[purpose] = { totalSales: 0, totalCount: 0, managerCount: 0 };
                    purposeAvgMap[purpose].totalSales += data.sales || 0;
                    purposeAvgMap[purpose].totalCount += data.count || 0;
                    if ((data.sales || 0) > 0) purposeAvgMap[purpose].managerCount++;
                });
            });
            // 평균 계산
            Object.keys(purposeAvgMap).forEach(purpose => {
                const p = purposeAvgMap[purpose];
                p.avgSales = p.totalSales / (managerCount || 1);
                p.avgCount = p.totalCount / (managerCount || 1);
            });
            const overallAvgSales = totalSales / (managerCount || 1);

            // 담당자별 상세 정보 준비
            const top15 = managers.slice(0, 15);
            const managerInfoMap = {};
            allManagers.forEach(m => {
                managerInfoMap[m[0]] = m[1];
            });

            const ctx = document.getElementById('managerChart').getContext('2d');
            if (charts.manager) charts.manager.destroy();

            const datasets = [{ label: currentData.year + '년', data: top15.map(d => d[1].sales), backgroundColor: 'rgba(99, 102, 241, 0.8)', borderRadius: 6 }];

            if (compareData && purposeFilter === '전체') {
                const compareMap = Object.fromEntries(compareData.by_manager || []);
                datasets.push({ label: compareData.year + '년', data: top15.map(d => compareMap[d[0]]?.sales || 0), backgroundColor: 'rgba(139, 92, 246, 0.5)', borderRadius: 6 });
                document.getElementById('managerLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background: rgba(99, 102, 241, 0.8);"></div><span>${currentData.year}년</span></div><div class="legend-item"><div class="legend-color" style="background: rgba(139, 92, 246, 0.5);"></div><span>${compareData.year}년</span></div>`;
                document.getElementById('managerLegend').style.display = 'flex';
            } else {
                document.getElementById('managerLegend').style.display = 'none';
            }

            charts.manager = new Chart(ctx, {
                type: 'bar',
                data: { labels: top15.map(d => d[0]), datasets },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                title: function(context) {
                                    return context[0].label;
                                },
                                label: function(context) {
                                    if (context.datasetIndex > 0) {
                                        return context.dataset.label + ': ' + formatCurrency(context.raw);
                                    }
                                    const idx = context.dataIndex;
                                    const name = top15[idx][0];
                                    const info = managerInfoMap[name] || top15[idx][1];
                                    const sales = info.sales || 0;
                                    const count = info.count || 0;
                                    const perCase = count > 0 ? sales / count : 0;

                                    const lines = [];
                                    lines.push('매출: ' + formatCurrency(sales));
                                    lines.push('건수: ' + count + '건, 건당: ' + formatCurrency(perCase));

                                    // 평균 대비
                                    const diffPct = overallAvgSales > 0 ? ((sales - overallAvgSales) / overallAvgSales * 100).toFixed(0) : 0;
                                    lines.push('평균(' + formatCurrency(overallAvgSales) + ') 대비: ' + (diffPct >= 0 ? '+' : '') + diffPct + '%');

                                    return lines;
                                },
                                afterBody: function(context) {
                                    if (context[0].datasetIndex > 0) return [];
                                    if (purposeFilter !== '전체') return [];

                                    const idx = context[0].dataIndex;
                                    const name = top15[idx][0];
                                    const info = managerInfoMap[name];
                                    if (!info) return [];

                                    const sales = info.sales || 0;
                                    const byPurpose = info.by_purpose || {};
                                    const isAboveAvg = sales >= overallAvgSales;

                                    const lines = [''];

                                    // 강점/약점 분석
                                    const strengths = [];
                                    const weaknesses = [];

                                    Object.entries(byPurpose).forEach(([purpose, data]) => {
                                        const avg = purposeAvgMap[purpose]?.avgSales || 0;
                                        if (avg > 0) {
                                            const diff = (data.sales || 0) - avg;
                                            const pct = (diff / avg * 100).toFixed(0);
                                            if (diff > 0) {
                                                strengths.push({ purpose, sales: data.sales, avg, diff, pct: '+' + pct });
                                            } else if (diff < 0) {
                                                weaknesses.push({ purpose, sales: data.sales, avg, diff, pct });
                                            }
                                        }
                                    });

                                    if (isAboveAvg) {
                                        // 평균 이상인 사람: 강점 표시
                                        if (strengths.length > 0) {
                                            strengths.sort((a, b) => b.diff - a.diff);
                                            lines.push('── 강점 (평균 대비 높음) ──');
                                            strengths.slice(0, 3).forEach(s => {
                                                lines.push('▲ ' + s.purpose + ': ' + formatCurrency(s.sales) + ' (' + s.pct + '%)');
                                            });
                                        }
                                    } else {
                                        // 평균 이하인 사람: 약점 표시
                                        if (weaknesses.length > 0) {
                                            weaknesses.sort((a, b) => a.diff - b.diff);
                                            lines.push('── 약점 (평균 대비 낮음) ──');
                                            weaknesses.slice(0, 3).forEach(w => {
                                                lines.push('▼ ' + w.purpose + ': ' + formatCurrency(w.sales) + ' (' + w.pct + '%)');
                                            });
                                        }
                                    }

                                    return lines;
                                }
                            }
                        }
                    },
                    scales: {
                        y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } },
                        x: { grid: { display: false } }
                    }
                }
            });
        }

        function initManagerChartPurposeFilter() {
            const purposes = new Set(['전체']);
            (currentData.by_purpose || []).forEach(p => {
                if (p[0] !== '접수취소') purposes.add(p[0]);
            });
            const select = document.getElementById('managerChartPurposeFilter');
            if (select) {
                select.innerHTML = '<option value="전체">전체 검사목적</option>' +
                    Array.from(purposes).filter(p => p !== '전체').map(p =>
                        `<option value="${p}">${p}</option>`
                    ).join('');
            }
        }

        function updateBranchChart() {
            const purposeFilter = document.getElementById('branchChartPurposeFilter')?.value || '전체';
            const branches = currentData.by_branch || [];

            // 뱃지 업데이트
            document.getElementById('branchChartBadge').textContent = currentData.year + '년';

            const ctx = document.getElementById('branchChart').getContext('2d');
            if (charts.branch) charts.branch.destroy();

            // 검사목적 필터 적용
            const branchData = branches.map(b => {
                let sales = 0, count = 0;
                if (purposeFilter === '전체') {
                    sales = b[1].sales || 0;
                    count = b[1].count || 0;
                } else {
                    const purposeData = b[1].by_purpose?.[purposeFilter];
                    if (purposeData) {
                        sales = purposeData.sales || 0;
                        count = purposeData.count || 0;
                    }
                }
                return { name: b[0], sales, count };
            }).filter(d => d.sales > 0).sort((a, b) => b.sales - a.sales);

            // 총계/평균 계산
            const totalSales = branchData.reduce((sum, d) => sum + d.sales, 0);
            const totalCount = branchData.reduce((sum, d) => sum + d.count, 0);
            const avgSales = branchData.length > 0 ? totalSales / branchData.length : 0;
            const avgCount = branchData.length > 0 ? totalCount / branchData.length : 0;

            if (compareData) {
                // 비교 데이터 처리 (검사목적 필터 적용)
                const compareBranches = compareData.by_branch || [];
                const compareMap = {};
                let compTotalSales = 0, compTotalCount = 0;
                compareBranches.forEach(b => {
                    let sales = 0, count = 0;
                    if (purposeFilter === '전체') {
                        sales = b[1].sales || 0;
                        count = b[1].count || 0;
                    } else {
                        const purposeData = b[1].by_purpose?.[purposeFilter];
                        if (purposeData) {
                            sales = purposeData.sales || 0;
                            count = purposeData.count || 0;
                        }
                    }
                    compareMap[b[0]] = { sales, count };
                    compTotalSales += sales;
                    compTotalCount += count;
                });

                document.getElementById('branchLegend').innerHTML = `
                    <div class="legend-item"><div class="legend-color" style="background: rgba(99, 102, 241, 0.8);"></div><span>${currentData.year}년</span></div>
                    <div class="legend-item"><div class="legend-color" style="background: rgba(139, 92, 246, 0.5);"></div><span>${compareData.year}년</span></div>
                    <div style="margin-left: auto; display: flex; gap: 20px; font-size: 12px; color: #666;">
                        <span>총매출: <strong>${formatCurrency(totalSales)}</strong> (평균 ${formatCurrency(avgSales)})</span>
                        <span>총건수: <strong>${totalCount.toLocaleString()}건</strong> (평균 ${Math.round(avgCount).toLocaleString()}건)</span>
                    </div>`;
                document.getElementById('branchLegend').style.display = 'flex';
                charts.branch = new Chart(ctx, {
                    type: 'bar',
                    data: {
                        labels: branchData.map(d => d.name),
                        datasets: [
                            { label: currentData.year + '년', data: branchData.map(d => d.sales), backgroundColor: 'rgba(99, 102, 241, 0.8)', borderRadius: 6 },
                            { label: compareData.year + '년', data: branchData.map(d => compareMap[d.name]?.sales || 0), backgroundColor: 'rgba(139, 92, 246, 0.5)', borderRadius: 6 }
                        ]
                    },
                    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
                });
            } else {
                document.getElementById('branchLegend').innerHTML = `
                    <div style="display: flex; gap: 20px; font-size: 12px; color: #666;">
                        <span>총매출: <strong>${formatCurrency(totalSales)}</strong> (평균 ${formatCurrency(avgSales)})</span>
                        <span>총건수: <strong>${totalCount.toLocaleString()}건</strong> (평균 ${Math.round(avgCount).toLocaleString()}건)</span>
                    </div>`;
                document.getElementById('branchLegend').style.display = 'flex';
                charts.branch = new Chart(ctx, { type: 'bar', data: { labels: branchData.map(d => d.name), datasets: [{ data: branchData.map(d => d.sales), backgroundColor: 'rgba(99, 102, 241, 0.8)', borderRadius: 6 }] }, options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } }, x: { grid: { display: false } } } } });
            }
        }

        // 월별 탭 전체 업데이트
        function updateMonthlyTab() {
            updateMonthlyKPI();
            updateMonthlyChart();
            updateMonthlyCountChart();
            updateQuarterlyChart();
            updateAvgPriceChart();
            updateYoyChart();
            updateHeatmap();
            updateMonthlyDetailTable();
        }

        // 월별 KPI 업데이트
        function updateMonthlyKPI() {
            const monthly = currentData.by_month || [];
            const monthMap = Object.fromEntries(monthly);
            const monthNames = ['', '1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월'];

            let maxMonth = 0, maxSales = 0, minMonth = 0, minSales = Infinity;
            let totalSales = 0, totalCount = 0, monthCount = 0;

            for (let m = 1; m <= 12; m++) {
                const data = monthMap[m];
                if (data && data.sales > 0) {
                    totalSales += data.sales;
                    totalCount += data.count;
                    monthCount++;
                    if (data.sales > maxSales) { maxSales = data.sales; maxMonth = m; }
                    if (data.sales < minSales) { minSales = data.sales; minMonth = m; }
                }
            }

            document.getElementById('monthlyMaxMonth').textContent = maxMonth > 0 ? monthNames[maxMonth] : '-';
            document.getElementById('monthlyMaxValue').textContent = maxMonth > 0 ? formatCurrency(maxSales) : '-';
            document.getElementById('monthlyMinMonth').textContent = minMonth > 0 && minMonth < 13 ? monthNames[minMonth] : '-';
            document.getElementById('monthlyMinValue').textContent = minMonth < Infinity ? formatCurrency(minSales) : '-';
            document.getElementById('monthlyAvgSales').textContent = monthCount > 0 ? formatCurrency(totalSales / monthCount) : '-';
            document.getElementById('monthlyAvgCount').textContent = monthCount > 0 ? `월평균 ${Math.round(totalCount / monthCount).toLocaleString()}건` : '-';
            document.getElementById('monthlyYtdSales').textContent = formatCurrency(totalSales);
            document.getElementById('monthlyYtdCount').textContent = `총 ${totalCount.toLocaleString()}건`;
        }

        // 월별 매출 차트
        function updateMonthlyChart() {
            const monthly = currentData.by_month || [];
            const ctx = document.getElementById('monthlyChart').getContext('2d');
            if (charts.monthly) charts.monthly.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const monthMap = Object.fromEntries(monthly);
            const datasets = [{ label: currentData.year + '년', data: labels.map((_, i) => monthMap[i+1]?.sales || 0), borderColor: '#6366f1', backgroundColor: 'rgba(99, 102, 241, 0.1)', fill: true, tension: 0.4, pointRadius: 4 }];

            if (compareData) {
                const compMap = Object.fromEntries(compareData.by_month || []);
                datasets.push({ label: compareData.year + '년', data: labels.map((_, i) => compMap[i+1]?.sales || 0), borderColor: '#8b5cf6', backgroundColor: 'rgba(139, 92, 246, 0.1)', fill: true, tension: 0.4, pointRadius: 4 });
                document.getElementById('monthlyLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background: #6366f1;"></div><span>${currentData.year}년</span></div><div class="legend-item"><div class="legend-color" style="background: #8b5cf6;"></div><span>${compareData.year}년</span></div>`;
                document.getElementById('monthlyLegend').style.display = 'flex';
            } else {
                document.getElementById('monthlyLegend').style.display = 'none';
            }

            charts.monthly = new Chart(ctx, { type: 'line', data: { labels, datasets }, options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } }, x: { grid: { display: false } } } } });
        }

        // 월별 건수 차트
        function updateMonthlyCountChart() {
            const monthly = currentData.by_month || [];
            const ctx = document.getElementById('monthlyCountChart').getContext('2d');
            if (charts.monthlyCount) charts.monthlyCount.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const monthMap = Object.fromEntries(monthly);
            const data = labels.map((_, i) => monthMap[i+1]?.count || 0);

            charts.monthlyCount = new Chart(ctx, {
                type: 'bar',
                data: { labels, datasets: [{ label: '건수', data, backgroundColor: 'rgba(34, 197, 94, 0.7)', borderRadius: 6 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true }, x: { grid: { display: false } } } }
            });
        }

        // 분기별 차트
        function updateQuarterlyChart() {
            const monthly = currentData.by_month || [];
            const ctx = document.getElementById('quarterlyChart').getContext('2d');
            if (charts.quarterly) charts.quarterly.destroy();

            const monthMap = Object.fromEntries(monthly);
            const quarters = [0, 0, 0, 0];
            for (let m = 1; m <= 12; m++) {
                const q = Math.floor((m - 1) / 3);
                quarters[q] += monthMap[m]?.sales || 0;
            }

            charts.quarterly = new Chart(ctx, {
                type: 'bar',
                data: { labels: ['1분기', '2분기', '3분기', '4분기'], datasets: [{ data: quarters, backgroundColor: ['#6366f1', '#8b5cf6', '#a855f7', '#d946ef'], borderRadius: 6 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
            });
        }

        // 월별 평균단가 차트
        function updateAvgPriceChart() {
            const monthly = currentData.by_month || [];
            const ctx = document.getElementById('monthlyAvgPriceChart').getContext('2d');
            if (charts.avgPrice) charts.avgPrice.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const monthMap = Object.fromEntries(monthly);
            const data = labels.map((_, i) => {
                const m = monthMap[i+1];
                return m && m.count > 0 ? m.sales / m.count : 0;
            });

            charts.avgPrice = new Chart(ctx, {
                type: 'line',
                data: { labels, datasets: [{ label: '평균단가', data, borderColor: '#f59e0b', backgroundColor: 'rgba(245, 158, 11, 0.1)', fill: true, tension: 0.4 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
            });
        }

        // 전년 대비 차트
        function updateYoyChart() {
            if (!compareData) {
                const ctx = document.getElementById('yoyChart').getContext('2d');
                if (charts.yoy) charts.yoy.destroy();
                charts.yoy = new Chart(ctx, {
                    type: 'bar',
                    data: { labels: ['비교 데이터 없음'], datasets: [{ data: [0], backgroundColor: '#ccc' }] },
                    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } } }
                });
                return;
            }

            const monthly = currentData.by_month || [];
            const compMonthly = compareData.by_month || [];
            const ctx = document.getElementById('yoyChart').getContext('2d');
            if (charts.yoy) charts.yoy.destroy();

            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            const monthMap = Object.fromEntries(monthly);
            const compMap = Object.fromEntries(compMonthly);
            const data = labels.map((_, i) => {
                const curr = monthMap[i+1]?.sales || 0;
                const prev = compMap[i+1]?.sales || 0;
                return prev > 0 ? ((curr - prev) / prev * 100) : 0;
            });

            charts.yoy = new Chart(ctx, {
                type: 'bar',
                data: { labels, datasets: [{ label: '전년대비 (%)', data, backgroundColor: data.map(v => v >= 0 ? 'rgba(34, 197, 94, 0.7)' : 'rgba(239, 68, 68, 0.7)'), borderRadius: 4 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => v + '%' } } } }
            });
        }

        // 히트맵 업데이트
        function updateHeatmap() {
            const monthly = currentData.by_month || [];
            const monthMap = Object.fromEntries(monthly);
            const purposes = {};

            // 모든 목적과 월별 데이터 수집
            for (let m = 1; m <= 12; m++) {
                const byPurpose = monthMap[m]?.byPurpose || {};
                for (const [purpose, data] of Object.entries(byPurpose)) {
                    if (!purposes[purpose]) purposes[purpose] = {};
                    purposes[purpose][m] = data.sales;
                }
            }

            // 헤더 구성
            const headerRow = document.getElementById('heatmapHeader');
            headerRow.innerHTML = '<th>검사목적</th>' + ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'].map(m => `<th class="text-center">${m}</th>`).join('') + '<th class="text-right">합계</th>';

            // 최대값 계산 (색상 스케일용)
            let maxVal = 0;
            for (const purpose of Object.keys(purposes)) {
                for (let m = 1; m <= 12; m++) {
                    const val = purposes[purpose][m] || 0;
                    if (val > maxVal) maxVal = val;
                }
            }

            // 본문 구성
            const tbody = document.getElementById('heatmapBody');
            const purposeEntries = Object.entries(purposes).sort((a, b) => {
                const aSum = Object.values(a[1]).reduce((s, v) => s + v, 0);
                const bSum = Object.values(b[1]).reduce((s, v) => s + v, 0);
                return bSum - aSum;
            });

            tbody.innerHTML = purposeEntries.map(([purpose, months]) => {
                const cells = [];
                let sum = 0;
                for (let m = 1; m <= 12; m++) {
                    const val = months[m] || 0;
                    sum += val;
                    const intensity = maxVal > 0 ? val / maxVal : 0;
                    const bgColor = val > 0 ? `rgba(99, 102, 241, ${0.1 + intensity * 0.7})` : '';
                    const textColor = intensity > 0.5 ? '#fff' : '#333';
                    cells.push(`<td class="text-center" style="background: ${bgColor}; color: ${textColor};">${val > 0 ? formatCurrency(val) : '-'}</td>`);
                }
                return `<tr><td>${purpose}</td>${cells.join('')}<td class="text-right font-bold">${formatCurrency(sum)}</td></tr>`;
            }).join('');
        }

        // 월별 상세 테이블
        function updateMonthlyDetailTable() {
            const monthly = currentData.by_month || [];
            const monthMap = Object.fromEntries(monthly);
            const totalSales = currentData.total_sales || 1;

            const tbody = document.querySelector('#monthlyDetailTable tbody');
            let rows = [];
            let activeMonths = 0;

            for (let m = 1; m <= 12; m++) {
                const data = monthMap[m];
                if (data && data.sales > 0) {
                    activeMonths++;
                    const avgPrice = data.count > 0 ? data.sales / data.count : 0;
                    const percent = (data.sales / totalSales * 100).toFixed(1);
                    rows.push(`<tr>
                        <td>${m}월</td>
                        <td class="text-right">${formatCurrency(data.sales)}</td>
                        <td class="text-right">${data.count.toLocaleString()}건</td>
                        <td class="text-right">${formatCurrency(avgPrice)}</td>
                        <td class="text-right">${percent}%</td>
                        <td class="text-center"><button class="btn btn-sm" onclick="showMonthDetail(${m})">상세</button></td>
                    </tr>`);
                }
            }

            tbody.innerHTML = rows.join('');
            document.getElementById('monthlyTableBadge').textContent = `${activeMonths}개월`;
        }

        // 월 상세 모달
        let monthPurposeChart = null;
        let monthManagerChart = null;

        function showMonthDetail(month) {
            const monthly = currentData.by_month || [];
            const monthMap = Object.fromEntries(monthly);
            const data = monthMap[month];

            if (!data) return;

            document.getElementById('monthModalTitle').textContent = `${month}월 상세 분석`;
            document.getElementById('monthModal').style.display = 'flex';

            // 검사목적별 도넛 차트
            const purposeCtx = document.getElementById('monthPurposeChart').getContext('2d');
            if (monthPurposeChart) monthPurposeChart.destroy();

            const byPurpose = data.byPurpose || {};
            const purposeLabels = Object.keys(byPurpose);
            const purposeValues = purposeLabels.map(p => byPurpose[p].sales);

            monthPurposeChart = new Chart(purposeCtx, {
                type: 'doughnut',
                data: { labels: purposeLabels, datasets: [{ data: purposeValues, backgroundColor: ['#6366f1', '#8b5cf6', '#a855f7', '#d946ef', '#ec4899', '#f43f5e', '#f59e0b', '#22c55e'] }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { boxWidth: 12 } } } }
            });

            // 담당자별 도넛 차트
            const managerCtx = document.getElementById('monthManagerChart').getContext('2d');
            if (monthManagerChart) monthManagerChart.destroy();

            const byManager = data.byManager || {};
            const managerEntries = Object.entries(byManager).sort((a, b) => b[1].sales - a[1].sales).slice(0, 8);
            const managerLabels = managerEntries.map(e => e[0]);
            const managerValues = managerEntries.map(e => e[1].sales);

            monthManagerChart = new Chart(managerCtx, {
                type: 'doughnut',
                data: { labels: managerLabels, datasets: [{ data: managerValues, backgroundColor: ['#6366f1', '#8b5cf6', '#a855f7', '#d946ef', '#ec4899', '#f43f5e', '#f59e0b', '#22c55e'] }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { boxWidth: 12 } } } }
            });

            // 주요 지표
            const avgPrice = data.count > 0 ? data.sales / data.count : 0;
            const statsHtml = `
                <div style="background: #f1f5f9; padding: 15px; border-radius: 8px; text-align: center;">
                    <div style="font-size: 0.85rem; color: #64748b;">매출액</div>
                    <div style="font-size: 1.3rem; font-weight: bold; color: #6366f1;">${formatCurrency(data.sales)}</div>
                </div>
                <div style="background: #f1f5f9; padding: 15px; border-radius: 8px; text-align: center;">
                    <div style="font-size: 0.85rem; color: #64748b;">건수</div>
                    <div style="font-size: 1.3rem; font-weight: bold; color: #22c55e;">${data.count.toLocaleString()}건</div>
                </div>
                <div style="background: #f1f5f9; padding: 15px; border-radius: 8px; text-align: center;">
                    <div style="font-size: 0.85rem; color: #64748b;">평균단가</div>
                    <div style="font-size: 1.3rem; font-weight: bold; color: #f59e0b;">${formatCurrency(avgPrice)}</div>
                </div>
                <div style="background: #f1f5f9; padding: 15px; border-radius: 8px; text-align: center;">
                    <div style="font-size: 0.85rem; color: #64748b;">검사목적 수</div>
                    <div style="font-size: 1.3rem; font-weight: bold; color: #8b5cf6;">${Object.keys(byPurpose).length}개</div>
                </div>
            `;
            document.getElementById('monthDetailStats').innerHTML = statsHtml;
        }

        function closeMonthModal() {
            document.getElementById('monthModal').style.display = 'none';
        }

        // 테이블 정렬 함수
        function sortManagerTable(column) {
            if (managerTableSort.column === column) {
                managerTableSort.direction = managerTableSort.direction === 'asc' ? 'desc' : 'asc';
            } else {
                managerTableSort.column = column;
                managerTableSort.direction = 'desc';
            }
            updateManagerTable();
        }

        function updateManagerTable() {
            const purposeFilter = document.getElementById('managerPurposeFilter')?.value || '전체';
            let managers = [];

            // 원본 by_manager 데이터 맵 (긴급 데이터 참조용)
            const originalManagerMap = Object.fromEntries((currentData.by_manager || []).map(m => [m[0], m[1]]));

            // 검사목적 필터 적용
            if (purposeFilter === '전체') {
                managers = [...(currentData.by_manager || [])];
            } else {
                // purpose_managers에서 해당 목적의 담당자 데이터 가져오기
                const purposeManagerData = currentData.purpose_managers?.[purposeFilter] || [];
                managers = purposeManagerData.map(m => {
                    // 원본 데이터에서 해당 목적의 긴급 건수 가져오기
                    const originalData = originalManagerMap[m.name] || {};
                    const urgentByPurpose = originalData.urgent_by_purpose || {};
                    const urgentCount = urgentByPurpose[purposeFilter] || 0;
                    return [m.name, { sales: m.sales, count: m.count, urgent: urgentCount }];
                });
            }

            const tbody = document.querySelector('#managerTable tbody');
            const total = purposeFilter === '전체' ? (currentData.total_sales || 1) : managers.reduce((sum, m) => sum + m[1].sales, 0) || 1;
            const workingDays = 250;
            const compareMap = compareData ? Object.fromEntries(compareData.by_manager || []) : {};

            // 정렬 적용
            if (managerTableSort.column) {
                const col = managerTableSort.column;
                const dir = managerTableSort.direction === 'asc' ? 1 : -1;
                managers.sort((a, b) => {
                    let aVal, bVal;
                    const aComp = compareMap[a[0]] || {};
                    const bComp = compareMap[b[0]] || {};
                    switch(col) {
                        case 'name': aVal = a[0]; bVal = b[0]; return dir * aVal.localeCompare(bVal); break;
                        case 'sales': aVal = a[1].sales || 0; bVal = b[1].sales || 0; break;
                        case 'count': aVal = a[1].count || 0; bVal = b[1].count || 0; break;
                        case 'avgPrice': aVal = (a[1].count > 0 ? a[1].sales / a[1].count : 0); bVal = (b[1].count > 0 ? b[1].sales / b[1].count : 0); break;
                        case 'dailyAvg': aVal = a[1].sales / workingDays; bVal = b[1].sales / workingDays; break;
                        case 'urgent': aVal = a[1].urgent || 0; bVal = b[1].urgent || 0; break;
                        case 'compSales': aVal = aComp.sales || 0; bVal = bComp.sales || 0; break;
                        case 'compAvgPrice': aVal = aComp.count > 0 ? aComp.sales / aComp.count : 0; bVal = bComp.count > 0 ? bComp.sales / bComp.count : 0; break;
                        case 'change':
                            const aCompS = aComp.sales || 0;
                            const bCompS = bComp.sales || 0;
                            aVal = aCompS > 0 ? ((a[1].sales - aCompS) / aCompS * 100) : 0;
                            bVal = bCompS > 0 ? ((b[1].sales - bCompS) / bCompS * 100) : 0;
                            break;
                        case 'percent': aVal = a[1].sales / total; bVal = b[1].sales / total; break;
                        default: aVal = a[1].sales || 0; bVal = b[1].sales || 0;
                    }
                    return dir * (aVal - bVal);
                });
            }

            // 테이블 배지 업데이트
            const badgeEl = document.getElementById('managerTableBadge');
            if (badgeEl) badgeEl.textContent = managers.length + '명';

            // 정렬 클래스 헬퍼
            const sortClass = (col) => {
                if (managerTableSort.column === col) return `sortable ${managerTableSort.direction}`;
                return 'sortable';
            };

            if (compareData) {
                document.getElementById('managerTableHead').innerHTML = `<tr>
                    <th class="${sortClass('name')}" onclick="sortManagerTable('name')">담당자</th>
                    <th class="text-right ${sortClass('sales')}" onclick="sortManagerTable('sales')">${currentData.year}년</th>
                    <th class="text-right ${sortClass('avgPrice')}" onclick="sortManagerTable('avgPrice')">${currentData.year}년 평균단가</th>
                    <th class="text-right ${sortClass('compSales')}" onclick="sortManagerTable('compSales')">${compareData.year}년</th>
                    <th class="text-right ${sortClass('compAvgPrice')}" onclick="sortManagerTable('compAvgPrice')">${compareData.year}년 평균단가</th>
                    <th class="text-right ${sortClass('urgent')}" onclick="sortManagerTable('urgent')">긴급</th>
                    <th class="text-right ${sortClass('change')}" onclick="sortManagerTable('change')">증감</th>
                    <th class="${sortClass('percent')}" onclick="sortManagerTable('percent')">비중</th>
                    <th class="text-center">상세</th>
                </tr>`;
                tbody.innerHTML = managers.map(d => {
                    const compData = compareMap[d[0]] || {};
                    const compSales = compData.sales || 0;
                    const compCount = compData.count || 0;
                    const diff = d[1].sales - compSales;
                    const diffRate = compSales > 0 ? ((diff / compSales) * 100).toFixed(1) : 0;
                    const percent = (d[1].sales / total * 100).toFixed(1);
                    const avgPrice = (d[1].count || 0) > 0 ? d[1].sales / d[1].count : 0;
                    const compAvgPrice = compCount > 0 ? compSales / compCount : 0;
                    const urgent = d[1].urgent || 0;
                    return `<tr>
                        <td><strong>${d[0]}</strong></td>
                        <td class="text-right">${formatCurrency(d[1].sales)}</td>
                        <td class="text-right">${formatCurrency(avgPrice)}</td>
                        <td class="text-right" style="color: var(--gray-400);">${formatCurrency(compSales)}</td>
                        <td class="text-right" style="color: var(--gray-400);">${formatCurrency(compAvgPrice)}</td>
                        <td class="text-right"><span class="urgent-badge">🚨 ${urgent}건</span></td>
                        <td class="text-right"><span class="change-badge ${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${diffRate}%</span></td>
                        <td><div class="progress-cell"><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%;"></div></div><span class="progress-value">${percent}%</span></div></td>
                        <td class="text-center"><button class="btn-detail" onclick="showManagerDetail('${d[0]}')">상세</button></td>
                    </tr>`;
                }).join('');
            } else {
                document.getElementById('managerTableHead').innerHTML = `<tr>
                    <th class="${sortClass('name')}" onclick="sortManagerTable('name')">담당자</th>
                    <th class="text-right ${sortClass('sales')}" onclick="sortManagerTable('sales')">매출액</th>
                    <th class="text-right ${sortClass('count')}" onclick="sortManagerTable('count')">건수</th>
                    <th class="text-right ${sortClass('avgPrice')}" onclick="sortManagerTable('avgPrice')">평균단가</th>
                    <th class="text-right ${sortClass('dailyAvg')}" onclick="sortManagerTable('dailyAvg')">일평균</th>
                    <th class="text-right ${sortClass('urgent')}" onclick="sortManagerTable('urgent')">긴급</th>
                    <th class="${sortClass('percent')}" onclick="sortManagerTable('percent')">비중</th>
                    <th class="text-center">상세</th>
                </tr>`;
                tbody.innerHTML = managers.map(d => {
                    const percent = (d[1].sales / total * 100).toFixed(1);
                    const avgPrice = (d[1].count || 0) > 0 ? d[1].sales / d[1].count : 0;
                    const dailyAvg = d[1].sales / workingDays;
                    const urgent = d[1].urgent || 0;
                    return `<tr>
                        <td><strong>${d[0]}</strong></td>
                        <td class="text-right">${formatCurrency(d[1].sales)}</td>
                        <td class="text-right">${(d[1].count || 0).toLocaleString()}</td>
                        <td class="text-right">${formatCurrency(avgPrice)}</td>
                        <td class="text-right">${formatCurrency(dailyAvg)}</td>
                        <td class="text-right"><span class="urgent-badge">🚨 ${urgent}건</span></td>
                        <td><div class="progress-cell"><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%;"></div></div><span class="progress-value">${percent}%</span></div></td>
                        <td class="text-center"><button class="btn-detail" onclick="showManagerDetail('${d[0]}')">상세</button></td>
                    </tr>`;
                }).join('');
            }
        }

        function updateBranchTable() {
            const purposeFilter = document.getElementById('branchTablePurposeFilter')?.value || '전체';
            const branches = currentData.by_branch || [];
            const tbody = document.querySelector('#branchTable tbody');

            // 검사목적 필터 적용
            const branchData = branches.map(b => {
                let sales = 0, count = 0;
                if (purposeFilter === '전체') {
                    sales = b[1].sales || 0;
                    count = b[1].count || 0;
                } else {
                    const purposeData = b[1].by_purpose?.[purposeFilter];
                    if (purposeData) {
                        sales = purposeData.sales || 0;
                        count = purposeData.count || 0;
                    }
                }
                return { name: b[0], sales, count, managers: b[1].managers };
            }).filter(d => d.sales > 0).sort((a, b) => b.sales - a.sales);

            const total = branchData.reduce((sum, b) => sum + b.sales, 0) || 1;
            document.getElementById('branchTableBadge').textContent = branchData.length + '개 팀';

            if (compareData) {
                // 비교 데이터 처리 (검사목적 필터 적용)
                const compareBranches = compareData.by_branch || [];
                const compareMap = {};
                compareBranches.forEach(b => {
                    let sales = 0, count = 0;
                    if (purposeFilter === '전체') {
                        sales = b[1].sales || 0;
                        count = b[1].count || 0;
                    } else {
                        const purposeData = b[1].by_purpose?.[purposeFilter];
                        if (purposeData) {
                            sales = purposeData.sales || 0;
                            count = purposeData.count || 0;
                        }
                    }
                    compareMap[b[0]] = { sales, count };
                });

                document.getElementById('branchTableHead').innerHTML = `<tr><th>팀명</th><th class="text-right">${currentData.year}년</th><th class="text-right">${compareData.year}년</th><th class="text-right">평균단가</th><th class="text-right">증감</th><th>비중</th></tr>`;
                tbody.innerHTML = branchData.map(d => {
                    const compSales = compareMap[d.name]?.sales || 0;
                    const diff = d.sales - compSales;
                    const diffRate = compSales > 0 ? ((diff / compSales) * 100).toFixed(1) : 0;
                    const avgPrice = d.count > 0 ? d.sales / d.count : 0;
                    const percent = (d.sales / total * 100).toFixed(1);
                    return `<tr>
                        <td><strong>${d.name}</strong></td>
                        <td class="text-right">${formatCurrency(d.sales)}</td>
                        <td class="text-right" style="color: var(--gray-400);">${formatCurrency(compSales)}</td>
                        <td class="text-right">${formatCurrency(avgPrice)}</td>
                        <td class="text-right"><span class="change-badge ${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${diffRate}%</span></td>
                        <td><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%"></div><span>${percent}%</span></div></td>
                    </tr>`;
                }).join('');
            } else {
                document.getElementById('branchTableHead').innerHTML = `<tr><th>팀명</th><th class="text-right">매출액</th><th class="text-right">건수</th><th class="text-right">평균단가</th><th class="text-right">담당자수</th><th>비중</th></tr>`;
                tbody.innerHTML = branchData.map(d => {
                    const avgPrice = d.count > 0 ? d.sales / d.count : 0;
                    const percent = (d.sales / total * 100).toFixed(1);
                    return `<tr>
                        <td><strong>${d.name}</strong></td>
                        <td class="text-right">${formatCurrency(d.sales)}</td>
                        <td class="text-right">${d.count.toLocaleString()}건</td>
                        <td class="text-right">${formatCurrency(avgPrice)}</td>
                        <td class="text-right">${d.managers?.size || d.managers || '-'}명</td>
                        <td><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%"></div><span>${percent}%</span></div></td>
                    </tr>`;
                }).join('');
            }
        }

        // 업체별 탭 전역 변수
        let clientTableMode = 'new';
        let clientAnalysisData = null;

        function updateClientTab() {
            const clients = currentData.by_client || [];
            const compareClients = compareData?.by_client || [];

            // 비교 데이터로 신규/유지/이탈 분류
            const currentClientMap = Object.fromEntries(clients.map(c => [c[0], c[1]]));
            const compareClientMap = Object.fromEntries(compareClients.map(c => [c[0], c[1]]));

            const newClients = [];      // 신규: 올해만 있음
            const retainedClients = []; // 유지: 양쪽 모두 있음
            const churnedClients = [];  // 이탈: 전년만 있음

            // 현재 연도 업체 분류
            clients.forEach(c => {
                const name = c[0];
                const data = c[1];
                if (compareClientMap[name]) {
                    retainedClients.push({
                        name,
                        ...data,
                        lastYearSales: compareClientMap[name].sales,
                        lastYearCount: compareClientMap[name].count,
                        growth: data.sales - compareClientMap[name].sales,
                        growthRate: compareClientMap[name].sales > 0 ? ((data.sales - compareClientMap[name].sales) / compareClientMap[name].sales * 100) : 0,
                        status: 'retained'
                    });
                } else {
                    newClients.push({ name, ...data, status: 'new' });
                }
            });

            // 이탈 업체 (전년만 있고 올해 없음)
            compareClients.forEach(c => {
                if (!currentClientMap[c[0]]) {
                    churnedClients.push({
                        name: c[0],
                        lastYearSales: c[1].sales,
                        lastYearCount: c[1].count,
                        manager: c[1].manager || '미지정',
                        purpose: c[1].purpose || '',
                        status: 'churned'
                    });
                }
            });

            // VIP 업체 (1억 이상)
            const vipClients = clients.filter(c => c[1].sales >= 100000000);

            // 분석 데이터 저장
            clientAnalysisData = { newClients, retainedClients, churnedClients, vipClients, clients };

            // KPI 업데이트
            document.getElementById('clientTotalCount').textContent = clients.length + '개';
            document.getElementById('clientTotalCompare').textContent = '전년: ' + compareClients.length + '개';
            document.getElementById('clientNewCount').textContent = newClients.length + '개';
            document.getElementById('clientRetainedCount').textContent = retainedClients.length + '개';
            document.getElementById('clientChurnedCount').textContent = churnedClients.length + '개';
            document.getElementById('clientVipCount').textContent = vipClients.length + '개';

            // 담당자별 통계 계산
            updateManagerKPIs(clients, newClients, retainedClients, churnedClients, vipClients, compareClientMap);

            // 차트 업데이트
            updateClientSalesChart(clients, newClients, retainedClients);
            updateClientCountChart(clients, newClients, retainedClients);

            // 테이블 업데이트
            updateRetainedClientTable(retainedClients);
            updateNewChurnClientTable();
            updateClientByPurposeTable(clients);
            updateClientByManagerTable(clients, newClients, retainedClients, churnedClients, compareClientMap);

            // 버튼 카운트
            document.getElementById('newClientsBtnCount').textContent = newClients.length;
            document.getElementById('churnedClientsBtnCount').textContent = churnedClients.length;
        }

        function updateManagerKPIs(clients, newClients, retainedClients, churnedClients, vipClients, compareClientMap) {
            // 담당자별 집계
            const managerStats = {};

            clients.forEach(c => {
                const manager = c[1].manager || '미지정';
                if (!managerStats[manager]) {
                    managerStats[manager] = {
                        totalClients: 0, newClients: 0, retainedClients: 0, vipClients: 0,
                        totalSales: 0, lastYearSales: 0, totalTradeMonths: 0, activeClients: 0
                    };
                }
                managerStats[manager].totalClients++;
                managerStats[manager].totalSales += c[1].sales;
                managerStats[manager].totalTradeMonths += c[1].tradeMonths || 0;
                if (c[1].count >= 36) managerStats[manager].activeClients++;  // 월3회 이상
                if (c[1].sales >= 100000000) managerStats[manager].vipClients++;
            });

            newClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (managerStats[manager]) managerStats[manager].newClients++;
            });

            retainedClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (managerStats[manager]) {
                    managerStats[manager].retainedClients++;
                    managerStats[manager].lastYearSales += c.lastYearSales || 0;
                }
            });

            // 이탈 업체 담당자별 집계
            const churnByManager = {};
            churnedClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (!churnByManager[manager]) churnByManager[manager] = 0;
                churnByManager[manager]++;
            });

            // 배열로 변환 및 정렬
            const managerArray = Object.entries(managerStats).map(([name, stats]) => ({
                name,
                ...stats,
                avgTradeMonths: stats.totalClients > 0 ? stats.totalTradeMonths / stats.totalClients : 0,
                activeRate: stats.totalClients > 0 ? (stats.activeClients / stats.totalClients * 100) : 0,
                salesGrowth: stats.totalSales - stats.lastYearSales,
                retentionRate: (stats.retainedClients + stats.newClients) > 0 ? (stats.retainedClients / (stats.retainedClients + (churnByManager[name] || 0)) * 100) : 0,
                churnedClients: churnByManager[name] || 0
            }));

            // 평균 계산
            const avgClients = managerArray.reduce((s, m) => s + m.totalClients, 0) / (managerArray.length || 1);
            const avgNew = managerArray.reduce((s, m) => s + m.newClients, 0) / (managerArray.length || 1);
            const avgGrowth = managerArray.reduce((s, m) => s + m.salesGrowth, 0) / (managerArray.length || 1);
            const avgVip = managerArray.reduce((s, m) => s + m.vipClients, 0) / (managerArray.length || 1);
            const avgRetention = managerArray.reduce((s, m) => s + m.retentionRate, 0) / (managerArray.length || 1);
            const avgTradeMonths = managerArray.reduce((s, m) => s + m.avgTradeMonths, 0) / (managerArray.length || 1);
            const avgActiveRate = managerArray.reduce((s, m) => s + m.activeRate, 0) / (managerArray.length || 1);
            const avgChurn = managerArray.reduce((s, m) => s + m.churnedClients, 0) / (managerArray.length || 1);

            // KPI 카드 업데이트 함수
            const updateKpiCard = (id, data, valueFormatter, avgValue, isLowerBetter = false) => {
                const sorted = [...data].sort((a, b) => isLowerBetter ? a.value - b.value : b.value - a.value);
                const qualified = sorted.filter(d => isLowerBetter ? d.value <= avgValue : d.value >= avgValue);

                const nameEl = document.getElementById(id + 'Name');
                const valueEl = document.getElementById(id + 'Value');

                if (qualified.length === 0) {
                    nameEl.textContent = '-';
                    valueEl.textContent = '-';
                } else if (qualified.length === 1) {
                    nameEl.textContent = qualified[0].name;
                    valueEl.textContent = valueFormatter(qualified[0].value);
                } else {
                    nameEl.textContent = qualified[0].name + ' 외 ' + (qualified.length - 1) + '명';
                    valueEl.textContent = '평균 ' + valueFormatter(avgValue) + '↑';
                }

                // 오버레이 생성
                const overlay = document.getElementById(id + 'Overlay');
                if (overlay) {
                    overlay.innerHTML = `
                        <div style="font-weight: 600; margin-bottom: 8px;">평균: ${valueFormatter(avgValue)}</div>
                        <div style="border-top: 1px dashed #e2e8f0; margin: 8px 0;"></div>
                        ${sorted.map((d, i) => {
                            const isAboveAvg = isLowerBetter ? d.value <= avgValue : d.value >= avgValue;
                            return `<div style="display: flex; justify-content: space-between; padding: 4px 0; ${i === sorted.findIndex(x => isLowerBetter ? x.value > avgValue : x.value < avgValue) ? 'border-top: 1px dashed #94a3b8; margin-top: 4px; padding-top: 8px;' : ''}">
                                <span>${i + 1}. ${d.name}</span>
                                <span>${valueFormatter(d.value)} ${isAboveAvg ? '⭐' : ''}</span>
                            </div>`;
                        }).join('')}
                    `;
                }
            };

            // 각 KPI 업데이트
            updateKpiCard('kpiClientKing', managerArray.map(m => ({ name: m.name, value: m.totalClients })), v => v + '개', avgClients);
            updateKpiCard('kpiNewKing', managerArray.map(m => ({ name: m.name, value: m.newClients })), v => v + '개 유치', avgNew);
            updateKpiCard('kpiGrowthKing', managerArray.map(m => ({ name: m.name, value: m.salesGrowth })), v => (v >= 0 ? '+' : '') + formatCurrency(v), avgGrowth);
            updateKpiCard('kpiVipKing', managerArray.map(m => ({ name: m.name, value: m.vipClients })), v => v + '개 VIP', avgVip);
            updateKpiCard('kpiRetentionKing', managerArray.map(m => ({ name: m.name, value: m.retentionRate })), v => v.toFixed(0) + '% 유지', avgRetention);
            updateKpiCard('kpiSteadyKing', managerArray.map(m => ({ name: m.name, value: m.avgTradeMonths })), v => '평균 ' + v.toFixed(1) + '월', avgTradeMonths);
            updateKpiCard('kpiActiveKing', managerArray.map(m => ({ name: m.name, value: m.activeRate })), v => v.toFixed(0) + '% 활성', avgActiveRate);
            updateKpiCard('kpiChurnWarning', managerArray.map(m => ({ name: m.name, value: m.churnedClients })), v => v + '개 이탈', avgChurn, true);

            // 오버레이 이벤트 등록
            document.querySelectorAll('.manager-kpi-card').forEach(card => {
                const overlay = card.querySelector('.manager-kpi-overlay');
                if (overlay) {
                    card.addEventListener('mouseenter', () => {
                        overlay.style.display = 'block';
                        overlay.style.position = 'absolute';
                        overlay.style.top = '100%';
                        overlay.style.left = '0';
                        overlay.style.width = '220px';
                        overlay.style.background = 'white';
                        overlay.style.border = '2px solid #6366f1';
                        overlay.style.borderRadius = '8px';
                        overlay.style.padding = '12px';
                        overlay.style.boxShadow = '0 10px 40px rgba(0,0,0,0.2)';
                        overlay.style.zIndex = '1000';
                        overlay.style.fontSize = '12px';
                    });
                    card.addEventListener('mouseleave', () => { overlay.style.display = 'none'; });
                }
            });
        }

        function updateClientSalesChart(clients, newClients, retainedClients) {
            const top10 = clients.slice(0, 10);
            const newClientNames = new Set(newClients.map(c => c.name));
            const retainedMap = Object.fromEntries(retainedClients.map(c => [c.name, c]));

            document.getElementById('clientSalesChartBadge').textContent = currentData.year + '년';

            const ctx = document.getElementById('clientSalesChart');
            if (!ctx) return;
            if (charts.clientSales) charts.clientSales.destroy();

            charts.clientSales = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: top10.map(c => c[0].length > 8 ? c[0].substring(0, 8) + '..' : c[0]),
                    datasets: [{
                        label: '매출',
                        data: top10.map(c => c[1].sales),
                        backgroundColor: top10.map(c => newClientNames.has(c[0]) ? 'rgba(16, 185, 129, 0.8)' : 'rgba(99, 102, 241, 0.8)'),
                        borderRadius: 6
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                title: ctx => {
                                    const idx = ctx[0].dataIndex;
                                    const name = top10[idx][0];
                                    const status = newClientNames.has(name) ? '신규' : '유지';
                                    const rank = idx === 0 ? '🥇' : idx === 1 ? '🥈' : idx === 2 ? '🥉' : (idx + 1) + '위';
                                    return rank + ' ' + name + '  ' + status;
                                },
                                label: ctx => {
                                    const idx = ctx.dataIndex;
                                    const c = top10[idx];
                                    const lines = [];
                                    lines.push('💰 연간 매출: ' + formatCurrency(c[1].sales));
                                    lines.push('📋 연간 건수: ' + c[1].count.toLocaleString() + '건');
                                    lines.push('📊 건당 매출: ' + formatCurrency(c[1].avg));
                                    return lines;
                                },
                                afterBody: ctx => {
                                    const idx = ctx[0].dataIndex;
                                    const name = top10[idx][0];
                                    const retained = retainedMap[name];
                                    const c = top10[idx][1];
                                    const lines = [];
                                    if (retained) {
                                        lines.push('');
                                        lines.push('전년 매출: ' + formatCurrency(retained.lastYearSales));
                                        const growthPct = retained.growthRate.toFixed(1);
                                        lines.push('증감률: ' + (growthPct >= 0 ? '+' : '') + growthPct + '%');
                                    }
                                    lines.push('');
                                    lines.push('📌 상세 정보');
                                    lines.push('담당자: ' + (c.manager || '미지정'));
                                    lines.push('주요 검사: ' + (c.purpose || '-'));
                                    return lines;
                                }
                            }
                        }
                    },
                    scales: {
                        y: { beginAtZero: true, ticks: { callback: v => formatCurrency(v) } },
                        x: { ticks: { maxRotation: 45, minRotation: 45 } }
                    }
                }
            });
        }

        function updateClientCountChart(clients, newClients, retainedClients) {
            const sorted = [...clients].sort((a, b) => b[1].count - a[1].count);
            const top10 = sorted.slice(0, 10);
            const newClientNames = new Set(newClients.map(c => c.name));
            const retainedMap = Object.fromEntries(retainedClients.map(c => [c.name, c]));

            document.getElementById('clientCountChartBadge').textContent = currentData.year + '년';

            const ctx = document.getElementById('clientCountChart');
            if (!ctx) return;
            if (charts.clientCount) charts.clientCount.destroy();

            charts.clientCount = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: top10.map(c => c[0].length > 8 ? c[0].substring(0, 8) + '..' : c[0]),
                    datasets: [{
                        label: '건수',
                        data: top10.map(c => c[1].count),
                        backgroundColor: top10.map(c => newClientNames.has(c[0]) ? 'rgba(16, 185, 129, 0.8)' : 'rgba(99, 102, 241, 0.8)'),
                        borderRadius: 6
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                title: ctx => {
                                    const idx = ctx[0].dataIndex;
                                    const name = top10[idx][0];
                                    const status = newClientNames.has(name) ? '신규' : '유지';
                                    const rank = idx === 0 ? '🥇' : idx === 1 ? '🥈' : idx === 2 ? '🥉' : (idx + 1) + '위';
                                    return rank + ' ' + name + '  ' + status;
                                },
                                label: ctx => {
                                    const idx = ctx.dataIndex;
                                    const c = top10[idx];
                                    const lines = [];
                                    lines.push('📋 연간 건수: ' + c[1].count.toLocaleString() + '건');
                                    lines.push('💰 연간 매출: ' + formatCurrency(c[1].sales));
                                    lines.push('📊 건당 매출: ' + formatCurrency(c[1].avg));
                                    return lines;
                                },
                                afterBody: ctx => {
                                    const idx = ctx[0].dataIndex;
                                    const c = top10[idx][1];
                                    const lines = [];
                                    lines.push('');
                                    lines.push('📌 상세 정보');
                                    lines.push('담당자: ' + (c.manager || '미지정'));
                                    lines.push('주요 검사: ' + (c.purpose || '-'));
                                    return lines;
                                }
                            }
                        }
                    },
                    scales: {
                        y: { beginAtZero: true },
                        x: { ticks: { maxRotation: 45, minRotation: 45 } }
                    }
                }
            });
        }

        function updateRetainedClientTable(retainedClients) {
            const sorted = [...retainedClients].sort((a, b) => b.growthRate - a.growthRate);
            document.getElementById('retainedTableBadge').textContent = sorted.length + '개';

            const tbody = document.querySelector('#retainedClientTable tbody');
            tbody.innerHTML = sorted.map(c => {
                const growthClass = c.growthRate >= 0 ? 'color: var(--success)' : 'color: var(--danger)';
                const growthSign = c.growthRate >= 0 ? '+' : '';
                return `<tr>
                    <td><strong>${c.name}</strong></td>
                    <td>${c.manager || '-'}</td>
                    <td class="text-right">${formatCurrency(c.sales)}</td>
                    <td class="text-right">${formatCurrency(c.lastYearSales)}</td>
                    <td class="text-right"><span style="${growthClass}; font-weight: 600;">${growthSign}${c.growthRate.toFixed(1)}%</span></td>
                </tr>`;
            }).join('');
        }

        function setClientTableMode(mode) {
            clientTableMode = mode;
            document.getElementById('btnNewClients').classList.toggle('active', mode === 'new');
            document.getElementById('btnChurnedClients').classList.toggle('active', mode === 'churned');
            document.getElementById('newChurnTableTitle').textContent = mode === 'new' ? '🆕 신규 업체' : '📤 이탈 업체';

            const thead = document.getElementById('newChurnTableHead');
            if (mode === 'new') {
                thead.innerHTML = '<tr><th>업체명</th><th>담당자</th><th class="text-right">매출액</th><th class="text-right">건수</th><th>주요 검사</th></tr>';
            } else {
                thead.innerHTML = '<tr><th>업체명</th><th>담당자</th><th class="text-right">전년 매출</th><th>주요 검사</th></tr>';
            }
            updateNewChurnClientTable();
        }

        function updateNewChurnClientTable() {
            if (!clientAnalysisData) return;
            const tbody = document.querySelector('#newChurnClientTable tbody');

            if (clientTableMode === 'new') {
                const sorted = [...clientAnalysisData.newClients].sort((a, b) => b.sales - a.sales);
                tbody.innerHTML = sorted.map(c => `<tr>
                    <td><strong>${c.name}</strong></td>
                    <td>${c.manager || '-'}</td>
                    <td class="text-right">${formatCurrency(c.sales)}</td>
                    <td class="text-right">${c.count.toLocaleString()}</td>
                    <td>${c.purpose || '-'}</td>
                </tr>`).join('');
            } else {
                const sorted = [...clientAnalysisData.churnedClients].sort((a, b) => b.lastYearSales - a.lastYearSales);
                tbody.innerHTML = sorted.map(c => `<tr>
                    <td><strong>${c.name}</strong></td>
                    <td>${c.manager || '-'}</td>
                    <td class="text-right">${formatCurrency(c.lastYearSales)}</td>
                    <td>${c.purpose || '-'}</td>
                </tr>`).join('');
            }
        }

        function updateClientByPurposeTable(clients) {
            // 검사목적별 업체 집계
            const purposeStats = {};
            clients.forEach(c => {
                const purposes = c[1].purposes || {};
                Object.entries(purposes).forEach(([purpose, data]) => {
                    if (!purposeStats[purpose]) {
                        purposeStats[purpose] = { clients: new Set(), totalSales: 0, topClient: null, topClientSales: 0 };
                    }
                    purposeStats[purpose].clients.add(c[0]);
                    purposeStats[purpose].totalSales += data.sales;
                    if (data.sales > purposeStats[purpose].topClientSales) {
                        purposeStats[purpose].topClient = c[0];
                        purposeStats[purpose].topClientSales = data.sales;
                    }
                });
            });

            const sorted = Object.entries(purposeStats)
                .map(([purpose, stats]) => ({
                    purpose,
                    clientCount: stats.clients.size,
                    totalSales: stats.totalSales,
                    avgSales: stats.totalSales / stats.clients.size,
                    topClient: stats.topClient
                }))
                .sort((a, b) => b.totalSales - a.totalSales);

            const tbody = document.querySelector('#clientByPurposeTable tbody');
            tbody.innerHTML = sorted.slice(0, 15).map(p => `<tr>
                <td><strong>${p.purpose}</strong></td>
                <td class="text-right">${p.clientCount}개</td>
                <td class="text-right">${formatCurrency(p.totalSales)}</td>
                <td class="text-right">${formatCurrency(p.avgSales)}</td>
                <td>${p.topClient || '-'}</td>
            </tr>`).join('');
        }

        function updateClientByManagerTable(clients, newClients, retainedClients, churnedClients, compareClientMap) {
            const managerStats = {};

            clients.forEach(c => {
                const manager = c[1].manager || '미지정';
                if (!managerStats[manager]) {
                    managerStats[manager] = { total: 0, newCount: 0, retained: 0, churned: 0, sales: 0, lastYearSales: 0 };
                }
                managerStats[manager].total++;
                managerStats[manager].sales += c[1].sales;
            });

            newClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (managerStats[manager]) managerStats[manager].newCount++;
            });

            retainedClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (managerStats[manager]) {
                    managerStats[manager].retained++;
                    managerStats[manager].lastYearSales += c.lastYearSales || 0;
                }
            });

            churnedClients.forEach(c => {
                const manager = c.manager || '미지정';
                if (!managerStats[manager]) {
                    managerStats[manager] = { total: 0, newCount: 0, retained: 0, churned: 0, sales: 0, lastYearSales: 0 };
                }
                managerStats[manager].churned++;
            });

            const sorted = Object.entries(managerStats)
                .map(([name, stats]) => ({
                    name,
                    ...stats,
                    growthRate: stats.lastYearSales > 0 ? ((stats.sales - stats.lastYearSales) / stats.lastYearSales * 100) : 0
                }))
                .sort((a, b) => b.sales - a.sales);

            const tbody = document.querySelector('#clientByManagerTable tbody');
            tbody.innerHTML = sorted.map(m => {
                const growthClass = m.growthRate >= 0 ? 'color: var(--success)' : 'color: var(--danger)';
                const growthSign = m.growthRate >= 0 ? '+' : '';
                return `<tr>
                    <td><strong>${m.name}</strong></td>
                    <td class="text-right">${m.total}개</td>
                    <td class="text-right" style="color: var(--success);">${m.newCount}</td>
                    <td class="text-right">${m.retained}</td>
                    <td class="text-right" style="color: var(--danger);">${m.churned}</td>
                    <td class="text-right">${formatCurrency(m.sales)}</td>
                    <td class="text-right"><span style="${growthClass}; font-weight: 600;">${growthSign}${m.growthRate.toFixed(1)}%</span></td>
                </tr>`;
            }).join('');
        }

        // 지역별 탭 전역 변수
        let regionAnalysisData = null;
        let selectedRegion = null;

        function updateRegionTab() {
            const regions = currentData.by_region || [];
            const compareRegions = compareData?.by_region || [];
            const regionTopManagers = currentData.region_top_managers || {};
            const managerRegions = currentData.manager_regions || {};
            const clients = currentData.by_client || [];

            const total = regions.reduce((s, r) => s + r[1].sales, 0) || 1;

            // 비교 데이터 맵 생성
            const currentRegionMap = Object.fromEntries(regions.map(r => [r[0], r[1]]));
            const compareRegionMap = Object.fromEntries(compareRegions.map(r => [r[0], r[1]]));

            // 지역별 분석 데이터 생성
            const regionData = regions.map(r => {
                const name = r[0];
                const data = r[1];
                const lastYear = compareRegionMap[name] || { sales: 0, count: 0 };
                const growth = data.sales - lastYear.sales;
                const growthRate = lastYear.sales > 0 ? ((growth / lastYear.sales) * 100) : (data.sales > 0 ? 100 : 0);
                return {
                    name,
                    sales: data.sales,
                    count: data.count,
                    sido: data.sido,
                    lastYearSales: lastYear.sales,
                    lastYearCount: lastYear.count,
                    growth,
                    growthRate,
                    isNew: !compareRegionMap[name] && data.sales > 0,
                    percent: (data.sales / total * 100)
                };
            });

            // 신규 지역 (전년 없고 올해 있음)
            const newRegions = regionData.filter(r => r.isNew);
            // 성장 지역 (성장률 높은 순)
            const growthRegions = [...regionData].filter(r => r.lastYearSales > 0).sort((a, b) => b.growthRate - a.growthRate);
            // 감소 지역 (공략 필요)
            const weakRegions = [...regionData].filter(r => r.growthRate < 0 && r.lastYearSales > 0).sort((a, b) => a.growthRate - b.growthRate);
            // 주력 지역 (매출 높은 순)
            const mainRegions = [...regionData].sort((a, b) => b.sales - a.sales);

            regionAnalysisData = { regionData, newRegions, growthRegions, weakRegions, mainRegions, regionTopManagers, managerRegions };

            // KPI 업데이트
            updateRegionKPIs(mainRegions, growthRegions, newRegions, weakRegions);

            // SVG 맵 업데이트
            updateKoreaMap(regionData);

            // 차트 업데이트
            updateRegionSalesChart(regionData);
            updateRegionGrowthChart(regionData);

            // 테이블 업데이트
            updateRegionHeatmapTable(regionData);
            updateRegionTopClientTable(regions, clients);
            updateManagerRegionTable(managerRegions);

            // 맵 클릭 이벤트 등록
            setupMapClickEvents(regionData, clients);
        }

        function updateRegionKPIs(mainRegions, growthRegions, newRegions, weakRegions) {
            // 주력 지역
            if (mainRegions.length > 0) {
                const main = mainRegions[0];
                document.getElementById('mainRegionName').textContent = main.name;
                document.getElementById('mainRegionValue').textContent = formatCurrency(main.sales) + ' (' + main.percent.toFixed(1) + '%)';

                const overlay = document.getElementById('mainRegionOverlay');
                overlay.innerHTML = `
                    <div style="font-weight: 600; margin-bottom: 8px;">📊 매출 TOP 5 지역</div>
                    ${mainRegions.slice(0, 5).map((r, i) => `
                        <div style="display: flex; justify-content: space-between; padding: 4px 0;">
                            <span>${i + 1}. ${r.name}</span>
                            <span>${formatCurrency(r.sales)} (${r.percent.toFixed(1)}%)</span>
                        </div>
                    `).join('')}
                `;
            }

            // 성장 지역
            if (growthRegions.length > 0) {
                const growth = growthRegions[0];
                document.getElementById('growthRegionName').textContent = growth.name;
                document.getElementById('growthRegionValue').textContent = '+' + growth.growthRate.toFixed(1) + '% 성장';

                const overlay = document.getElementById('growthRegionOverlay');
                overlay.innerHTML = `
                    <div style="font-weight: 600; margin-bottom: 8px;">📈 성장률 TOP 5 지역</div>
                    ${growthRegions.slice(0, 5).map((r, i) => `
                        <div style="display: flex; justify-content: space-between; padding: 4px 0;">
                            <span>${i + 1}. ${r.name}</span>
                            <span style="color: var(--success);">+${r.growthRate.toFixed(1)}%</span>
                        </div>
                    `).join('')}
                `;
            } else {
                document.getElementById('growthRegionName').textContent = '-';
                document.getElementById('growthRegionValue').textContent = '비교 데이터 없음';
            }

            // 신규 진출
            if (newRegions.length > 0) {
                document.getElementById('newRegionName').textContent = newRegions.length + '개 지역';
                document.getElementById('newRegionValue').textContent = '올해 첫 거래';

                const overlay = document.getElementById('newRegionOverlay');
                overlay.innerHTML = `
                    <div style="font-weight: 600; margin-bottom: 8px;">🆕 신규 진출 지역</div>
                    ${newRegions.map((r, i) => `
                        <div style="display: flex; justify-content: space-between; padding: 4px 0;">
                            <span>${i + 1}. ${r.name}</span>
                            <span>${formatCurrency(r.sales)}</span>
                        </div>
                    `).join('')}
                `;
            } else {
                document.getElementById('newRegionName').textContent = '-';
                document.getElementById('newRegionValue').textContent = '신규 없음';
            }

            // 공략 필요
            if (weakRegions.length > 0) {
                const weak = weakRegions[0];
                document.getElementById('weakRegionName').textContent = weak.name;
                document.getElementById('weakRegionValue').textContent = weak.growthRate.toFixed(1) + '% 감소';

                const overlay = document.getElementById('weakRegionOverlay');
                overlay.innerHTML = `
                    <div style="font-weight: 600; margin-bottom: 8px;">⚠️ 공략 필요 지역</div>
                    ${weakRegions.slice(0, 5).map((r, i) => `
                        <div style="display: flex; justify-content: space-between; padding: 4px 0;">
                            <span>${i + 1}. ${r.name}</span>
                            <span style="color: var(--danger);">${r.growthRate.toFixed(1)}%</span>
                        </div>
                    `).join('')}
                `;
            } else {
                document.getElementById('weakRegionName').textContent = '-';
                document.getElementById('weakRegionValue').textContent = '감소 지역 없음';
            }

            // KPI 오버레이 이벤트
            document.querySelectorAll('.region-kpi').forEach(card => {
                const overlay = card.querySelector('.region-kpi-overlay');
                if (overlay) {
                    card.addEventListener('mouseenter', () => {
                        overlay.style.display = 'block';
                    });
                    card.addEventListener('mouseleave', () => { overlay.style.display = 'none'; });
                }
            });
        }

        function updateKoreaMap(regionData) {
            const maxSales = Math.max(...regionData.map(r => r.sales), 1);

            // 시/도 이름 매핑 (데이터 지역명 → SVG ID)
            const sidoMap = {
                '서울': '서울', '서울특별시': '서울',
                '경기': '경기', '경기도': '경기',
                '인천': '인천', '인천광역시': '인천',
                '강원': '강원', '강원도': '강원', '강원특별자치도': '강원',
                '충북': '충북', '충청북도': '충북',
                '충남': '충남', '충청남도': '충남',
                '대전': '대전', '대전광역시': '대전',
                '세종': '세종', '세종특별자치시': '세종',
                '전북': '전북', '전라북도': '전북', '전북특별자치도': '전북',
                '전남': '전남', '전라남도': '전남',
                '광주': '광주', '광주광역시': '광주',
                '경북': '경북', '경상북도': '경북',
                '경남': '경남', '경상남도': '경남',
                '대구': '대구', '대구광역시': '대구',
                '울산': '울산', '울산광역시': '울산',
                '부산': '부산', '부산광역시': '부산',
                '제주': '제주', '제주특별자치도': '제주', '제주도': '제주'
            };

            // 지역별 매출 합산 (시/도 기준)
            const sidoSales = {};
            regionData.forEach(r => {
                const sido = sidoMap[r.sido] || sidoMap[r.name] || r.sido;
                if (sido) {
                    if (!sidoSales[sido]) sidoSales[sido] = 0;
                    sidoSales[sido] += r.sales;
                }
            });

            const maxSidoSales = Math.max(...Object.values(sidoSales), 1);

            // SVG 경로 색상 업데이트
            document.querySelectorAll('.region-path').forEach(path => {
                const regionName = path.dataset.region;
                const sales = sidoSales[regionName] || 0;

                // 색상 레벨 결정
                path.classList.remove('level-1', 'level-2', 'level-3', 'level-4', 'selected');
                if (sales >= 1000000000) { // 10억 이상
                    path.classList.add('level-4');
                } else if (sales >= 500000000) { // 5억 이상
                    path.classList.add('level-3');
                } else if (sales >= 100000000) { // 1억 이상
                    path.classList.add('level-2');
                } else {
                    path.classList.add('level-1');
                }
            });
        }

        function setupMapClickEvents(regionData, clients) {
            const regionDataMap = Object.fromEntries(regionData.map(r => [r.name, r]));
            const sidoDataMap = {};

            // 시도별 데이터 집계
            regionData.forEach(r => {
                const sido = r.sido || r.name;
                if (!sidoDataMap[sido]) {
                    sidoDataMap[sido] = { sales: 0, count: 0, growth: 0, lastYearSales: 0, regions: [] };
                }
                sidoDataMap[sido].sales += r.sales;
                sidoDataMap[sido].count += r.count;
                sidoDataMap[sido].lastYearSales += r.lastYearSales;
                sidoDataMap[sido].growth += r.growth;
                sidoDataMap[sido].regions.push(r.name);
            });

            document.querySelectorAll('.region-path').forEach(path => {
                path.addEventListener('click', function() {
                    const regionName = this.dataset.region;

                    // 선택 상태 토글
                    document.querySelectorAll('.region-path').forEach(p => p.classList.remove('selected'));
                    this.classList.add('selected');
                    selectedRegion = regionName;

                    showRegionDetail(regionName, sidoDataMap[regionName] || regionDataMap[regionName], clients);
                });
            });
        }

        function showRegionDetail(regionName, data, clients) {
            const regionTopManagers = currentData.region_top_managers || {};
            const managers = regionTopManagers[regionName] || [];

            // 해당 지역 업체 필터링
            const regionClients = clients.filter(c => {
                const clientData = c[1];
                // 업체 주소에서 지역 추출
                return (clientData.address && clientData.address.includes(regionName)) ||
                       (clientData.sido && clientData.sido.includes(regionName));
            }).sort((a, b) => b[1].sales - a[1].sales).slice(0, 5);

            // 성장률 계산
            const growthRate = data.lastYearSales > 0
                ? ((data.sales - data.lastYearSales) / data.lastYearSales * 100)
                : (data.sales > 0 ? 100 : 0);

            // AI 분석 의견 생성
            let aiOpinion = '';
            if (growthRate > 20) {
                aiOpinion = `${regionName} 지역은 전년 대비 ${growthRate.toFixed(1)}% 성장으로 매우 양호한 실적을 보이고 있습니다. 현재 영업 전략을 유지하고 추가 고객 확보에 집중하세요.`;
            } else if (growthRate > 0) {
                aiOpinion = `${regionName} 지역은 전년 대비 ${growthRate.toFixed(1)}% 소폭 성장 중입니다. 기존 고객 유지와 함께 신규 업체 발굴이 필요합니다.`;
            } else if (growthRate < -10) {
                aiOpinion = `${regionName} 지역은 전년 대비 ${Math.abs(growthRate).toFixed(1)}% 감소로 주의가 필요합니다. 이탈 고객 분석 및 경쟁사 동향 파악이 시급합니다.`;
            } else {
                aiOpinion = `${regionName} 지역은 현상 유지 상태입니다. 신규 고객 유치 전략 강화를 권장합니다.`;
            }

            document.getElementById('regionDetailTitle').textContent = '📍 ' + regionName + ' 상세 정보';
            document.getElementById('regionDetailBadge').textContent = currentData.year + '년';

            const body = document.getElementById('regionDetailBody');
            body.innerHTML = `
                <div class="region-stat-grid">
                    <div class="region-stat-item">
                        <div class="region-stat-value">${formatCurrency(data.sales || 0)}</div>
                        <div class="region-stat-label">총 매출액</div>
                    </div>
                    <div class="region-stat-item">
                        <div class="region-stat-value">${(data.count || 0).toLocaleString()}건</div>
                        <div class="region-stat-label">총 건수</div>
                    </div>
                    <div class="region-stat-item">
                        <div class="region-stat-value" style="color: ${growthRate >= 0 ? 'var(--success)' : 'var(--danger)'};">${growthRate >= 0 ? '+' : ''}${growthRate.toFixed(1)}%</div>
                        <div class="region-stat-label">전년 대비 성장률</div>
                    </div>
                    <div class="region-stat-item">
                        <div class="region-stat-value">${formatCurrency(data.lastYearSales || 0)}</div>
                        <div class="region-stat-label">전년 매출</div>
                    </div>
                </div>

                <div class="region-detail-section">
                    <div class="region-detail-title">👤 담당자별 현황</div>
                    <div class="region-manager-list">
                        ${managers.length > 0 ? managers.slice(0, 5).map(m => `
                            <div class="region-manager-item">
                                <span><strong>${m.name}</strong></span>
                                <span>${formatCurrency(m.sales)} (${m.count}건)</span>
                            </div>
                        `).join('') : '<div style="color: #94a3b8; text-align: center; padding: 16px;">담당자 데이터 없음</div>'}
                    </div>
                </div>

                <div class="region-detail-section">
                    <div class="region-detail-title">🤖 AI 분석 의견</div>
                    <div class="region-ai-opinion">${aiOpinion}</div>
                </div>

                <div class="region-detail-section">
                    <div class="region-detail-title">🏢 주요 업체 TOP 5</div>
                    <div class="region-top-clients">
                        ${regionClients.length > 0 ? regionClients.map(c => `
                            <div class="region-client-item">
                                <span><strong>${c[0]}</strong></span>
                                <span>${formatCurrency(c[1].sales)}</span>
                            </div>
                        `).join('') : '<div style="color: #94a3b8; text-align: center; padding: 16px;">해당 지역 업체 데이터 없음</div>'}
                    </div>
                </div>
            `;
        }

        function updateRegionSalesChart(regionData) {
            const sorted = [...regionData].sort((a, b) => b.sales - a.sales).slice(0, 15);

            document.getElementById('regionSalesChartBadge').textContent = currentData.year + '년';

            const ctx = document.getElementById('regionSalesChart');
            if (!ctx) return;
            if (charts.regionSales) charts.regionSales.destroy();

            charts.regionSales = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: sorted.map(r => r.name),
                    datasets: [{
                        label: '매출',
                        data: sorted.map(r => r.sales),
                        backgroundColor: sorted.map(r => r.growthRate >= 0 ? 'rgba(99, 102, 241, 0.8)' : 'rgba(239, 68, 68, 0.6)'),
                        borderRadius: 6
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    indexAxis: 'y',
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                label: ctx => {
                                    const r = sorted[ctx.dataIndex];
                                    return [
                                        '매출: ' + formatCurrency(r.sales),
                                        '건수: ' + r.count.toLocaleString() + '건',
                                        '비중: ' + r.percent.toFixed(1) + '%',
                                        '성장률: ' + (r.growthRate >= 0 ? '+' : '') + r.growthRate.toFixed(1) + '%'
                                    ];
                                }
                            }
                        }
                    },
                    scales: {
                        x: { beginAtZero: true, ticks: { callback: v => (v / 100000000).toFixed(0) + '억' } }
                    }
                }
            });
        }

        function updateRegionGrowthChart(regionData) {
            const sorted = [...regionData].filter(r => r.lastYearSales > 0)
                .sort((a, b) => b.growthRate - a.growthRate);

            const ctx = document.getElementById('regionGrowthChart');
            if (!ctx) return;
            if (charts.regionGrowth) charts.regionGrowth.destroy();

            charts.regionGrowth = new Chart(ctx.getContext('2d'), {
                type: 'bar',
                data: {
                    labels: sorted.map(r => r.name),
                    datasets: [{
                        label: '성장률',
                        data: sorted.map(r => r.growthRate),
                        backgroundColor: sorted.map(r => r.growthRate >= 0 ? 'rgba(16, 185, 129, 0.8)' : 'rgba(239, 68, 68, 0.8)'),
                        borderRadius: 6
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    indexAxis: 'y',
                    plugins: {
                        legend: { display: false },
                        tooltip: {
                            callbacks: {
                                label: ctx => {
                                    const r = sorted[ctx.dataIndex];
                                    return [
                                        '성장률: ' + (r.growthRate >= 0 ? '+' : '') + r.growthRate.toFixed(1) + '%',
                                        '올해: ' + formatCurrency(r.sales),
                                        '전년: ' + formatCurrency(r.lastYearSales),
                                        '증감: ' + (r.growth >= 0 ? '+' : '') + formatCurrency(r.growth)
                                    ];
                                }
                            }
                        }
                    },
                    scales: {
                        x: {
                            ticks: { callback: v => v + '%' }
                        }
                    }
                }
            });
        }

        function updateRegionHeatmapTable(regionData) {
            const sorted = [...regionData].sort((a, b) => b.sales - a.sales);
            const total = sorted.reduce((s, r) => s + r.sales, 0) || 1;
            const avgGrowth = sorted.reduce((s, r) => s + r.growthRate, 0) / (sorted.length || 1);

            const tbody = document.querySelector('#regionHeatmapTable tbody');
            tbody.innerHTML = sorted.map(r => {
                const percent = (r.sales / total * 100).toFixed(1);
                const growthClass = r.growthRate > avgGrowth ? 'heatmap-high' :
                                   r.growthRate < 0 ? 'heatmap-low' : 'heatmap-medium';
                return `<tr onclick="showRegionDetailFromTable('${r.name}')" style="cursor: pointer;">
                    <td><strong>${r.name}</strong></td>
                    <td class="text-right">${formatCurrency(r.sales)}</td>
                    <td class="text-right">${r.count.toLocaleString()}</td>
                    <td class="text-right"><span class="${growthClass}">${r.growthRate >= 0 ? '+' : ''}${r.growthRate.toFixed(1)}%</span></td>
                    <td><div class="progress-cell"><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%;"></div></div><span class="progress-value">${percent}%</span></div></td>
                </tr>`;
            }).join('');
        }

        function showRegionDetailFromTable(regionName) {
            // 지도에서 해당 지역 선택
            const path = document.querySelector(`.region-path[data-region="${regionName}"]`);
            if (path) {
                path.click();
            }
        }

        function updateRegionTopClientTable(regions, clients) {
            const regionClientMap = {};

            // 지역별 TOP 업체 집계
            clients.forEach(c => {
                const clientName = c[0];
                const clientData = c[1];

                // 간단한 지역 매칭 (업체 데이터에 지역 정보가 있으면 사용)
                regions.forEach(r => {
                    const regionName = r[0];
                    // 지역명이 업체 데이터에 포함되어 있는지 확인 (간단한 매칭)
                    if (!regionClientMap[regionName]) {
                        regionClientMap[regionName] = [];
                    }
                });
            });

            // region_top_managers를 이용해 지역별 주요 업체 표시
            const regionTopManagers = currentData.region_top_managers || {};
            const rows = [];

            regions.slice(0, 10).forEach(r => {
                const regionName = r[0];
                const managers = regionTopManagers[regionName] || [];
                if (managers.length > 0) {
                    const topManager = managers[0];
                    rows.push(`<tr>
                        <td><strong>${regionName}</strong></td>
                        <td>-</td>
                        <td class="text-right">${formatCurrency(r[1].sales)}</td>
                        <td>${topManager.name}</td>
                    </tr>`);
                } else {
                    rows.push(`<tr>
                        <td><strong>${regionName}</strong></td>
                        <td>-</td>
                        <td class="text-right">${formatCurrency(r[1].sales)}</td>
                        <td>-</td>
                    </tr>`);
                }
            });

            const tbody = document.querySelector('#regionTopClientTable tbody');
            tbody.innerHTML = rows.join('');
        }

        function updateManagerRegionTable(managerRegions) {
            const managers = Object.entries(managerRegions).map(([name, regions]) => {
                const totalSales = regions.reduce((s, r) => s + r.sales, 0);
                const mainRegion = regions[0]?.region || '-';
                return { name, regions, totalSales, mainRegion, regionCount: regions.length };
            }).sort((a, b) => b.totalSales - a.totalSales);

            document.getElementById('managerRegionBadge').textContent = managers.length + '명';

            const tbody = document.querySelector('#managerRegionTable tbody');
            tbody.innerHTML = managers.map(m => `<tr>
                <td><strong>${m.name}</strong></td>
                <td>${m.mainRegion}</td>
                <td class="text-right">${m.regionCount}개</td>
                <td class="text-right">${formatCurrency(m.totalSales)}</td>
                <td>
                    <div class="region-distribution">
                        ${m.regions.slice(0, 5).map(r => `<span class="region-chip">${r.region}</span>`).join('')}
                        ${m.regions.length > 5 ? `<span class="region-chip">+${m.regions.length - 5}</span>` : ''}
                    </div>
                </td>
            </tr>`).join('');
        }

        function updateSampleTypeTab() {
            const types = currentData.by_sample_type || [];
            const colors = ['blue', 'green', 'orange', 'purple', 'pink', 'info', 'teal', 'amber'];
            const icons = ['📦', '🌿', '🥩', '🐟', '💊', '🥤', '🧀', '📁'];
            const total = types.reduce((s, t) => s + t[1].sales, 0) || 1;

            document.getElementById('sampleTypeCount').textContent = types.length + '개 유형';

            const grid = document.getElementById('sampleTypeGrid');
            grid.innerHTML = types.map((t, i) => `
                <div class="purpose-kpi-card" data-color="${colors[i % colors.length]}">
                    <div class="purpose-kpi-header"><div class="purpose-kpi-icon">${icons[i % icons.length]}</div></div>
                    <div class="purpose-kpi-name">${t[0]}</div>
                    <div class="purpose-kpi-value">${formatCurrency(t[1].sales)}</div>
                    <div class="purpose-kpi-sub">건수: <span>${t[1].count.toLocaleString()}건</span></div>
                </div>
            `).join('');

            const ctx = document.getElementById('sampleTypeChart').getContext('2d');
            if (charts.sampleType) charts.sampleType.destroy();
            charts.sampleType = new Chart(ctx, { type: 'pie', data: { labels: types.map(t => t[0]), datasets: [{ data: types.map(t => t[1].sales), backgroundColor: ['#6366f1', '#10b981', '#f97316', '#8b5cf6', '#ec4899', '#06b6d4', '#14b8a6', '#f59e0b'] }] }, options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right' } } } });

            const tbody = document.querySelector('#sampleTypeTable tbody');
            tbody.innerHTML = types.map(t => {
                const percent = (t[1].sales / total * 100).toFixed(1);
                return `<tr><td><strong>${t[0]}</strong></td><td class="text-right">${formatCurrency(t[1].sales)}</td><td class="text-right">${t[1].count.toLocaleString()}</td><td><div class="progress-cell"><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%;"></div></div><span class="progress-value">${percent}%</span></div></td></tr>`;
            }).join('');
        }

        function updateDefectTab() {
            const defects = currentData.by_defect || [];
            const total = defects.reduce((s, d) => s + d[1].count, 0) || 1;

            const ctx = document.getElementById('defectChart').getContext('2d');
            if (charts.defect) charts.defect.destroy();
            charts.defect = new Chart(ctx, { type: 'bar', data: { labels: defects.slice(0, 10).map(d => d[0]), datasets: [{ data: defects.slice(0, 10).map(d => d[1].count), backgroundColor: 'rgba(239, 68, 68, 0.8)', borderRadius: 6 }] }, options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } } });

            const tbody = document.querySelector('#defectTable tbody');
            tbody.innerHTML = defects.map(d => {
                const percent = (d[1].count / total * 100).toFixed(1);
                return `<tr><td><strong>${d[0]}</strong></td><td class="text-right">${d[1].count.toLocaleString()}</td><td><div class="progress-cell"><div class="progress-bar" style="background: var(--danger-light);"><div class="progress-fill" style="width: ${percent}%; background: var(--danger);"></div></div><span class="progress-value">${percent}%</span></div></td></tr>`;
            }).join('');
        }

        function updatePurposeTab() {
            const purposes = currentData.by_purpose || [];
            const total = purposes.reduce((s, p) => s + p[1].sales, 0) || 1;

            const ctx = document.getElementById('purposeMonthlyChart').getContext('2d');
            if (charts.purposeMonthly) charts.purposeMonthly.destroy();

            const monthMap = Object.fromEntries(currentData.by_month || []);
            const labels = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
            charts.purposeMonthly = new Chart(ctx, { type: 'line', data: { labels, datasets: [{ label: '매출', data: labels.map((_, i) => monthMap[i+1]?.sales || 0), borderColor: '#6366f1', backgroundColor: 'rgba(99, 102, 241, 0.1)', fill: true, tension: 0.4 }] }, options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } } });

            const tbody = document.querySelector('#purposeTable tbody');
            tbody.innerHTML = purposes.map(p => {
                const percent = (p[1].sales / total * 100).toFixed(1);
                return `<tr><td><strong>${p[0]}</strong></td><td class="text-right">${formatCurrency(p[1].sales)}</td><td class="text-right">${p[1].count.toLocaleString()}</td><td><div class="progress-cell"><div class="progress-bar"><div class="progress-fill" style="width: ${percent}%;"></div></div><span class="progress-value">${percent}%</span></div></td></tr>`;
            }).join('');
        }

        // AI 분석
        function setAiQuery(text) { document.getElementById('aiQueryInput').value = text; }

        async function runAiAnalysis() {
            const query = document.getElementById('aiQueryInput').value.trim();
            if (!query) { alert('질문을 입력해주세요.'); return; }

            const btn = document.getElementById('aiBtn');
            const loading = document.getElementById('aiLoading');
            const error = document.getElementById('aiError');
            const content = document.getElementById('aiContent');
            const result = document.getElementById('aiResult');

            btn.disabled = true;
            loading.style.display = 'block';
            error.style.display = 'none';
            content.innerHTML = '';
            result.classList.add('show');

            try {
                const res = await fetch('/api/ai/analyze', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ query })
                });
                const data = await res.json();

                loading.style.display = 'none';

                if (data.error) {
                    error.textContent = data.error;
                    error.style.display = 'block';
                } else {
                    let html = `<p style="margin-bottom: 12px;"><strong>📝 ${data.description || '분석 결과'}</strong></p>`;

                    if (data.analysis_type === 'year_comparison' && data.comparison) {
                        const c = data.comparison;
                        html += `<table class="ai-result-table"><thead><tr><th>구분</th><th>건수</th><th>매출</th></tr></thead><tbody>`;
                        html += `<tr><td>${c.main_year?.year || '2025'}년</td><td>${(c.main_year?.count || 0).toLocaleString()}</td><td>${formatCurrency(c.main_year?.sales || 0)}</td></tr>`;
                        html += `<tr><td>${c.compare_year?.year || '2024'}년</td><td>${(c.compare_year?.count || 0).toLocaleString()}</td><td>${formatCurrency(c.compare_year?.sales || 0)}</td></tr>`;
                        const diff = c.difference || {};
                        const color = (diff.sales || 0) >= 0 ? 'var(--success)' : 'var(--danger)';
                        const sign = (diff.sales || 0) >= 0 ? '+' : '';
                        html += `<tr style="font-weight: bold; color: ${color};"><td>차이</td><td>${sign}${(diff.count || 0).toLocaleString()}</td><td>${sign}${formatCurrency(diff.sales || 0)} (${sign}${diff.growth_rate || 0}%)</td></tr>`;
                        html += `</tbody></table>`;
                    } else if (data.top_items) {
                        html += `<table class="ai-result-table"><thead><tr><th>순위</th><th>항목</th><th>매출</th></tr></thead><tbody>`;
                        data.top_items.forEach((item, i) => {
                            html += `<tr><td>${i+1}</td><td>${item.name}</td><td>${formatCurrency(item.sales || item.fee || 0)}</td></tr>`;
                        });
                        html += `</tbody></table>`;
                    } else if (data.summary) {
                        html += `<p>총 건수: <strong>${data.summary.total_count?.toLocaleString() || 0}건</strong></p>`;
                        html += `<p>총 매출: <strong>${formatCurrency(data.summary.total_sales || data.summary.total_fee || 0)}</strong></p>`;
                    } else if (data.direct_answer) {
                        html += `<p>${data.direct_answer}</p>`;
                    }

                    content.innerHTML = html;
                }
            } catch (e) {
                loading.style.display = 'none';
                error.textContent = '분석 실패: ' + e.message;
                error.style.display = 'block';
            } finally {
                btn.disabled = false;
                loadTokenUsage();
            }
        }

        // 초기화
        loadTokenUsage();
        showToast('조회 버튼을 클릭하세요.', 'loading', 3000);
    </script>
</body>
</html>

'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

def filter_data_by_date(data, year, month=None, day=None, end_year=None, end_month=None, end_day=None):
    """날짜 조건으로 데이터 필터링"""
    from datetime import datetime, date

    filtered = []
    year = int(year)
    month = int(month) if month else None
    day = int(day) if day else None
    end_year = int(end_year) if end_year else None
    end_month = int(end_month) if end_month else None
    end_day = int(end_day) if end_day else None

    # 범위 모드인 경우
    if end_year:
        # 시작 날짜 결정
        if month and day:
            start_date = date(year, month, day)
        elif month:
            start_date = date(year, month, 1)
        else:
            start_date = date(year, 1, 1)

        # 종료 날짜 결정
        if end_month and end_day:
            end_date = date(end_year, end_month, end_day)
        elif end_month:
            # 해당 월의 마지막 날
            import calendar
            last_day = calendar.monthrange(end_year, end_month)[1]
            end_date = date(end_year, end_month, last_day)
        else:
            end_date = date(end_year, 12, 31)

        for row in data:
            row_date = row.get('접수일자')
            if not row_date:
                continue

            # datetime 또는 date 객체로 변환
            if hasattr(row_date, 'date'):
                row_date = row_date.date()
            elif hasattr(row_date, 'year'):
                row_date = date(row_date.year, row_date.month, row_date.day)
            else:
                try:
                    parts = str(row_date).split('-')
                    row_date = date(int(parts[0]), int(parts[1]), int(parts[2][:2]))
                except:
                    continue

            if start_date <= row_date <= end_date:
                filtered.append(row)
    else:
        # 단일 날짜 모드
        for row in data:
            row_date = row.get('접수일자')
            if not row_date:
                continue

            # 연도 확인
            if hasattr(row_date, 'year'):
                row_year = row_date.year
                row_month = row_date.month
                row_day = row_date.day
            else:
                try:
                    parts = str(row_date).split('-')
                    row_year = int(parts[0])
                    row_month = int(parts[1])
                    row_day = int(parts[2][:2])
                except:
                    continue

            if row_year != year:
                continue

            if month and row_month != month:
                continue

            if day and row_day != day:
                continue

            filtered.append(row)

    return filtered

@app.route('/api/data')
def get_data():
    year = request.args.get('year', '2025')
    month = request.args.get('month', '')
    day = request.args.get('day', '')
    end_year = request.args.get('end_year', '')
    end_month = request.args.get('end_month', '')
    end_day = request.args.get('end_day', '')
    purpose = request.args.get('purpose', '전체')

    # 로그 출력
    date_info = f"year={year}"
    if month: date_info += f", month={month}"
    if day: date_info += f", day={day}"
    if end_year: date_info += f" ~ end_year={end_year}"
    if end_month: date_info += f", end_month={end_month}"
    if end_day: date_info += f", end_day={end_day}"
    print(f"[API] 요청: {date_info}, purpose={purpose}")

    # 기본 데이터 로드 (연도별)
    years_to_load = {year}
    if end_year and end_year != year:
        years_to_load.add(end_year)

    all_data = []
    for y in years_to_load:
        all_data.extend(load_excel_data(y))

    print(f"[API] 로드된 원본 데이터: {len(all_data)}건")

    # 날짜 필터링 적용
    filtered_data = filter_data_by_date(all_data, year, month, day, end_year, end_month, end_day)
    print(f"[API] 날짜 필터링 후 데이터: {len(filtered_data)}건")

    processed = process_data(filtered_data, purpose)
    print(f"[API] 처리 완료: total_count={processed['total_count']}")
    return jsonify(processed)

@app.route('/api/food_item')
def get_food_item_data():
    """검사항목 데이터 API"""
    year = request.args.get('year', '2025')
    purpose = request.args.get('purpose', '전체')
    sample_type = request.args.get('sample_type', '전체')
    item = request.args.get('item', '전체')
    manager = request.args.get('manager', '전체')

    print(f"[API] food_item 요청: year={year}, purpose={purpose}, sample_type={sample_type}, item={item}, manager={manager}")

    # 데이터 로드
    data = load_food_item_data(year)
    print(f"[API] food_item 로드: {len(data)}건")

    # 데이터 처리
    processed = process_food_item_data(
        data,
        purpose_filter=purpose if purpose != '전체' else None,
        sample_type_filter=sample_type if sample_type != '전체' else None,
        item_filter=item if item != '전체' else None,
        manager_filter=manager if manager != '전체' else None
    )

    processed['year'] = int(year)
    print(f"[API] food_item 처리 완료: total_count={processed['total_count']}")
    return jsonify(processed)

@app.route('/api/columns')
def get_columns():
    """Excel 파일의 컬럼명 조회"""
    year = request.args.get('year', '2025')
    from openpyxl import load_workbook

    data_path = DATA_DIR / str(year)
    if not data_path.exists():
        return jsonify({'error': f'{year}년 데이터 폴더가 없습니다.', 'columns': []})

    files = sorted(data_path.glob("*.xlsx"))
    if not files:
        return jsonify({'error': f'{year}년 데이터 파일이 없습니다.', 'columns': []})

    try:
        wb = load_workbook(files[0], read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        wb.close()

        # 주소 관련 컬럼 표시
        address_cols = [h for h in headers if h and any(k in str(h) for k in ['주소', '지역', '시', '도', '군', '구', '동', '장소'])]

        return jsonify({
            'year': year,
            'file': files[0].name,
            'total_columns': len(headers),
            'columns': headers,
            'address_columns': address_cols
        })
    except Exception as e:
        return jsonify({'error': str(e), 'columns': []})

@app.route('/api/cache/refresh')
def refresh_cache():
    """캐시 새로고침"""
    global DATA_CACHE, CACHE_TIME, AI_SUMMARY_CACHE, FILE_MTIME
    DATA_CACHE = {}
    CACHE_TIME = {}
    AI_SUMMARY_CACHE = {}
    FILE_MTIME = {}
    print("[CACHE] 모든 캐시 초기화됨")
    # 데이터 미리 로드
    for year in ['2024', '2025']:
        load_excel_data(year, use_cache=False)
    # AI 요약 캐시도 미리 생성
    get_ai_data_summary(force_refresh=True)
    return jsonify({'status': 'ok', 'message': '캐시가 새로고침되었습니다.'})


@app.route('/api/debug/urgent')
def debug_urgent():
    """긴급여부 필드 값 확인용 디버그 API"""
    import pandas as pd
    urgent_values = {}

    for year in ['2024', '2025']:
        data_path = DATA_DIR / str(year)
        if not data_path.exists():
            continue
        urgent_values[year] = set()
        for f in sorted(data_path.glob("*.xlsx")):
            try:
                df = pd.read_excel(f)
                if '긴급여부' in df.columns:
                    values = df['긴급여부'].dropna().unique()
                    for v in values:
                        urgent_values[year].add(str(v).strip())
            except Exception as e:
                pass
        urgent_values[year] = list(urgent_values[year])

    return jsonify({'urgent_values': urgent_values})


@app.route('/api/token-usage')
def api_token_usage():
    """토큰 사용량 조회 API"""
    try:
        stats = get_token_usage_stats()
        return jsonify({
            'success': True,
            'this_month': stats['this_month'],
            'last_month': stats['last_month']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# 기업 정보 파일 경로
COMPANY_INFO_FILE = os.path.join(DATA_DIR, 'company_info.json')

@app.route('/api/company-info', methods=['GET'])
def get_company_info():
    """기업 정보 조회"""
    try:
        if os.path.exists(COMPANY_INFO_FILE):
            with open(COMPANY_INFO_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            print(f"[CompanyInfo] 기업 정보 로드 성공: {data.get('companyName', 'N/A')}")
            return jsonify({'success': True, 'data': data})
        else:
            print("[CompanyInfo] 저장된 기업 정보 없음")
            return jsonify({'success': True, 'data': None})
    except Exception as e:
        print(f"[CompanyInfo] 로드 오류: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/company-info', methods=['POST'])
def save_company_info():
    """기업 정보 저장"""
    try:
        data = request.json
        with open(COMPANY_INFO_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"[CompanyInfo] 기업 정보 저장 완료: {data.get('companyName', 'N/A')}")
        return jsonify({'success': True})
    except Exception as e:
        print(f"[CompanyInfo] 저장 오류: {e}")
        return jsonify({'success': False, 'error': str(e)})

def get_company_context():
    """AI 분석용 기업 정보 컨텍스트 생성"""
    try:
        if os.path.exists(COMPANY_INFO_FILE):
            with open(COMPANY_INFO_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 부서 정보 요약
            dept_summary = []
            total_employees = 0
            if data.get('departments'):
                for dept_name, dept_info in data['departments'].items():
                    count = dept_info.get('count', 0)
                    if count > 0:
                        total_employees += count
                        role = dept_info.get('role', '')
                        dept_summary.append(f"{dept_name}({count}명): {role}")

            # 영업부 인력 요약
            sales_summary = []
            if data.get('salesPersonnel'):
                for person in data['salesPersonnel']:
                    if person.get('name'):
                        sales_summary.append(f"{person['name']}({person.get('region', '')})")

            # 지사 인력 요약
            branch_summary = []
            if data.get('branchPersonnel'):
                for person in data['branchPersonnel']:
                    if person.get('name'):
                        branch_summary.append(f"{person['name']}({person.get('region', '')})")

            context = f"""[기업 정보]
- 기업명: {data.get('companyName', '미입력')}
- 설립연도: {data.get('foundedYear', '미입력')}
- 사업분야: {data.get('businessField', '미입력')}
- 주요서비스: {data.get('mainServices', '미입력')}
- 연간매출목표: {data.get('revenueTarget', '미입력')}억원
- 연간검사목표: {data.get('inspectionTarget', '미입력')}건
- KPI: {data.get('kpiDescription', '미입력')}
- 경영전략: {data.get('businessStrategy', '미입력')}
- 총인원: {total_employees}명
- 조직구성: {'; '.join(dept_summary[:5]) if dept_summary else '미입력'}
- 영업담당자: {', '.join(sales_summary) if sales_summary else '미입력'}
- 지사담당자: {', '.join(branch_summary) if branch_summary else '미입력'}"""
            return context
        return ""
    except Exception as e:
        print(f"[CompanyInfo] 컨텍스트 생성 오류: {e}")
        return ""

@app.route('/api/ai/analyze', methods=['POST'])
def ai_analyze():
    """AI 분석 API - Claude 또는 Gemini로 자연어 질문 분석"""
    import urllib.request
    import urllib.error
    import time

    query = request.json.get('query', '')
    print(f"[AI] === 분석 요청 시작 ===")
    print(f"[AI] 질문: {query}")
    print(f"[AI] 사용 API: {'Claude' if USE_CLAUDE else 'Gemini'}")

    if not query:
        print(f"[AI] 오류: 질문 없음")
        return jsonify({'error': '질문을 입력해주세요.'})

    # 캐시된 데이터 요약 사용 (변경 감지 포함)
    data_summary = get_ai_data_summary()
    filter_values = data_summary['filter_values']
    print(f"[AI] 캐시된 요약 사용: 목적 {len(filter_values['purposes'])}개, 유형 {len(filter_values['sample_types'])}개")

    # 2025년 주요 통계 요약
    stats_2025 = data_summary['2025']
    top_purposes = sorted(stats_2025['by_purpose'].items(), key=lambda x: x[1]['fee'], reverse=True)[:5]
    top_managers = sorted(stats_2025['by_manager'].items(), key=lambda x: x[1]['fee'], reverse=True)[:5]

    stats_text = f"""2025년 현황:
- 총 건수: {stats_2025['total_count']:,}건
- 총 매출: {stats_2025['total_fee']/100000000:.2f}억원
- TOP 검사목적: {', '.join([f"{p[0]}({p[1]['fee']/10000:.0f}만)" for p in top_purposes])}
- TOP 영업담당: {', '.join([f"{m[0]}({m[1]['fee']/10000:.0f}만)" for m in top_managers])}"""

    # 기업 정보 컨텍스트 추가
    company_context = get_company_context()
    if company_context:
        stats_text = company_context + "\n\n" + stats_text
        print(f"[AI] 기업 정보 컨텍스트 추가됨")

    # Claude API 사용
    if USE_CLAUDE and CLAUDE_API_KEY:
        print(f"[AI] Claude API 사용 (모델: {CLAUDE_MODEL})")

        system_prompt = f"""당신은 경영 데이터 분석 전문가입니다. 사용자의 질문을 분석하여 JSON 형식으로 응답하세요.

{stats_text}

사용 가능한 필터 값:
- 연도: 2024, 2025
- 월: 1~12 (특정 월 분석 시)
- 검사목적: {', '.join(filter_values['purposes'][:10])}
- 검체유형: {', '.join(filter_values['sample_types'][:10])}
- 영업담당: {', '.join(filter_values['managers'][:10])}

분석 유형:
- year_comparison: 연도간 비교 분석 (예: 2025년 vs 2024년)
- monthly_trend: 월별 추이 분석
- top_managers: 영업담당별 TOP N 분석
- top_purposes: 검사목적별 TOP N 분석
- summary: 요약 통계
- direct_answer: 직접 답변 (계산 없이 바로 답변 가능한 경우)

중요: 연도 비교 질문(예: "2025년 1월과 2024년 1월 비교")은 반드시 year_comparison 타입을 사용하고 compare_year를 설정하세요.

반드시 JSON 형식만 응답하세요:
{{"analysis_type":"타입","year":"2025","compare_year":"2024","month":null,"purpose":null,"sample_type":null,"manager":null,"top_n":10,"description":"분석 설명","direct_answer":"직접 답변 가능시 여기에 작성"}}"""

        claude_result = call_claude_api(f"질문: {query}", system_prompt=system_prompt, max_tokens=800)

        if claude_result['success']:
            ai_response = claude_result['text']
            print(f"[AI] Claude 응답: {ai_response[:300]}...")

            # JSON 파싱
            try:
                json_str = ai_response.strip()
                if '```json' in json_str:
                    json_str = json_str.split('```json')[1].split('```')[0]
                elif '```' in json_str:
                    json_str = json_str.split('```')[1].split('```')[0]

                parsed = json.loads(json_str.strip())
                print(f"[AI] 파싱 성공: {parsed}")

                # direct_answer 타입이면 바로 응답 반환
                if parsed.get('analysis_type') == 'direct_answer' and parsed.get('direct_answer'):
                    return jsonify({
                        'success': True,
                        'analysis_type': 'direct_answer',
                        'description': parsed.get('description', ''),
                        'direct_answer': parsed.get('direct_answer'),
                        'parsed_query': parsed,
                        'ai_model': 'Claude Sonnet 4'
                    })

                # 데이터 조회 및 분석 실행
                food_2024 = load_food_item_data('2024')
                food_2025 = load_food_item_data('2025')
                data_2024 = load_excel_data('2024')
                data_2025 = load_excel_data('2025')

                analysis_result = execute_analysis(parsed, food_2024, food_2025, data_2024, data_2025)
                analysis_result['parsed_query'] = parsed
                analysis_result['ai_model'] = 'Claude Sonnet 4'

                print(f"[AI] 분석 완료: {analysis_result.get('analysis_type')}")
                return jsonify(analysis_result)

            except json.JSONDecodeError as e:
                print(f"[AI] Claude JSON 파싱 오류: {e}")
                return jsonify({
                    'error': 'Claude 응답 파싱 실패',
                    'raw_response': ai_response[:500]
                })
        else:
            print(f"[AI] Claude API 실패: {claude_result.get('error')}")
            # Claude 실패 시 Gemini로 폴백
            print(f"[AI] Gemini로 폴백...")

    # Gemini API 사용 (폴백 또는 기본)
    global current_api_key_index
    if not GEMINI_API_KEYS:
        print(f"[AI] 오류: API 키 없음")
        return jsonify({'error': 'API 키가 설정되지 않았습니다.'})

    print(f"[AI] 사용 가능한 Gemini API 키: {len(GEMINI_API_KEYS)}개")

    # 간소화된 Gemini 프롬프트 (토큰 절약)
    system_prompt = f"""데이터 분석 도우미입니다. 질문을 JSON으로 변환하세요.

{stats_text}

가능한 값:
- 연도: 2024, 2025
- 월: 1~12
- 검사목적: {', '.join(filter_values['purposes'][:10])}
- 검체유형: {', '.join(filter_values['sample_types'][:10])}
- 영업담당: {', '.join(filter_values['managers'][:10])}

분석유형: year_comparison(연도비교), monthly_trend(월별추이), top_managers(담당자TOP), top_purposes(목적별TOP), summary(요약), direct_answer(직접답변)

연도 비교 질문은 year_comparison 사용, compare_year 설정 필수

JSON 형식만 응답:
{{"analysis_type":"타입","year":"2025","compare_year":"2024","month":null,"purpose":null,"sample_type":null,"manager":null,"top_n":10,"description":"설명","direct_answer":"직접 답변이 가능하면 여기에 작성"}}"""

    print(f"[AI] 프롬프트 길이: {len(system_prompt)}자")

    payload = {
        "contents": [{"parts": [{"text": system_prompt + f"\n\n질문: {query}"}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 500}
    }

    # Gemini API 호출 (여러 키로 429 대응)
    total_keys = len(GEMINI_API_KEYS)
    keys_tried = 0

    while keys_tried < total_keys:
        api_key = GEMINI_API_KEYS[current_api_key_index]
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"

        print(f"[AI] API 키 {current_api_key_index + 1}/{total_keys} 사용: {api_key[:15]}...")

        try:
            req = urllib.request.Request(
                url,
                data=json.dumps(payload).encode('utf-8'),
                headers={'Content-Type': 'application/json'},
                method='POST'
            )

            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode('utf-8'))

            print(f"[AI] Gemini API 응답 수신 성공")

            # 토큰 사용량 기록
            try:
                usage_metadata = result.get('usageMetadata', {})
                input_tokens = usage_metadata.get('promptTokenCount', len(system_prompt) // 4)
                output_tokens = usage_metadata.get('candidatesTokenCount', 100)
                record_token_usage('gemini-2.0-flash', input_tokens, output_tokens)
                print(f"[AI] 토큰 사용: 입력={input_tokens}, 출력={output_tokens}")
            except Exception as te:
                print(f"[AI] 토큰 기록 오류: {te}")

            # 라운드 로빈: 성공 후에도 다음 키로 전환 (부하 분산)
            current_api_key_index = (current_api_key_index + 1) % total_keys

            # Gemini 응답에서 JSON 추출
            ai_response = result['candidates'][0]['content']['parts'][0]['text']
            print(f"[AI] Gemini 원본 응답: {ai_response[:200]}...")

            # JSON 파싱 (코드블록 제거)
            json_str = ai_response.strip()
            if '```json' in json_str:
                json_str = json_str.split('```json')[1].split('```')[0]
            elif '```' in json_str:
                json_str = json_str.split('```')[1].split('```')[0]

            parsed = json.loads(json_str.strip())
            print(f"[AI] 파싱 성공: {parsed}")

            # direct_answer 타입이면 바로 응답 반환
            if parsed.get('analysis_type') == 'direct_answer' and parsed.get('direct_answer'):
                print(f"[AI] 직접 답변 반환")
                return jsonify({
                    'success': True,
                    'analysis_type': 'direct_answer',
                    'description': parsed.get('description', ''),
                    'direct_answer': parsed.get('direct_answer'),
                    'parsed_query': parsed
                })

            # 데이터 조회 및 분석 실행 (캐시된 데이터 사용)
            food_2024 = load_food_item_data('2024')
            food_2025 = load_food_item_data('2025')
            data_2024 = load_excel_data('2024')
            data_2025 = load_excel_data('2025')

            analysis_result = execute_analysis(parsed, food_2024, food_2025, data_2024, data_2025)
            analysis_result['parsed_query'] = parsed

            print(f"[AI] 분석 완료: {analysis_result.get('analysis_type')}, 건수: {analysis_result.get('total_count')}")
            return jsonify(analysis_result)

        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8') if e.fp else ''
            print(f"[AI] HTTP 오류 {e.code}: {e.reason}")
            print(f"[AI] 오류 상세: {error_body[:300]}")

            if e.code == 429:  # Too Many Requests - 다음 키로 전환
                keys_tried += 1
                current_api_key_index = (current_api_key_index + 1) % total_keys
                print(f"[AI] 429 오류 - 다음 API 키로 전환 (키 {current_api_key_index + 1})")
                time.sleep(1)  # 짧은 대기 후 다음 키 시도
                continue
            elif e.code == 404:
                return jsonify({'error': f'API 모델을 찾을 수 없습니다 (404). 모델명 확인 필요.'})
            else:
                return jsonify({'error': f'API 오류 {e.code}: {e.reason}'})

        except urllib.error.URLError as e:
            print(f"[AI] URL 오류: {e.reason}")
            return jsonify({'error': f'API 연결 실패: {str(e.reason)}'})

        except json.JSONDecodeError as e:
            print(f"[AI] JSON 파싱 오류: {e}")
            print(f"[AI] 파싱 시도한 문자열: {json_str[:300] if 'json_str' in locals() else 'N/A'}")
            return jsonify({
                'error': f'응답 파싱 실패. Gemini가 올바른 JSON을 반환하지 않았습니다.',
                'raw_response': ai_response[:500] if 'ai_response' in locals() else ''
            })

        except Exception as e:
            import traceback
            print(f"[AI] 예외 발생: {e}")
            print(f"[AI] 트레이스백: {traceback.format_exc()}")
            return jsonify({'error': f'분석 실패: {str(e)}'})

    return jsonify({'error': f'모든 API 키({total_keys}개)가 할당량을 초과했습니다. 잠시 후 다시 시도해주세요.'})


def execute_analysis(params, food_2024, food_2025, data_2024, data_2025):
    """파싱된 조건으로 실제 데이터 분석 실행 - 대시보드와 동일한 데이터(공급가액) 사용"""
    analysis_type = params.get('analysis_type', 'summary')
    year = params.get('year', '2025')
    compare_year = params.get('compare_year')  # 비교 연도 (예: 2024)
    month = params.get('month')  # 월 필터
    purpose = params.get('purpose')
    sample_type = params.get('sample_type')
    manager = params.get('manager')
    top_n = params.get('top_n', 10)
    description = params.get('description', '')

    def get_sales(row):
        """공급가액 추출 (대시보드와 동일)"""
        sales = row.get('공급가액', 0) or 0
        if isinstance(sales, str):
            sales = float(sales.replace(',', '').replace('원', '')) if sales else 0
        return sales

    def get_month(row):
        """월 추출"""
        date = row.get('접수일자')
        if date and hasattr(date, 'month'):
            return date.month
        return 0

    def filter_data(data, month_filter=None, purpose_filter=None, sample_type_filter=None, manager_filter=None):
        """데이터 필터링"""
        filtered = []
        for row in data:
            if month_filter:
                row_month = get_month(row)
                if row_month != int(month_filter):
                    continue
            if purpose_filter and str(row.get('검사목적', '')).strip() != purpose_filter:
                continue
            if sample_type_filter and str(row.get('검체유형', '')).strip() != sample_type_filter:
                continue
            if manager_filter and str(row.get('영업담당', '')).strip() != manager_filter:
                continue
            filtered.append(row)
        return filtered

    # 대시보드와 동일한 데이터 소스 사용 (공급가액 기준)
    main_data = data_2025 if year == '2025' else data_2024
    compare_data = data_2024 if compare_year == '2024' else (data_2025 if compare_year == '2025' else None)

    # 메인 데이터 필터링
    filtered = filter_data(main_data, month, purpose, sample_type, manager)

    # 비교 데이터 필터링
    filtered_compare = filter_data(compare_data, month, purpose, sample_type, manager) if compare_data else []

    # compare_year가 있으면 무조건 year_comparison 타입으로 처리
    if compare_year:
        analysis_type = 'year_comparison'

    result = {
        'success': True,
        'description': description,
        'analysis_type': analysis_type,
        'total_count': len(filtered),
        'year': year
    }

    if analysis_type == 'year_comparison':
        # 연도간 비교 분석
        main_total = sum(get_sales(row) for row in filtered)
        main_count = len(filtered)
        compare_total = sum(get_sales(row) for row in filtered_compare)
        compare_count = len(filtered_compare)

        diff_sales = main_total - compare_total
        diff_count = main_count - compare_count
        growth_rate = ((main_total - compare_total) / compare_total * 100) if compare_total > 0 else 0

        result['comparison'] = {
            'main_year': {'year': year, 'count': main_count, 'sales': main_total},
            'compare_year': {'year': compare_year, 'count': compare_count, 'sales': compare_total},
            'difference': {'count': diff_count, 'sales': diff_sales, 'growth_rate': round(growth_rate, 1)}
        }
        result['total_fee'] = main_total

        # 월별 비교 차트 데이터
        if month:
            result['month'] = int(month)
        else:
            # 전체 월별 추이 비교
            monthly_main = {}
            monthly_compare = {}
            for row in filtered:
                m = get_month(row)
                if m > 0:
                    monthly_main[m] = monthly_main.get(m, 0) + get_sales(row)
            for row in filtered_compare:
                m = get_month(row)
                if m > 0:
                    monthly_compare[m] = monthly_compare.get(m, 0) + get_sales(row)

            result['chart_data'] = {
                'labels': [f'{m}월' for m in range(1, 13)],
                'datasets': [
                    {'label': f'{year}년', 'data': [monthly_main.get(m, 0) for m in range(1, 13)]},
                    {'label': f'{compare_year}년', 'data': [monthly_compare.get(m, 0) for m in range(1, 13)]}
                ]
            }

    elif analysis_type == 'monthly_trend':
        # 월별 추이
        monthly = {}
        for row in filtered:
            m = get_month(row)
            if m > 0:
                monthly[m] = monthly.get(m, 0) + get_sales(row)

        result['chart_data'] = {
            'labels': [f'{m}월' for m in range(1, 13)],
            'datasets': [
                {'label': f'{year}년 매출', 'data': [monthly.get(m, 0) for m in range(1, 13)]}
            ]
        }
        result['total_fee'] = sum(monthly.values())

    elif analysis_type == 'top_managers':
        # 영업담당별 TOP N
        manager_stats = {}
        for row in filtered:
            mgr = str(row.get('영업담당', '미지정')).strip()
            if mgr not in manager_stats:
                manager_stats[mgr] = {'count': 0, 'sales': 0}
            manager_stats[mgr]['count'] += 1
            manager_stats[mgr]['sales'] += get_sales(row)

        sorted_managers = sorted(manager_stats.items(), key=lambda x: x[1]['sales'], reverse=True)[:top_n]
        result['top_items'] = [{'name': k, 'count': v['count'], 'sales': v['sales']} for k, v in sorted_managers]
        result['chart_data'] = {
            'labels': [m[0] for m in sorted_managers],
            'datasets': [{'label': '매출', 'data': [m[1]['sales'] for m in sorted_managers]}]
        }
        result['total_fee'] = sum(get_sales(row) for row in filtered)

    elif analysis_type == 'top_purposes':
        # 검사목적별 TOP N
        purpose_stats = {}
        for row in filtered:
            p = str(row.get('검사목적', '미지정')).strip()
            if p not in purpose_stats:
                purpose_stats[p] = {'count': 0, 'sales': 0}
            purpose_stats[p]['count'] += 1
            purpose_stats[p]['sales'] += get_sales(row)

        sorted_purposes = sorted(purpose_stats.items(), key=lambda x: x[1]['sales'], reverse=True)[:top_n]
        result['top_items'] = [{'name': k, 'count': v['count'], 'sales': v['sales']} for k, v in sorted_purposes]
        result['chart_data'] = {
            'labels': [p[0][:15] for p in sorted_purposes],
            'datasets': [{'label': '매출', 'data': [p[1]['sales'] for p in sorted_purposes]}]
        }
        result['total_fee'] = sum(get_sales(row) for row in filtered)

    else:  # summary
        total_sales = sum(get_sales(row) for row in filtered)
        result['summary'] = {
            'total_count': len(filtered),
            'total_sales': total_sales,
            'avg_sales': total_sales / len(filtered) if filtered else 0
        }
        result['total_fee'] = total_sales

    return result


@app.route('/api/ai/goal-analysis', methods=['POST'])
def goal_analysis():
    """목표 달성 분석 API - 데이터 기반 종합 분석"""
    try:
        target_revenue = request.json.get('target', 7000000000)  # 기본 70억
        target_year = request.json.get('year', 2026)

        # 필터 옵션 (체크박스 선택)
        filters = request.json.get('filters', {})
        selected_managers = filters.get('managers', [])  # 빈 배열 = 전체
        selected_teams = filters.get('teams', [])
        selected_months = filters.get('months', [])
        selected_purposes = filters.get('purposes', [])
        selected_regions = filters.get('regions', [])
        selected_sample_types = filters.get('sample_types', [])
        selected_items = filters.get('items', [])
        selected_analyzers = filters.get('analyzers', [])

        # 데이터 로드 (메인 Excel 데이터 사용 - 공급가액 기준)
        data_2024 = load_excel_data('2024')
        data_2025 = load_excel_data('2025')

        def get_fee(row):
            """공급가액 추출"""
            fee = row.get('공급가액', 0) or 0
            if isinstance(fee, str):
                fee = float(fee.replace(',', '').replace('원', '')) if fee else 0
            return float(fee)

        def match_filter(row, managers, teams, months, purposes, regions, sample_types, items, analyzers):
            """필터 조건 매칭"""
            # 빈 배열이면 전체 선택으로 처리
            if managers and str(row.get('영업담당', '')).strip() not in managers:
                return False
            if teams:
                manager = str(row.get('영업담당', '')).strip()
                team = MANAGER_TO_BRANCH.get(manager, '기타')
                if team not in teams:
                    return False
            if months:
                date = row.get('접수일자')
                if date and hasattr(date, 'month'):
                    if date.month not in months:
                        return False
            if purposes and str(row.get('검사목적', '')).strip() not in purposes:
                return False
            if regions and str(row.get('지역', '')).strip() not in regions:
                return False
            if sample_types and str(row.get('검체유형', '')).strip() not in sample_types:
                return False
            if items and str(row.get('항목명', '')).strip() not in items:
                return False
            if analyzers and str(row.get('결과입력자', '')).strip() not in analyzers:
                return False
            return True

        # 연도별 매출 계산 (공급가액 기준)
        revenue_2024 = sum(get_fee(row) for row in data_2024 if match_filter(
            row, selected_managers, selected_teams, selected_months, selected_purposes,
            selected_regions, selected_sample_types, selected_items, selected_analyzers))
        revenue_2025 = sum(get_fee(row) for row in data_2025 if match_filter(
            row, selected_managers, selected_teams, selected_months, selected_purposes,
            selected_regions, selected_sample_types, selected_items, selected_analyzers))

        # 성장률 계산
        growth_rate = ((revenue_2025 - revenue_2024) / revenue_2024 * 100) if revenue_2024 > 0 else 0

        # 목표 달성에 필요한 추가 매출
        gap = target_revenue - revenue_2025
        required_growth = ((target_revenue - revenue_2025) / revenue_2025 * 100) if revenue_2025 > 0 else 0

        result = {
            'success': True,
            'target': target_revenue,
            'target_year': target_year,
            'current_status': {
                'revenue_2024': revenue_2024,
                'revenue_2025': revenue_2025,
                'growth_rate': round(growth_rate, 1),
                'gap_to_target': gap,
                'required_growth': round(required_growth, 1)
            },
            'analysis': {},
            'recommendations': []
        }

        # 1. 영업담당별 분석
        by_manager = {}
        for row in data_2025:
            if not match_filter(row, [], selected_teams, selected_months, selected_purposes,
                               selected_regions, selected_sample_types, selected_items, selected_analyzers):
                continue
            manager = str(row.get('영업담당', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if manager not in by_manager:
                by_manager[manager] = {'revenue_2025': 0, 'count_2025': 0, 'revenue_2024': 0, 'count_2024': 0}
            by_manager[manager]['revenue_2025'] += revenue
            by_manager[manager]['count_2025'] += 1

        for row in data_2024:
            if not match_filter(row, [], selected_teams, selected_months, selected_purposes,
                               selected_regions, selected_sample_types, selected_items, selected_analyzers):
                continue
            manager = str(row.get('영업담당', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if manager not in by_manager:
                by_manager[manager] = {'revenue_2025': 0, 'count_2025': 0, 'revenue_2024': 0, 'count_2024': 0}
            by_manager[manager]['revenue_2024'] += revenue
            by_manager[manager]['count_2024'] += 1

        # 영업담당별 성장률 계산 (ISA, IBK 등 제외)
        manager_analysis = []
        for manager, data in by_manager.items():
            # 제외 대상 확인
            if manager in EXCLUDED_MANAGERS:
                continue
            if data['revenue_2024'] > 0:
                mgr_growth = ((data['revenue_2025'] - data['revenue_2024']) / data['revenue_2024'] * 100)
            else:
                mgr_growth = 100 if data['revenue_2025'] > 0 else 0
            manager_analysis.append({
                'name': manager,
                'revenue_2024': data['revenue_2024'],
                'revenue_2025': data['revenue_2025'],
                'growth': round(mgr_growth, 1),
                'count_2025': data['count_2025'],
                'potential': data['revenue_2025'] * (required_growth / 100) if mgr_growth < required_growth else 0
            })

        manager_analysis.sort(key=lambda x: x['revenue_2025'], reverse=True)
        result['analysis']['by_manager'] = manager_analysis[:15]

        # 성장률 낮은 영업담당 (개선 필요) - 제외 대상 빼고
        underperforming_managers = [m for m in manager_analysis if m['growth'] < growth_rate and m['revenue_2024'] > 10000000 and m['name'] not in EXCLUDED_MANAGERS]
        underperforming_managers.sort(key=lambda x: x['growth'])

        # 2. 검사목적별 분석
        by_purpose = {}
        for row in data_2025:
            if not match_filter(row, selected_managers, selected_teams, selected_months, [],
                               selected_regions, selected_sample_types, selected_items, selected_analyzers):
                continue
            purpose = str(row.get('검사목적', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if purpose not in by_purpose:
                by_purpose[purpose] = {'revenue_2025': 0, 'count_2025': 0, 'revenue_2024': 0, 'count_2024': 0}
            by_purpose[purpose]['revenue_2025'] += revenue
            by_purpose[purpose]['count_2025'] += 1

        for row in data_2024:
            if not match_filter(row, selected_managers, selected_teams, selected_months, [],
                               selected_regions, selected_sample_types, selected_items, selected_analyzers):
                continue
            purpose = str(row.get('검사목적', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if purpose not in by_purpose:
                by_purpose[purpose] = {'revenue_2025': 0, 'count_2025': 0, 'revenue_2024': 0, 'count_2024': 0}
            by_purpose[purpose]['revenue_2024'] += revenue
            by_purpose[purpose]['count_2024'] += 1

        purpose_analysis = []
        for purpose, data in by_purpose.items():
            if data['revenue_2024'] > 0:
                purp_growth = ((data['revenue_2025'] - data['revenue_2024']) / data['revenue_2024'] * 100)
            else:
                purp_growth = 100 if data['revenue_2025'] > 0 else 0
            purpose_analysis.append({
                'name': purpose,
                'revenue_2024': data['revenue_2024'],
                'revenue_2025': data['revenue_2025'],
                'growth': round(purp_growth, 1),
                'count_2025': data['count_2025'],
                'share': round(data['revenue_2025'] / revenue_2025 * 100, 1) if revenue_2025 > 0 else 0
            })

        purpose_analysis.sort(key=lambda x: x['revenue_2025'], reverse=True)
        result['analysis']['by_purpose'] = purpose_analysis[:10]

        # 3. 검체유형별 분석
        by_sample_type = {}
        for row in data_2025:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               selected_regions, [], selected_items, selected_analyzers):
                continue
            sample_type = str(row.get('검체유형', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if sample_type not in by_sample_type:
                by_sample_type[sample_type] = {'revenue_2025': 0, 'revenue_2024': 0}
            by_sample_type[sample_type]['revenue_2025'] += revenue

        for row in data_2024:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               selected_regions, [], selected_items, selected_analyzers):
                continue
            sample_type = str(row.get('검체유형', '') or '').strip() or '미지정'
            revenue = get_fee(row)
            if sample_type not in by_sample_type:
                by_sample_type[sample_type] = {'revenue_2025': 0, 'revenue_2024': 0}
            by_sample_type[sample_type]['revenue_2024'] += revenue

        sample_analysis = []
        for st, data in by_sample_type.items():
            if data['revenue_2024'] > 0:
                st_growth = ((data['revenue_2025'] - data['revenue_2024']) / data['revenue_2024'] * 100)
            else:
                st_growth = 100 if data['revenue_2025'] > 0 else 0
            sample_analysis.append({
                'name': st,
                'revenue_2024': data['revenue_2024'],
                'revenue_2025': data['revenue_2025'],
                'growth': round(st_growth, 1)
            })

        sample_analysis.sort(key=lambda x: x['revenue_2025'], reverse=True)
        result['analysis']['by_sample_type'] = sample_analysis[:15]

        # 4. 지역별 분석
        by_region = {}
        for row in data_2025:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               [], selected_sample_types, selected_items, selected_analyzers):
                continue
            address = str(row.get('업체주소', '') or '').strip()
            region = extract_sido(address)
            if not region:
                region = '미지정'
            revenue = get_fee(row)
            if region not in by_region:
                by_region[region] = {'revenue_2025': 0, 'revenue_2024': 0, 'count_2025': 0}
            by_region[region]['revenue_2025'] += revenue
            by_region[region]['count_2025'] += 1

        for row in data_2024:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               [], selected_sample_types, selected_items, selected_analyzers):
                continue
            address = str(row.get('업체주소', '') or '').strip()
            region = extract_sido(address)
            if not region:
                region = '미지정'
            revenue = get_fee(row)
            if region not in by_region:
                by_region[region] = {'revenue_2025': 0, 'revenue_2024': 0, 'count_2025': 0}
            by_region[region]['revenue_2024'] += revenue

        region_analysis = []
        for region, data in by_region.items():
            if data['revenue_2024'] > 0:
                reg_growth = ((data['revenue_2025'] - data['revenue_2024']) / data['revenue_2024'] * 100)
            else:
                reg_growth = 100 if data['revenue_2025'] > 0 else 0
            region_analysis.append({
                'name': region,
                'revenue_2024': data['revenue_2024'],
                'revenue_2025': data['revenue_2025'],
                'growth': round(reg_growth, 1),
                'count_2025': data['count_2025']
            })

        region_analysis.sort(key=lambda x: x['revenue_2025'], reverse=True)
        result['analysis']['by_region'] = region_analysis

        # 5. 항목별 분석 (food_item 데이터)
        by_item = {}
        for row in data_2025:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               selected_regions, selected_sample_types, [], selected_analyzers):
                continue
            item = str(row.get('항목명', '') or '').strip()
            if not item:
                continue
            fee = get_fee(row)
            if item not in by_item:
                by_item[item] = {'fee_2025': 0, 'count_2025': 0, 'fee_2024': 0, 'count_2024': 0}
            by_item[item]['fee_2025'] += fee
            by_item[item]['count_2025'] += 1

        for row in data_2024:
            if not match_filter(row, selected_managers, selected_teams, selected_months, selected_purposes,
                               selected_regions, selected_sample_types, [], selected_analyzers):
                continue
            item = str(row.get('항목명', '') or '').strip()
            if not item:
                continue
            fee = get_fee(row)
            if item not in by_item:
                by_item[item] = {'fee_2025': 0, 'count_2025': 0, 'fee_2024': 0, 'count_2024': 0}
            by_item[item]['fee_2024'] += fee
            by_item[item]['count_2024'] += 1

        item_analysis = []
        for item, data in by_item.items():
            if data['fee_2024'] > 0:
                item_growth = ((data['fee_2025'] - data['fee_2024']) / data['fee_2024'] * 100)
            else:
                item_growth = 100 if data['fee_2025'] > 0 else 0
            item_analysis.append({
                'name': item,
                'fee_2024': data['fee_2024'],
                'fee_2025': data['fee_2025'],
                'growth': round(item_growth, 1),
                'count_2025': data['count_2025']
            })

        item_analysis.sort(key=lambda x: x['fee_2025'], reverse=True)
        result['analysis']['by_item'] = item_analysis[:20]

        # 감소 항목 (위험 요소)
        declining_items = [i for i in item_analysis if i['growth'] < 0 and i['fee_2024'] > 5000000]
        declining_items.sort(key=lambda x: x['growth'])

        # ===== 추천사항 생성 =====
        recommendations = []

        # 1. 전체 목표 분석
        recommendations.append({
            'category': '📊 목표 분석',
            'title': f'{target_year}년 {target_revenue/100000000:.0f}억 달성 가능성',
            'content': f'현재 추세(연 {growth_rate:.1f}% 성장) 유지 시 {target_year}년 예상 매출: {revenue_2025 * (1 + growth_rate/100)/100000000:.1f}억원',
            'action': f'목표 달성을 위해 추가 {gap/100000000:.1f}억원 ({required_growth:.1f}% 성장) 필요',
            'priority': 'high' if required_growth > growth_rate * 1.5 else 'medium'
        })

        # 2. 영업담당 개선
        if underperforming_managers:
            top_under = underperforming_managers[:3]
            potential_gain = sum(m['potential'] for m in top_under)
            recommendations.append({
                'category': '👤 영업담당',
                'title': '성장률 개선 필요 담당자',
                'content': ', '.join([f"{m['name']}({m['growth']:+.1f}%)" for m in top_under]),
                'action': f'이 담당자들이 평균 성장률 달성 시 약 {potential_gain/10000:.0f}만원 추가 가능',
                'evidence': [{'name': m['name'], 'current': m['revenue_2025'], 'growth': m['growth']} for m in top_under],
                'priority': 'high'
            })

        # 3. 고성장 영업담당 (롤모델)
        high_growth_managers = [m for m in manager_analysis if m['growth'] > growth_rate * 1.5 and m['revenue_2025'] > 50000000]
        if high_growth_managers:
            recommendations.append({
                'category': '⭐ 우수 사례',
                'title': '고성장 영업담당 (벤치마킹 대상)',
                'content': ', '.join([f"{m['name']}({m['growth']:+.1f}%)" for m in high_growth_managers[:3]]),
                'action': '이들의 영업 전략 분석 및 공유 권장',
                'priority': 'medium'
            })

        # 4. 검사목적별 기회
        growing_purposes = [p for p in purpose_analysis if p['growth'] > 10 and p['revenue_2025'] > 100000000]
        if growing_purposes:
            recommendations.append({
                'category': '🎯 검사목적',
                'title': '성장 중인 검사목적 (집중 공략)',
                'content': ', '.join([f"{p['name']}({p['growth']:+.1f}%)" for p in growing_purposes[:3]]),
                'action': '이 분야 마케팅 강화 및 전문성 확보',
                'evidence': growing_purposes[:3],
                'priority': 'high'
            })

        # 5. 감소 항목 경고
        if declining_items:
            total_decline = sum(abs(i['fee_2025'] - i['fee_2024']) for i in declining_items[:5])
            recommendations.append({
                'category': '⚠️ 위험 요소',
                'title': '매출 감소 항목',
                'content': ', '.join([f"{i['name']}({i['growth']:.1f}%)" for i in declining_items[:5]]),
                'action': f'감소 원인 분석 필요 (총 감소액: {total_decline/10000:.0f}만원)',
                'evidence': declining_items[:5],
                'priority': 'high'
            })

        # 6. 지역별 기회
        growing_regions = [r for r in region_analysis if r['growth'] > growth_rate and r['revenue_2025'] > 50000000]
        weak_regions = [r for r in region_analysis if r['growth'] < 0 and r['revenue_2024'] > 50000000]

        if growing_regions:
            recommendations.append({
                'category': '📍 지역',
                'title': '성장 지역 (확대 공략)',
                'content': ', '.join([f"{r['name']}({r['growth']:+.1f}%)" for r in growing_regions[:5]]),
                'action': '해당 지역 영업 인력/마케팅 확대 검토',
                'priority': 'medium'
            })

        if weak_regions:
            recommendations.append({
                'category': '📍 지역',
                'title': '감소 지역 (원인 분석 필요)',
                'content': ', '.join([f"{r['name']}({r['growth']:.1f}%)" for r in weak_regions[:5]]),
                'action': '경쟁사 동향 및 고객 이탈 원인 파악',
                'priority': 'medium'
            })

        # 7. 실행 계획 제안
        monthly_target = gap / 12 if gap > 0 else 0
        active_managers = len([m for m in manager_analysis if m['revenue_2025'] > 0])
        per_manager_target = (monthly_target / active_managers / 10000) if active_managers > 0 else 0
        recommendations.append({
            'category': '📋 실행 계획',
            'title': '월별 추가 목표',
            'content': f'목표 달성을 위해 월 평균 {monthly_target/10000:.0f}만원 추가 매출 필요',
            'action': f'영업담당 1인당 월 {per_manager_target:.0f}만원 추가 목표 설정 ({active_managers}명 기준)',
            'priority': 'high'
        })

        result['recommendations'] = recommendations

        # ===== Claude AI 인사이트 생성 =====
        if USE_CLAUDE and CLAUDE_API_KEY:
            try:
                # 분석 데이터 요약
                analysis_summary = f"""
## 사업 성과 분석 데이터

### 기본 현황
- 2024년 매출: {revenue_2024/100000000:.2f}억원
- 2025년 매출: {revenue_2025/100000000:.2f}억원
- 전년 대비 성장률: {growth_rate:+.1f}%
- {target_year}년 목표: {target_revenue/100000000:.0f}억원
- 목표 달성 격차: {gap/100000000:.2f}억원 (추가 {required_growth:.1f}% 성장 필요)

### 영업담당별 현황 (상위 5명)
{chr(10).join([f"- {m['name']}: {m['revenue_2025']/10000:.0f}만원 (성장률 {m['growth']:+.1f}%)" for m in manager_analysis[:5]])}

### 성장률 부진 담당자
{chr(10).join([f"- {m['name']}: 성장률 {m['growth']:+.1f}% (전체 평균 {growth_rate:.1f}% 미달)" for m in underperforming_managers[:3]]) if underperforming_managers else '- 없음'}

### 검사목적별 현황 (상위 5개)
{chr(10).join([f"- {p['name']}: {p['revenue_2025']/10000:.0f}만원 (비중 {p['share']:.1f}%, 성장률 {p['growth']:+.1f}%)" for p in purpose_analysis[:5]])}

### 지역별 현황 (상위 5개)
{chr(10).join([f"- {r['name']}: {r['revenue_2025']/10000:.0f}만원 (성장률 {r['growth']:+.1f}%)" for r in region_analysis[:5]])}

### 매출 감소 항목
{chr(10).join([f"- {i['name']}: {i['growth']:.1f}% 감소" for i in declining_items[:5]]) if declining_items else '- 없음'}
"""

                ai_prompt = f"""당신은 사업 분석 전문가입니다. 아래 데이터를 분석하여 목표 달성을 위한 구체적인 전략적 인사이트를 제공해주세요.

{analysis_summary}

다음 형식으로 분석해주세요:

1. **핵심 진단** (3줄 이내): 현재 상황의 핵심 문제점 또는 기회
2. **우선순위 전략** (3개): 가장 효과적인 매출 증대 전략
3. **위험 요소** (2개): 주의해야 할 리스크
4. **실행 제안** (3개): 구체적인 실행 방안

한국어로 간결하고 실행 가능한 조언을 제공해주세요."""

                ai_result = call_claude_api(ai_prompt, max_tokens=1024)

                if ai_result and 'response' in ai_result:
                    result['ai_insight'] = {
                        'content': ai_result['response'],
                        'model': 'Claude Opus 4',
                        'generated_at': datetime.now().isoformat()
                    }
                    if 'tokens' in ai_result:
                        result['ai_insight']['tokens'] = ai_result['tokens']
            except Exception as ai_error:
                result['ai_insight'] = {
                    'error': str(ai_error),
                    'content': None
                }

        # 필터 옵션 추가 (선택 가능한 값들)
        all_managers = set()
        all_purposes = set()
        all_sample_types = set()
        all_items = set()
        all_analyzers = set()
        all_regions = set()

        for row in data_2025:
            if row.get('영업담당'): all_managers.add(str(row.get('영업담당')).strip())
            if row.get('검사목적'): all_purposes.add(str(row.get('검사목적')).strip())
            if row.get('검체유형'): all_sample_types.add(str(row.get('검체유형')).strip())
            if row.get('항목명'): all_items.add(str(row.get('항목명')).strip())
            if row.get('결과입력자'): all_analyzers.add(str(row.get('결과입력자')).strip())
            address = str(row.get('업체주소', '') or '').strip()
            region = extract_sido(address)
            if region: all_regions.add(region)

        # 팀 목록 생성
        teams = set(MANAGER_TO_BRANCH.values())

        result['filter_options'] = {
            'managers': sorted([m for m in all_managers if m not in EXCLUDED_MANAGERS]),  # ISA, IBK 등 제외
            'teams': sorted(teams),
            'months': list(range(1, 13)),
            'purposes': sorted(all_purposes),
            'regions': sorted(all_regions),
            'sample_types': sorted(all_sample_types),
            'items': sorted(all_items)[:100],  # 상위 100개만
            'analyzers': sorted(all_analyzers)
        }

        # 적용된 필터 정보
        result['applied_filters'] = {
            'managers': selected_managers,
            'teams': selected_teams,
            'months': selected_months,
            'purposes': selected_purposes,
            'regions': selected_regions,
            'sample_types': selected_sample_types,
            'items': selected_items,
            'analyzers': selected_analyzers
        }

        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()})


def extract_sido(address):
    """주소에서 시/도 추출"""
    if not address:
        return None
    sido_patterns = ['서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                    '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
    for pattern in sido_patterns:
        if pattern in address:
            return pattern
    return None


def preload_data():
    """서버 시작 시 데이터 미리 로드 (SQLite 우선)"""
    import time
    start_time = time.time()

    # 1. SQLite 모드인 경우
    if USE_SQLITE:
        print("[PRELOAD] SQLite 모드로 시작...")

        # SQLite DB 업데이트 필요 여부 확인
        if check_sqlite_needs_update():
            print("[PRELOAD] SQLite DB 업데이트 필요 - Excel 변환 시작...")
            convert_excel_to_sqlite()
        else:
            print("[PRELOAD] SQLite DB 최신 상태 유지")

        # SQLite에서 빠르게 로드
        for year in ['2024', '2025']:
            load_excel_data(year)
            load_food_item_data(year)

        # AI 요약 캐시 생성
        get_ai_data_summary(force_refresh=True)

        elapsed = time.time() - start_time
        print(f"[PRELOAD] SQLite 로드 완료! ({elapsed:.1f}초)")
        return

    # 2. 기존 방식: 파일 캐시에서 로드 시도
    if load_cache_from_file():
        elapsed = time.time() - start_time
        print(f"[PRELOAD] 파일 캐시에서 로드 완료! ({elapsed:.1f}초)")
        return

    # 3. 파일 캐시가 없거나 무효 -> Excel에서 로드
    print("[PRELOAD] Excel에서 데이터 로드 시작...")
    for year in ['2024', '2025']:
        load_excel_data(year)
        load_food_item_data(year)

    # 4. AI 요약 캐시도 미리 생성
    get_ai_data_summary(force_refresh=True)

    # 5. 파일로 캐시 저장
    save_cache_to_file()

    elapsed = time.time() - start_time
    print(f"[PRELOAD] 완료! ({elapsed:.1f}초)")


# ========== 웹 터미널 API ==========
@app.route('/api/terminal/auth', methods=['POST'])
def terminal_auth():
    """터미널 인증 API"""
    try:
        password = request.json.get('password', '')

        if password == TERMINAL_PASSWORD:
            # 세션 토큰 생성
            token = secrets.token_hex(32)
            terminal_sessions[token] = {
                'created': datetime.now(),
                'ip': request.remote_addr
            }
            return jsonify({'success': True, 'token': token})
        else:
            return jsonify({'success': False, 'error': '비밀번호가 틀렸습니다'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/terminal/exec', methods=['POST'])
def terminal_exec():
    """터미널 명령어 실행 API"""
    try:
        token = request.json.get('token', '')
        command = request.json.get('command', '')

        # 토큰 검증
        if token not in terminal_sessions:
            return jsonify({'success': False, 'error': '인증이 필요합니다'})

        # 세션 만료 확인 (1시간)
        session = terminal_sessions[token]
        if (datetime.now() - session['created']).seconds > 3600:
            del terminal_sessions[token]
            return jsonify({'success': False, 'error': '세션이 만료되었습니다. 다시 인증해주세요.'})

        if not command.strip():
            return jsonify({'success': False, 'error': '명령어를 입력하세요'})

        # 위험한 명령어 차단
        dangerous_commands = ['rm -rf /', 'mkfs', 'dd if=', ':(){:|:&};:', '> /dev/sda']
        for dangerous in dangerous_commands:
            if dangerous in command:
                return jsonify({'success': False, 'error': f'위험한 명령어가 차단되었습니다: {dangerous}'})

        # 명령어 실행
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=60,  # 60초 타임아웃
            cwd='/home/biofl/business_metrics'  # 작업 디렉토리
        )

        output = result.stdout
        if result.stderr:
            output += '\n[STDERR]\n' + result.stderr

        return jsonify({
            'success': True,
            'output': output if output else '(출력 없음)',
            'returncode': result.returncode
        })

    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': '명령어 실행 시간 초과 (60초)'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    # 서버 시작 시 데이터 미리 로드
    preload_data()
    app.run(host='0.0.0.0', port=6001, debug=False)
