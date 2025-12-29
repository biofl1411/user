"""
경영지표 대시보드 (Flask 버전)
- 오래된 CPU에서도 작동
- Chart.js 사용
- 연도 비교, 검사목적 필터, 업체별 분석, 부적합항목 분석
"""
from flask import Flask, render_template_string, jsonify, request
import os
from pathlib import Path
from datetime import datetime

app = Flask(__name__)

# 경로 설정 - 절대 경로 사용
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

# 데이터 캐시 (메모리에 저장)
DATA_CACHE = {}
CACHE_TIME = {}

# 설정
MANAGER_TO_BRANCH = {
    "장동욱": "충청지사", "지병훈": "충청지사", "박은태": "충청지사",
    "도준구": "경북지사",
    "이강현": "전북지사",
    "엄은정": "경기지사", "정유경": "경기지사",
    "이성복": "서울지사",
    "조봉현": "서울센터", "오세중": "서울센터", "장동주": "서울센터", "오석현": "서울센터",
    "엄상흠": "경북센터",
    "마케팅": "마케팅",
    "본사접수": "본사접수",
}

def load_excel_data(year, use_cache=True):
    """openpyxl로 직접 엑셀 로드 (캐시 사용)"""
    import time
    from openpyxl import load_workbook

    # 캐시 확인 (1시간 유효)
    cache_key = str(year)
    if use_cache and cache_key in DATA_CACHE:
        cache_age = time.time() - CACHE_TIME.get(cache_key, 0)
        if cache_age < 3600:  # 1시간
            print(f"[CACHE] {year}년 데이터 캐시 사용 ({len(DATA_CACHE[cache_key])}건)")
            return DATA_CACHE[cache_key]

    data_path = DATA_DIR / str(year)
    if not data_path.exists():
        return []

    print(f"[LOAD] {year}년 데이터 로딩 시작...")
    start_time = time.time()

    all_data = []
    files = sorted(data_path.glob("*.xlsx"))

    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                all_data.append(row_dict)
            wb.close()
            print(f"[LOAD] {f.name} 완료")
        except Exception as e:
            print(f"[ERROR] Loading {f}: {e}")

    elapsed = time.time() - start_time
    print(f"[LOAD] {year}년 완료: {len(all_data)}건, {elapsed:.1f}초 소요")

    # 캐시 저장
    DATA_CACHE[cache_key] = all_data
    CACHE_TIME[cache_key] = time.time()

    return all_data

def extract_region(address):
    """주소에서 시/도, 시/군/구 추출"""
    if not address:
        return None, None

    addr = str(address).strip()
    if not addr:
        return None, None

    # 시/도 추출
    sido = None
    sigungu = None

    # 광역시/특별시/도 패턴
    sido_patterns = [
        '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
        '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
    ]

    for pattern in sido_patterns:
        if pattern in addr:
            sido = pattern
            break

    # 시/군/구 추출 (첫 번째 시/군/구 단위)
    import re
    # 시, 군, 구 패턴 매칭
    match = re.search(r'([가-힣]+(?:시|군|구))', addr)
    if match:
        sigungu = match.group(1)
        # 시도명이 시군구에 포함되어 있으면 다음 매칭 찾기
        if sido and (sigungu == sido + '시' or sigungu == sido + '도'):
            matches = re.findall(r'([가-힣]+(?:시|군|구))', addr)
            if len(matches) > 1:
                sigungu = matches[1]

    return sido, sigungu

def process_data(data, purpose_filter=None):
    """데이터 처리"""
    by_manager = {}
    by_branch = {}
    by_month = {}
    by_client = {}
    by_purpose = {}
    by_defect = {}
    by_defect_month = {}
    by_region = {}  # 지역별 데이터
    by_region_manager = {}  # 지역-담당자별 데이터
    by_purpose_manager = {}  # 목적별-담당자 데이터
    by_purpose_region = {}  # 목적별-지역 데이터
    purposes = set()
    total_sales = 0
    total_count = 0

    # 주소 컬럼 자동 감지
    address_columns = ['거래처 주소', '채품지주소', '채품장소', '주소', '시료주소', '업체주소', '거래처주소', '검체주소', '시료채취장소']

    for row in data:
        purpose = str(row.get('검사목적', '') or '').strip()
        purposes.add(purpose) if purpose else None

        # 검사목적 필터 적용
        if purpose_filter and purpose_filter != '전체' and purpose != purpose_filter:
            continue

        manager = row.get('영업담당', '미지정')
        sales = row.get('공급가액', 0) or 0
        date = row.get('접수일자')
        client = str(row.get('거래처', '') or '').strip() or '미지정'
        defect = str(row.get('부적합항목', '') or '').strip()

        if isinstance(sales, str):
            sales = float(sales.replace(',', '').replace('원', '')) if sales else 0

        # 매니저별
        if manager not in by_manager:
            by_manager[manager] = {'sales': 0, 'count': 0, 'clients': {}}
        by_manager[manager]['sales'] += sales
        by_manager[manager]['count'] += 1
        if client not in by_manager[manager]['clients']:
            by_manager[manager]['clients'][client] = {'sales': 0, 'count': 0}
        by_manager[manager]['clients'][client]['sales'] += sales
        by_manager[manager]['clients'][client]['count'] += 1

        # 지사별
        branch = MANAGER_TO_BRANCH.get(manager, '기타')
        if branch not in by_branch:
            by_branch[branch] = {'sales': 0, 'count': 0, 'managers': set()}
        by_branch[branch]['sales'] += sales
        by_branch[branch]['count'] += 1
        by_branch[branch]['managers'].add(manager)

        # 월별
        month = 0
        if date:
            if hasattr(date, 'month'):
                month = date.month
            else:
                try:
                    month = int(str(date).split('-')[1])
                except:
                    month = 0

        if month > 0:
            if month not in by_month:
                by_month[month] = {'sales': 0, 'count': 0}
            by_month[month]['sales'] += sales
            by_month[month]['count'] += 1

        # 거래처별
        if client not in by_client:
            by_client[client] = {'sales': 0, 'count': 0, 'purposes': {}}
        by_client[client]['sales'] += sales
        by_client[client]['count'] += 1
        if purpose:
            if purpose not in by_client[client]['purposes']:
                by_client[client]['purposes'][purpose] = {'sales': 0, 'count': 0}
            by_client[client]['purposes'][purpose]['sales'] += sales
            by_client[client]['purposes'][purpose]['count'] += 1

        # 검사목적별
        if purpose:
            if purpose not in by_purpose:
                by_purpose[purpose] = {'sales': 0, 'count': 0}
            by_purpose[purpose]['sales'] += sales
            by_purpose[purpose]['count'] += 1

            # 목적별-담당자 데이터
            if purpose not in by_purpose_manager:
                by_purpose_manager[purpose] = {}
            if manager not in by_purpose_manager[purpose]:
                by_purpose_manager[purpose][manager] = {'sales': 0, 'count': 0}
            by_purpose_manager[purpose][manager]['sales'] += sales
            by_purpose_manager[purpose][manager]['count'] += 1

        # 부적합항목별
        if defect:
            if defect not in by_defect:
                by_defect[defect] = {'count': 0}
            by_defect[defect]['count'] += 1

            # 부적합항목 월별
            if month > 0:
                if defect not in by_defect_month:
                    by_defect_month[defect] = {}
                if month not in by_defect_month[defect]:
                    by_defect_month[defect][month] = 0
                by_defect_month[defect][month] += 1

        # 지역별 분석
        address = None
        for col in address_columns:
            if row.get(col):
                address = row.get(col)
                break

        sido, sigungu = extract_region(address)

        if sido:
            region_key = sido
            if sigungu:
                region_key = f"{sido} {sigungu}"

            # 지역별 통계
            if region_key not in by_region:
                by_region[region_key] = {'sales': 0, 'count': 0, 'sido': sido, 'sigungu': sigungu or '', 'managers': {}}
            by_region[region_key]['sales'] += sales
            by_region[region_key]['count'] += 1

            # 지역-담당자별 통계
            if manager not in by_region[region_key]['managers']:
                by_region[region_key]['managers'][manager] = {'sales': 0, 'count': 0}
            by_region[region_key]['managers'][manager]['sales'] += sales
            by_region[region_key]['managers'][manager]['count'] += 1

            # 담당자-지역별 통계
            if manager not in by_region_manager:
                by_region_manager[manager] = {}
            if region_key not in by_region_manager[manager]:
                by_region_manager[manager][region_key] = {'sales': 0, 'count': 0, 'sido': sido, 'sigungu': sigungu or ''}
            by_region_manager[manager][region_key]['sales'] += sales
            by_region_manager[manager][region_key]['count'] += 1

            # 목적별-지역 데이터
            if purpose:
                if purpose not in by_purpose_region:
                    by_purpose_region[purpose] = {}
                if region_key not in by_purpose_region[purpose]:
                    by_purpose_region[purpose][region_key] = {'sales': 0, 'count': 0}
                by_purpose_region[purpose][region_key]['sales'] += sales
                by_purpose_region[purpose][region_key]['count'] += 1

        total_sales += sales
        total_count += 1

    # 정렬
    sorted_managers = sorted(by_manager.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_branches = sorted(by_branch.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_clients = sorted(by_client.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_purposes = sorted(by_purpose.items(), key=lambda x: x[1]['sales'], reverse=True)
    sorted_defects = sorted(by_defect.items(), key=lambda x: x[1]['count'], reverse=True)

    # 매니저별 TOP 10 거래처
    manager_top_clients = {}
    for mgr, data in sorted_managers:
        clients = sorted(data['clients'].items(), key=lambda x: x[1]['sales'], reverse=True)[:10]
        manager_top_clients[mgr] = clients

    # 고효율 업체 (높은 단가)
    high_efficiency = [(c, d) for c, d in sorted_clients if d['count'] > 0]
    high_efficiency = sorted(high_efficiency, key=lambda x: x[1]['sales'] / x[1]['count'] if x[1]['count'] > 0 else 0, reverse=True)[:20]

    # 대량 업체 (많은 건수)
    high_volume = sorted(by_client.items(), key=lambda x: x[1]['count'], reverse=True)[:20]

    # 지역별 정렬 (매출 기준)
    sorted_regions = sorted(by_region.items(), key=lambda x: x[1]['sales'], reverse=True)

    # 지역별 TOP 담당자
    region_top_managers = {}
    for region, data in sorted_regions:
        managers = sorted(data['managers'].items(), key=lambda x: x[1]['sales'], reverse=True)
        region_top_managers[region] = [
            {'name': m, 'sales': d['sales'], 'count': d['count']}
            for m, d in managers[:5]
        ]

    # 담당자별 지역 분포
    manager_regions = {}
    for mgr, regions in by_region_manager.items():
        sorted_mgr_regions = sorted(regions.items(), key=lambda x: x[1]['sales'], reverse=True)
        manager_regions[mgr] = [
            {'region': r, 'sales': d['sales'], 'count': d['count'], 'sido': d['sido'], 'sigungu': d['sigungu']}
            for r, d in sorted_mgr_regions[:10]
        ]

    # 목적별 담당자 데이터 정리
    purpose_managers = {}
    for purpose, managers in by_purpose_manager.items():
        sorted_pm = sorted(managers.items(), key=lambda x: x[1]['sales'], reverse=True)
        purpose_managers[purpose] = [
            {'name': m, 'sales': d['sales'], 'count': d['count']}
            for m, d in sorted_pm[:20]
        ]

    # 목적별 지역 데이터 정리
    purpose_regions = {}
    for purpose, regions in by_purpose_region.items():
        sorted_pr = sorted(regions.items(), key=lambda x: x[1]['sales'], reverse=True)
        purpose_regions[purpose] = [
            {'region': r, 'sales': d['sales'], 'count': d['count']}
            for r, d in sorted_pr[:20]
        ]

    return {
        'by_manager': [(m, {'sales': d['sales'], 'count': d['count']}) for m, d in sorted_managers],
        'by_branch': [(k, {'sales': v['sales'], 'count': v['count'], 'managers': len(v['managers'])})
                      for k, v in sorted_branches],
        'by_month': sorted(by_month.items()),
        'by_client': [(c, {'sales': d['sales'], 'count': d['count'], 'avg': d['sales']/d['count'] if d['count'] > 0 else 0})
                      for c, d in sorted_clients[:50]],
        'by_purpose': sorted_purposes,
        'by_defect': sorted_defects[:30],
        'by_defect_month': {d: sorted(months.items()) for d, months in by_defect_month.items()},
        'manager_top_clients': manager_top_clients,
        'high_efficiency': [(c, {'sales': d['sales'], 'count': d['count'], 'avg': d['sales']/d['count'] if d['count'] > 0 else 0})
                           for c, d in high_efficiency],
        'high_volume': [(c, {'sales': d['sales'], 'count': d['count'], 'avg': d['sales']/d['count'] if d['count'] > 0 else 0})
                       for c, d in high_volume],
        'by_region': [(r, {'sales': d['sales'], 'count': d['count'], 'sido': d['sido'], 'sigungu': d['sigungu']})
                      for r, d in sorted_regions[:50]],
        'region_top_managers': region_top_managers,
        'manager_regions': manager_regions,
        'purpose_managers': purpose_managers,
        'purpose_regions': purpose_regions,
        'purposes': sorted(list(purposes)),
        'total_sales': total_sales,
        'total_count': total_count
    }

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>경영지표 대시보드</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Malgun Gothic', sans-serif; background: #f5f7fa; padding: 20px; }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; padding: 20px; border-radius: 10px; margin-bottom: 20px;
        }
        .header h1 { font-size: 24px; }
        .controls { display: flex; gap: 10px; margin: 15px 0; flex-wrap: wrap; align-items: center; }
        .controls select { padding: 8px 15px; border-radius: 5px; border: 1px solid #ddd; font-size: 14px; }
        .compare-box {
            display: flex; align-items: center; gap: 8px;
            background: rgba(255,255,255,0.2); padding: 8px 15px; border-radius: 5px;
        }
        .compare-box input[type="checkbox"] { width: 18px; height: 18px; cursor: pointer; }
        .compare-box label { color: white; cursor: pointer; }
        .compare-box select { padding: 5px 10px; }
        .summary { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px; }
        .card { background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .card h3 { color: #666; font-size: 14px; margin-bottom: 10px; }
        .card .value { font-size: 28px; font-weight: bold; color: #333; }
        .card .compare-value { font-size: 14px; color: #764ba2; margin-top: 5px; padding-top: 5px; border-top: 1px dashed #ddd; }
        .card .diff { font-size: 12px; margin-top: 3px; }
        .card .diff.positive { color: #2ecc71; }
        .card .diff.negative { color: #e74c3c; }
        .charts { display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 20px; }
        .chart-container { background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .chart-container h3 { margin-bottom: 15px; color: #333; }
        .chart-container.full { grid-column: 1 / -1; }
        table { width: 100%; border-collapse: collapse; margin-top: 10px; }
        th, td { padding: 8px 10px; text-align: left; border-bottom: 1px solid #eee; font-size: 13px; }
        th { background: #f8f9fa; font-weight: 600; }
        tr:hover { background: #f8f9fa; }
        .tabs { display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; }
        .tab { padding: 10px 20px; background: white; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; }
        .tab.active { background: #667eea; color: white; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        .btn-search {
            padding: 8px 20px; background: #fff; color: #667eea;
            border: 2px solid #fff; border-radius: 5px; font-size: 14px; font-weight: bold; cursor: pointer;
        }
        .btn-search:hover { background: rgba(255,255,255,0.9); }
        .btn-search:disabled { opacity: 0.6; cursor: not-allowed; }
        .toast {
            position: fixed; top: 20px; right: 20px; padding: 15px 25px;
            background: #2ecc71; color: white; border-radius: 8px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2); z-index: 1000; display: none;
        }
        .toast.error { background: #e74c3c; }
        .toast.loading { background: #3498db; }
        .legend-custom { display: flex; gap: 20px; margin-bottom: 10px; font-size: 13px; }
        .legend-item { display: flex; align-items: center; gap: 5px; }
        .legend-color { width: 12px; height: 12px; border-radius: 2px; }
        .sub-select { margin-bottom: 15px; }
        .sub-select select { padding: 8px 15px; border-radius: 5px; border: 1px solid #ddd; }
        .scroll-table { max-height: 400px; overflow-y: auto; }
    </style>
</head>
<body>
    <div id="toast" class="toast"></div>
    <div class="header">
        <h1>📊 경영지표 대시보드</h1>
        <div class="controls">
            <select id="yearSelect">
                <option value="2025">2025년</option>
                <option value="2024">2024년</option>
            </select>
            <div class="compare-box">
                <input type="checkbox" id="compareCheck" onchange="toggleCompare()">
                <label for="compareCheck">비교</label>
                <select id="compareYearSelect" disabled>
                    <option value="2024">2024년</option>
                    <option value="2025">2025년</option>
                </select>
            </div>
            <select id="purposeSelect">
                <option value="전체">검사목적: 전체</option>
            </select>
            <button id="btnSearch" class="btn-search" onclick="loadData()">조회하기</button>
        </div>
    </div>

    <div class="summary" id="summary">
        <div class="card">
            <h3>총 매출</h3>
            <div class="value" id="totalSales">-</div>
            <div class="compare-value" id="compareTotalSales" style="display:none;"></div>
            <div class="diff" id="diffTotalSales"></div>
        </div>
        <div class="card">
            <h3>총 건수</h3>
            <div class="value" id="totalCount">-</div>
            <div class="compare-value" id="compareTotalCount" style="display:none;"></div>
            <div class="diff" id="diffTotalCount"></div>
        </div>
        <div class="card">
            <h3>평균 단가</h3>
            <div class="value" id="avgPrice">-</div>
            <div class="compare-value" id="compareAvgPrice" style="display:none;"></div>
            <div class="diff" id="diffAvgPrice"></div>
        </div>
    </div>

    <div class="tabs">
        <button class="tab active" onclick="showTab('personal')">👤 개인별</button>
        <button class="tab" onclick="showTab('team')">🏢 팀별</button>
        <button class="tab" onclick="showTab('monthly')">📅 월별</button>
        <button class="tab" onclick="showTab('client')">🏭 업체별</button>
        <button class="tab" onclick="showTab('region')">📍 지역별</button>
        <button class="tab" onclick="showTab('purpose')">🎯 목적별</button>
        <button class="tab" onclick="showTab('defect')">⚠️ 부적합</button>
    </div>

    <!-- 개인별 탭 -->
    <div id="personal" class="tab-content active">
        <div class="charts">
            <div class="chart-container">
                <h3>영업담당별 매출 TOP 15</h3>
                <div id="managerLegend" class="legend-custom" style="display:none;"></div>
                <canvas id="managerChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>영업담당별 상세</h3>
                <div class="scroll-table">
                    <table id="managerTable">
                        <thead id="managerTableHead"><tr><th>담당자</th><th>매출액</th><th>건수</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- 팀별 탭 -->
    <div id="team" class="tab-content">
        <div class="charts">
            <div class="chart-container">
                <h3>지사/센터별 매출</h3>
                <div id="branchLegend" class="legend-custom" style="display:none;"></div>
                <canvas id="branchChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>지사/센터별 상세</h3>
                <table id="branchTable">
                    <thead id="branchTableHead"><tr><th>지사/센터</th><th>매출액</th><th>건수</th><th>담당자수</th></tr></thead>
                    <tbody></tbody>
                </table>
            </div>
        </div>
    </div>

    <!-- 월별 탭 -->
    <div id="monthly" class="tab-content">
        <div class="charts">
            <div class="chart-container full">
                <h3>월별 매출 추이</h3>
                <div id="monthlyLegend" class="legend-custom" style="display:none;"></div>
                <div style="height: 300px;"><canvas id="monthlyChart"></canvas></div>
            </div>
        </div>
    </div>

    <!-- 업체별 탭 -->
    <div id="client" class="tab-content">
        <div class="sub-select" style="margin-bottom: 20px; padding: 15px; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 20px; flex-wrap: wrap;">
            <div>
                <span id="clientYearLabel" style="font-weight: bold; color: #667eea; font-size: 16px;">📅 2025년</span>
            </div>
            <div>
                <label style="margin-right: 10px; font-weight: bold;">👤 담당자 필터:</label>
                <select id="clientManagerFilter" onchange="updateClientTables()">
                    <option value="">전체 담당자</option>
                </select>
            </div>
        </div>
        <div class="charts">
            <div class="chart-container">
                <h3>🏆 매출 TOP 20 업체</h3>
                <div class="scroll-table">
                    <table id="clientTopTable">
                        <thead id="clientTopTableHead"><tr><th>순위</th><th>거래처</th><th>매출액</th><th>건수</th><th>평균단가</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container">
                <h3>💎 고효율 업체 (높은 단가)</h3>
                <div class="scroll-table">
                    <table id="clientEffTable">
                        <thead id="clientEffTableHead"><tr><th>거래처</th><th>평균단가</th><th>매출액</th><th>건수</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container">
                <h3>📦 대량 업체 (많은 건수)</h3>
                <div class="scroll-table">
                    <table id="clientVolTable">
                        <thead id="clientVolTableHead"><tr><th>거래처</th><th>건수</th><th>매출액</th><th>평균단가</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- 지역별 탭 -->
    <div id="region" class="tab-content">
        <div class="sub-select" style="margin-bottom: 20px; padding: 15px; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 20px; flex-wrap: wrap;">
            <div>
                <span id="regionYearLabel" style="font-weight: bold; color: #667eea; font-size: 16px;">📅 2025년</span>
            </div>
            <div>
                <label style="margin-right: 10px; font-weight: bold;">👤 담당자 필터:</label>
                <select id="regionManagerFilter" onchange="updateRegionTables()">
                    <option value="">전체 담당자</option>
                </select>
            </div>
        </div>
        <div class="charts">
            <div class="chart-container">
                <h3>📍 지역별 매출 TOP 20</h3>
                <canvas id="regionChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>지역별 상세 (시/도, 시/군/구)</h3>
                <div class="scroll-table">
                    <table id="regionTable">
                        <thead><tr><th>순위</th><th>지역</th><th>매출액</th><th>건수</th><th>평균단가</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container full">
                <h3>🏆 지역별 TOP 담당자</h3>
                <div class="sub-select">
                    <select id="regionSelect" onchange="updateRegionManagers()">
                        <option value="">지역 선택</option>
                    </select>
                </div>
                <div class="scroll-table">
                    <table id="regionManagerTable">
                        <thead><tr><th>순위</th><th>담당자</th><th>매출액</th><th>건수</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container full">
                <h3>👤 담당자별 지역 분포</h3>
                <div class="sub-select">
                    <select id="managerRegionSelect" onchange="updateManagerRegions()">
                        <option value="">담당자 선택</option>
                    </select>
                </div>
                <div class="scroll-table">
                    <table id="managerRegionTable">
                        <thead><tr><th>순위</th><th>지역</th><th>매출액</th><th>건수</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- 목적별 탭 -->
    <div id="purpose" class="tab-content">
        <div class="sub-select" style="margin-bottom: 20px; padding: 15px; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="display: flex; align-items: center; gap: 20px; flex-wrap: wrap; margin-bottom: 15px;">
                <span id="purposeYearLabel" style="font-weight: bold; color: #667eea; font-size: 16px;">📅 2025년</span>
                <button onclick="selectAllPurposes()" style="padding: 5px 10px; background: #667eea; color: white; border: none; border-radius: 5px; cursor: pointer;">전체선택</button>
                <button onclick="clearAllPurposes()" style="padding: 5px 10px; background: #999; color: white; border: none; border-radius: 5px; cursor: pointer;">선택해제</button>
            </div>
            <div id="purposeCheckboxes" style="display: flex; flex-wrap: wrap; gap: 10px; max-height: 100px; overflow-y: auto; padding: 10px; background: #f8f9fa; border-radius: 5px;">
                <!-- 검사목적 체크박스들이 여기에 동적으로 추가됨 -->
            </div>
        </div>
        <div class="charts">
            <div class="chart-container">
                <h3>🎯 선택된 목적별 매출</h3>
                <canvas id="purposeChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>목적별 상세</h3>
                <div class="scroll-table">
                    <table id="purposeTable">
                        <thead><tr><th>검사목적</th><th>매출액</th><th>건수</th><th>평균단가</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container">
                <h3>👤 목적별 담당자 실적</h3>
                <div class="scroll-table">
                    <table id="purposeManagerTable">
                        <thead><tr><th>담당자</th><th>매출액</th><th>건수</th><th>평균단가</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container">
                <h3>📍 목적별 지역 실적</h3>
                <div class="scroll-table">
                    <table id="purposeRegionTable">
                        <thead><tr><th>지역</th><th>매출액</th><th>건수</th><th>평균단가</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- 부적합 탭 -->
    <div id="defect" class="tab-content">
        <div class="charts">
            <div class="chart-container">
                <h3>⚠️ 부적합항목 TOP 20</h3>
                <canvas id="defectChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>부적합항목 상세</h3>
                <div class="scroll-table">
                    <table id="defectTable">
                        <thead><tr><th>순위</th><th>부적합항목</th><th>건수</th><th>비중</th></tr></thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
            <div class="chart-container full">
                <h3>부적합항목 월별 추이</h3>
                <div class="sub-select">
                    <select id="defectSelect" onchange="updateDefectMonthly()">
                        <option value="">항목 선택</option>
                    </select>
                </div>
                <div style="height: 250px;"><canvas id="defectMonthlyChart"></canvas></div>
            </div>
        </div>
    </div>

    <script>
        let charts = {};
        let currentData = null;
        let compareData = null;

        function formatCurrency(value) {
            if (value >= 100000000) return (value/100000000).toFixed(1) + '억';
            if (value >= 10000) return (value/10000).toFixed(0) + '만';
            return value.toLocaleString();
        }

        function formatDiff(current, compare) {
            if (!compare) return '';
            const diff = current - compare;
            const percent = compare > 0 ? ((diff / compare) * 100).toFixed(1) : 0;
            const sign = diff >= 0 ? '+' : '';
            return { diff, percent, sign, text: `${sign}${formatCurrency(Math.abs(diff))} (${sign}${percent}%)` };
        }

        function showToast(message, type = 'success', duration = 3000) {
            const toast = document.getElementById('toast');
            toast.textContent = message;
            toast.className = 'toast ' + type;
            toast.style.display = 'block';
            if (type !== 'loading') setTimeout(() => { toast.style.display = 'none'; }, duration);
        }

        function hideToast() { document.getElementById('toast').style.display = 'none'; }

        function toggleCompare() {
            document.getElementById('compareYearSelect').disabled = !document.getElementById('compareCheck').checked;
        }

        function showTab(tabId) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
            document.querySelector(`[onclick="showTab('${tabId}')"]`).classList.add('active');
            document.getElementById(tabId).classList.add('active');
        }

        async function loadData() {
            const year = document.getElementById('yearSelect').value;
            const compareEnabled = document.getElementById('compareCheck').checked;
            const compareYear = document.getElementById('compareYearSelect').value;
            const purpose = document.getElementById('purposeSelect').value;
            const btn = document.getElementById('btnSearch');

            btn.disabled = true;
            btn.textContent = '로딩중...';
            showToast('데이터를 불러오는 중입니다...', 'loading');

            try {
                const response = await fetch(`/api/data?year=${year}&purpose=${encodeURIComponent(purpose)}`);
                currentData = await response.json();
                currentData.year = year;

                // 검사목적 드롭다운 업데이트
                updatePurposeSelect(currentData.purposes);

                if (compareEnabled && compareYear !== year) {
                    const compareResponse = await fetch(`/api/data?year=${compareYear}&purpose=${encodeURIComponent(purpose)}`);
                    compareData = await compareResponse.json();
                    compareData.year = compareYear;
                } else {
                    compareData = null;
                }

                updateAll();

                let msg = `${year}년 데이터 로드 완료 (${currentData.total_count.toLocaleString()}건)`;
                if (compareData) msg = `${year}년 vs ${compareYear}년 비교 로드 완료`;
                showToast(msg, 'success');

            } catch (error) {
                console.error('Error:', error);
                showToast('데이터 로드 중 오류가 발생했습니다.', 'error');
            } finally {
                btn.disabled = false;
                btn.textContent = '조회하기';
            }
        }

        function updatePurposeSelect(purposes) {
            const select = document.getElementById('purposeSelect');
            const currentValue = select.value;
            select.innerHTML = '<option value="전체">검사목적: 전체</option>';
            purposes.forEach(p => {
                if (p) select.innerHTML += `<option value="${p}">${p}</option>`;
            });
            if (purposes.includes(currentValue)) select.value = currentValue;
        }

        function updateAll() {
            const steps = [
                ['updateSummary', updateSummary],
                ['updateManagerChart', updateManagerChart],
                ['updateBranchChart', updateBranchChart],
                ['updateMonthlyChart', updateMonthlyChart],
                ['updateManagerTable', updateManagerTable],
                ['updateBranchTable', updateBranchTable],
                ['updateClientTables', updateClientTables],
                ['updateRegionTables', updateRegionTables],
                ['updateRegionSelects', updateRegionSelects],
                ['updatePurposeCheckboxes', updatePurposeCheckboxes],
                ['updatePurposeTab', updatePurposeTab],
                ['updateDefectChart', updateDefectChart],
                ['updateDefectTable', updateDefectTable],
                ['updateDefectSelect', updateDefectSelect]
            ];

            for (const [name, fn] of steps) {
                try {
                    console.log(`[UPDATE] ${name} 시작...`);
                    fn();
                    console.log(`[UPDATE] ${name} 완료 ✓`);
                } catch (e) {
                    console.error(`[UPDATE ERROR] ${name} 실패:`, e);
                    throw e;
                }
            }
            console.log('[UPDATE] 모든 업데이트 완료');
        }

        function updateSummary() {
            document.getElementById('totalSales').textContent = formatCurrency(currentData.total_sales);
            document.getElementById('totalCount').textContent = currentData.total_count.toLocaleString() + '건';
            const avgPrice = currentData.total_count > 0 ? currentData.total_sales / currentData.total_count : 0;
            document.getElementById('avgPrice').textContent = formatCurrency(avgPrice);

            if (compareData) {
                const compAvg = compareData.total_count > 0 ? compareData.total_sales / compareData.total_count : 0;
                document.getElementById('compareTotalSales').textContent = `${compareData.year}년: ${formatCurrency(compareData.total_sales)}`;
                document.getElementById('compareTotalSales').style.display = 'block';
                const salesDiff = formatDiff(currentData.total_sales, compareData.total_sales);
                document.getElementById('diffTotalSales').textContent = salesDiff.text;
                document.getElementById('diffTotalSales').className = 'diff ' + (salesDiff.diff >= 0 ? 'positive' : 'negative');

                document.getElementById('compareTotalCount').textContent = `${compareData.year}년: ${compareData.total_count.toLocaleString()}건`;
                document.getElementById('compareTotalCount').style.display = 'block';
                const countDiff = formatDiff(currentData.total_count, compareData.total_count);
                document.getElementById('diffTotalCount').textContent = countDiff.text;
                document.getElementById('diffTotalCount').className = 'diff ' + (countDiff.diff >= 0 ? 'positive' : 'negative');

                document.getElementById('compareAvgPrice').textContent = `${compareData.year}년: ${formatCurrency(compAvg)}`;
                document.getElementById('compareAvgPrice').style.display = 'block';
                const avgDiff = formatDiff(avgPrice, compAvg);
                document.getElementById('diffAvgPrice').textContent = avgDiff.text;
                document.getElementById('diffAvgPrice').className = 'diff ' + (avgDiff.diff >= 0 ? 'positive' : 'negative');
            } else {
                ['compareTotalSales', 'compareTotalCount', 'compareAvgPrice'].forEach(id => {
                    document.getElementById(id).style.display = 'none';
                });
                ['diffTotalSales', 'diffTotalCount', 'diffAvgPrice'].forEach(id => {
                    document.getElementById(id).textContent = '';
                });
            }
        }

        function updateManagerChart() {
            const top15 = currentData.by_manager.slice(0, 15);
            const ctx = document.getElementById('managerChart').getContext('2d');
            if (charts.manager) charts.manager.destroy();

            const datasets = [{ label: currentData.year + '년', data: top15.map(d => d[1].sales), backgroundColor: 'rgba(102, 126, 234, 0.8)' }];

            if (compareData) {
                const compareMap = Object.fromEntries(compareData.by_manager);
                datasets.push({ label: compareData.year + '년', data: top15.map(d => compareMap[d[0]]?.sales || 0), backgroundColor: 'rgba(118, 75, 162, 0.6)' });
                document.getElementById('managerLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background:rgba(102,126,234,0.8)"></div>${currentData.year}년</div><div class="legend-item"><div class="legend-color" style="background:rgba(118,75,162,0.6)"></div>${compareData.year}년</div>`;
                document.getElementById('managerLegend').style.display = 'flex';
            } else {
                document.getElementById('managerLegend').style.display = 'none';
            }

            charts.manager = new Chart(ctx, {
                type: 'bar',
                data: { labels: top15.map(d => d[0]), datasets },
                options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
            });

            // 업체별 탭 담당자 필터 드롭다운 업데이트
            const clientManagerFilter = document.getElementById('clientManagerFilter');
            const currentFilter = clientManagerFilter.value;
            clientManagerFilter.innerHTML = '<option value="">전체 담당자</option>';
            currentData.by_manager.forEach(m => {
                clientManagerFilter.innerHTML += `<option value="${m[0]}">${m[0]}</option>`;
            });
            if (currentFilter) clientManagerFilter.value = currentFilter;
        }

        function updateBranchChart() {
            const ctx = document.getElementById('branchChart').getContext('2d');
            if (charts.branch) charts.branch.destroy();

            if (compareData) {
                const labels = currentData.by_branch.map(d => d[0]);
                const compareMap = Object.fromEntries(compareData.by_branch);
                document.getElementById('branchLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background:rgba(102,126,234,0.8)"></div>${currentData.year}년</div><div class="legend-item"><div class="legend-color" style="background:rgba(118,75,162,0.6)"></div>${compareData.year}년</div>`;
                document.getElementById('branchLegend').style.display = 'flex';
                charts.branch = new Chart(ctx, {
                    type: 'bar',
                    data: { labels, datasets: [
                        { label: currentData.year + '년', data: currentData.by_branch.map(d => d[1].sales), backgroundColor: 'rgba(102, 126, 234, 0.8)' },
                        { label: compareData.year + '년', data: labels.map(l => compareMap[l]?.sales || 0), backgroundColor: 'rgba(118, 75, 162, 0.6)' }
                    ]},
                    options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
                });
            } else {
                document.getElementById('branchLegend').style.display = 'none';
                charts.branch = new Chart(ctx, {
                    type: 'pie',
                    data: { labels: currentData.by_branch.map(d => d[0]), datasets: [{ data: currentData.by_branch.map(d => d[1].sales), backgroundColor: ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe', '#43e97b', '#fa709a', '#fee140'] }] },
                    options: { responsive: true, plugins: { legend: { position: 'right' } } }
                });
            }
        }

        function updateMonthlyChart() {
            const ctx = document.getElementById('monthlyChart').getContext('2d');
            if (charts.monthly) charts.monthly.destroy();

            const labels = []; for (let i = 1; i <= 12; i++) labels.push(i + '월');
            const currentMap = Object.fromEntries(currentData.by_month);
            const datasets = [{ label: currentData.year + '년', data: labels.map((_, i) => currentMap[i+1]?.sales || 0), borderColor: '#667eea', backgroundColor: 'rgba(102, 126, 234, 0.1)', fill: true, tension: 0.4 }];

            if (compareData) {
                const compareMap = Object.fromEntries(compareData.by_month);
                datasets.push({ label: compareData.year + '년', data: labels.map((_, i) => compareMap[i+1]?.sales || 0), borderColor: '#764ba2', backgroundColor: 'rgba(118, 75, 162, 0.1)', fill: true, tension: 0.4 });
                document.getElementById('monthlyLegend').innerHTML = `<div class="legend-item"><div class="legend-color" style="background:#667eea"></div>${currentData.year}년</div><div class="legend-item"><div class="legend-color" style="background:#764ba2"></div>${compareData.year}년</div>`;
                document.getElementById('monthlyLegend').style.display = 'flex';
            } else {
                document.getElementById('monthlyLegend').style.display = 'none';
            }

            charts.monthly = new Chart(ctx, {
                type: 'line', data: { labels, datasets },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { callback: v => formatCurrency(v) } } } }
            });
        }

        function updateManagerTable() {
            const thead = document.getElementById('managerTableHead');
            const tbody = document.querySelector('#managerTable tbody');

            if (compareData) {
                thead.innerHTML = `<tr><th>담당자</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th><th>비중</th></tr>`;
                const compareMap = Object.fromEntries(compareData.by_manager);
                tbody.innerHTML = currentData.by_manager.map(d => {
                    const compSales = compareMap[d[0]]?.sales || 0;
                    const diff = d[1].sales - compSales;
                    return `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(compSales)}</td><td class="${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${formatCurrency(diff)}</td><td>${(d[1].sales / currentData.total_sales * 100).toFixed(1)}%</td></tr>`;
                }).join('');
            } else {
                thead.innerHTML = `<tr><th>담당자</th><th>매출액</th><th>건수</th><th>비중</th></tr>`;
                tbody.innerHTML = currentData.by_manager.map(d => `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${d[1].count}</td><td>${(d[1].sales / currentData.total_sales * 100).toFixed(1)}%</td></tr>`).join('');
            }
        }

        function updateBranchTable() {
            const thead = document.getElementById('branchTableHead');
            const tbody = document.querySelector('#branchTable tbody');

            if (compareData) {
                thead.innerHTML = `<tr><th>지사/센터</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th></tr>`;
                const compareMap = Object.fromEntries(compareData.by_branch);
                tbody.innerHTML = currentData.by_branch.map(d => {
                    const compSales = compareMap[d[0]]?.sales || 0;
                    const diff = d[1].sales - compSales;
                    return `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(compSales)}</td><td class="${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${formatCurrency(diff)}</td></tr>`;
                }).join('');
            } else {
                thead.innerHTML = `<tr><th>지사/센터</th><th>매출액</th><th>건수</th><th>담당자수</th></tr>`;
                tbody.innerHTML = currentData.by_branch.map(d => `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${d[1].count}</td><td>${d[1].managers}명</td></tr>`).join('');
            }
        }

        function updateClientTables() {
            const selectedManager = document.getElementById('clientManagerFilter').value;

            // 연도 라벨 업데이트
            const yearLabel = document.getElementById('clientYearLabel');
            if (compareData) {
                yearLabel.textContent = `📅 ${currentData.year}년 vs ${compareData.year}년`;
            } else {
                yearLabel.textContent = `📅 ${currentData.year}년`;
            }

            let clientData, effData, volData;
            let compareClientMap = {};

            // 비교 데이터 맵 생성
            if (compareData) {
                compareData.by_client.forEach(c => {
                    compareClientMap[c[0]] = c[1];
                });
            }

            if (selectedManager && currentData.manager_top_clients[selectedManager]) {
                // 담당자별 데이터 사용
                const managerClients = currentData.manager_top_clients[selectedManager];

                // 매출순 정렬
                clientData = managerClients.map(c => [c[0], {
                    sales: c[1].sales,
                    count: c[1].count,
                    avg: c[1].count > 0 ? c[1].sales / c[1].count : 0
                }]);

                // 고효율 (단가순)
                effData = [...clientData].sort((a, b) => b[1].avg - a[1].avg).slice(0, 20);

                // 대량 (건수순)
                volData = [...clientData].sort((a, b) => b[1].count - a[1].count).slice(0, 20);

                clientData = clientData.slice(0, 20);
            } else {
                // 전체 데이터 사용
                clientData = currentData.by_client.slice(0, 20);
                effData = currentData.high_efficiency;
                volData = currentData.high_volume;
            }

            // TOP 20 업체 (비교 모드 지원)
            const topThead = document.getElementById('clientTopTableHead');
            const topTbody = document.querySelector('#clientTopTable tbody');

            if (compareData) {
                topThead.innerHTML = `<tr><th>순위</th><th>거래처</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th><th>건수</th></tr>`;
                topTbody.innerHTML = clientData.map((d, i) => {
                    const compSales = compareClientMap[d[0]]?.sales || 0;
                    const diff = d[1].sales - compSales;
                    return `<tr><td>${i+1}</td><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(compSales)}</td><td class="${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${formatCurrency(diff)}</td><td>${d[1].count}</td></tr>`;
                }).join('') || '<tr><td colspan="6">데이터 없음</td></tr>';
            } else {
                topThead.innerHTML = `<tr><th>순위</th><th>거래처</th><th>매출액</th><th>건수</th><th>평균단가</th></tr>`;
                topTbody.innerHTML = clientData.map((d, i) =>
                    `<tr><td>${i+1}</td><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${d[1].count}</td><td>${formatCurrency(d[1].avg)}</td></tr>`
                ).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
            }

            // 고효율 업체 (비교 모드 지원)
            const effThead = document.getElementById('clientEffTableHead');
            const effTbody = document.querySelector('#clientEffTable tbody');

            if (compareData) {
                effThead.innerHTML = `<tr><th>거래처</th><th>평균단가</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th></tr>`;
                effTbody.innerHTML = effData.map(d => {
                    const compSales = compareClientMap[d[0]]?.sales || 0;
                    const diff = d[1].sales - compSales;
                    return `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].avg)}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(compSales)}</td><td class="${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${formatCurrency(diff)}</td></tr>`;
                }).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
            } else {
                effThead.innerHTML = `<tr><th>거래처</th><th>평균단가</th><th>매출액</th><th>건수</th></tr>`;
                effTbody.innerHTML = effData.map(d =>
                    `<tr><td>${d[0]}</td><td>${formatCurrency(d[1].avg)}</td><td>${formatCurrency(d[1].sales)}</td><td>${d[1].count}</td></tr>`
                ).join('') || '<tr><td colspan="4">데이터 없음</td></tr>';
            }

            // 대량 업체 (비교 모드 지원)
            const volThead = document.getElementById('clientVolTableHead');
            const volTbody = document.querySelector('#clientVolTable tbody');

            if (compareData) {
                volThead.innerHTML = `<tr><th>거래처</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th><th>매출액</th></tr>`;
                volTbody.innerHTML = volData.map(d => {
                    const compCount = compareClientMap[d[0]]?.count || 0;
                    const diff = d[1].count - compCount;
                    return `<tr><td>${d[0]}</td><td>${d[1].count}</td><td>${compCount}</td><td class="${diff >= 0 ? 'positive' : 'negative'}">${diff >= 0 ? '+' : ''}${diff}</td><td>${formatCurrency(d[1].sales)}</td></tr>`;
                }).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
            } else {
                volThead.innerHTML = `<tr><th>거래처</th><th>건수</th><th>매출액</th><th>평균단가</th></tr>`;
                volTbody.innerHTML = volData.map(d =>
                    `<tr><td>${d[0]}</td><td>${d[1].count}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(d[1].avg)}</td></tr>`
                ).join('') || '<tr><td colspan="4">데이터 없음</td></tr>';
            }
        }

        function updateDefectChart() {
            const ctx = document.getElementById('defectChart').getContext('2d');
            if (charts.defect) charts.defect.destroy();

            const top20 = currentData.by_defect.slice(0, 20);
            charts.defect = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: top20.map(d => d[0].length > 15 ? d[0].substring(0, 15) + '...' : d[0]),
                    datasets: [{ label: '건수', data: top20.map(d => d[1].count), backgroundColor: 'rgba(231, 76, 60, 0.7)' }]
                },
                options: { indexAxis: 'y', responsive: true, plugins: { legend: { display: false } } }
            });
        }

        function updateDefectTable() {
            const tbody = document.querySelector('#defectTable tbody');
            const totalDefects = currentData.by_defect.reduce((sum, d) => sum + d[1].count, 0);
            tbody.innerHTML = currentData.by_defect.map((d, i) =>
                `<tr><td>${i+1}</td><td>${d[0]}</td><td>${d[1].count}</td><td>${(d[1].count / totalDefects * 100).toFixed(1)}%</td></tr>`
            ).join('');
        }

        function updateDefectSelect() {
            const select = document.getElementById('defectSelect');
            select.innerHTML = '<option value="">항목 선택</option>';
            currentData.by_defect.slice(0, 20).forEach(d => {
                select.innerHTML += `<option value="${d[0]}">${d[0]}</option>`;
            });
        }

        function updateDefectMonthly() {
            const defect = document.getElementById('defectSelect').value;
            const ctx = document.getElementById('defectMonthlyChart').getContext('2d');
            if (charts.defectMonthly) charts.defectMonthly.destroy();

            if (!defect || !currentData.by_defect_month[defect]) {
                return;
            }

            const labels = []; for (let i = 1; i <= 12; i++) labels.push(i + '월');
            const monthData = Object.fromEntries(currentData.by_defect_month[defect] || []);
            const values = labels.map((_, i) => monthData[i+1] || 0);

            charts.defectMonthly = new Chart(ctx, {
                type: 'line',
                data: {
                    labels,
                    datasets: [{ label: defect, data: values, borderColor: '#e74c3c', backgroundColor: 'rgba(231, 76, 60, 0.1)', fill: true, tension: 0.4 }]
                },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: true } } }
            });
        }

        // 지역별 함수들
        function updateRegionChart() {
            if (!currentData.by_region || currentData.by_region.length === 0) {
                // 지역 데이터가 없으면 안내 메시지 표시
                const ctx = document.getElementById('regionChart').getContext('2d');
                if (charts.region) charts.region.destroy();
                ctx.font = '14px Malgun Gothic';
                ctx.fillStyle = '#999';
                ctx.textAlign = 'center';
                ctx.fillText('지역 데이터가 없습니다. (주소 컬럼 확인 필요)', ctx.canvas.width / 2, ctx.canvas.height / 2);
                return;
            }

            const ctx = document.getElementById('regionChart').getContext('2d');
            if (charts.region) charts.region.destroy();

            const top20 = currentData.by_region.slice(0, 20);
            charts.region = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: top20.map(d => d[0].length > 12 ? d[0].substring(0, 12) + '...' : d[0]),
                    datasets: [{ label: '매출', data: top20.map(d => d[1].sales), backgroundColor: 'rgba(52, 152, 219, 0.7)' }]
                },
                options: { indexAxis: 'y', responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { callback: v => formatCurrency(v) } } } }
            });
        }

        function updateRegionTables() {
            if (!currentData.by_region) return;

            // 연도 라벨 업데이트
            const yearLabel = document.getElementById('regionYearLabel');
            if (compareData) {
                yearLabel.textContent = `📅 ${currentData.year}년 vs ${compareData.year}년`;
            } else {
                yearLabel.textContent = `📅 ${currentData.year}년`;
            }

            // 담당자 필터 확인
            const selectedManager = document.getElementById('regionManagerFilter').value;
            let regionData = currentData.by_region;
            let compareRegionData = compareData ? compareData.by_region : null;

            // 담당자가 선택된 경우 해당 담당자의 지역 데이터만 표시
            if (selectedManager && currentData.manager_regions && currentData.manager_regions[selectedManager]) {
                const managerRegions = currentData.manager_regions[selectedManager];
                regionData = managerRegions.map(r => [r.region, {sales: r.sales, count: r.count}]);
                // 비교 데이터도 담당자 필터 적용
                if (compareData && compareData.manager_regions && compareData.manager_regions[selectedManager]) {
                    const compareManagerRegions = compareData.manager_regions[selectedManager];
                    compareRegionData = compareManagerRegions.map(r => [r.region, {sales: r.sales, count: r.count}]);
                } else {
                    compareRegionData = null;
                }
            }

            const thead = document.querySelector('#regionTable thead');
            const tbody = document.querySelector('#regionTable tbody');

            // 비교 모드일 때 테이블 헤더 및 데이터 변경
            if (compareData && compareRegionData) {
                thead.innerHTML = `<tr><th>순위</th><th>지역</th><th>${currentData.year}년</th><th>${compareData.year}년</th><th>증감</th><th>건수</th></tr>`;
                const compareMap = Object.fromEntries(compareRegionData);

                tbody.innerHTML = regionData.map((d, i) => {
                    const compData = compareMap[d[0]] || {sales: 0, count: 0};
                    const diff = formatDiff(d[1].sales, compData.sales);
                    const diffClass = diff.diff >= 0 ? 'positive' : 'negative';
                    const diffText = diff.text ? `<span class="${diffClass}">${diff.text}</span>` : '-';
                    return `<tr><td>${i+1}</td><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${formatCurrency(compData.sales)}</td><td>${diffText}</td><td>${d[1].count}</td></tr>`;
                }).join('') || '<tr><td colspan="6">지역 데이터 없음</td></tr>';
            } else {
                thead.innerHTML = `<tr><th>순위</th><th>지역</th><th>매출액</th><th>건수</th><th>평균단가</th></tr>`;
                tbody.innerHTML = regionData.map((d, i) => {
                    const avg = d[1].count > 0 ? d[1].sales / d[1].count : 0;
                    return `<tr><td>${i+1}</td><td>${d[0]}</td><td>${formatCurrency(d[1].sales)}</td><td>${d[1].count}</td><td>${formatCurrency(avg)}</td></tr>`;
                }).join('') || '<tr><td colspan="5">지역 데이터 없음</td></tr>';
            }

            // 차트도 업데이트
            updateRegionChart(regionData, compareRegionData);
        }

        function updateRegionChart(regionData, compareRegionData) {
            const top20 = regionData.slice(0, 20);
            if (regionChart) {
                regionChart.data.labels = top20.map(d => d[0]);

                if (compareData && compareRegionData) {
                    const compareMap = Object.fromEntries(compareRegionData);
                    regionChart.data.datasets = [
                        { label: currentData.year + '년', data: top20.map(d => d[1].sales), backgroundColor: 'rgba(102, 126, 234, 0.8)' },
                        { label: compareData.year + '년', data: top20.map(d => (compareMap[d[0]]?.sales || 0)), backgroundColor: 'rgba(118, 75, 162, 0.6)' }
                    ];
                    regionChart.options.plugins.legend = { display: true };
                } else {
                    regionChart.data.datasets = [
                        { label: '매출액', data: top20.map(d => d[1].sales), backgroundColor: 'rgba(102, 126, 234, 0.8)' }
                    ];
                    regionChart.options.plugins.legend = { display: false };
                }
                regionChart.update();
            }
        }

        function updateRegionSelects() {
            if (!currentData.by_region) return;

            // 지역 선택 드롭다운
            const regionSelect = document.getElementById('regionSelect');
            regionSelect.innerHTML = '<option value="">지역 선택</option>';
            currentData.by_region.forEach(d => {
                regionSelect.innerHTML += `<option value="${d[0]}">${d[0]}</option>`;
            });

            // 담당자 선택 드롭다운 (담당자별 지역 분포용)
            const managerRegionSelect = document.getElementById('managerRegionSelect');
            managerRegionSelect.innerHTML = '<option value="">담당자 선택</option>';
            currentData.by_manager.forEach(m => {
                managerRegionSelect.innerHTML += `<option value="${m[0]}">${m[0]}</option>`;
            });

            // 지역별 탭 담당자 필터
            const regionManagerFilter = document.getElementById('regionManagerFilter');
            const currentFilter = regionManagerFilter.value;
            regionManagerFilter.innerHTML = '<option value="">전체 담당자</option>';
            currentData.by_manager.forEach(m => {
                regionManagerFilter.innerHTML += `<option value="${m[0]}">${m[0]}</option>`;
            });
            if (currentFilter) regionManagerFilter.value = currentFilter;
        }

        function updateRegionManagers() {
            const region = document.getElementById('regionSelect').value;
            const tbody = document.querySelector('#regionManagerTable tbody');

            if (!region || !currentData.region_top_managers || !currentData.region_top_managers[region]) {
                tbody.innerHTML = '<tr><td colspan="5">지역을 선택해주세요</td></tr>';
                return;
            }

            const managers = currentData.region_top_managers[region];
            const totalSales = managers.reduce((sum, m) => sum + m.sales, 0);

            tbody.innerHTML = managers.map((m, i) =>
                `<tr><td>${i+1}</td><td>${m.name}</td><td>${formatCurrency(m.sales)}</td><td>${m.count}</td><td>${(m.sales / totalSales * 100).toFixed(1)}%</td></tr>`
            ).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
        }

        function updateManagerRegions() {
            const manager = document.getElementById('managerRegionSelect').value;
            const tbody = document.querySelector('#managerRegionTable tbody');

            if (!manager || !currentData.manager_regions || !currentData.manager_regions[manager]) {
                tbody.innerHTML = '<tr><td colspan="5">담당자를 선택해주세요</td></tr>';
                return;
            }

            const regions = currentData.manager_regions[manager];
            const totalSales = regions.reduce((sum, r) => sum + r.sales, 0);

            tbody.innerHTML = regions.map((r, i) =>
                `<tr><td>${i+1}</td><td>${r.region}</td><td>${formatCurrency(r.sales)}</td><td>${r.count}</td><td>${(r.sales / totalSales * 100).toFixed(1)}%</td></tr>`
            ).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
        }

        // 목적별 탭 함수들
        function updatePurposeCheckboxes() {
            const container = document.getElementById('purposeCheckboxes');
            container.innerHTML = '';

            if (!currentData.purposes) return;

            currentData.purposes.forEach(p => {
                if (!p) return;
                const label = document.createElement('label');
                label.style.cssText = 'display: flex; align-items: center; gap: 5px; background: white; padding: 5px 10px; border-radius: 5px; cursor: pointer; border: 1px solid #ddd;';
                label.innerHTML = `<input type="checkbox" value="${p}" onchange="updatePurposeTab()" checked> ${p}`;
                container.appendChild(label);
            });
        }

        function selectAllPurposes() {
            document.querySelectorAll('#purposeCheckboxes input[type="checkbox"]').forEach(cb => cb.checked = true);
            updatePurposeTab();
        }

        function clearAllPurposes() {
            document.querySelectorAll('#purposeCheckboxes input[type="checkbox"]').forEach(cb => cb.checked = false);
            updatePurposeTab();
        }

        function getSelectedPurposes() {
            const checkboxes = document.querySelectorAll('#purposeCheckboxes input[type="checkbox"]:checked');
            return Array.from(checkboxes).map(cb => cb.value);
        }

        function updatePurposeTab() {
            // 연도 라벨 업데이트
            const yearLabel = document.getElementById('purposeYearLabel');
            yearLabel.textContent = `📅 ${currentData.year}년`;

            const selectedPurposes = getSelectedPurposes();

            if (selectedPurposes.length === 0) {
                document.querySelector('#purposeTable tbody').innerHTML = '<tr><td colspan="5">검사목적을 선택해주세요</td></tr>';
                document.querySelector('#purposeManagerTable tbody').innerHTML = '<tr><td colspan="5">검사목적을 선택해주세요</td></tr>';
                document.querySelector('#purposeRegionTable tbody').innerHTML = '<tr><td colspan="5">검사목적을 선택해주세요</td></tr>';
                if (charts.purpose) charts.purpose.destroy();
                return;
            }

            // 선택된 목적별 데이터 필터링
            const filteredPurposes = currentData.by_purpose.filter(p => selectedPurposes.includes(p[0]));
            const totalSales = filteredPurposes.reduce((sum, p) => sum + p[1].sales, 0);

            // 목적별 차트
            const ctx = document.getElementById('purposeChart').getContext('2d');
            if (charts.purpose) charts.purpose.destroy();
            charts.purpose = new Chart(ctx, {
                type: 'pie',
                data: {
                    labels: filteredPurposes.map(p => p[0]),
                    datasets: [{
                        data: filteredPurposes.map(p => p[1].sales),
                        backgroundColor: ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe', '#43e97b', '#fa709a', '#fee140', '#a8edea', '#fed6e3']
                    }]
                },
                options: { responsive: true, plugins: { legend: { position: 'right' } } }
            });

            // 목적별 테이블
            document.querySelector('#purposeTable tbody').innerHTML = filteredPurposes.map(p => {
                const avg = p[1].count > 0 ? p[1].sales / p[1].count : 0;
                const ratio = totalSales > 0 ? (p[1].sales / totalSales * 100).toFixed(1) : 0;
                return `<tr><td>${p[0]}</td><td>${formatCurrency(p[1].sales)}</td><td>${p[1].count}</td><td>${formatCurrency(avg)}</td><td>${ratio}%</td></tr>`;
            }).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';

            // 목적별 담당자 테이블 (선택된 목적에 해당하는 담당자 데이터)
            if (currentData.purpose_managers) {
                const managerData = {};
                selectedPurposes.forEach(purpose => {
                    if (currentData.purpose_managers[purpose]) {
                        currentData.purpose_managers[purpose].forEach(m => {
                            if (!managerData[m.name]) {
                                managerData[m.name] = { sales: 0, count: 0 };
                            }
                            managerData[m.name].sales += m.sales;
                            managerData[m.name].count += m.count;
                        });
                    }
                });
                const sortedManagers = Object.entries(managerData).sort((a, b) => b[1].sales - a[1].sales);
                const managerTotalSales = sortedManagers.reduce((sum, m) => sum + m[1].sales, 0);

                document.querySelector('#purposeManagerTable tbody').innerHTML = sortedManagers.slice(0, 20).map(([name, data]) => {
                    const avg = data.count > 0 ? data.sales / data.count : 0;
                    const ratio = managerTotalSales > 0 ? (data.sales / managerTotalSales * 100).toFixed(1) : 0;
                    return `<tr><td>${name}</td><td>${formatCurrency(data.sales)}</td><td>${data.count}</td><td>${formatCurrency(avg)}</td><td>${ratio}%</td></tr>`;
                }).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
            } else {
                document.querySelector('#purposeManagerTable tbody').innerHTML = '<tr><td colspan="5">담당자 데이터 없음</td></tr>';
            }

            // 목적별 지역 테이블
            if (currentData.purpose_regions) {
                const regionData = {};
                selectedPurposes.forEach(purpose => {
                    if (currentData.purpose_regions[purpose]) {
                        currentData.purpose_regions[purpose].forEach(r => {
                            if (!regionData[r.region]) {
                                regionData[r.region] = { sales: 0, count: 0 };
                            }
                            regionData[r.region].sales += r.sales;
                            regionData[r.region].count += r.count;
                        });
                    }
                });
                const sortedRegions = Object.entries(regionData).sort((a, b) => b[1].sales - a[1].sales);
                const regionTotalSales = sortedRegions.reduce((sum, r) => sum + r[1].sales, 0);

                document.querySelector('#purposeRegionTable tbody').innerHTML = sortedRegions.slice(0, 20).map(([region, data]) => {
                    const avg = data.count > 0 ? data.sales / data.count : 0;
                    const ratio = regionTotalSales > 0 ? (data.sales / regionTotalSales * 100).toFixed(1) : 0;
                    return `<tr><td>${region}</td><td>${formatCurrency(data.sales)}</td><td>${data.count}</td><td>${formatCurrency(avg)}</td><td>${ratio}%</td></tr>`;
                }).join('') || '<tr><td colspan="5">데이터 없음</td></tr>';
            } else {
                document.querySelector('#purposeRegionTable tbody').innerHTML = '<tr><td colspan="5">지역 데이터 없음</td></tr>';
            }
        }

        showToast('연도를 선택하고 [조회하기] 버튼을 클릭하세요.', 'loading', 5000);
        setTimeout(() => hideToast(), 5000);
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/data')
def get_data():
    year = request.args.get('year', '2025')
    purpose = request.args.get('purpose', '전체')
    print(f"[API] 요청: year={year}, purpose={purpose}")
    data = load_excel_data(year)
    print(f"[API] 로드된 데이터: {len(data)}건")
    processed = process_data(data, purpose)
    print(f"[API] 처리 완료: total_count={processed['total_count']}")
    return jsonify(processed)

@app.route('/api/columns')
def get_columns():
    """Excel 파일의 컬럼명 조회"""
    year = request.args.get('year', '2025')
    from openpyxl import load_workbook

    data_path = DATA_DIR / str(year)
    if not data_path.exists():
        return jsonify({'error': f'{year}년 데이터 폴더가 없습니다.', 'columns': []})

    files = sorted(data_path.glob("*.xlsx"))
    if not files:
        return jsonify({'error': f'{year}년 데이터 파일이 없습니다.', 'columns': []})

    try:
        wb = load_workbook(files[0], read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        wb.close()

        # 주소 관련 컬럼 표시
        address_cols = [h for h in headers if h and any(k in str(h) for k in ['주소', '지역', '시', '도', '군', '구', '동', '장소'])]

        return jsonify({
            'year': year,
            'file': files[0].name,
            'total_columns': len(headers),
            'columns': headers,
            'address_columns': address_cols
        })
    except Exception as e:
        return jsonify({'error': str(e), 'columns': []})

@app.route('/api/cache/refresh')
def refresh_cache():
    """캐시 새로고침"""
    global DATA_CACHE, CACHE_TIME
    DATA_CACHE = {}
    CACHE_TIME = {}
    print("[CACHE] 캐시 초기화됨")
    # 데이터 미리 로드
    for year in ['2024', '2025']:
        load_excel_data(year, use_cache=False)
    return jsonify({'status': 'ok', 'message': '캐시가 새로고침되었습니다.'})

def preload_data():
    """서버 시작 시 데이터 미리 로드"""
    print("[PRELOAD] 데이터 미리 로드 시작...")
    for year in ['2024', '2025']:
        load_excel_data(year)
    print("[PRELOAD] 완료!")

if __name__ == '__main__':
    # 서버 시작 시 데이터 미리 로드
    preload_data()
    app.run(host='0.0.0.0', port=6001, debug=False)
